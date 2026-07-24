#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════╗
║     RENTAL BOOKING & INVENTORY MANAGEMENT SYSTEM v3.0       ║
╠══════════════════════════════════════════════════════════════╣
║  • Accept/Deny workflow with Stripe payment integration      ║
║  • Invoice + contract emailed automatically on Accept        ║
║  • Stripe webhook auto-confirms booking on deposit payment   ║
║  • Real-time inventory — first-come-first-PAID               ║
║  • Admin panel with full booking management                  ║
╚══════════════════════════════════════════════════════════════╝
"""

import os, re, json, logging, smtplib, secrets, decimal, io, hmac, hashlib
import urllib.parse
from datetime import datetime, timezone, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders as _email_encoders
from functools import wraps
from difflib import SequenceMatcher

import requests
import psycopg2
import psycopg2.extras
import stripe
try:
    from webauthn import (
        generate_registration_options, verify_registration_response,
        generate_authentication_options, verify_authentication_response,
        options_to_json, base64url_to_bytes,
    )
    from webauthn import parse_registration_credential_json, parse_authentication_credential_json
    from webauthn.helpers.structs import (
        AuthenticatorSelectionCriteria, UserVerificationRequirement,
        ResidentKeyRequirement, PublicKeyCredentialDescriptor,
    )
    WEBAUTHN_AVAILABLE = True
except ImportError:
    WEBAUTHN_AVAILABLE = False
from flask import (Flask, request, render_template_string,
                   redirect, url_for, jsonify, session, Response, send_file)
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=90)  # stay logged in 90 days


# ══════════════════════════════════════════════════════════════════════════════
#  RENTAL CATALOG  —  edit prices/quantities here anytime
# ══════════════════════════════════════════════════════════════════════════════

PRODUCTS = [
    {"id": "chairs",         "name": "White Folding Plastic Chairs",     "price": 2.75,  "total": 200},
    {"id": "tables_6ft",     "name": "6ft White Folding Plastic Tables", "price": 8.00,  "total": 30},
    {"id": "banquet_tables", "name": "8x30 Wood Banquet Tables",         "price": 15.00, "total": 10},
    {"id": "round_tables",   "name": "60in Wood Round Tables",           "price": 15.00, "total": 10},
    {"id": "cocktail_30",    "name": "30in Cocktail Tables",             "price": 15.00, "total": 10},
    {"id": "cocktail_cloth", "name": "Cocktail Table Cloths",            "price": 8.00,  "total": 10},
]

EXACT_TIME_FEE  = 175.00
DEPOSIT_PERCENT = 0.25   # 25% deposit required
CT_TAX_RATE     = 0.0635 # Connecticut sales tax 6.35%


# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  (set in Render.com → Environment)
# ══════════════════════════════════════════════════════════════════════════════

def _float(key, default):
    try:
        return float(os.getenv(key, "") or default)
    except ValueError:
        return float(default)

# ── PWA Icons (base64-encoded PNGs) ──────────────────────────────────────────
import base64 as _b64
_ICON_192_B64 = "iVBORw0KGgoAAAANSUhEUgAAAMAAAADACAIAAADdvvtQAAAE8klEQVR4nO3dP05UURiG8YGQkBBjovb0bIFVuAO2QOEarCiIhbWtnYWdKzCxdyVi7CxOcnMzEOWe95zv331+vc7wncfvzDAGTi6vbw5Ar1PvJ4DcCAgSAoKEgCAhIEgICBICgoSAICEgSAgIEgKChIAgISBICAiSM+8nEML51V3Hn/rz893wZ5LOHgPqy+U5f88Ok9pFQKOK2fpAe+ipckBm3fz3CRQuqVpA7tE8af2sisVUJ6CY6TzWnmeZjNIHlKWbI2Vut8QBJU3nSPaFlDKgGums5c0oWUD10lnLmFGagGqns5Yroxyfhe2nnkWWLzn6BsoyxxlSrKK4Ae05nbXgGQW9wqjnSNiBRAwo7LB8xRxLrCss5oziCHidBdpA1PNMoQYVJaBQQ4kvzrj8r7A4s8glyHXmvIGoR+Q+QM+A3L/4GnzH6BYQ9QzkOEyfgKhnOK+ROgREPZO4DNY6IOqZyn68pgFRjwHjIdsFRD1mLEdtFBD1GDMbeJSPMpCURUCsHxc2Y58eEPU4Mhj+3ICox93sI+A1ECQTA2L9BDH1IGYFRD2hzDuOKQFRT0CTDoXXQJCMD4j1E9aMo2EDQTI4INZPcMMPaGRA1JPC2GPiCoNkWECsn0QGHhYbCJIxAbF+0hl1ZGwgSAYExPpJasjBsYEgUQNi/aSmHx8bCBICgkT6AVMx76+XFyc/PrzxfhbH3n/+9enbb+9n8YTzqzvlp1SxgSDpDyjm+kEH5SjZQJAQECQEBEnnu7BcL4Duvzx8/Ppg+YivXpx+v39t+Yii7vdibCBICAiSnoBy3V94pr5jZQNBQkCQEBAkBATJ5oB4BV1Yx+GygSAhIEgICBICgoSAINkWEG/Bytt6xGwgSAgIEgKChIAgISBICAgSAoKEgCAhIEgICBICgoSAICEgSAgIEgKChIAgISBItgWk/DhPpLD1iNlAkBAQJAQECQFBQkCQbA6IN2KFdRwuGwgSAoKEgCAhIEh6AuJ1dEn8qgM4ICBIOgPiFium+0DZQJAQECQEBEl/QLwMKoPfGw83UkAsoQLEQ2QDQUJAkKgBcYulph8fGwiSAQGxhJIacnBsIEjGBMQSSmfUkbGBIBkWEEsokYGHxQaCZGRALKEUxh7T4A1EQ8ENPyCuMEjGB8QSCmvG0bCBIJkSEEsooEmHMmsD0VAo845j4hVGQ0FMPQheA0EyNyCWkLvZRzB9A9GQI4PhW1xhNOTCZuxnBo/h7vbtxe3bC+9nUZPRi2iWkDGzgdu9C6MhM5ajNn0bT0MGjIds/X0gGprKfrwO30ikoUlcBntyeX1j/6iHw+H86s7lcavy+mfp9lEGe2ggx2F6fhZGQ0P4jtH5w1QaErkP0P870W0EvCTayj2dJsp/5wgyjizijCtKQIdIQwku1KD8r7A1rrN/C5VOE2gDLQKOKYKYY4kY0CHqsByFHUisK2yN66wJm04TN6BmzxkFT6cJeoUdSTHKsbJ8ydE30GI/qyhLOk2agJraGeVKp0kWUFMvo4zpNCkDampklDedJnFAzXIAuUrK3s0ifUCLLAupTDpNnYCa9fHEialYNGvVAlpzv90Kd7OoHNDi6CDn9bSHYo7sIqAjj4+5L6kd5vLYHgN6jBS65fgsDGERECQEBAkBQUJAkBAQJAQECQFBQkCQEBAkBAQJAUFCQJAQECR/AWuBXRYBgcQbAAAAAElFTkSuQmCC"
_ICON_512_B64 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAIAAAB7GkOtAAAOd0lEQVR4nO3dMY5sVxWF4bJlZCECGIFT5AkQOCYnJUGehIkIEDEeBhNgGIzAImYCEFoQELT13O7XXV1169579t7r+wYADXXO+t+pfoZPvvjq6wsAeT5d/QMAsIYAAIQSAIBQAgAQSgAAQgkAQCgBAAglAAChBAAglAAAhBIAgFACABBKAABCCQBAKAEACCUAAKEEACCUAACEEgCAUAIAEEoAAEIJAEAoAQAIJQAAoQQAIJQAAIQSAIBQAgAQSgAAQgkAQCgBAAglAAChBAAglAAAhBIAgFACABBKAABCCQBAKAEACCUAAKEEACCUAACEEgCAUAIAEEoAAEIJAECoZQH4/rtvVv1bA9SxcAy9AABCCQBAKAEACCUAAKEEACCUAACEEgCAUAIAEEoAAEKtDIB/GBgIt3YGvQAAQgkAQCgBAAglAAChBAAglAAAhBIAgFCLA+AfBQBiLR9ALwCAUAIAEEoAAEIJAEAoAQAItT4Ay38PDnC+CtO3PgAALCEAAKEEACCUAACEKhGACr8MAThNkdErEQAAzicAAKEEACBUlQAU+UYM4Gh15q5KAAA4mQAAhBIAgFACABCqUADq/GIE4CClhq5QAAA4kwAAhKoVgFKPI4B9VZu4WgEA4DQCABBKAABClQtAte/IAHZRcNzKBQCAcwgAQKiKASj4UAJ4RM1ZqxgAAE4gAAChigag5nMJYIOyg1Y0AAAcTQAAQtUNQNlHE8DtKk9Z3QAAcCgBAAhVOgCVn04A7yo+YqUDAMBxBAAgVPUAFH9AAbyl/nxVDwAABxEAgFANAlD/GQXwQovhahAAAI7QIwAtWgrwpMtk9QgAALsIAMDhageg7EMNJO84Sf0CAMAuWgagY2mBwZqOUssAAPC4rgFo2ltgnr5z1DUAADyocQD6VhcYo/UQNQ4AAI/oHYDW7QW66z5BvQMAwGbtA9C9wEBTA8anfQAA2GZCAAZ0GOhlxuxMCAAAGwwJwIwaAy2MGZwhAbgM+kiAyiZNzZwAAHCXUQGYVGagoGEjMyoAANxuWgCG9RmoY968TAsAADcaGIB5lQaWGzksAwNwGfpRAatMnZSZAQDgXWMDMLXYwMkGj8nYAFxGf2zAOWbPyOQAAHDF8ADMrjdwqPEDMjwAl4CPEDhCwnTMDwAAr4oIQELJgR2FjEZEAC4xHyfwuJy5SAkAAC8EBSCn6sBmUUMRFIBL2EcL3CttIrICAMAHcQFIKzxwo8BxiAvAJfJjBq7LnIXEAFxSP2zgVbGDEBoAAHIDENt84LnkKcgNwCX7gwcu8SMQHYBL/McPyVz/9AAAxBIAfwqARC7+RQCeOAoQxZV/IgA/cCAghMv+gQD8yLGA8Vzz5wQAIJQA/IQ/HcBgLvgLAvCSIwIjudof+2z1D1DR99998/mX367+Keb4+59/tfpHIJ31f5UXwOscFxjDdX6LALzJoYEBXOQrBOAaRwdac4WvE4B3OEDQlMv7LgEACCUA7/PnCGjHtb2FANzEYYJGXNgbCcCtHClowVW9nQDcwcGC4lzSuwjAfRwvKMv1vJcA3M0hg4JczA0EYAtHDUpxJbcRgI0cOCjCZdxMALZz7GA51/ARn3zx1derf4be/A9HwyrW/0FeAI9yBGEJV+9xArADBxFO5tLtwv8j2D6ejqOvg+Bopn9HXgB7cjThUK7YvgRgZw4oHMTl2p0A7M8xhd25VkcQgEM4rLAjF+ogAnAURxZ24SodRwAO5ODCg1yiQ/lroMfy10NhG9N/Ai+AMzjKcBdX5hwCcBIHGm7kspxGAM7jWMO7XJMz+R3AqfxKAN5i+s/nBbCAgw4vuBRLCMAajjt84Dqs4iugZXwdBKZ/LS+AxVwAYjn8ywnAeq4BgRz7CnwFVIKvg8hh+uvwAijExWA8h7wUL4BaPAWYyvQX5AVQkavCMI50TV4ARXkKMIPpr8wLoDSXh9Yc4OK8AKrzFKAj09+CF0APrhONOK5deAG04SlAfaa/Fy+AZlwwynI42/EC6MdTgGpMf1MC0JUMUIHpb00AepMBVjH9A/gdwASuIidz5GbwAhjCU4BzmP5JBGAUGeA4pn8eARhIBtiX6Z9KAMaSAR5n+mcTgOFkgG1MfwIBiCAD3M705xCAIDLAdaY/jQDE+XDJlYAndj+WAOTyIMD0hxOAdDKQyfRzEQCe+F4ohN3nOQHgJzwIpjL9fEwAeIUHwRh2nysEgGuUoCm7zy0EgJsoQQt2n7sIAPdRgoLsPtsIABspwXJ2nwcJAI9SgpPZffYiAOzm+TCJwb6MPkcQAA7hWbALu8+hBIBjvZgwPbjO4nMmAeBUvib6mNFnFQFgmdjHgcWnCAGgio9ncUYSzD1lCQB1vTWdNcNg6GlHAOjn+tQelwcTzzACwDRmGm706eofAIA1BAAglAAAhBIAgFACABBKAABCCQBAKAEACCUAAKEEACCUAACEEgCAUAIAEEoAAEIJAEAoAQAIJQAAoQQAIJQAAIQSAIBQAgAQSgAAQgkAQCgBAAglAAChBAAglAAAhBIAgFACABBKAABCCQBAKAEACCUAAKEEACCUAACEEgCAUAIAEEoAAEIJAECo/wO4sn30VF69pwAAAABJRU5ErkJggg=="
_ICON_192 = _b64.b64decode(_ICON_192_B64)
_ICON_512 = _b64.b64decode(_ICON_512_B64)

def _log_payment(booking_id, amount, method="stripe", note=None, recorded_by="system"):
    """Record a payment event in payment_logs."""
    conn = get_db()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO payment_logs (booking_id, amount, method, note, recorded_by) VALUES (%s,%s,%s,%s,%s)",
            (booking_id, round(float(amount), 2), method, note, recorded_by)
        )
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        log.error(f"_log_payment error: {e}")

def _row(row):
    """Convert a psycopg2 DictRow to a plain dict, converting Decimal→float."""
    if row is None:
        return None
    return {k: float(v) if isinstance(v, decimal.Decimal) else v for k, v in dict(row).items()}

def _fmt_date(d):
    """Format a date as MM/DD/YYYY. Accepts datetime.date, datetime, or YYYY-MM-DD string."""
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%m/%d/%Y')
    s = str(d)[:10]
    if len(s) == 10 and s[4] == '-':
        return f"{s[5:7]}/{s[8:10]}/{s[0:4]}"
    return s

BUSINESS_NAME    = os.getenv("BUSINESS_NAME",    "Rent a Party, LLC")
DEPOT_ADDRESS    = os.getenv("DEPOT_ADDRESS",    "799 New Haven Rd, Naugatuck, CT 06770")
CALENDAR_TOKEN   = os.getenv("CALENDAR_TOKEN",   secrets.token_hex(16))
BUSINESS_PHONE   = os.getenv("BUSINESS_PHONE",   "")
BUSINESS_EMAIL   = os.getenv("BUSINESS_EMAIL",   "")
BUSINESS_ADDRESS = os.getenv("BUSINESS_ADDRESS", "")

DELIVERY_THRESHOLD = _float("DELIVERY_THRESHOLD", "15")
DELIVERY_BASE_FEE  = _float("DELIVERY_BASE_FEE",  "55")
DELIVERY_RATE      = _float("DELIVERY_RATE",       "3.80")

GOOGLE_MAPS_KEY    = os.getenv("GOOGLE_MAPS_KEY",    "")
DATABASE_URL       = os.getenv("DATABASE_URL",       "")
OWNER_EMAIL        = os.getenv("OWNER_EMAIL",        "")
GMAIL_USER         = os.getenv("GMAIL_USER",         "")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
ADMIN_PASSWORD     = os.getenv("ADMIN_PASSWORD",     "admin123")

STRIPE_SECRET_KEY     = os.getenv("STRIPE_SECRET_KEY",     "")
STRIPE_WEBHOOK_SECRET = os.getenv("STRIPE_WEBHOOK_SECRET", "")
BASE_URL              = os.getenv("BASE_URL", "").rstrip("/")
CRON_SECRET           = os.getenv("CRON_SECRET", "")
WEBAUTHN_RP_ID        = os.getenv("WEBAUTHN_RP_ID", "")
TWILIO_ACCOUNT_SID    = os.getenv("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN     = os.getenv("TWILIO_AUTH_TOKEN", "")
TWILIO_FROM_NUMBER    = os.getenv("TWILIO_FROM_NUMBER", "")
CRON_SECRET_KEY       = os.getenv("CRON_SECRET_KEY", CRON_SECRET)

if STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY


# ══════════════════════════════════════════════════════════════════════════════
#  DATABASE
# ══════════════════════════════════════════════════════════════════════════════

def get_db():
    if not DATABASE_URL:
        log.warning("DATABASE_URL not set")
        return None
    try:
        return psycopg2.connect(DATABASE_URL, sslmode="require")
    except Exception as e:
        log.error(f"DB connection failed: {e}")
        return None


def init_db():
    """Create tables and run column migrations on startup."""
    conn = get_db()
    if not conn:
        return
    try:
        cur = conn.cursor()

        # Create bookings table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS bookings (
                id               SERIAL PRIMARY KEY,
                created_at       TIMESTAMPTZ DEFAULT NOW(),
                status           VARCHAR(20) DEFAULT 'pending',

                full_name        VARCHAR(255),
                company_name     VARCHAR(255),
                renter_street    VARCHAR(255),
                renter_city      VARCHAR(100),
                renter_state     VARCHAR(50),
                renter_zip       VARCHAR(20),
                phone            VARCHAR(50),
                email            VARCHAR(255),

                event_start_date DATE,
                event_end_date   DATE,
                event_start_time VARCHAR(20),
                event_end_time   VARCHAR(20),
                setup_time       VARCHAR(20),
                venue_type       VARCHAR(20),
                venue_latest_pickup VARCHAR(20),

                event_street     VARCHAR(255),
                event_city       VARCHAR(100),
                event_state      VARCHAR(50),
                event_zip        VARCHAR(20),

                exact_time_delivery BOOLEAN DEFAULT FALSE,
                delivery_location   TEXT,
                delivery_fee        DECIMAL(10,2) DEFAULT 0,
                distance_miles      DECIMAL(6,1),

                items_json       TEXT,
                items_subtotal   DECIMAL(10,2) DEFAULT 0,
                exact_time_fee   DECIMAL(10,2) DEFAULT 0,
                grand_total      DECIMAL(10,2) DEFAULT 0,
                notes            TEXT,

                stripe_payment_link TEXT,
                stripe_session_id   TEXT
            )
        """)

        # Payment links tracker table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS payment_links (
                id             SERIAL PRIMARY KEY,
                booking_id     INTEGER REFERENCES bookings(id) ON DELETE CASCADE,
                label          TEXT,
                amount         DECIMAL(10,2),
                url            TEXT,
                stripe_link_id TEXT,
                status         VARCHAR(20) DEFAULT 'active',
                created_at     TIMESTAMPTZ DEFAULT NOW()
            )
        """)

        # Migrations: add new columns to existing tables (safe to run every time)
        migrations = [
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS stripe_payment_link TEXT",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS stripe_session_id TEXT",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS final_payment_link TEXT",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS final_reminder_sent BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS auto_invoice_sent BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS setup_date DATE DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS archived BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS tax_rate DECIMAL(5,4) DEFAULT 0",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS tax_amount DECIMAL(10,2) DEFAULT 0",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS tax_exempt BOOLEAN DEFAULT FALSE",
            "ALTER TABLE customers ADD COLUMN IF NOT EXISTS tax_exempt BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS delivery_status TEXT DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS late_night_fee DECIMAL(10,2) DEFAULT 0",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS amount_paid DECIMAL(10,2) DEFAULT 0",
            # Clean up literal 'None' strings left by Booqable import
            "UPDATE bookings SET phone         = NULL WHERE phone         = 'None'",
            "UPDATE bookings SET renter_street = NULL WHERE renter_street = 'None'",
            "UPDATE bookings SET renter_city   = NULL WHERE renter_city   = 'None'",
            "UPDATE bookings SET renter_state  = NULL WHERE renter_state  = 'None'",
            "UPDATE bookings SET renter_zip    = NULL WHERE renter_zip    = 'None'",
            "UPDATE bookings SET company_name  = NULL WHERE company_name  = 'None'",
            # Set all Marquee Letter items to $85 (name = "Marquee A" through "Marquee Z")
            "UPDATE inventory SET price = 85 WHERE TRIM(name) SIMILAR TO 'Marquee [A-Za-z]'",
            # Admin-only private notes
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS admin_notes TEXT",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS view_token TEXT DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS discount_type VARCHAR(10) DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS discount_value DECIMAL(10,2) DEFAULT 0",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS discount_amount DECIMAL(10,2) DEFAULT 0",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS delivered_at TIMESTAMPTZ DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS picked_up_at TIMESTAMPTZ DEFAULT NULL",
            "ALTER TABLE customers ADD COLUMN IF NOT EXISTS street2 VARCHAR(255) DEFAULT NULL",
            # Tax transfer tracking
            """CREATE TABLE IF NOT EXISTS tax_transfers (
                id           SERIAL PRIMARY KEY,
                created_at   TIMESTAMPTZ DEFAULT NOW(),
                amount       DECIMAL(10,2) NOT NULL,
                note         TEXT,
                period_label VARCHAR(100)
            )""",
            # New status/payment system
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS payment_status VARCHAR(20) DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS payment_method VARCHAR(30) DEFAULT NULL",
            """CREATE TABLE IF NOT EXISTS payment_logs (
                id          SERIAL PRIMARY KEY,
                booking_id  INTEGER NOT NULL,
                paid_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                amount      DECIMAL(10,2) NOT NULL DEFAULT 0,
                method      VARCHAR(50) NOT NULL DEFAULT 'stripe',
                note        TEXT,
                recorded_by VARCHAR(50) DEFAULT 'system'
            )""",
            "CREATE INDEX IF NOT EXISTS payment_logs_booking_idx ON payment_logs(booking_id)",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS route_override BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS delivery_date DATE DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS delivery_time VARCHAR(10) DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS pickup_date DATE DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS pickup_time VARCHAR(10) DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS reminder_sent_at TIMESTAMP DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS confirmation_sent_at TIMESTAMP DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS pickup_reminder_sent_at TIMESTAMP DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS agreement_signed BOOLEAN DEFAULT FALSE",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS agreement_signed_at TIMESTAMP DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS agreement_signer_name VARCHAR(200) DEFAULT NULL",
            "ALTER TABLE bookings ADD COLUMN IF NOT EXISTS delivered_at TIMESTAMP DEFAULT NULL",
            # Migrate: confirmed → accepted + paid
            "UPDATE bookings SET status='accepted', payment_status='paid' WHERE status='confirmed' AND amount_paid IS NOT NULL AND amount_paid > 0 AND amount_paid >= grand_total - 0.50",
            "UPDATE bookings SET status='accepted', payment_status='waiting' WHERE status='confirmed' AND (amount_paid IS NULL OR amount_paid < 0.50)",
            "UPDATE bookings SET status='accepted', payment_status='paid'     WHERE status='paid'",
            # Migrate: partial → accepted + partial
            "UPDATE bookings SET status='accepted', payment_status='partial'  WHERE status='partial'",
            # Migrate accepted bookings: set payment_status based on amount_paid
            "UPDATE bookings SET payment_status='paid'    WHERE status='accepted' AND payment_status IS NULL AND amount_paid >= grand_total - 0.50 AND grand_total > 0",
            "UPDATE bookings SET payment_status='partial' WHERE status='accepted' AND payment_status IS NULL AND amount_paid > 0 AND amount_paid < grand_total - 0.50",
            "UPDATE bookings SET payment_status='waiting' WHERE status='accepted' AND payment_status IS NULL",
            # Fix bookings already migrated to 'waiting' that actually have amount_paid
            "UPDATE bookings SET payment_status='paid'    WHERE status='accepted' AND payment_status='waiting' AND amount_paid >= grand_total - 0.50 AND grand_total > 0 AND amount_paid > 0",
            "UPDATE bookings SET payment_status='partial' WHERE status='accepted' AND payment_status='waiting' AND amount_paid > 0 AND amount_paid < grand_total - 0.50 AND grand_total > 0",
            # Auto-conclude: picked up 2+ days ago
            """UPDATE bookings SET status='concluded'
               WHERE delivery_status='picked_up'
                 AND status NOT IN ('concluded','cancelled','denied')
                 AND (
                   (picked_up_at IS NOT NULL AND picked_up_at <= NOW() - INTERVAL '2 days')
                   OR (picked_up_at IS NULL AND event_end_date IS NOT NULL AND event_end_date <= CURRENT_DATE - 2)
                 )""",
        ]
        for m in migrations:
            try:
                cur.execute("SAVEPOINT mig")
                cur.execute(m)
                cur.execute("RELEASE SAVEPOINT mig")
            except Exception as me:
                cur.execute("ROLLBACK TO SAVEPOINT mig")
                log.warning(f"Migration warning: {me}")

        # Create customers table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id           SERIAL PRIMARY KEY,
                created_at   TIMESTAMPTZ DEFAULT NOW(),
                full_name    VARCHAR(255) NOT NULL,
                company_name VARCHAR(255),
                email        VARCHAR(255),
                phone        VARCHAR(50),
                street       VARCHAR(255),
                city         VARCHAR(100),
                state        VARCHAR(50),
                zip          VARCHAR(20),
                notes        TEXT
            )
        """)
        # Unique index on email so ON CONFLICT works for upserts
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS customers_email_idx
            ON customers (email) WHERE email IS NOT NULL
        """)

        # Import all existing booking customers into customers table (safe to run every time)
        try:
            cur.execute("SAVEPOINT cust_import")
            cur.execute("""
                INSERT INTO customers (full_name, company_name, email, phone, street, city, state, zip)
                SELECT DISTINCT ON (email)
                    full_name, company_name, email, phone,
                    renter_street, renter_city, renter_state, renter_zip
                FROM bookings
                WHERE email IS NOT NULL
                ORDER BY email, created_at DESC
                ON CONFLICT (email) WHERE email IS NOT NULL DO UPDATE SET
                    full_name    = EXCLUDED.full_name,
                    company_name = EXCLUDED.company_name,
                    phone        = EXCLUDED.phone,
                    street       = EXCLUDED.street,
                    city         = EXCLUDED.city,
                    state        = EXCLUDED.state,
                    zip          = EXCLUDED.zip
            """)
            cur.execute("RELEASE SAVEPOINT cust_import")
            log.info("Existing booking customers synced to customers table")
        except Exception as ie:
            cur.execute("ROLLBACK TO SAVEPOINT cust_import")
            log.warning(f"Customer import warning: {ie}")

        # Create inventory table
        cur.execute("""
            CREATE TABLE IF NOT EXISTS webauthn_credentials (
                id            SERIAL PRIMARY KEY,
                credential_id TEXT UNIQUE NOT NULL,
                public_key    BYTEA NOT NULL,
                sign_count    INT NOT NULL DEFAULT 0,
                created_at    TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS inventory (
                id         VARCHAR(100) PRIMARY KEY,
                name       VARCHAR(255) NOT NULL,
                price      DECIMAL(10,2) NOT NULL DEFAULT 0,
                total      INT NOT NULL DEFAULT 0,
                sort_order INT NOT NULL DEFAULT 0
            )
        """)
        # Seed from PRODUCTS constant if inventory table is empty
        cur.execute("SELECT COUNT(*) FROM inventory")
        if cur.fetchone()[0] == 0:
            for i, p in enumerate(PRODUCTS):
                cur.execute(
                    "INSERT INTO inventory (id, name, price, total, sort_order) VALUES (%s, %s, %s, %s, %s)",
                    (p["id"], p["name"], p["price"], p["total"], i)
                )
            log.info("Inventory seeded from PRODUCTS default list")

        conn.commit()
        cur.close()
        conn.close()
        log.info("Database ready")
    except Exception as e:
        log.error(f"DB init error: {e}")


# ── Run on startup ─────────────────────────────────────────────────────────
with app.app_context():
    init_db()


# ══════════════════════════════════════════════════════════════════════════════
#  PRODUCT CATALOG — load from DB (falls back to hardcoded PRODUCTS)
# ══════════════════════════════════════════════════════════════════════════════

def get_products():
    """Load product catalog from inventory table; fall back to PRODUCTS constant."""
    conn = get_db()
    if not conn:
        return PRODUCTS
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM inventory ORDER BY sort_order, name")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if rows:
            products = []
            for r in rows:
                p = dict(r)
                p["price"] = float(p["price"])
                p["total"] = int(p["total"])
                products.append(p)
            return products
    except Exception as e:
        log.error(f"get_products error: {e}")
    return PRODUCTS


# ══════════════════════════════════════════════════════════════════════════════
#  INVENTORY CHECKING
# ══════════════════════════════════════════════════════════════════════════════

def get_available(start_date_str, end_date_str, exclude_id=None):
    """
    Returns {product_id: available_qty} for a date range.
    Locks inventory for both CONFIRMED (paid in full) and ACCEPTED (deposit paid / partially paid).
    Also matches items by name for Booqable-imported bookings that lack a product id.
    """
    products   = get_products()
    available  = {p["id"]: p["total"] for p in products}
    # Build name → id map so name-only items (e.g. Booqable imports) are matched
    name_to_id = {p["name"].lower(): p["id"] for p in products}
    conn = get_db()
    if not conn:
        return available
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # Window = setup_date (est. delivery) → event_end_date (est. pickup).
        query = """
            SELECT items_json FROM bookings
            WHERE (
              (status = 'accepted' AND payment_status IN ('partial','paid'))
              OR status = 'agree_to_pay'
            )
              AND (delivery_status IS NULL OR delivery_status != 'picked_up')
              AND setup_date     IS NOT NULL
              AND event_end_date IS NOT NULL
              AND setup_date     <= %s
              AND event_end_date >= %s
        """
        params = [end_date_str, start_date_str]
        if exclude_id:
            query += " AND id != %s"
            params.append(exclude_id)
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        for row in rows:
            try:
                items = json.loads(row["items_json"] or "[]")
                for item in items:
                    # Match by product id first, fall back to name lookup
                    pid = item.get("id") or name_to_id.get((item.get("name") or "").lower())
                    qty = int(item.get("qty") or 0)
                    if pid and pid in available and qty > 0:
                        available[pid] = max(0, available[pid] - qty)
            except Exception:
                pass
    except Exception as e:
        log.error(f"Inventory check error: {e}")
    return available


def get_inventory_conflicts():
    """
    Dashboard alert: flag pending/accepted bookings that can't be fulfilled
    given what confirmed/partial bookings have already locked up.
    Confirmed bookings are never flagged — they own their inventory.
    """
    products    = get_products()
    prod_totals = {p["id"]: int(p["total"]) for p in products}
    name_to_pid = {p["name"].lower(): p["id"] for p in products}

    conn = get_db()
    if not conn:
        return []
    conflicts = []
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # Window = setup_date (est. delivery) → event_end_date (est. pickup).
        # Only bookings with BOTH dates count.
        cur.execute("""
            SELECT id, full_name, status, payment_status, delivery_status,
                   setup_date, event_end_date, items_json
            FROM bookings
            WHERE status NOT IN ('cancelled','denied','concluded')
              AND (archived IS NULL OR archived = FALSE)
              AND setup_date     IS NOT NULL
              AND event_end_date IS NOT NULL
        """)
        all_bookings = [_row(r) for r in cur.fetchall()]
        cur.close(); conn.close()

        # Only paid/partial bookings lock inventory
        confirmed   = [b for b in all_bookings
                       if b.get("status") == "accepted"
                       and b.get("payment_status") in ("paid", "partial")]
        # Waiting/pending bookings might be short
        unconfirmed = [b for b in all_bookings
                       if b.get("status") in ("pending",)
                       or (b.get("status") == "accepted" and b.get("payment_status") == "waiting")]

        for b in unconfirmed:
            b_start = str(b.get("setup_date")    or "")[:10]
            b_end   = str(b.get("event_end_date") or "")[:10]
            if not b_start or not b_end:
                continue

            # Confirmed bookings whose est. delivery→pickup window overlaps this booking
            confirmed_reserved = {}
            for c in confirmed:
                c_start  = str(c.get("setup_date")    or "")[:10]
                c_end    = str(c.get("event_end_date") or "")[:10]
                c_picked = (c.get("delivery_status") or "") == "picked_up"
                c_concluded = c.get("status") == "concluded"
                if c_start and c_end and not c_picked and not c_concluded and c_start <= b_end and c_end >= b_start:
                    for item in json.loads(c.get("items_json") or "[]"):
                        pid = item.get("id") or name_to_pid.get((item.get("name") or "").lower())
                        qty = int(item.get("qty") or 0)
                        if pid and qty > 0:
                            confirmed_reserved[pid] = confirmed_reserved.get(pid, 0) + qty

            # Flag if this booking needs more than what's left after confirmed bookings
            for item in json.loads(b.get("items_json") or "[]"):
                pid = item.get("id") or name_to_pid.get((item.get("name") or "").lower())
                qty = int(item.get("qty") or 0)
                if pid and qty > 0 and pid in prod_totals:
                    avail = max(0, prod_totals[pid] - confirmed_reserved.get(pid, 0))
                    if qty > avail:
                        conflicts.append({
                            "booking_id": b["id"],
                            "customer":   b.get("full_name", "Unknown"),
                            "delivery_date": b_start,
                            "item":       item.get("name", ""),
                            "needed":     qty,
                            "available":  avail,
                            "shortfall":  qty - avail,
                        })
    except Exception as e:
        log.error(f"Conflict check error: {e}")
    return conflicts


def get_booking_inventory_check(booking_id):
    """
    For a specific booking (any status), check whether its items can be fulfilled
    given other confirmed/accepted bookings on the same dates.
    Returns list of {item, needed, available, shortfall}
    """
    products    = get_products()
    prod_totals = {p["id"]: int(p["total"]) for p in products}
    name_to_pid = {p["name"].lower(): p["id"] for p in products}
    log.info(f"[INV CHECK] booking #{booking_id} — prod_totals={prod_totals}")

    conn = get_db()
    if not conn:
        return []
    issues = []
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return []
        b       = _row(row)
        # Use setup_date (drop-off) → event_end_date (pickup) as the physical window
        b_start = str(b.get("setup_date",     "") or "")[:10]
        b_end   = str(b.get("event_end_date", "") or "")[:10]
        has_dates = bool(b_start or b_end)

        # Absolute inventory check (no dates needed) — catch orders > total stock
        b_items = json.loads(b.get("items_json") or "[]")
        log.info(f"[INV CHECK] #{booking_id} start={b_start} end={b_end} items={b_items}")
        for item in b_items:
            iname = (item.get("name") or "").strip()
            pid   = item.get("id") or name_to_pid.get(iname.lower())
            qty   = int(item.get("qty") or 0)
            if pid and qty > 0 and pid in prod_totals:
                total_stock = prod_totals[pid]
                if qty > total_stock:
                    log.info(f"[INV CHECK] #{booking_id} ABSOLUTE SHORTAGE: {iname} needs {qty}, total stock={total_stock}")
                    issues.append({
                        "item":                 iname,
                        "needed":               qty,
                        "available":            total_stock,
                        "shortfall":            qty - total_stock,
                        "conflicting_bookings": [],
                        "absolute":             True,
                    })

        if not has_dates:
            cur.close(); conn.close()
            return issues  # return absolute issues even without dates

        # Window = setup_date (est. delivery) → event_end_date (est. pickup).
        # Only bookings with BOTH dates set count against inventory.
        cur.execute("""
            SELECT id, full_name, items_json, setup_date, event_end_date
            FROM bookings
            WHERE (
              (status = 'accepted' AND payment_status IN ('partial','paid'))
              OR status = 'agree_to_pay'
            )
              AND (delivery_status IS NULL OR delivery_status != 'picked_up')
              AND id != %s
              AND setup_date     IS NOT NULL
              AND event_end_date IS NOT NULL
              AND setup_date     <= %s
              AND event_end_date >= %s
        """, (booking_id, b_end or b_start, b_start or b_end))
        others = [_row(r) for r in cur.fetchall()]
        cur.close(); conn.close()

        others_reserved  = {}   # pid -> total qty reserved by others
        pid_to_bookings  = {}   # pid -> list of booking dicts (primary, most reliable)
        name_to_bookings = {}   # name_lower -> list of booking dicts (fallback)

        for o in others:
            o_entry = {
                "id":    o["id"],
                "name":  o.get("full_name", "Unknown"),
                "start": str(o.get("setup_date",     ""))[:10],
                "end":   str(o.get("event_end_date", ""))[:10],
            }
            for item in json.loads(o.get("items_json") or "[]"):
                iname       = (item.get("name") or "").strip()
                iname_lower = iname.lower()
                pid         = item.get("id") or name_to_pid.get(iname_lower)
                qty         = int(item.get("qty") or 0)
                if qty <= 0:
                    continue
                entry = dict(o_entry, qty=qty)
                if pid:
                    others_reserved[pid] = others_reserved.get(pid, 0) + qty
                    if pid not in pid_to_bookings:
                        pid_to_bookings[pid] = []
                    if not any(e["id"] == o["id"] for e in pid_to_bookings[pid]):
                        pid_to_bookings[pid].append(entry)
                # Also track by name as fallback
                if iname_lower:
                    if iname_lower not in name_to_bookings:
                        name_to_bookings[iname_lower] = []
                    if not any(e["id"] == o["id"] for e in name_to_bookings[iname_lower]):
                        name_to_bookings[iname_lower].append(entry)

        already_flagged = {iss["item"] for iss in issues}
        for item in json.loads(b.get("items_json") or "[]"):
            iname       = (item.get("name") or "").strip()
            iname_lower = iname.lower()
            pid         = item.get("id") or name_to_pid.get(iname_lower)
            qty         = int(item.get("qty") or 0)
            if pid and qty > 0 and pid in prod_totals:
                avail = max(0, prod_totals[pid] - others_reserved.get(pid, 0))
                if qty > avail and iname not in already_flagged:
                    # Use pid lookup first; fall back to name lookup
                    conflicting = pid_to_bookings.get(pid) or name_to_bookings.get(iname_lower, [])
                    # Last resort: all overlapping bookings (so user always sees something)
                    if not conflicting:
                        conflicting = [{
                            "id":    o["id"],
                            "name":  o.get("full_name", "Unknown"),
                            "start": str(o.get("event_start_date", ""))[:10],
                            "end":   str(o.get("event_end_date",   ""))[:10],
                            "qty":   "?",
                        } for o in others]
                    issues.append({
                        "item":                 iname,
                        "needed":               qty,
                        "available":            avail,
                        "shortfall":            qty - avail,
                        "conflicting_bookings": conflicting,
                    })
    except Exception as e:
        log.error(f"Booking inventory check error: {e}")
    return issues


# ══════════════════════════════════════════════════════════════════════════════
#  DISTANCE & DELIVERY FEE
# ══════════════════════════════════════════════════════════════════════════════

def get_distance_miles(destination):
    if not GOOGLE_MAPS_KEY or not BUSINESS_ADDRESS or not destination:
        return None
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/distancematrix/json",
            params={"origins": BUSINESS_ADDRESS, "destinations": destination,
                    "units": "imperial", "key": GOOGLE_MAPS_KEY},
            timeout=10,
        )
        el = r.json()["rows"][0]["elements"][0]
        if el.get("status") != "OK":
            return None
        return round(el["distance"]["value"] / 1609.344, 1)
    except Exception as e:
        log.error(f"Distance error: {e}")
        return None


def calc_delivery_fee(miles):
    if miles is None:
        return DELIVERY_BASE_FEE, "flat fee (distance could not be verified)"
    if miles <= DELIVERY_THRESHOLD:
        return DELIVERY_BASE_FEE, f"{miles} mi — flat delivery fee"
    fee = round(miles * DELIVERY_RATE, 2)
    return fee, f"{miles} mi x ${DELIVERY_RATE:.2f}/mi"


# ══════════════════════════════════════════════════════════════════════════════
#  STRIPE
# ══════════════════════════════════════════════════════════════════════════════

def _deactivate_booking_links(booking_id):
  """Deactivate all active Stripe Payment Links for a booking to prevent double-charging."""
  conn = get_db()
  if not conn:
    return
  try:
    cur = conn.cursor()
    cur.execute(
      "SELECT stripe_link_id FROM payment_links WHERE booking_id=%s AND status='active' AND stripe_link_id IS NOT NULL",
      (booking_id,)
    )
    rows = cur.fetchall()
    for row in rows:
      link_id = row[0]
      try:
        stripe.PaymentLink.modify(link_id, active=False)
        log.info(f"Deactivated Stripe link {link_id} for booking #{booking_id}")
      except Exception as e:
        log.warning(f"Could not deactivate Stripe link {link_id}: {e}")
    cur.execute(
      "UPDATE payment_links SET status='inactive' WHERE booking_id=%s AND status='active'",
      (booking_id,)
    )
    conn.commit()
    cur.close()
    conn.close()
  except Exception as e:
    log.error(f"_deactivate_booking_links error: {e}")
def create_stripe_payment_link(booking_id, deposit_amount, customer_email, items_desc, product_name=None):
  """Create a Stripe Payment Link. Returns (url, stripe_link_id, error)."""
  if not STRIPE_SECRET_KEY:
      log.warning("STRIPE_SECRET_KEY not set — cannot create payment link")
      return None, None, "Stripe not configured"
  _deactivate_booking_links(booking_id)
  try:
      name = product_name or f"25% Deposit — Booking #{booking_id}"
      product = stripe.Product.create(
          name=name,
          description=(items_desc[:500] if items_desc else "Rental deposit"),
      )
      price = stripe.Price.create(
          unit_amount=int(round(deposit_amount * 100)),
          currency="usd",
          product=product.id,
      )
      kwargs = {
          "line_items": [{"price": price.id, "quantity": 1}],
          "metadata": {
              "booking_id": str(booking_id),
              "payment_type": "deposit",
              "amount": str(deposit_amount),
          },
      }
      if BASE_URL:
          kwargs["after_completion"] = {
              "type": "redirect",
              "redirect": {"url": f"{BASE_URL}/payment/success/{booking_id}"}
          }
      link = stripe.PaymentLink.create(**kwargs)
      log.info(f"Stripe Payment Link created for booking #{booking_id}: {link.url}")
      return link.url, link.id, None
  except Exception as e:
      log.error(f"Stripe Payment Link error: {e}")
      return None, None, str(e)


def save_payment_link(booking_id, label, amount, url, stripe_link_id):
    """Record a payment link in the payment_links table."""
    conn = get_db()
    if not conn:
        return
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO payment_links (booking_id, label, amount, url, stripe_link_id, status)
            VALUES (%s, %s, %s, %s, %s, 'active')
        """, (booking_id, label, amount, url, stripe_link_id))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"save_payment_link error: {e}")


def get_payment_links(booking_id):
    """Return all payment links for a booking, newest first."""
    conn = get_db()
    if not conn:
        return []
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("""
            SELECT id, label, amount, url, stripe_link_id, status, created_at
            FROM payment_links WHERE booking_id=%s ORDER BY created_at DESC
        """, (booking_id,))
        rows = [_row(r) for r in cur.fetchall()]
        cur.close(); conn.close()
        return rows
    except Exception as e:
        log.error(f"get_payment_links error: {e}")
        return []


# ══════════════════════════════════════════════════════════════════════════════
#  CONTRACT TEXT
# ══════════════════════════════════════════════════════════════════════════════

def build_contract_html(b, deposit_amount):
    """Build formatted HTML contract with booking details filled in."""
    customer_name = b.get('full_name', '')
    items = json.loads(b.get('items_json') or '[]')
    items_list = ', '.join(f"{i.get('qty','')}x {i.get('name','')}" for i in items)
    deposit_str = f"${deposit_amount:.2f}"
    event_date = str(b.get('event_start_date', ''))
    today_str = date.today().strftime("%B %d, %Y")

    return f"""
<div style="font-size:.84rem;color:#374151;line-height:1.75;font-family:-apple-system,sans-serif">

  <h3 style="font-size:1rem;font-weight:700;color:#1a365d;margin:0 0 1rem;border-bottom:2px solid #e2e8f0;padding-bottom:.5rem">
    NON-REFUNDABLE DEPOSIT AGREEMENT
  </h3>

  <p>This Non-Refundable Deposit Agreement is made and entered into by and between <strong>Rent a Party, LLC</strong>
  and <strong>{customer_name}</strong> ("Deposit Recipient") for the purpose of securing the date and time of
  <em>{items_list}</em> for event on <strong>{event_date}</strong>. The Deposit Provider agrees to provide a
  non-refundable deposit in the amount of <strong>{deposit_str}</strong> to secure the reservation.
  Throughout this agreement, <strong>{customer_name}</strong> shall also be referred to as "The Renter."</p>

  <p>Deposit Recipient acknowledges that the deposit is non-refundable and will be forfeited if any of the
  conditions outlined in this agreement occur. The conditions for forfeiture are as follows:</p>

  <ol style="margin:.75rem 0 .75rem 1.25rem;padding:0">
    <li style="margin-bottom:.4rem"><strong>DEPOSIT NEEDED IS TWENTY-FIVE PERCENT (25%) AND IS NOT REFUNDABLE UNDER ANY CIRCUMSTANCES.</strong></li>
    <li style="margin-bottom:.4rem">If canceled within 20 days of the scheduled event, 50% of all items will be charged.</li>
    <li style="margin-bottom:.4rem">If canceled within 10 days of the scheduled event, 75% of all items will be charged.</li>
    <li style="margin-bottom:.4rem">If canceled within 24 hours of the scheduled event, there will still be a 100% charge.</li>
    <li style="margin-bottom:.4rem">The scheduled event is not secured until deposit is paid in full. Full payment is required for bookings
    made within one week of the event. If Rent a Party, LLC is prevented or delayed in delivering or picking up equipment at the
    agreed-upon time and location due to the negligence of the Renter, the Renter shall be responsible for a fee of $75 per hour
    for any additional time required. The Renter agrees to pay the remaining balance 48 hours before the scheduled pick-up/drop-off.
    Failure to make payment 48 hours prior may result in the order being considered canceled. Renter agrees that a person 18 years
    or older must be present at time of delivery. Rent a Party, LLC does not offer refunds; postponed events due to inclement weather
    will receive store credits.</li>
  </ol>

  <p>Deposit Recipient agrees to the terms of this Agreement and acknowledges that they have read and understood all terms and
  conditions. This Agreement shall be governed by the laws of the State of Connecticut. <strong>By making the non-refundable payment,
  you agree to the terms and conditions stated in this agreement. No signature is required for this agreement to be legally binding.</strong></p>

  <h3 style="font-size:1rem;font-weight:700;color:#1a365d;margin:1.25rem 0 1rem;border-bottom:2px solid #e2e8f0;padding-bottom:.5rem">
    EQUIPMENT RENTAL TERMS
  </h3>

  <p>With the consensual agreement of the Owner leasing equipment(s) described above, the Renter agrees to the Terms and Conditions as follows:</p>

  <ol style="margin:.75rem 0 .75rem 1.25rem;padding:0">
    <li style="margin-bottom:.4rem">In the event any equipment upon its return is not in good repair, condition and working order (ordinary wear and tear excepted),
    the renter will be obligated to pay Owner for reasonable out-of-pocket expenses to restore such equipment.</li>
    <li style="margin-bottom:.4rem">If the Renter declines delivery service, the Renter agrees to return all equipment on or before the specified time.
    Late fees of $75 per hour will be charged for returns made after the specified time. The Renter assumes all responsibility for the equipment.</li>
    <li style="margin-bottom:.4rem">If Rent a Party, LLC is prevented or delayed in delivering or picking up equipment due to the negligence of the Renter,
    the Renter shall be responsible for a fee of $75 per hour for any additional time required.</li>
    <li style="margin-bottom:.4rem">The Renter has obtained authorization from the venue to use all equipment on their premises.</li>
    <li style="margin-bottom:.4rem">Renter may only use and operate any equipment for its intended purpose.</li>
    <li style="margin-bottom:.4rem">Renter shall install all equipment in a manner that allows for removal without damage.</li>
    <li style="margin-bottom:.4rem">Renter shall not make any additions, attachments, alterations or improvements to any equipment without prior written consent of Owner.</li>
    <li style="margin-bottom:.4rem">Renter agrees that all equipment received is in safe and proper order.</li>
    <li style="margin-bottom:.4rem">Renter may only use and operate all equipment for its intended purpose.</li>
    <li style="margin-bottom:.4rem">The Renter acknowledges that all equipment provided by Rent a Party, LLC is accurate and corresponds exactly to the equipment rental list.</li>
    <li style="margin-bottom:.4rem">Renter acknowledges that the quantity of equipment is accurate to what is stated in the equipment rental list.</li>
    <li style="margin-bottom:.4rem">Regarding deliveries without setup/breakdown package: all chairs must be stacked with the black circle facing up. If chairs are not stacked properly, a fee of $1 per rented chair may be charged.</li>
    <li style="margin-bottom:.4rem">Marquee items should not be exposed to moisture or rain, left outside overnight, or stood upon. Keep all marquee items dry at all times.</li>
    <li style="margin-bottom:.4rem">Items with electrical or battery-operated systems (speakers, microphones, etc.) should not be left outside overnight or exposed to moisture or rain.</li>
    <li style="margin-bottom:.4rem">Any water damage to rental products will result in the renter being responsible for the cost of repairing or replacing the damaged item(s).</li>
    <li style="margin-bottom:.4rem"><strong>OVERNIGHT RENTALS:</strong> Lessee understands that all equipment is to be locked up in a secure location overnight. The Renter is fully responsible for all equipment until it is returned or picked up by Rent a Party, LLC.</li>
  </ol>

  <p style="margin-top:1rem;font-style:italic;color:#6b7280;font-size:.8rem">
    Agreement date: {today_str} &nbsp;|&nbsp; Booking #{b.get('id')} &nbsp;|&nbsp; {customer_name}
  </p>
</div>
"""


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL
# ══════════════════════════════════════════════════════════════════════════════

OWNER_BCC = "rentapartyct@gmail.com"

def _send_email(to, subject, html, plain, reply_to=None):
    if not all([GMAIL_USER, GMAIL_APP_PASSWORD]):
        log.warning("Gmail not configured")
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["From"]    = f"{BUSINESS_NAME} <{GMAIL_USER}>"
        msg["To"]      = to
        msg["Subject"] = subject
        msg["Bcc"]     = OWNER_BCC
        if reply_to:
            msg["Reply-To"] = reply_to
        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html,  "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            s.send_message(msg)
        log.info(f"Email sent → {to} (bcc: {OWNER_BCC})")
    except Exception as e:
        log.error(f"Email error: {e}")


def send_sms(to_number, message):
    """Send an SMS via Twilio. Returns True on success, False on failure."""
    if not (TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_FROM_NUMBER):
        log.warning("Twilio not configured — SMS not sent")
        return False
    if not to_number:
        return False
    # Normalize number: strip non-digits, add +1 if 10-digit US
    digits = re.sub(r"\D", "", to_number)
    if len(digits) == 10:
        digits = "1" + digits
    if not digits.startswith("1") or len(digits) != 11:
        log.warning(f"SMS skipped — bad number: {to_number}")
        return False
    to_e164 = "+" + digits
    try:
        resp = requests.post(
            f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Messages.json",
            auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
            data={"From": TWILIO_FROM_NUMBER, "To": to_e164, "Body": message},
            timeout=10,
        )
        if resp.status_code in (200, 201):
            log.info(f"SMS sent → {to_e164}")
            return True
        else:
            log.error(f"Twilio error {resp.status_code}: {resp.text[:200]}")
            return False
    except Exception as e:
        log.error(f"SMS exception: {e}")
        return False


def send_push(title, body, click_url=""):
    """Send a push notification to the owner via ntfy.sh (free, no account needed).
    Tapping the notification opens click_url directly in the browser."""
    topic = os.getenv("NTFY_TOPIC", "")
    if not topic:
        log.warning("Push notification skipped — NTFY_TOPIC not set")
        return
    try:
        headers = {
            "Title":    title,
            "Priority": "high",
            "Tags":     "bell",
        }
        if click_url:
            headers["Click"] = click_url
        resp = requests.post(
            f"https://ntfy.sh/{topic}",
            data=body.encode("utf-8"),
            headers=headers,
            timeout=10
        )
        if resp.status_code >= 400:
            log.warning(f"Push notification failed ({resp.status_code}): {resp.text[:200]}")
        else:
            log.info(f"Push notification sent → ntfy.sh/{topic}")
    except Exception as e:
        log.error(f"Push notification error: {e}")


def send_owner_email(b):
    """Send detailed booking notification to owner."""
    if not OWNER_EMAIL:
        return
    items = json.loads(b.get("items_json") or "[]")
    exact = b.get("exact_time_delivery", False)
    event_addr = f"{b.get('event_street','')}, {b.get('event_city','')}, {b.get('event_state','')} {b.get('event_zip','')}"
    renter_addr = f"{b.get('renter_street','')}, {b.get('renter_city','')}, {b.get('renter_state','')} {b.get('renter_zip','')}"
    item_rows = ""
    for it in items:
        _it_up  = float(it.get("unit_price") or 0)
        _it_tot = float(it.get("total") or round(_it_up * int(it.get("qty") or 1), 2))
        item_rows += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">{it.get('name','')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center">{it.get('qty','')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right">${_it_up:.2f}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right;font-weight:600">${_it_tot:.2f}</td>
        </tr>"""
    subject = f"New Booking #{b.get('id')} — {b.get('full_name')} | {_fmt_date(b.get('event_start_date'))}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:640px;margin:0 auto">
  <div style="background:linear-gradient(135deg,#1a365d,#2b6cb0);border-radius:12px 12px 0 0;padding:1.5rem 2rem;color:white">
    <h2 style="margin:0">New Booking Request #{b.get('id')}</h2>
    <p style="margin:.4rem 0 0;opacity:.85">{BUSINESS_NAME} — Review in Admin Panel</p>
  </div>
  <div style="background:white;padding:2rem;border-radius:0 0 12px 12px;box-shadow:0 4px 16px rgba(0,0,0,.08)">
    <table style="width:100%;border-collapse:collapse;margin-bottom:1.5rem">
      <tr style="background:#ebf4ff"><td colspan="2" style="padding:10px 12px;font-weight:700;color:#2b6cb0;text-transform:uppercase;font-size:.85rem">Customer</td></tr>
      <tr><td style="padding:8px 12px;color:#718096;width:160px">Name</td><td style="padding:8px 12px;font-weight:600">{b.get('full_name')}</td></tr>
      <tr style="background:#f7fafc"><td style="padding:8px 12px;color:#718096">Address</td><td style="padding:8px 12px">{renter_addr}</td></tr>
      <tr><td style="padding:8px 12px;color:#718096">Phone</td><td style="padding:8px 12px">{b.get('phone')}</td></tr>
      <tr style="background:#f7fafc"><td style="padding:8px 12px;color:#718096">Email</td><td style="padding:8px 12px"><a href="mailto:{b.get('email')}">{b.get('email')}</a></td></tr>
    </table>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1.5rem">
      <tr style="background:#ebf4ff"><td colspan="2" style="padding:10px 12px;font-weight:700;color:#2b6cb0;text-transform:uppercase;font-size:.85rem">Event</td></tr>
      <tr><td style="padding:8px 12px;color:#718096;width:160px">Dates</td><td style="padding:8px 12px;font-weight:600">{_fmt_date(b.get('event_start_date'))} to {_fmt_date(b.get('event_end_date'))}</td></tr>
      <tr style="background:#f7fafc"><td style="padding:8px 12px;color:#718096">Event Address</td><td style="padding:8px 12px">{event_addr}</td></tr>
    </table>
    <table style="width:100%;border-collapse:collapse;margin-bottom:1.5rem">
      <tr style="background:#ebf4ff">
        <th style="padding:10px 12px;text-align:left;color:#2b6cb0;font-size:.85rem">Item</th>
        <th style="padding:10px 12px;text-align:center;color:#2b6cb0;font-size:.85rem">Qty</th>
        <th style="padding:10px 12px;text-align:right;color:#2b6cb0;font-size:.85rem">Price</th>
        <th style="padding:10px 12px;text-align:right;color:#2b6cb0;font-size:.85rem">Total</th>
      </tr>
      {item_rows}
      {"<tr><td colspan='3' style='padding:8px 12px;border-bottom:1px solid #e2e8f0'>Exact Time Delivery</td><td style='padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0'>$175.00</td></tr>" if exact else ""}
      <tr style="background:#1a365d;color:white">
        <td colspan="3" style="padding:12px;font-weight:700">ESTIMATED TOTAL</td>
        <td style="padding:12px;text-align:right;font-weight:700;font-size:1.2rem">${b.get('grand_total',0):.2f}</td>
      </tr>
    </table>
    <div style="background:#ebf4ff;border-radius:8px;padding:1rem;text-align:center">
      <p style="margin:0;font-weight:700;color:#1a365d">Log in to Admin Panel to Accept or Deny</p>
    </div>
  </div>
</div></body></html>"""
    plain = f"NEW BOOKING #{b.get('id')}\n{b.get('full_name')} | {b.get('email')} | {b.get('phone')}\nEvent: {_fmt_date(b.get('event_start_date'))}\nTotal: ${b.get('grand_total',0):.2f}\n"
    _send_email(OWNER_EMAIL, subject, html, plain, reply_to=b.get("email"))


def send_customer_email(b):
    """Send initial confirmation to customer (booking received, pending review)."""
    email = b.get("email")
    first = b.get("full_name", "").split()[0]
    if not email:
        return
    token = b.get("view_token") or ""
    base_url = os.environ.get("APP_BASE_URL", "").rstrip("/")
    view_url = f"{base_url}/booking/view/{token}" if token and base_url else ""
    view_btn = f"""
    <div style="text-align:center;margin:1.25rem 0">
      <a href="{view_url}"
         style="display:inline-block;background:#2b6cb0;color:white;padding:.75rem 2rem;border-radius:8px;font-weight:700;font-size:.95rem;text-decoration:none">
        &#x1F4CB; View Your Order
      </a>
      <p style="margin:.5rem 0 0;font-size:.78rem;color:#718096">Bookmark this link to check your order anytime</p>
    </div>""" if view_url else ""

    # ── Build items table rows ──────────────────────────────────────────
    items = json.loads(b.get("items_json") or "[]")
    item_rows = ""
    for it in items:
        item_rows += (
            f"<tr>"
            f"<td style='padding:.45rem .6rem;border-bottom:1px solid #e2e8f0'>{it.get('name','')}</td>"
            f"<td style='padding:.45rem .6rem;border-bottom:1px solid #e2e8f0;text-align:center'>{it.get('qty',1)}</td>"
            f"<td style='padding:.45rem .6rem;border-bottom:1px solid #e2e8f0;text-align:right'>${float(it.get('total',0)):.2f}</td>"
            f"</tr>"
        )

    subtotal     = float(b.get("items_subtotal") or 0)
    delivery_fee = float(b.get("delivery_fee") or 0)
    exact_fee    = 175.0 if b.get("exact_time_delivery") else 0.0
    tax          = float(b.get("tax_amount") or 0)
    grand_total  = float(b.get("grand_total") or 0)

    delivery_row = (
        f"<tr><td colspan='2' style='padding:.4rem .6rem;color:#4a5568'>Delivery Fee</td>"
        f"<td style='padding:.4rem .6rem;text-align:right'>${delivery_fee:.2f}</td></tr>"
    ) if delivery_fee else ""
    exact_row = (
        f"<tr><td colspan='2' style='padding:.4rem .6rem;color:#4a5568'>Exact Time Delivery</td>"
        f"<td style='padding:.4rem .6rem;text-align:right'>${exact_fee:.2f}</td></tr>"
    ) if exact_fee else ""

    # ── Event details ───────────────────────────────────────────────────
    start_date   = b.get("event_start_date") or ""
    end_date     = b.get("event_end_date") or ""
    start_time   = b.get("event_start_time") or ""
    end_time     = b.get("event_end_time") or ""       # pickup time
    setup_time   = b.get("setup_time") or ""           # delivery time
    venue_type   = b.get("venue_type") or ""
    event_addr   = ", ".join(filter(None, [
        b.get("event_street",""), b.get("event_city",""),
        b.get("event_state",""), b.get("event_zip","")
    ]))
    delivery_addr = b.get("delivery_address") or ""

    date_display = f"{start_date} - {end_date}" if end_date and end_date != start_date else start_date

    time_rows_html = ""
    if start_time:
        time_rows_html += f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Event Start</td><td style='padding:.3rem .5rem;font-size:.88rem'>{start_time}</td></tr>"
    if setup_time:
        time_rows_html += f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Delivery Time</td><td style='padding:.3rem .5rem;font-size:.88rem'>{setup_time}</td></tr>"
    if end_time:
        time_rows_html += f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Pickup Time</td><td style='padding:.3rem .5rem;font-size:.88rem'>{end_time}</td></tr>"

    venue_row   = f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Venue Type</td><td style='padding:.3rem .5rem;font-size:.88rem'>{venue_type}</td></tr>" if venue_type else ""
    addr_row    = f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Event Address</td><td style='padding:.3rem .5rem;font-size:.88rem'>{event_addr}</td></tr>" if event_addr else ""
    del_row     = f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Delivery To</td><td style='padding:.3rem .5rem;font-size:.88rem'>{delivery_addr}</td></tr>" if delivery_addr and delivery_addr != event_addr else ""
    cust_addr   = b.get("customer_address") or ""
    cust_addr_row = f"<tr><td style='padding:.3rem .5rem;color:#718096;font-size:.88rem'>Your Address</td><td style='padding:.3rem .5rem;font-size:.88rem'>{cust_addr}</td></tr>" if cust_addr else ""

    subject = f"We received your rental request! - {BUSINESS_NAME}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:580px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.08)">
  <div style="background:linear-gradient(135deg,#1a365d,#2b6cb0);padding:2rem;color:white;text-align:center">
    <h2 style="margin:0">Request Received!</h2>
    <p style="margin:.5rem 0 0;opacity:.85">{BUSINESS_NAME}</p>
  </div>
  <div style="padding:2rem">
    <p>Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7;margin:.75rem 0">
      Thank you for your rental inquiry! We have received your request and will review it shortly.
      Below is a full summary of everything you submitted for your records.
    </p>

    <div style="background:#f0f4f8;border-radius:8px;padding:1rem;margin:1rem 0;text-align:center">
      <p style="margin:0;font-weight:600;color:#2d3748">Booking Reference</p>
      <p style="margin:.3rem 0 0;font-size:1.5rem;font-weight:700;color:#2b6cb0">#{b.get('id')}</p>
    </div>
    {view_btn}

    <h3 style="margin:1.5rem 0 .5rem;color:#1a365d;font-size:.9rem;text-transform:uppercase;letter-spacing:.06em;border-bottom:2px solid #e2e8f0;padding-bottom:.4rem">Contact Information</h3>
    <table style="width:100%;border-collapse:collapse">
      <tr><td style="padding:.3rem .5rem;color:#718096;font-size:.88rem;width:38%">Name</td><td style="padding:.3rem .5rem;font-size:.88rem">{b.get('full_name','')}</td></tr>
      <tr><td style="padding:.3rem .5rem;color:#718096;font-size:.88rem">Email</td><td style="padding:.3rem .5rem;font-size:.88rem">{b.get('email','')}</td></tr>
      <tr><td style="padding:.3rem .5rem;color:#718096;font-size:.88rem">Phone</td><td style="padding:.3rem .5rem;font-size:.88rem">{b.get('phone','')}</td></tr>
      {cust_addr_row}
    </table>

    <h3 style="margin:1.5rem 0 .5rem;color:#1a365d;font-size:.9rem;text-transform:uppercase;letter-spacing:.06em;border-bottom:2px solid #e2e8f0;padding-bottom:.4rem">Event Details</h3>
    <table style="width:100%;border-collapse:collapse">
      <tr><td style="padding:.3rem .5rem;color:#718096;font-size:.88rem;width:38%">Date(s)</td><td style="padding:.3rem .5rem;font-size:.88rem">{date_display}</td></tr>
      {time_rows_html}
      {venue_row}
      {addr_row}
      {del_row}
    </table>

    <h3 style="margin:1.5rem 0 .5rem;color:#1a365d;font-size:.9rem;text-transform:uppercase;letter-spacing:.06em;border-bottom:2px solid #e2e8f0;padding-bottom:.4rem">Items Ordered</h3>
    <table style="width:100%;border-collapse:collapse">
      <thead>
        <tr style="background:#f7fafc">
          <th style="padding:.45rem .6rem;text-align:left;color:#4a5568;font-size:.85rem">Item</th>
          <th style="padding:.45rem .6rem;text-align:center;color:#4a5568;font-size:.85rem">Qty</th>
          <th style="padding:.45rem .6rem;text-align:right;color:#4a5568;font-size:.85rem">Price</th>
        </tr>
      </thead>
      <tbody>{item_rows}</tbody>
    </table>
    <table style="width:100%;border-collapse:collapse;margin-top:.25rem">
      <tr><td colspan="2" style="padding:.4rem .6rem;color:#4a5568;font-size:.88rem">Items Subtotal</td>
          <td style="padding:.4rem .6rem;text-align:right;font-size:.88rem">${subtotal:.2f}</td></tr>
      {delivery_row}
      {exact_row}
      <tr><td colspan="2" style="padding:.4rem .6rem;color:#4a5568;font-size:.88rem">CT Sales Tax (6.35%)</td>
          <td style="padding:.4rem .6rem;text-align:right;font-size:.88rem">${tax:.2f}</td></tr>
      <tr style="background:#f0f4f8;font-weight:700">
        <td colspan="2" style="padding:.55rem .6rem;border-top:2px solid #cbd5e0;font-size:.9rem">Estimated Total</td>
        <td style="padding:.55rem .6rem;border-top:2px solid #cbd5e0;text-align:right;font-size:.9rem">${grand_total:.2f}</td>
      </tr>
    </table>

    <div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;padding:.85rem 1rem;margin:1.25rem 0">
      <p style="margin:0;font-size:.88rem;color:#78350f;line-height:1.6">
        <strong>Please note:</strong> This is an estimate based on your submission.
        The final invoice will be sent to you before any payment is required.
      </p>
    </div>
    <p style="color:#4a5568;line-height:1.7;font-size:.9rem">
      Keep your booking reference handy.{f' Questions? Call <strong>{BUSINESS_PHONE}</strong>.' if BUSINESS_PHONE else ''}
    </p>
    <p style="color:#2d3748;font-weight:600;margin-top:1.5rem">- The {BUSINESS_NAME} Team</p>
  </div>
</div></body></html>"""

    item_lines = "\n".join(
        f"  {it.get('name','')} x{it.get('qty',1)}  ${float(it.get('total',0)):.2f}"
        for it in items
    )
    plain = (
        f"Hi {first},\n\n"
        f"Thank you for your rental inquiry! Here is a summary of your submission.\n\n"
        f"BOOKING REFERENCE: #{b.get('id')}\n"
        + (f"View your order: {view_url}\n" if view_url else "") +
        f"\n--- CONTACT ---\n"
        f"Name:  {b.get('full_name','')}\n"
        f"Email: {b.get('email','')}\n"
        f"Phone: {b.get('phone','')}\n"
        + (f"Address: {cust_addr}\n" if cust_addr else "") +
        f"\n--- EVENT ---\n"
        f"Date(s): {date_display}\n"
        + (f"Event Start:   {start_time}\n" if start_time else "")
        + (f"Delivery Time: {setup_time}\n" if setup_time else "")
        + (f"Pickup Time:   {end_time}\n" if end_time else "")
        + (f"Venue: {venue_type}\n" if venue_type else "")
        + (f"Address: {event_addr}\n" if event_addr else "") +
        f"\n--- ITEMS ORDERED ---\n{item_lines}\n\n"
        f"Subtotal:    ${subtotal:.2f}\n"
        + (f"Delivery:    ${delivery_fee:.2f}\n" if delivery_fee else "")
        + (f"Exact Time:  ${exact_fee:.2f}\n" if exact_fee else "") +
        f"Tax (6.35%): ${tax:.2f}\n"
        f"Est. Total:  ${grand_total:.2f}\n\n"
        f"Please note: this is an estimate. The final invoice will be sent before any payment is required.\n\n"
        f"- {BUSINESS_NAME}"
    )
    _send_email(email, subject, html, plain)


def send_accepted_email(b, charge_amount, payment_type="deposit"):
    """
    Send invoice + contract + Stripe payment link to customer.
    payment_type: "deposit" (25% now, rest later) or "full" (100% required now).
    """
    email = b.get("email")
    first = b.get("full_name", "").split()[0]
    if not email:
        return

    payment_link = b.get("stripe_payment_link", "")
    items = json.loads(b.get("items_json") or "[]")
    exact = b.get("exact_time_delivery", False)
    grand_total = float(b.get("grand_total") or 0)
    remaining = round(grand_total - charge_amount, 2)
    event_addr = f"{b.get('event_street','')}, {b.get('event_city','')}, {b.get('event_state','')} {b.get('event_zip','')}"

    is_deposit = (payment_type == "deposit")

    # Labels based on payment type
    if is_deposit:
        due_label      = "25% Deposit Due Now"
        pay_btn_label  = f"Pay ${charge_amount:.2f} Deposit Now"
        header_sub     = "Pay your 25% deposit to secure your date"
        urgency_msg    = "Your booking is <strong>not secured</strong> until the deposit is paid. Pay now to lock in your date."
        balance_line   = f'<div style="border-top:1px solid #c6f6d5;margin-top:1rem;padding-top:1rem;font-size:.87rem;color:#4a5568"><p style="margin:0"><strong>Remaining balance:</strong> ${remaining:.2f} — due <strong>48 hours before</strong> your delivery on {b.get("setup_date") or b.get("event_start_date")}</p></div>'
        balance_plain  = f"Remaining balance: ${remaining:.2f} — due 48 hours before your delivery."
    else:
        due_label      = "Full Payment Required"
        pay_btn_label  = f"Pay Full Amount ${charge_amount:.2f}"
        header_sub     = "Full payment required — your delivery day is within 7 days"
        urgency_msg    = "Because your delivery day is <strong>within 7 days</strong>, full payment is required to secure your booking."
        balance_line   = ""
        balance_plain  = "Full payment required — no remaining balance."

    tax_amount_val  = float(b.get("tax_amount") or 0)
    tax_rate_val    = float(b.get("tax_rate") or 0)
    is_tax_exempt_b = bool(b.get("tax_exempt"))
    disc_amount     = float(b.get("discount_amount") or 0)
    disc_type       = b.get("discount_type") or ""
    disc_value      = float(b.get("discount_value") or 0)

    item_rows = ""
    for it in items:
        _it_up  = float(it.get("unit_price") or 0)
        _it_tot = float(it.get("total") or round(_it_up * int(it.get("qty") or 1), 2))
        item_rows += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">{it.get('name','')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center">{it.get('qty','')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right">${_it_up:.2f}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right;font-weight:600">${_it_tot:.2f}</td>
        </tr>"""

    if disc_amount > 0:
        disc_label = f"{disc_value:.1f}% Discount" if disc_type == "percent" else f"${disc_value:.2f} Discount"
        disc_row_html  = f'<tr style="background:#f0fdf4"><td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#166534;font-weight:600">🏷️ {disc_label}</td><td style="padding:8px 12px;text-align:right;font-weight:700;border-bottom:1px solid #e2e8f0;color:#16a34a">- ${disc_amount:.2f}</td></tr>'
        disc_row_plain = f"  {disc_label}:    -${disc_amount:.2f}\n"
    else:
        disc_row_html  = ""
        disc_row_plain = ""

    if is_tax_exempt_b:
        tax_row_html  = '<tr><td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#276749">CT Sales Tax <span style="font-size:.76rem;background:#c6f6d5;color:#276749;border-radius:4px;padding:.1rem .35rem;margin-left:.3rem">TAX EXEMPT</span></td><td style="padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0;color:#276749">$0.00</td></tr>'
        tax_row_plain = "  CT Sales Tax:       $0.00 (TAX EXEMPT)\n"
    elif tax_amount_val:
        tax_row_html  = f'<tr><td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#718096">CT Sales Tax ({tax_rate_val*100:.2f}%)</td><td style="padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0">${tax_amount_val:.2f}</td></tr>'
        tax_row_plain = f"  CT Sales Tax ({tax_rate_val*100:.2f}%): ${tax_amount_val:.2f}\n"
    else:
        tax_row_html  = ""
        tax_row_plain = ""

    payment_btn = f"""
    <div style="text-align:center;margin:1.5rem 0">
      <a href="{payment_link}"
         style="display:inline-block;background:linear-gradient(135deg,#276749,#38a169);color:white;padding:1.1rem 2.75rem;border-radius:10px;font-weight:700;font-size:1.15rem;text-decoration:none;letter-spacing:.3px;box-shadow:0 4px 12px rgba(39,103,73,.35)">
        {pay_btn_label}
      </a>
      <p style="margin:.6rem 0 0;font-size:.82rem;color:#718096">Secure payment powered by Stripe</p>
    </div>""" if payment_link else f"""
    <div style="background:#fffaf0;border:2px solid #ed8936;border-radius:10px;padding:1.25rem;text-align:center;margin:1.5rem 0">
      <p style="font-weight:700;color:#744210">Amount Due: ${charge_amount:.2f}</p>
      <p style="color:#744210;font-size:.9rem">We will send your payment link shortly.{f" Questions? Call {BUSINESS_PHONE}" if BUSINESS_PHONE else ""}</p>
    </div>"""

    contract_html = build_contract_html(b, charge_amount)

    subject_tag = "Deposit Required" if is_deposit else "Full Payment Required"
    subject = f"Booking Accepted — {subject_tag} | {BUSINESS_NAME}"

    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:640px;margin:0 auto">

  <div style="background:linear-gradient(135deg,#276749,#38a169);border-radius:12px 12px 0 0;padding:1.75rem 2rem;color:white;text-align:center">
    <div style="font-size:2.2rem;margin-bottom:.4rem">&#127881; Booking Accepted!</div>
    <h2 style="margin:0;font-weight:700;font-size:1.2rem">Booking #{b.get('id')} &mdash; {BUSINESS_NAME}</h2>
    <p style="margin:.5rem 0 0;opacity:.88;font-size:.95rem">{header_sub}</p>
  </div>

  <div style="background:white;padding:2rem;border-radius:0 0 12px 12px;box-shadow:0 4px 16px rgba(0,0,0,.08)">

    <p style="color:#2d3748;font-size:1.05rem;margin-bottom:.75rem">Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7;margin-bottom:1.25rem">
      Thank you for choosing Rent a Party, LLC. We are pleased to confirm that your rental request has been reviewed and we have availability for your event.
      Please carefully review your invoice below, read the rental agreement at the bottom of this email, and submit your payment at your earliest convenience to secure your reservation.
    </p>

    <!-- Event Summary box -->
    <div style="background:#f0fff4;border:1.5px solid #68d391;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1.5rem;font-size:.9rem;color:#2d3748">
      <div style="margin-bottom:.3rem"><strong>Event Date:</strong> {_fmt_date(b.get('event_start_date'))} &rarr; {_fmt_date(b.get('event_end_date'))}</div>
      <div style="margin-bottom:.3rem"><strong>Location:</strong> {event_addr}</div>
      <div><strong>Deliver to:</strong> {b.get('delivery_location','')}</div>
    </div>

    <!-- Invoice table -->
    <h3 style="color:#1a365d;font-size:.95rem;font-weight:700;margin:0 0 .75rem;text-transform:uppercase;letter-spacing:.5px">Invoice</h3>
    <table style="width:100%;border-collapse:collapse;font-size:.9rem;margin-bottom:1.5rem">
      <thead>
        <tr style="background:#ebf4ff">
          <th style="padding:9px 12px;text-align:left;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Item</th>
          <th style="padding:9px 12px;text-align:center;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Qty</th>
          <th style="padding:9px 12px;text-align:right;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Unit Price</th>
          <th style="padding:9px 12px;text-align:right;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Total</th>
        </tr>
      </thead>
      <tbody>
        {item_rows}
        {"<tr><td colspan='3' style='padding:8px 12px;border-bottom:1px solid #e2e8f0'>Exact Time Delivery</td><td style='padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0'>$175.00</td></tr>" if exact else ""}
        <tr>
          <td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#718096">Delivery Fee ({b.get('distance_miles','?')} mi)</td>
          <td style="padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0">${b.get('delivery_fee',0):.2f}</td>
        </tr>
        {disc_row_html}
        {tax_row_html}
        <tr style="background:#1a365d;color:white">
          <td colspan="3" style="padding:11px 12px;font-weight:700;font-size:1rem">TOTAL</td>
          <td style="padding:11px 12px;text-align:right;font-weight:700;font-size:1.2rem">${grand_total:.2f}</td>
        </tr>
      </tbody>
    </table>

    <!-- Payment section -->
    <div style="background:#f0fff4;border:2px solid #38a169;border-radius:12px;padding:1.5rem;margin-bottom:1.5rem;text-align:center">
      <p style="font-size:.82rem;color:#276749;margin:0 0 .3rem;font-weight:700;text-transform:uppercase;letter-spacing:.6px">{due_label}</p>
      <p style="font-size:2.75rem;font-weight:800;color:#276749;margin:.2rem 0 .1rem;line-height:1">${charge_amount:.2f}</p>
      {f'<p style="font-size:.85rem;color:#718096;margin:.2rem 0 0">of ${grand_total:.2f} total</p>' if is_deposit else ""}
      {payment_btn}
      {balance_line}
    </div>

    <!-- Urgency note -->
    <div style="background:#fffaf0;border-left:4px solid #ed8936;padding:1rem 1.25rem;border-radius:0 8px 8px 0;margin-bottom:1.75rem;font-size:.88rem;color:#744210">
      <strong>Important:</strong> {urgency_msg}
      Inventory is first-come-first-paid &mdash; we cannot hold your reservation without payment.
    </div>

    <!-- Contract -->
    <div style="border-top:2px solid #e2e8f0;padding-top:1.5rem">
      <h3 style="color:#1a365d;font-size:.95rem;font-weight:700;margin:0 0 .4rem;text-transform:uppercase;letter-spacing:.5px">Rental Agreement</h3>
      <p style="font-size:.82rem;color:#718096;margin:0 0 1rem">
        Please read the following agreement carefully.
        By completing your payment above, you agree to all terms below. No additional signature is required.
      </p>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:1.25rem">
        {contract_html}
      </div>
    </div>

    <p style="color:#2d3748;font-weight:600;margin-top:1.75rem">&mdash; The {BUSINESS_NAME} Team</p>
    {f'<p style="font-size:.85rem;color:#718096">Questions? Call us at {BUSINESS_PHONE}</p>' if BUSINESS_PHONE else ""}
  </div>
</div></body></html>"""

    _plain_items = "".join("  " + str(i.get('qty','')) + "x " + i.get('name','') + " @ $" + f"{float(i.get('unit_price') or 0):.2f}" + " = $" + f"{float(i.get('total') or 0):.2f}" + "\n" for i in items)
    _plain_exact = "  Exact Time Delivery: $175.00\n" if exact else ""
    plain = f"""Hi {first},

GREAT NEWS — Your rental request (Booking #{b.get('id')}) has been ACCEPTED!

EVENT DETAILS
  Date:       {_fmt_date(b.get('event_start_date'))} - {_fmt_date(b.get('event_end_date'))}
  Location:   {event_addr}
  Deliver to: {b.get('delivery_location','')}

INVOICE
{_plain_items}{_plain_exact}  Delivery Fee: ${b.get('delivery_fee',0):.2f}
{disc_row_plain}{tax_row_plain}  ─────────────────────────────
  TOTAL: ${grand_total:.2f}

PAYMENT REQUIRED
  {due_label}: ${charge_amount:.2f}
  {f"Pay here: {payment_link}" if payment_link else "We will send your payment link shortly."}
  {balance_plain}

IMPORTANT: Your booking is NOT secured until payment is received.
{f"Call us at {BUSINESS_PHONE} with any questions." if BUSINESS_PHONE else ""}

By completing payment you agree to the rental terms and contract.

— {BUSINESS_NAME}"""

    _send_email(email, subject, html, plain)


def send_receipt_email(b):
    """Send a detailed receipt to the customer after payment."""
    email = b.get("email")
    if not email:
        return
    first  = (b.get("full_name") or "").split()[0] or "Valued Customer"
    bid    = b.get("id")
    paid   = float(b.get("amount_paid") or 0)
    total  = float(b.get("grand_total") or 0)
    balance = max(round(total - paid, 2), 0)

    # Build items table rows
    items = b.get("items_json") or []
    if isinstance(items, str):
        try: items = json.loads(items)
        except Exception: items = []
    item_rows = ""
    for it in items:
        name  = it.get("name", "")
        qty   = it.get("qty", 1)
        up    = float(it.get("unit_price") or 0)
        tot   = float(it.get("total") or round(up * qty, 2))
        item_rows += f"""<tr>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0">{name}</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:center">{qty}</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:right">${up:,.2f}</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:right">${tot:,.2f}</td>
        </tr>"""

    delivery_fee  = float(b.get("delivery_fee") or 0)
    late_fee      = float(b.get("late_night_fee") or 0)
    r_disc_amount = float(b.get("discount_amount") or 0)
    r_disc_type   = b.get("discount_type") or ""
    r_disc_value  = float(b.get("discount_value") or 0)
    if delivery_fee:
        item_rows += f"""<tr><td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0" colspan="3">Delivery Fee</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:right">${delivery_fee:,.2f}</td></tr>"""
    if late_fee:
        item_rows += f"""<tr><td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0" colspan="3">Late Night Fee</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:right">${late_fee:,.2f}</td></tr>"""
    if r_disc_amount > 0:
        r_disc_label = f"{r_disc_value:.1f}% Discount" if r_disc_type == "percent" else f"${r_disc_value:.2f} Discount"
        item_rows += f"""<tr style="background:#f0fdf4"><td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;color:#166534;font-weight:600" colspan="3">🏷️ {r_disc_label}</td>
          <td style="padding:.5rem .75rem;border-bottom:1px solid #e2e8f0;text-align:right;font-weight:700;color:#16a34a">- ${r_disc_amount:,.2f}</td></tr>"""

    balance_row = ""
    if balance > 0.01:
        balance_row = f"""<tr style="background:#fff5f5"><td colspan="3" style="padding:.5rem .75rem;font-weight:600;color:#991b1b">Balance Remaining</td>
          <td style="padding:.5rem .75rem;text-align:right;font-weight:700;color:#991b1b">${balance:,.2f}</td></tr>"""

    subject = f"Receipt for Order #{bid} — {BUSINESS_NAME}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:560px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.08)">
  <div style="background:linear-gradient(135deg,#1a365d,#2b6cb0);padding:2rem;color:white;text-align:center">
    <h2 style="margin:0">Payment Receipt</h2>
    <p style="margin:.4rem 0 0;opacity:.85">{BUSINESS_NAME}</p>
  </div>
  <div style="padding:2rem">
    <p>Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7">Thank you so much for your business! We truly appreciate you choosing {BUSINESS_NAME} and look forward to making your event a great one.</p>
    <div style="background:#f0f4f8;border-radius:8px;padding:1rem;margin:1.25rem 0;text-align:center">
      <p style="margin:0;font-size:.8rem;font-weight:600;color:#718096;text-transform:uppercase;letter-spacing:.05em">Order Number</p>
      <p style="margin:.25rem 0 0;font-size:1.6rem;font-weight:800;color:#2b6cb0">#{bid}</p>
      <p style="margin:.25rem 0 0;font-size:.85rem;color:#718096">{_fmt_date(b.get('event_start_date'))} — {_fmt_date(b.get('event_end_date'))}</p>
    </div>
    <table style="width:100%;border-collapse:collapse;font-size:.9rem;margin:1rem 0">
      <thead>
        <tr style="background:#f0f4f8">
          <th style="padding:.5rem .75rem;text-align:left;font-weight:600;color:#4a5568">Item</th>
          <th style="padding:.5rem .75rem;text-align:center;font-weight:600;color:#4a5568">Qty</th>
          <th style="padding:.5rem .75rem;text-align:right;font-weight:600;color:#4a5568">Unit Price</th>
          <th style="padding:.5rem .75rem;text-align:right;font-weight:600;color:#4a5568">Total</th>
        </tr>
      </thead>
      <tbody>{item_rows}</tbody>
      <tfoot>
        <tr style="background:#f0fdf4">
          <td colspan="3" style="padding:.6rem .75rem;font-weight:700;color:#166534">Grand Total</td>
          <td style="padding:.6rem .75rem;text-align:right;font-weight:800;color:#166534">${total:,.2f}</td>
        </tr>
        <tr style="background:#f0fdf4">
          <td colspan="3" style="padding:.5rem .75rem;color:#166534">✅ Amount Paid</td>
          <td style="padding:.5rem .75rem;text-align:right;font-weight:700;color:#166534">${paid:,.2f}</td>
        </tr>
        {balance_row}
      </tfoot>
    </table>
    <p style="color:#4a5568;line-height:1.7;font-size:.9rem">Please keep this receipt for your records. Order #{bid} is your reference for any questions or changes.</p>
    {f'<p style="color:#4a5568;font-size:.9rem">Questions? Call <strong>{BUSINESS_PHONE}</strong> or reply to this email.</p>' if BUSINESS_PHONE else ''}
    <p style="color:#2d3748;font-weight:600;margin-top:1.5rem">With gratitude,<br>— The {BUSINESS_NAME} Team</p>
  </div>
</div></body></html>"""
    plain = (f"RECEIPT — Order #{bid}\n{BUSINESS_NAME}\n\n"
             f"Hi {first},\n\nThank you for your business! We appreciate you choosing {BUSINESS_NAME}.\n\n"
             f"Event: {_fmt_date(b.get('event_start_date'))} — {_fmt_date(b.get('event_end_date'))}\n"
             f"Grand Total: ${total:,.2f}\nAmount Paid: ${paid:,.2f}\n"
             + (f"Balance Due: ${balance:,.2f}\n" if balance > 0.01 else "Paid in Full\n") +
             f"\nOrder Reference: #{bid}\n\n— {BUSINESS_NAME}")
    _send_email(email, subject, html, plain)


def send_denied_email(b):
    """Send polite denial email to customer."""
    email = b.get("email")
    first = b.get("full_name", "").split()[0]
    if not email:
        return
    subject = f"Regarding Your Rental Request — {BUSINESS_NAME}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:500px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.08)">
  <div style="background:linear-gradient(135deg,#1a365d,#2b6cb0);padding:1.75rem 2rem;color:white;text-align:center">
    <h2 style="margin:0">{BUSINESS_NAME}</h2>
  </div>
  <div style="padding:2rem">
    <p>Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7;margin:.75rem 0">
      Thank you for thinking of us for your event on <strong>{_fmt_date(b.get('event_start_date'))}</strong>.
      Unfortunately, we are unable to accommodate your rental request at this time.
      This may be due to availability, date conflicts, or other circumstances.
    </p>
    <p style="color:#4a5568;line-height:1.7;margin:.75rem 0">
      We're sorry for any inconvenience and hope to have the opportunity to serve you in the future.
      Please don't hesitate to reach out if you have a different date in mind or would like to discuss other options.
    </p>
    {f'<p style="color:#4a5568;margin:.75rem 0">You can reach us at <strong>{BUSINESS_PHONE}</strong>.</p>' if BUSINESS_PHONE else ""}
    <p style="color:#2d3748;font-weight:600;margin-top:1.5rem">— The {BUSINESS_NAME} Team</p>
  </div>
</div></body></html>"""
    plain = f"Hi {first},\n\nThank you for your interest in {BUSINESS_NAME}. Unfortunately, we are unable to accommodate your rental request for {_fmt_date(b.get('event_start_date'))} at this time.\n\nWe hope to serve you in the future.{f' Please call {BUSINESS_PHONE} if you have questions.' if BUSINESS_PHONE else ''}\n\n— {BUSINESS_NAME}"
    _send_email(email, subject, html, plain)


def send_denied_inventory_email(b, short_items):
    """Send professional denial email specifically citing inventory shortage."""
    email = b.get("email")
    first = (b.get("full_name") or "").split()[0]
    if not email:
        return

    event_date_fmt = _fmt_date(b.get("event_start_date"))

    item_rows_html = ""
    item_rows_plain = ""
    for name in short_items:
        item_rows_html += f"""
        <div style="display:flex;justify-content:space-between;font-size:.88rem;
                    color:#2d3748;padding:.4rem 0;border-bottom:1px solid #fbd38d">
          <span>{name}</span>
          <span style="color:#c05621;font-weight:600">Unavailable for your date</span>
        </div>"""
        item_rows_plain += f"  - {name}: unavailable for your date\n"

    subject = f"Regarding Your Rental Request — {BUSINESS_NAME}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:520px;margin:0 auto">

  <div style="background:#1a365d;border-radius:12px 12px 0 0;padding:1.75rem 2rem;color:white;text-align:center">
    <div style="font-size:.95rem;opacity:.8;margin-bottom:.3rem">{BUSINESS_NAME}</div>
    <h2 style="margin:0;font-weight:600;font-size:1.3rem">Regarding Your Rental Request</h2>
  </div>

  <div style="background:white;padding:2rem;border-radius:0 0 12px 12px;box-shadow:0 4px 16px rgba(0,0,0,.08)">

    <p style="color:#2d3748;font-size:1rem;margin:0 0 1rem">Hi <strong>{first}</strong>,</p>

    <p style="color:#4a5568;line-height:1.75;margin:0 0 1rem;font-size:.95rem">
      Thank you for choosing {BUSINESS_NAME} for your upcoming event. We truly appreciate
      your interest and the time you took to submit your request.
    </p>

    <p style="color:#4a5568;line-height:1.75;margin:0 0 1.25rem;font-size:.95rem">
      Unfortunately, after reviewing your booking for <strong style="color:#2d3748">{event_date_fmt}</strong>,
      we are unable to fulfill your order at this time. The item(s) you requested are not available
      in sufficient quantity for your event date due to prior commitments already in place.
    </p>

    <div style="background:#fffbeb;border:1.5px solid #fbd38d;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1.25rem">
      <div style="font-size:.75rem;font-weight:700;color:#92400e;margin-bottom:.6rem;
                  text-transform:uppercase;letter-spacing:.05em">Items unavailable</div>
      {item_rows_html}
    </div>

    <p style="color:#4a5568;line-height:1.75;margin:0 0 1.25rem;font-size:.95rem">
      We sincerely apologize for any inconvenience this may cause. We would love the
      opportunity to serve you and encourage you to reach out if you have a flexible date,
      need adjusted quantities, or would like to be placed on a waitlist in case
      availability opens up.
    </p>

    <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;
                padding:.85rem 1.1rem;margin-bottom:1.5rem;font-size:.88rem;color:#4a5568;line-height:1.8">
      {f'<div>📞 <strong style="color:#2d3748">{BUSINESS_PHONE}</strong></div>' if BUSINESS_PHONE else ''}
      {f'<div>✉️ <strong style="color:#2d3748">{BUSINESS_EMAIL}</strong></div>' if BUSINESS_EMAIL else ''}
    </div>

    <p style="margin:0;color:#2d3748;font-size:.95rem;font-weight:600">— The {BUSINESS_NAME} Team</p>
  </div>
</div></body></html>"""

    plain = (
        f"Hi {first},\n\n"
        f"Thank you for choosing {BUSINESS_NAME} for your upcoming event.\n\n"
        f"Unfortunately, after reviewing your booking for {event_date_fmt}, we are unable to fulfill "
        f"your order at this time. The following item(s) are not available in sufficient quantity "
        f"due to prior commitments:\n\n{item_rows_plain}\n"
        f"We sincerely apologize for any inconvenience. Please don't hesitate to reach out if you "
        f"have a flexible date, need adjusted quantities, or would like to discuss other options."
        + (f"\n\nCall us at {BUSINESS_PHONE}." if BUSINESS_PHONE else "")
        + f"\n\n— The {BUSINESS_NAME} Team"
    )
    _send_email(email, subject, html, plain)


def send_final_payment_email(b, remaining_amount, payment_link):
    """Send final payment reminder 48 hours before event with Stripe link for remaining balance."""
    email = b.get("email")
    first = b.get("full_name", "").split()[0]
    if not email:
        return

    items = json.loads(b.get("items_json") or "[]")
    exact = b.get("exact_time_delivery", False)
    grand_total    = float(b.get("grand_total") or 0)
    deposit_paid   = round(grand_total - remaining_amount, 2)
    event_addr     = f"{b.get('event_street','')}, {b.get('event_city','')}, {b.get('event_state','')} {b.get('event_zip','')}"
    event_date     = str(b.get("event_start_date", ""))
    event_time     = str(b.get("event_start_time", ""))
    setup_time     = str(b.get("setup_time", ""))

    # Calculate actual days until delivery
    try:
        ev_dt = datetime.strptime(event_date[:10], "%Y-%m-%d").date()
        days_until = (ev_dt - date.today()).days
        if days_until <= 0:
            days_label = "today"
        elif days_until == 1:
            days_label = "tomorrow"
        else:
            days_label = f"{days_until} days away"
    except Exception:
        days_label = "soon"

    item_rows = ""
    for it in items:
        item_rows += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0">{it['name']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:center">{it['qty']}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right">${it['unit_price']:.2f}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #e2e8f0;text-align:right;font-weight:600">${it['total']:.2f}</td>
        </tr>"""

    contract_html = build_contract_html(b, remaining_amount)

    f_tax_amount   = float(b.get("tax_amount") or 0)
    f_tax_rate     = float(b.get("tax_rate") or 0)
    f_tax_exempt   = bool(b.get("tax_exempt"))
    if f_tax_exempt:
        f_tax_row_html  = '<tr><td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#276749">CT Sales Tax <span style="font-size:.76rem;background:#c6f6d5;color:#276749;border-radius:4px;padding:.1rem .35rem;margin-left:.3rem">TAX EXEMPT</span></td><td style="padding:8px 12px;text-align:right;color:#276749;font-weight:600;border-bottom:1px solid #e2e8f0">$0.00</td></tr>'
        f_tax_row_plain = "  CT Sales Tax:       $0.00 (TAX EXEMPT)\n"
    elif f_tax_amount:
        f_tax_row_html  = f'<tr><td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#718096">CT Sales Tax ({f_tax_rate*100:.2f}%)</td><td style="padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0">${f_tax_amount:.2f}</td></tr>'
        f_tax_row_plain = f"  CT Sales Tax ({f_tax_rate*100:.2f}%): ${f_tax_amount:.2f}\n"
    else:
        f_tax_row_html  = ""
        f_tax_row_plain = ""

    pay_btn = f"""
      <a href="{payment_link}"
         style="display:inline-block;background:linear-gradient(135deg,#c05621,#dd6b20);color:white;padding:1.1rem 2.75rem;border-radius:10px;font-weight:700;font-size:1.15rem;text-decoration:none;letter-spacing:.3px;box-shadow:0 4px 12px rgba(192,86,33,.35)">
        Pay Remaining Balance ${remaining_amount:.2f}
      </a>
      <p style="margin:.6rem 0 0;font-size:.82rem;color:#718096">Secure payment powered by Stripe</p>""" if payment_link else f"""
      <p style="font-weight:700;color:#c05621">Remaining Balance Due: ${remaining_amount:.2f}</p>
      <p style="color:#744210;font-size:.9rem">Please contact us to complete your payment.{f" Call {BUSINESS_PHONE}" if BUSINESS_PHONE else ""}</p>"""

    subject = f"Final Payment Due — Your Delivery is {days_label}! | {BUSINESS_NAME}"
    html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:640px;margin:0 auto">

  <div style="background:linear-gradient(135deg,#c05621,#dd6b20);border-radius:12px 12px 0 0;padding:1.75rem 2rem;color:white;text-align:center">
    <div style="font-size:2rem;margin-bottom:.4rem">&#8987; Your Delivery is {days_label}!</div>
    <h2 style="margin:0;font-weight:700;font-size:1.2rem">Final Payment Due — {BUSINESS_NAME}</h2>
    <p style="margin:.5rem 0 0;opacity:.88;font-size:.95rem">Booking #{b.get('id')} &bull; {event_date}</p>
  </div>

  <div style="background:white;padding:2rem;border-radius:0 0 12px 12px;box-shadow:0 4px 16px rgba(0,0,0,.08)">

    <p style="color:#2d3748;font-size:1.05rem;margin-bottom:.75rem">Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7;margin-bottom:1.25rem">
      This is your final payment reminder. Your delivery is <strong>{days_label}</strong> and your remaining balance
      is due now to ensure everything is ready for delivery.
    </p>

    <!-- Delivery Summary -->
    <div style="background:#fff8f3;border:1.5px solid #fbd38d;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1.5rem;font-size:.9rem;color:#2d3748">
      <div style="margin-bottom:.3rem"><strong>&#128197; Delivery Date:</strong> {event_date}</div>
      <div style="margin-bottom:.3rem"><strong>&#8986; Event Start Time:</strong> {event_time}</div>
      <div style="margin-bottom:.3rem"><strong>&#128666; Delivery Time:</strong> {setup_time}</div>
      <div style="margin-bottom:.3rem"><strong>&#128205; Location:</strong> {event_addr}</div>
      <div><strong>&#128204; Deliver to:</strong> {b.get('delivery_location','')}</div>
    </div>

    <!-- Invoice summary -->
    <h3 style="color:#1a365d;font-size:.95rem;font-weight:700;margin:0 0 .75rem;text-transform:uppercase;letter-spacing:.5px">Your Order</h3>
    <table style="width:100%;border-collapse:collapse;font-size:.9rem;margin-bottom:1.5rem">
      <thead>
        <tr style="background:#ebf4ff">
          <th style="padding:9px 12px;text-align:left;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Item</th>
          <th style="padding:9px 12px;text-align:center;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Qty</th>
          <th style="padding:9px 12px;text-align:right;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Unit</th>
          <th style="padding:9px 12px;text-align:right;color:#2b6cb0;font-size:.78rem;text-transform:uppercase">Total</th>
        </tr>
      </thead>
      <tbody>
        {item_rows}
        {"<tr><td colspan='3' style='padding:8px 12px;border-bottom:1px solid #e2e8f0'>Exact Time Delivery</td><td style='padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0'>$175.00</td></tr>" if exact else ""}
        <tr>
          <td colspan="3" style="padding:8px 12px;border-bottom:1px solid #e2e8f0;color:#718096">Delivery Fee</td>
          <td style="padding:8px 12px;text-align:right;font-weight:600;border-bottom:1px solid #e2e8f0">${b.get('delivery_fee',0):.2f}</td>
        </tr>
        {f_tax_row_html}
        <tr style="background:#f7fafc">
          <td colspan="3" style="padding:9px 12px;border-bottom:1px solid #e2e8f0;color:#718096">Total Invoice</td>
          <td style="padding:9px 12px;text-align:right;border-bottom:1px solid #e2e8f0">${grand_total:.2f}</td>
        </tr>
        <tr style="background:#f7fafc">
          <td colspan="3" style="padding:9px 12px;border-bottom:1px solid #e2e8f0;color:#718096">Deposit Paid</td>
          <td style="padding:9px 12px;text-align:right;color:#276749;font-weight:600;border-bottom:1px solid #e2e8f0">- ${deposit_paid:.2f}</td>
        </tr>
        <tr style="background:#1a365d;color:white">
          <td colspan="3" style="padding:11px 12px;font-weight:700;font-size:1rem">REMAINING BALANCE DUE</td>
          <td style="padding:11px 12px;text-align:right;font-weight:700;font-size:1.2rem">${remaining_amount:.2f}</td>
        </tr>
      </tbody>
    </table>

    <!-- Payment section -->
    <div style="background:#fff8f3;border:2px solid #dd6b20;border-radius:12px;padding:1.5rem;margin-bottom:1.5rem;text-align:center">
      <p style="font-size:.82rem;color:#c05621;margin:0 0 .3rem;font-weight:700;text-transform:uppercase;letter-spacing:.6px">Final Payment Due Now</p>
      <p style="font-size:2.75rem;font-weight:800;color:#c05621;margin:.2rem 0 .8rem;line-height:1">${remaining_amount:.2f}</p>
      {pay_btn}
    </div>

    <!-- Urgency note -->
    <div style="background:#fff5f5;border-left:4px solid #e53e3e;padding:1rem 1.25rem;border-radius:0 8px 8px 0;margin-bottom:1.5rem;font-size:.88rem;color:#742a2a">
      <strong>Important:</strong> Failure to make final payment may result in your order being considered canceled.
      Please complete payment as soon as possible to guarantee your delivery.
    </div>

    <p style="color:#4a5568;line-height:1.7;font-size:.9rem">
      If you have any questions or need assistance, please don't hesitate to reach out.
      {f"You can call us at <strong>{BUSINESS_PHONE}</strong>." if BUSINESS_PHONE else ""}
      We look forward to making your event a success!
    </p>

    <!-- Rental Agreement -->
    <div style="border-top:2px solid #e2e8f0;padding-top:1.5rem;margin-top:1.5rem">
      <h3 style="color:#1a365d;font-size:.95rem;font-weight:700;margin:0 0 .4rem;text-transform:uppercase;letter-spacing:.5px">Rental Agreement</h3>
      <p style="font-size:.82rem;color:#718096;margin:0 0 1rem">
        As a reminder, your rental is subject to the following terms and conditions.
        By completing your final payment above, you confirm your agreement to all terms below.
        No additional signature is required.
      </p>
      <div style="background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:1.25rem">
        {contract_html}
      </div>
    </div>

    <p style="color:#2d3748;font-weight:600;margin-top:1.75rem">&mdash; The {BUSINESS_NAME} Team</p>
  </div>
</div></body></html>"""

    plain = f"""Hi {first},

YOUR DELIVERY IS IN 2 DAYS — FINAL PAYMENT REQUIRED

Booking #{b.get('id')} | {event_date}
Location: {event_addr}
Event Start Time: {event_time} | Delivery Time: {setup_time}

PAYMENT SUMMARY
  Total Invoice:    ${grand_total:.2f}
  Deposit Paid:   - ${deposit_paid:.2f}
{f_tax_row_plain}  ──────────────────────────
  REMAINING DUE:   ${remaining_amount:.2f}

PAY NOW: {payment_link if payment_link else 'Contact us to complete payment.'}

Failure to make final payment may result in your order being canceled.
{f"Questions? Call {BUSINESS_PHONE}" if BUSINESS_PHONE else ""}

────────────────────────────────────────────────────────
RENTAL AGREEMENT — TERMS & CONDITIONS
────────────────────────────────────────────────────────

NON-REFUNDABLE DEPOSIT AGREEMENT
Deposit is 25% and is NOT REFUNDABLE under any circumstances.
- Canceled within 20 days of event: 50% of all items charged.
- Canceled within 10 days of event: 75% of all items charged.
- Canceled within 24 hours of event: 100% charge applies.
- $75/hr fee if Rent a Party LLC is delayed due to Renter negligence.
- Remaining balance due 48 hours before event; failure to pay may result in cancellation.
- Person 18+ must be present at time of delivery.
- No refunds; inclement weather postponements receive store credit.

EQUIPMENT RENTAL TERMS
1. Equipment returned damaged (beyond normal wear) — Renter pays repair/replacement costs.
2. Late returns charged at $75/hour after specified return time.
3. Renter must have venue authorization to use equipment on premises.
4. Equipment may only be used for its intended purpose.
5. No additions, attachments, or alterations without prior written consent.
6. All equipment must be installed to allow removal without damage.
7. Chairs must be stacked with the black circle facing up; improper stacking = $1/chair fee.
8. Marquee items: keep dry, do not leave outside overnight, do not stand on them.
9. Electrical/battery items (speakers, microphones, etc.) must not be left outside overnight or exposed to moisture.
10. Water damage to any rental item = Renter responsible for full repair/replacement cost.
11. OVERNIGHT RENTALS: All equipment must be secured in a locked location overnight.
    Renter is fully responsible for all equipment until returned/picked up by Rent a Party, LLC.

By making final payment you confirm your agreement to all terms above.
No signature required — payment constitutes acceptance of this agreement.

— {BUSINESS_NAME}"""

    _send_email(email, subject, html, plain)


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN AUTHENTICATION
# ══════════════════════════════════════════════════════════════════════════════

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
#  HTML TEMPLATES
# ══════════════════════════════════════════════════════════════════════════════

FORM_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/manifest.json">
  <meta name="theme-color" content="#1a365d">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Book a Rental">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Book a Rental — {{ business_name }}</title>
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f4f8;color:#1a202c;min-height:100vh}
    header{background:linear-gradient(135deg,#1a365d 0%,#2b6cb0 100%);color:white;padding:2.5rem 1.5rem 3.5rem;text-align:center;position:relative;overflow:hidden}
    header::after{content:'';position:absolute;bottom:-1px;left:0;right:0;height:40px;background:#f0f4f8;clip-path:ellipse(55% 100% at 50% 100%)}
    header img{height:90px;width:auto;object-fit:contain;margin-bottom:.75rem;filter:drop-shadow(0 4px 12px rgba(0,0,0,.3));position:relative;z-index:1}
    header h1{font-size:2rem;font-weight:800;letter-spacing:-.5px;position:relative;z-index:1}
    header p{margin-top:.4rem;opacity:.85;font-size:1.05rem;position:relative;z-index:1}
    .container{max-width:740px;margin:0 auto;padding:2rem 1rem 5rem}
    .card{background:white;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08);padding:1.75rem;margin-bottom:1.5rem}
    .card h2{font-size:1rem;font-weight:700;color:#2b6cb0;border-bottom:2px solid #ebf4ff;padding-bottom:.6rem;margin-bottom:1.25rem;text-transform:uppercase;letter-spacing:.5px}
    .field{margin-bottom:1rem}
    .field label{display:block;font-size:.85rem;font-weight:600;color:#4a5568;margin-bottom:.3rem}
    .field input,.field select,.field textarea{width:100%;padding:.65rem .85rem;border:1.5px solid #cbd5e0;border-radius:8px;font-size:.97rem;color:#1a202c;background:#fff;transition:border-color .15s}
    .field input:focus,.field select:focus,.field textarea:focus{outline:none;border-color:#2b6cb0;box-shadow:0 0 0 3px rgba(43,108,176,.12)}
    .field textarea{resize:vertical;min-height:80px}
    .row{display:grid;gap:1rem;grid-template-columns:1fr 1fr}
    .row3{display:grid;gap:1rem;grid-template-columns:2fr 1fr 1fr}
    @media(max-width:560px){.row,.row3{grid-template-columns:1fr}}
    .required{color:#e53e3e}
    .section-note{font-size:.82rem;color:#718096;margin-bottom:1rem;font-style:italic}
    .type-toggle{display:flex;gap:.75rem;margin-bottom:1rem}
    .type-btn{flex:1;padding:.65rem;border:2px solid #cbd5e0;border-radius:8px;background:white;font-size:.9rem;font-weight:600;color:#718096;cursor:pointer;text-align:center;transition:all .15s}
    .type-btn.active{border-color:#2b6cb0;background:#ebf4ff;color:#2b6cb0}
    .exact-toggle{display:flex;align-items:center;gap:.75rem;padding:1rem;background:#fffaf0;border:2px solid #ed8936;border-radius:10px;cursor:pointer;margin-bottom:.75rem}
    .exact-toggle input[type=checkbox]{width:20px;height:20px;cursor:pointer;accent-color:#2b6cb0}
    .exact-label{flex:1}
    .exact-label strong{display:block;font-size:.97rem;color:#1a202c}
    .exact-label span{font-size:.82rem;color:#718096}
    .exact-badge{background:#ed8936;color:white;padding:.2rem .6rem;border-radius:20px;font-size:.8rem;font-weight:700}
    .product-row{display:grid;grid-template-columns:1fr auto auto;align-items:center;gap:.75rem;padding:.85rem 0;border-bottom:1px solid #f0f4f8}
    .product-row:last-child{border-bottom:none}
    .product-name{font-weight:600;font-size:.95rem}
    .product-meta{display:flex;gap:1rem;font-size:.8rem;margin-top:.15rem}
    .product-price{color:#718096}
    .avail-badge{font-weight:600}
    .avail-badge.ok{color:#38a169}
    .avail-badge.low{color:#d69e2e}
    .avail-badge.out{color:#e53e3e}
    .qty-control{display:flex;align-items:center;border:1.5px solid #cbd5e0;border-radius:8px;overflow:hidden}
    .qty-btn{background:#f7fafc;border:none;width:34px;height:36px;font-size:1.1rem;color:#2b6cb0;cursor:pointer;transition:background .12s}
    .qty-btn:hover{background:#ebf4ff}
    .qty-input{width:52px;border:none;border-left:1.5px solid #cbd5e0;border-right:1.5px solid #cbd5e0;text-align:center;font-size:.95rem;font-weight:600;padding:.4rem .2rem;outline:none}
    .product-sub{text-align:right;min-width:70px;font-weight:600;color:#718096;font-size:.95rem}
    .product-sub.has-val{color:#2b6cb0}
    .total-bar{background:linear-gradient(135deg,#1a365d,#2b6cb0);border-radius:12px;padding:1.25rem 1.75rem;color:white;margin-bottom:1.5rem}
    .total-row{display:flex;justify-content:space-between;padding:.2rem 0;font-size:.95rem;opacity:.85}
    .total-row.grand{font-size:1.35rem;font-weight:700;opacity:1;border-top:1px solid rgba(255,255,255,.25);margin-top:.5rem;padding-top:.6rem}
    .total-note{font-size:.78rem;opacity:.7;margin-top:.5rem;font-style:italic}
    .alert{background:#fff5f5;border:1px solid #feb2b2;color:#c53030;padding:.85rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem}
    .submit-btn{width:100%;padding:1rem;background:linear-gradient(135deg,#1a365d,#2b6cb0);color:white;border:none;border-radius:10px;font-size:1.1rem;font-weight:700;cursor:pointer;transition:opacity .15s}
    .submit-btn:hover{opacity:.92}
    .submit-btn:disabled{opacity:.6;cursor:not-allowed}
  </style>
</head>
<body>
<header>
  <img src="/logo.png" alt="{{ business_name }}" style="height:90px;width:auto;object-fit:contain;margin-bottom:.6rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,.25))">
  <h1>{{ business_name }}</h1>
  <p>Request a rental quote — we'll respond quickly!</p>
</header>
<div class="container">
{% if error %}<div class="alert">{{ error }}</div>{% endif %}
<form method="POST" action="/submit" id="bookingForm">
  <div class="card">
    <h2>Your Information</h2>
    <div class="row">
      <div class="field"><label>Full Name <span class="required">*</span></label><input name="full_name" required placeholder="Jane Smith" value="{{ form.full_name or '' }}"></div>
      <div class="field"><label>Company Name <span style="color:#718096;font-weight:400">(if applicable)</span></label><input name="company_name" placeholder="ABC Events LLC" value="{{ form.company_name or '' }}"></div>
    </div>
    <div class="field"><label>Street Address <span class="required">*</span></label><input name="renter_street" required placeholder="123 Main Street" value="{{ form.renter_street or '' }}"></div>
    <div class="row3">
      <div class="field"><label>City <span class="required">*</span></label><input name="renter_city" required placeholder="Hartford" value="{{ form.renter_city or '' }}"></div>
      <div class="field"><label>State <span class="required">*</span></label><input name="renter_state" required placeholder="CT" maxlength="2" value="{{ form.renter_state or '' }}"></div>
      <div class="field"><label>Zip <span class="required">*</span></label><input name="renter_zip" required placeholder="06101" value="{{ form.renter_zip or '' }}"></div>
    </div>
    <div class="row">
      <div class="field"><label>Phone <span class="required">*</span></label><input name="phone" type="tel" required placeholder="(555) 000-0000" value="{{ form.phone or '' }}"></div>
      <div class="field"><label>Email <span class="required">*</span></label><input name="email" type="email" required placeholder="jane@email.com" value="{{ form.email or '' }}"></div>
    </div>
    <div style="margin-top:1rem;padding:.75rem 1rem;background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px;display:flex;align-items:flex-start;gap:.75rem">
      <input type="checkbox" name="tax_exempt_request" id="tax_exempt_request" value="1" onchange="updateTotals()" style="width:18px;height:18px;margin-top:.15rem;accent-color:#16a34a;cursor:pointer;flex-shrink:0">
      <label for="tax_exempt_request" style="cursor:pointer;font-size:.9rem;color:#166534;font-weight:600;line-height:1.4">
        I have a Connecticut Tax-Exempt Certificate
        <span style="display:block;font-weight:400;font-size:.8rem;color:#4b7c5a;margin-top:.1rem">Check this if your organization is tax-exempt. You will need to provide your certificate number to the rental office.</span>
      </label>
    </div>
  </div>

  <div class="card">
    <h2>Event Details</h2>

    {%- set time_opts -%}
    <option value="">-- Select --</option>
    <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
    <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
    <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
    <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
    <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
    <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
    <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
    <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
    <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
    <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
    <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
    <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
    <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
    <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
    <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
    <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
    <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
    <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
    {%- endset -%}

    <div class="row">
      <div class="field"><label>Event Start Date <span class="required">*</span></label><input id="event_start_date" name="event_start_date" type="date" required onchange="onDateChange()" value="{{ form.event_start_date or '' }}"></div>
      <div class="field"><label>Event Start Time <span class="required">*</span></label>
        <select name="event_start_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field"><label>Event End Date <span class="required">*</span></label><input id="event_end_date" name="event_end_date" type="date" required onchange="onDateChange()" value="{{ form.event_end_date or '' }}"></div>
      <div class="field"><label>Event End Time <span class="required">*</span></label>
        <select name="event_end_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field"><label>Setup Date <span class="required">*</span></label><input name="setup_date" type="date" required value="{{ form.setup_date or '' }}" id="setupDateEl" onchange="checkDeliveryBeforeEvent()"></div>
      <div class="field"><label>Setup Time <span class="required">*</span></label>
        <select name="setup_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>


    <!-- Early delivery acknowledgment — shown only when delivery date is before event date -->
    <div id="early-delivery-notice" style="display:none;background:#fff7ed;border:2px solid #f97316;border-radius:10px;padding:1rem 1.15rem;margin-bottom:1rem">
      <div style="display:flex;align-items:flex-start;gap:.75rem">
        <span style="font-size:1.5rem;line-height:1">📦</span>
        <div style="flex:1">
          <div style="font-weight:700;color:#c2410c;font-size:1rem;margin-bottom:.35rem">Your delivery is scheduled BEFORE your event date</div>
          <div style="font-size:.9rem;color:#7c2d12;line-height:1.5;margin-bottom:.75rem">
            Your event starts on <strong id="notice-event-date"></strong>, but you requested delivery on <strong id="notice-delivery-date"></strong>.
          </div>
          <label style="display:flex;align-items:flex-start;gap:.6rem;cursor:pointer;background:#ffedd5;border:1px solid #fb923c;border-radius:7px;padding:.65rem .85rem">
            <input type="checkbox" id="early_delivery_ack" name="early_delivery_ack" value="1"
              onchange="checkDeliveryAck()"
              style="width:20px;height:20px;margin-top:.1rem;accent-color:#ea580c;flex-shrink:0;cursor:pointer">
            <span style="font-size:.88rem;font-weight:600;color:#9a3412;line-height:1.45">
              I understand that my rental items will be delivered on <strong id="notice-ack-date"></strong> — before my event — and I approve this early delivery.
            </span>
          </label>
        </div>
      </div>
    </div>

    <div class="field">
      <label>Venue Type <span class="required">*</span></label>
      <div class="type-toggle">
        <div class="type-btn active" id="btn_venue" onclick="setVenue('venue')">Venue</div>
        <div class="type-btn" id="btn_residential" onclick="setVenue('residential')">Residential</div>
      </div>
      <input type="hidden" name="venue_type" id="venue_type_input" value="venue">
    </div>
    <div id="venue_pickup_row" class="field"><label>Latest Pickup Time at Venue <span class="required">*</span></label>
      <select id="venue_latest_pickup" name="venue_latest_pickup" style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
        {{ time_opts }}
      </select>
    </div>
  </div>

  <div class="card">
    <h2>Event Address</h2>
    <p class="section-note">Where will we deliver your rental items?</p>
    <div class="field"><label>Street Address <span class="required">*</span></label><input id="event_street" name="event_street" required placeholder="456 Venue Blvd" value="{{ form.event_street or '' }}" oninput="scheduleDistanceCalc()"></div>
    <div class="row3">
      <div class="field"><label>City <span class="required">*</span></label><input id="event_city" name="event_city" required placeholder="Hartford" value="{{ form.event_city or '' }}" oninput="scheduleDistanceCalc()"></div>
      <div class="field"><label>State <span class="required">*</span></label><input id="event_state" name="event_state" required placeholder="CT" maxlength="2" value="{{ form.event_state or '' }}" oninput="scheduleDistanceCalc()"></div>
      <div class="field"><label>Zip <span class="required">*</span></label><input id="event_zip" name="event_zip" required placeholder="06101" value="{{ form.event_zip or '' }}" oninput="scheduleDistanceCalc()"></div>
    </div>
  </div>

  <div class="card">
    <h2>Delivery Options</h2>
    <label class="exact-toggle">
      <input type="checkbox" id="exact_time_cb" name="exact_time_delivery" value="yes" onchange="updateTotals();toggleExactTimes(this.checked)">
      <div class="exact-label"><strong>Exact Time Delivery</strong><span>Guaranteed delivery at your specified setup time</span></div>
      <span class="exact-badge">+$175</span>
    </label>

    <div id="exact_time_fields" style="display:none;margin-top:1rem;background:#fff8f0;border:1px solid #f6ad55;border-radius:10px;padding:1rem 1.1rem">
      <p style="margin:0 0 .75rem;font-size:.88rem;color:#92400e;font-weight:600">⏰ Exact Time Details</p>
      <div class="row">
        <div class="field">
          <label>Exact Delivery Time <span class="required">*</span></label>
          <select name="exact_delivery_time" id="exact_delivery_time" style="width:100%;border:1px solid #f6ad55;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
            <option value="">-- Select delivery time --</option>
            <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
            <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
            <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
            <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
            <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
            <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
            <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
            <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
            <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
            <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
            <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
            <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
            <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
            <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
            <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
            <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
            <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
            <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
          </select>
        </div>
        <div class="field">
          <label>Exact Pickup Time <span class="required">*</span></label>
          <select name="exact_pickup_time" id="exact_pickup_time" style="width:100%;border:1px solid #f6ad55;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
            <option value="">-- Select pickup time --</option>
            <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
            <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
            <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
            <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
            <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
            <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
            <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
            <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
            <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
            <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
            <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
            <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
            <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
            <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
            <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
            <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
            <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
            <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
          </select>
        </div>
      </div>
    </div>

    <script>
    function toggleExactTimes(checked) {
      var el = document.getElementById('exact_time_fields');
      el.style.display = checked ? 'block' : 'none';
      document.getElementById('exact_delivery_time').required = checked;
      document.getElementById('exact_pickup_time').required = checked;
    }
    </script>

    <div class="field" style="margin-top:1rem"><label>Where on the premises will items be delivered? <span class="required">*</span></label><textarea name="delivery_location" required placeholder="e.g. Through the main entrance, set up in the ballroom on the left side...">{{ form.delivery_location or '' }}</textarea></div>
  </div>

  <!-- Hidden qty inputs — submitted with form -->
  {% for p in products %}
  <input type="hidden" class="qty-input" id="qty_{{ p.id }}" name="qty_{{ p.id }}" value="0" data-price="{{ p.price }}" data-max="{{ p.total }}">
  {% endfor %}

  <!-- Product data for JS -->
  <script>
  const ALL_PRODUCTS = [
    {% for p in products %}
    { id:"{{ p.id }}", name:{{ p.name | tojson }}, price:{{ p.price }}, max:{{ p.total }} },
    {% endfor %}
  ];
  </script>

  <div class="card">
    <h2>Select Your Items</h2>
    <p style="color:#6b7280;font-size:.88rem;margin-bottom:1.25rem">Click a category below to browse items. Select an item to add it to your order.</p>

    <!-- Category accordion dropdowns -->
    <div id="category-dropdowns"></div>

    <!-- Selected items list -->
    <div id="selected-items-wrap" style="display:none;margin-top:1.25rem;border-top:2px solid #e5e7eb;padding-top:1rem">
      <div style="font-weight:700;font-size:.95rem;color:#1a202c;margin-bottom:.75rem">🛒 Your Items</div>
      <div id="marquee-tier-notice" style="display:none;background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.88rem;color:#1e40af"></div>
      <div id="selected-items-list"></div>
    </div>
  </div>

  <!-- Stackable Marquee Option — shown only when marquee items are in cart -->
  <div id="stackable-section" style="display:none;margin:.75rem 0;padding:1.1rem 1.25rem;background:#faf5ff;border:1.5px solid #d8b4fe;border-radius:10px">
    <div style="font-weight:700;font-size:.95rem;color:#6b21a8;margin-bottom:.5rem">🔡 Stackable Marquee</div>
    <p style="font-size:.88rem;color:#4c1d95;margin-bottom:.85rem;line-height:1.6">
      Would you like your marquee letters/numbers to be <strong>stackable</strong> (stacked on top of each other)?
      <span style="background:#ede9fe;color:#7c3aed;border-radius:4px;padding:.1rem .45rem;font-size:.8rem;font-weight:700;margin-left:.3rem">+$75 fee</span>
    </p>
    <div style="display:flex;gap:1rem;margin-bottom:.85rem">
      <label style="display:flex;align-items:center;gap:.4rem;cursor:pointer;font-weight:600;color:#1a202c">
        <input type="radio" name="stackable_choice" value="yes" onchange="onStackableChange()"
          style="accent-color:#7c3aed;width:16px;height:16px"> Yes, I want stackable (+$75)
      </label>
      <label style="display:flex;align-items:center;gap:.4rem;cursor:pointer;font-weight:600;color:#1a202c">
        <input type="radio" name="stackable_choice" value="no" onchange="onStackableChange()" checked
          style="accent-color:#7c3aed;width:16px;height:16px"> No thanks
      </label>
    </div>
    <div id="stackable-top-wrap" style="display:none">
      <label style="font-size:.88rem;font-weight:600;color:#1a202c;display:block;margin-bottom:.35rem">
        Which letters or numbers go on top? <span style="color:#dc2626">*</span>
      </label>
      <input type="text" id="stackable_top_display" placeholder="e.g. A, E, O — or 2, 5"
        oninput="document.getElementById('stackable_top_input').value=this.value"
        style="width:100%;border:1.5px solid #c4b5fd;border-radius:8px;padding:.5rem .75rem;font-size:.9rem;box-sizing:border-box">
    </div>
  </div>
  <input type="hidden" name="stackable" id="stackable_input" value="no">
  <input type="hidden" name="stackable_top" id="stackable_top_input" value="">

  <input type="hidden" name="late_night_fee" id="late_night_fee_input" value="0">
  <div id="late_night_notice" style="display:none;margin:.75rem 0;padding:.75rem 1rem;background:#fef3c7;border:1.5px solid #fcd34d;border-radius:8px;font-size:.88rem;color:#92400e">
    <strong>⏰ Late Night / Early Morning Fee: $125.00</strong><br>
    Your pickup or dropoff time falls between 11:30 PM – 7:00 AM. A $125 fee applies for pickups or deliveries outside of standard hours.
  </div>
  <div class="total-bar">
    <div class="total-row"><span>Items Subtotal</span><span id="t_items">$0.00</span></div>
    <div class="total-row"><span>Exact Time Delivery</span><span id="t_exact">-</span></div>
    <div class="total-row" id="t_stackable_row" style="display:none"><span>Stackable Marquee</span><span id="t_stackable">$75.00</span></div>
    <div class="total-row" id="t_latenight_row" style="display:none"><span>Late Night / Early Morning Fee</span><span id="t_latenight">$125.00</span></div>
    <div class="total-row"><span>Delivery Fee</span><span id="t_delivery">Calculated after review</span></div>
    <div class="total-row"><span>CT Sales Tax (6.35%)</span><span id="t_tax">$0.00</span></div>
    <div class="total-row grand"><span>Estimated Total</span><span id="t_grand">$0.00</span></div>
    <p class="total-note">Final delivery fee and tax confirmed after we review your address. This is a quote request, not a charge.</p>
  </div>

  <div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1.25rem">
    <p style="margin:0;font-size:.9rem;color:#78350f;line-height:1.7">
      <strong>📋 Please note:</strong> This is a quote request, not a charge. The initial quote may not reflect the exact final price — pricing may vary based on your event details.
      Any differences will be clearly explained in the invoice we send you before any payment is required.
    </p>
  </div>

  <button type="submit" class="submit-btn" id="submitBtn">Send Quote Request</button>
</form>
</div>
<script>
const EXACT_FEE = {{ exact_time_fee }};
const LATE_NIGHT_FEE = 125;
const CT_TAX_RATE = 0.0635;
function isLateNight(timeStr){
  if(!timeStr) return false;
  const [h,m]=timeStr.split(':').map(Number);
  const mins=h*60+m;
  return mins>=1410||mins<420;
}
function checkLateNightFee(){
  const endTime=document.querySelector('[name="event_end_time"]').value;
  const late=isLateNight(endTime);
  document.getElementById('late_night_notice').style.display=late?'block':'none';
  document.getElementById('t_latenight_row').style.display=late?'flex':'none';
  document.getElementById('late_night_fee_input').value=late?LATE_NIGHT_FEE:0;
  return late?LATE_NIGHT_FEE:0;
}

// ── Marquee Tier Pricing ──────────────────────────────────────────
const MARQUEE_NUMBER_TIERS = [
  { count:1, total:80 },
  { count:2, total:150 },
  { count:3, total:215 },
  { count:4, total:275 },
];
function getMarqueeNumberTotal(n){
  if(n<=0) return 0;
  const t=MARQUEE_NUMBER_TIERS.find(x=>x.count===n);
  if(t) return t.total;
  return 275+(n-4)*55;
}
const MARQUEE_LETTER_TIERS = [
  { count:1, total:85 },
  { count:2, total:160 },
  { count:3, total:225 },
  { count:4, total:285 },
];
function getMarqueeLetterTotal(n){
  if(n<=0) return 0;
  const t=MARQUEE_LETTER_TIERS.find(x=>x.count===n);
  if(t) return t.total;
  return 285+(n-4)*55;
}
function isMarqueeNumber(name){ return /^marquee\s+#?\d/i.test(name); }
function isMarqueeLetter(name){ return /^marquee\s+[a-z]$/i.test(name); }

// ── Item Categories ───────────────────────────────────────────────
const ITEM_CATEGORIES = [
  { label:"🪑 Chairs",           keywords:["chair","stool","bench","seat","chiavari"] },
  { label:"🪣 Tables",           keywords:["table","tablecloth","linen","cloth","runner","overlay","skirt"] },
  { label:"🔢 Marquee Numbers",  keywords:["marquee number"] },
  { label:"🔤 Marquee Letters",  keywords:["marquee letter"] },
  { label:"💡 Lighting",         keywords:["light","lamp","led","glow","neon","bulb","lantern","fairy","chandelier","uplighting"] },
  { label:"🎭 Backdrops & Décor",keywords:["backdrop","banner","balloon","arch","flower","floral","decor","sign","drape","curtain","pillar","column","centerpiece","vase","frame","wall"] },
  { label:"🎪 Entertainment",    keywords:["bounce","slide","game","popcorn","cotton candy","machine","photo booth","casino","carnival","inflatable"] },
  { label:"⛺ Tents & Canopies", keywords:["tent","canopy","pergola","gazebo","umbrella"] },
];
function getCat(name){
  const n=name.trim();
  // Marquee Letter: "Marquee A", "Marquee B", etc. — marquee followed by a single letter
  if(/^marquee\s+[a-zA-Z]$/i.test(n) || /marquee\s+[a-zA-Z]\s*$/i.test(n)) return "🔤 Marquee Letters";
  // Marquee Number: "Marquee #5", "Marquee 3", etc. — marquee followed by # or digit
  if(/marquee\s+#?\d/i.test(n)) return "🔢 Marquee Numbers";
  const nl=n.toLowerCase();
  for(const c of ITEM_CATEGORIES){ if(c.keywords.some(k=>nl.includes(k))) return c.label; }
  return "📦 Other";
}

// ── Build Category Accordion Dropdowns ───────────────────────────
function buildDropdowns(){
  // Group products by category
  const groups={};
  ALL_PRODUCTS.forEach(p=>{
    const cat=getCat(p.name);
    if(!groups[cat]) groups[cat]=[];
    groups[cat].push(p);
  });
  const wrap=document.getElementById('category-dropdowns');
  wrap.innerHTML='';
  const CAT_ORDER=['🪑 Chairs','🪣 Tables','🔤 Marquee Letters','🔢 Marquee Numbers'];
  const sorted=Object.entries(groups).sort(([a],[b])=>{
    if(a==='📦 Other') return 1;
    if(b==='📦 Other') return -1;
    const ai=CAT_ORDER.indexOf(a), bi=CAT_ORDER.indexOf(b);
    if(ai>=0&&bi>=0) return ai-bi;
    if(ai>=0) return -1;
    if(bi>=0) return 1;
    return a.localeCompare(b);
  });
  sorted.forEach(([cat,items])=>{
    const sec=document.createElement('div');
    sec.style.cssText='border:1px solid #e5e7eb;border-radius:8px;margin-bottom:.5rem;overflow:hidden';
    sec.innerHTML=`
      <button type="button" onclick="toggleCat(this)"
        style="width:100%;text-align:left;background:#f9fafb;border:none;padding:.75rem 1rem;font-size:.95rem;font-weight:700;color:#1a202c;cursor:pointer;display:flex;justify-content:space-between;align-items:center">
        <span>${cat} <span style="font-size:.8rem;font-weight:400;color:#9ca3af">(${items.length} item${items.length!==1?'s':''})</span></span>
        <span class="chev" style="transition:transform .2s;font-size:.8rem">▼</span>
      </button>
      <div class="cat-body" style="display:none;padding:.75rem 1rem;background:white">
        ${cat==='🔢 Marquee Numbers'?`<div style="background:#fefce8;border:1px solid #fde047;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.83rem;color:#713f12">
          <strong>📋 Tier Pricing:</strong> 1 for $80 &nbsp;·&nbsp; 2 for $150 &nbsp;·&nbsp; 3 for $215 &nbsp;·&nbsp; 4 for $275 &nbsp;·&nbsp; 5+ = $275 + $55 each additional
        </div>`:''}
        ${cat==='🔤 Marquee Letters'?`<div style="background:#fefce8;border:1px solid #fde047;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.83rem;color:#713f12">
          <strong>📋 Tier Pricing:</strong> 1 for $85 &nbsp;·&nbsp; 2 for $160 &nbsp;·&nbsp; 3 for $225 &nbsp;·&nbsp; 4 for $285 &nbsp;·&nbsp; 5+ = $285 + $55 each additional
        </div>`:''}
        <div style="display:flex;flex-wrap:wrap;gap:.5rem">
          ${items.map(p=>`
            <button type="button" onclick="addToCart('${p.id}')"
              data-id="${p.id}"
              style="background:#eff6ff;color:#1d4ed8;border:1.5px solid #bfdbfe;border-radius:20px;padding:.35rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer;transition:background .15s"
              onmouseover="this.style.background='#dbeafe'" onmouseout="this.style.background='#eff6ff'">
              ${p.name} — $${p.price.toFixed(2)}
            </button>`).join('')}
        </div>
      </div>`;
    wrap.appendChild(sec);
  });
}
function toggleCat(btn){
  const body=btn.nextElementSibling;
  const open=body.style.display==='block';
  body.style.display=open?'none':'block';
  btn.querySelector('.chev').style.transform=open?'':'rotate(180deg)';
  btn.style.background=open?'#f9fafb':'#eff6ff';
}

// ── Cart ──────────────────────────────────────────────────────────
const cart={};  // id -> qty
function addToCart(id){
  const p=ALL_PRODUCTS.find(x=>x.id===id);
  if(!p) return;
  if(cart[id]){ setQty(id, cart[id]+1); }
  else { cart[id]=1; renderCart(); }
  setHiddenQty(id, cart[id]);
  updateTotals();
  // Highlight the button
  const btn=document.querySelector(`button[data-id="${id}"]`);
  if(btn){ btn.style.background='#bbf7d0'; btn.style.borderColor='#4ade80'; btn.style.color='#166534';
    setTimeout(()=>{ btn.style.background='#eff6ff'; btn.style.borderColor='#bfdbfe'; btn.style.color='#1d4ed8'; },600); }
}
function setQty(id, val){
  const p=ALL_PRODUCTS.find(x=>x.id===id);
  if(!p) return;
  const v=Math.max(0, val);
  if(v===0){ delete cart[id]; }
  else { cart[id]=v; }
  const inp=document.getElementById('cart-qty-'+id);
  if(inp) inp.value=v===0?'':v;
  setHiddenQty(id, v);
  renderCart();
  updateTotals();
}
function setHiddenQty(id, val){
  const h=document.getElementById('qty_'+id);
  if(h) h.value=val;
}
function removeFromCart(id){
  delete cart[id];
  setHiddenQty(id,0);
  renderCart();
  updateTotals();
}
const STACKABLE_FEE = 75;
function hasMarqueeInCart(){
  return Object.keys(cart).some(id=>{
    const p=ALL_PRODUCTS.find(x=>x.id===id);
    return p && (isMarqueeLetter(p.name)||isMarqueeNumber(p.name));
  });
}
function onStackableChange(){
  const sel=document.querySelector('input[name="stackable_choice"]:checked');
  const yes=sel&&sel.value==='yes';
  document.getElementById('stackable_input').value=yes?'yes':'no';
  document.getElementById('stackable-top-wrap').style.display=yes?'block':'none';
  if(!yes){ document.getElementById('stackable_top_input').value=''; const d=document.getElementById('stackable_top_display'); if(d) d.value=''; }
  updateTotals();
}
function renderCart(){
  const list=document.getElementById('selected-items-list');
  const wrap=document.getElementById('selected-items-wrap');
  const ids=Object.keys(cart);
  if(ids.length===0){
    wrap.style.display='none'; list.innerHTML='';
    // Hide stackable section and reset when cart is empty
    const ss=document.getElementById('stackable-section');
    if(ss) ss.style.display='none';
    document.getElementById('stackable_input').value='no';
    const noRadio=document.querySelector('input[name="stackable_choice"][value="no"]');
    if(noRadio){ noRadio.checked=true; }
    document.getElementById('stackable-top-wrap').style.display='none';
    return;
  }
  wrap.style.display='block';
  // Show stackable section only when marquee items are in cart
  const ss=document.getElementById('stackable-section');
  if(ss) ss.style.display=hasMarqueeInCart()?'block':'none';
  if(!hasMarqueeInCart()){
    document.getElementById('stackable_input').value='no';
    const noRadio=document.querySelector('input[name="stackable_choice"][value="no"]');
    if(noRadio){ noRadio.checked=true; }
    document.getElementById('stackable-top-wrap').style.display='none';
  }
  // Calculate total marquee numbers for proration
  let mnCount=0, mlCount=0;
  ids.forEach(id=>{ const p=ALL_PRODUCTS.find(x=>x.id===id); if(!p) return;
    if(isMarqueeNumber(p.name)) mnCount+=cart[id]||0;
    else if(isMarqueeLetter(p.name)) mlCount+=cart[id]||0;
  });
  const mnTierTotal=getMarqueeNumberTotal(mnCount);
  const mlTierTotal=getMarqueeLetterTotal(mlCount);
  const mnUnitPrice=mnCount>0?(mnTierTotal/mnCount):0;
  const mlUnitPrice=mlCount>0?(mlTierTotal/mlCount):0;
  list.innerHTML=ids.map(id=>{
    const p=ALL_PRODUCTS.find(x=>x.id===id);
    const q=cart[id]||1;
    const isMN=isMarqueeNumber(p.name);
    const isML=isMarqueeLetter(p.name);
    const unitPrice=isMN?mnUnitPrice:isML?mlUnitPrice:p.price;
    const lineTotal=(unitPrice*q).toFixed(2);
    const tierTag=(isMN||isML)?` <span style="font-size:.75rem;color:#2563eb;font-weight:600">(tier)</span>`:'';
    const unitLabel=`$${unitPrice.toFixed(2)} ea${tierTag}`;
    return `<div style="display:flex;align-items:center;gap:.75rem;padding:.6rem .5rem;border-bottom:1px solid #f3f4f6">
      <span style="flex:1;font-size:.92rem;font-weight:600;color:#1a202c">${p.name}</span>
      <span style="font-size:.82rem;color:#6b7280;white-space:nowrap">${unitLabel}</span>
      <div style="display:flex;align-items:center;gap:.3rem">
        <button type="button" onclick="setQty('${id}',${q-1})"
          style="width:28px;height:28px;border:1px solid #d1d5db;border-radius:6px;background:white;font-size:1rem;cursor:pointer;line-height:1">−</button>
        <input id="cart-qty-${id}" type="number" value="${q}" min="1" max="9999"
          onchange="setQty('${id}',parseInt(this.value)||1)"
          style="width:44px;text-align:center;border:1px solid #d1d5db;border-radius:6px;padding:.2rem;font-size:.9rem;font-weight:700">
        <button type="button" onclick="setQty('${id}',${q+1})"
          style="width:28px;height:28px;border:1px solid #d1d5db;border-radius:6px;background:white;font-size:1rem;cursor:pointer;line-height:1">+</button>
      </div>
      <span style="font-size:.9rem;font-weight:700;color:#2563eb;min-width:52px;text-align:right">$${lineTotal}</span>
      <button type="button" onclick="removeFromCart('${id}')"
        style="background:none;border:none;color:#9ca3af;font-size:1.1rem;cursor:pointer;padding:.1rem .3rem" title="Remove">✕</button>
    </div>`;
  }).join('');
}
function updateTotals(){
  let regularSub=0, mnCount=0, mlCount=0;
  ALL_PRODUCTS.forEach(p=>{
    const qty=cart[p.id]||0;
    if(!qty) return;
    if(isMarqueeNumber(p.name)) mnCount+=qty;
    else if(isMarqueeLetter(p.name)) mlCount+=qty;
    else regularSub+=qty*p.price;
  });
  const mnSub=getMarqueeNumberTotal(mnCount);
  const mlSub=getMarqueeLetterTotal(mlCount);
  const sub=regularSub+mnSub+mlSub;
  // Update tier notices in cart
  const tierEl=document.getElementById('marquee-tier-notice');
  if(tierEl){
    let html='';
    if(mnCount>0){
      const nextN=MARQUEE_NUMBER_TIERS.find(t=>t.count===mnCount+1);
      const savN=nextN?` <span style="color:#16a34a;font-size:.82rem">· Add 1 more for $${nextN.total} total</span>`:'';
      html+=`🔢 <strong>${mnCount} Marquee Number${mnCount!==1?'s':''}</strong> — Tier Price: <strong>$${mnSub.toFixed(2)}</strong>${savN}<br>`;
    }
    if(mlCount>0){
      const nextL=MARQUEE_LETTER_TIERS.find(t=>t.count===mlCount+1);
      const savL=nextL?` <span style="color:#16a34a;font-size:.82rem">· Add 1 more for $${nextL.total} total</span>`:'';
      html+=`🔤 <strong>${mlCount} Marquee Letter${mlCount!==1?'s':''}</strong> — Tier Price: <strong>$${mlSub.toFixed(2)}</strong>${savL}`;
    }
    tierEl.innerHTML=html;
    tierEl.style.display=(mnCount>0||mlCount>0)?'block':'none';
  }
  const exactCb=document.getElementById('exact_time_cb');
  const ef=exactCb&&exactCb.checked?EXACT_FEE:0;
  const lf=checkLateNightFee();
  const stackableSel=document.querySelector('input[name="stackable_choice"]:checked');
  const sf=(stackableSel&&stackableSel.value==='yes'&&hasMarqueeInCart())?STACKABLE_FEE:0;
  const stackRow=document.getElementById('t_stackable_row');
  if(stackRow) stackRow.style.display=sf>0?'flex':'none';
  const exemptCb=document.getElementById('tax_exempt_request');
  const exempt=exemptCb&&exemptCb.checked;
  const df=typeof _calcDeliveryFee!=='undefined'?_calcDeliveryFee:0;
  const tax=exempt?0:(sub+ef+sf+lf+df)*CT_TAX_RATE;
  const taxEl=document.getElementById('t_tax');
  if(taxEl){ taxEl.textContent='$'+tax.toFixed(2); taxEl.style.color=exempt?'#16a34a':'';
    const lbl=taxEl.previousElementSibling; if(lbl) lbl.textContent=exempt?'CT Sales Tax (EXEMPT)':'CT Sales Tax (6.35%)'; }
  document.getElementById('t_items').textContent='$'+sub.toFixed(2);
  document.getElementById('t_exact').textContent=ef>0?'$'+ef.toFixed(2):'-';
  document.getElementById('t_grand').textContent='$'+(sub+ef+sf+lf+df+tax).toFixed(2)+'+';
}
document.addEventListener('DOMContentLoaded', buildDropdowns);
function setVenue(type){document.getElementById('venue_type_input').value=type;document.getElementById('btn_venue').classList.toggle('active',type==='venue');document.getElementById('btn_residential').classList.toggle('active',type==='residential');const row=document.getElementById('venue_pickup_row');const inp=document.getElementById('venue_latest_pickup');row.style.display=type==='venue'?'block':'none';inp.required=type==='venue';}
setVenue('venue');
function onDateChange(){const start=document.getElementById('event_start_date').value;const end=document.getElementById('event_end_date').value;if(!start||!end||end<start)return;fetch('/availability?start='+start+'&end='+end).then(r=>r.json()).then(data=>{ALL_PRODUCTS.forEach(p=>{if(data[p.id]!==undefined){p.max=data[p.id];}});updateTotals();}).catch(()=>{});checkDeliveryBeforeEvent();}
function fmtDateNice(d){if(!d)return'';const p=d.split('-');if(p.length!==3)return d;const months=['January','February','March','April','May','June','July','August','September','October','November','December'];return months[parseInt(p[1],10)-1]+' '+parseInt(p[2],10)+', '+p[0];}
function checkDeliveryBeforeEvent(){
  const ed=document.getElementById('event_start_date').value;
  const sd=document.getElementById('setupDateEl').value;
  const notice=document.getElementById('early-delivery-notice');
  const ack=document.getElementById('early_delivery_ack');
  if(!notice)return;
  if(sd&&ed&&sd<ed){
    document.getElementById('notice-event-date').textContent=fmtDateNice(ed);
    document.getElementById('notice-delivery-date').textContent=fmtDateNice(sd);
    document.getElementById('notice-ack-date').textContent=fmtDateNice(sd);
    // Calculate days before
    const diff=Math.round((new Date(ed)-new Date(sd))/(1000*60*60*24));
    document.getElementById('notice-days-before').textContent=diff===1?'1 day':diff+' days';
    notice.style.display='';
  } else {
    notice.style.display='none';
    if(ack)ack.checked=false;
  }
}
function checkDeliveryAck(){
  // no-op, submit handler checks state
}
function applyWeekendDelivery(){
  const startDate = document.getElementById('event_start_date');
  if (!startDate || !startDate.value) {
    alert('Please enter the Event Start Date first.');
    return;
  }
  const esd = new Date(startDate.value + 'T00:00:00');
  const wd = esd.getDay(); // 0=Sun, 6=Sat
  let deliveryDate, pickupDate, label;
  if (wd === 6) { // Saturday
    deliveryDate = new Date(esd); deliveryDate.setDate(esd.getDate() - 1);
    pickupDate   = new Date(esd); pickupDate.setDate(esd.getDate() + 1);
    label = 'Saturday event → Deliver Friday, Pickup Sunday';
  } else if (wd === 0) { // Sunday
    deliveryDate = new Date(esd); deliveryDate.setDate(esd.getDate() - 2);
    pickupDate   = new Date(esd); pickupDate.setDate(esd.getDate() + 1);
    label = 'Sunday event → Deliver Friday, Pickup Monday';
  } else {
    alert('Event start date is not a Saturday or Sunday. Weekend schedule only applies to weekend events.');
    return;
  }
  const fmt = d => d.toISOString().split('T')[0];
  const ddEl = document.getElementById('deliveryDateEl');
  if (ddEl) ddEl.value = fmt(deliveryDate);
  const dtEl = document.getElementById('deliveryTimeEl');
  if (dtEl) { for (let o of dtEl.options) { if (o.value === '16:00') { dtEl.value = '16:00'; break; } } }
  const endDateEl = document.getElementById('event_end_date');
  if (endDateEl && !endDateEl.value) endDateEl.value = fmt(pickupDate);
  const msg = document.getElementById('weekend-msg');
  if (msg) { msg.textContent = '✓ ' + label; msg.style.display = 'inline'; }
}
let distTimer;
let _calcDeliveryFee=0;
function scheduleDistanceCalc(){clearTimeout(distTimer);distTimer=setTimeout(()=>{const street=document.getElementById('event_street').value;const city=document.getElementById('event_city').value;const state=document.getElementById('event_state').value;const zip=document.getElementById('event_zip').value;if(street&&city&&state&&zip){const addr=street+', '+city+', '+state+' '+zip;fetch('/delivery_fee?address='+encodeURIComponent(addr)).then(r=>r.json()).then(d=>{document.getElementById('t_delivery').textContent='$'+d.fee.toFixed(2)+' ('+d.note+')';_calcDeliveryFee=d.fee;updateTotals();}).catch(()=>{});}},800);}
// ── Date validation ───────────────────────────────────────────────────────
const today=new Date().toISOString().split('T')[0];
const startDateEl = document.getElementById('event_start_date');
const endDateEl   = document.getElementById('event_end_date');
const startTimeEl = document.querySelector('[name="event_start_time"]');
const endTimeEl   = document.querySelector('[name="event_end_time"]');
const setupTimeEl = document.querySelector('[name="setup_time"]');

startDateEl.min = today;
endDateEl.min   = today;

// When start date changes, end date must be >= start date
startDateEl.addEventListener('change', function() {
  endDateEl.min = this.value;
  if (endDateEl.value && endDateEl.value < this.value) {
    endDateEl.value = this.value;
  }
  onDateChange();
});

endDateEl.addEventListener('change', function() {
  if (this.value < startDateEl.value) {
    this.value = startDateEl.value;
    showTimeError('End date cannot be before start date.');
  }
  onDateChange();
});

// When start time changes:
//   - end time rule only applies if start date == end date (same-day event)
//   - delivery (setup) time must always be before start time (delivery is always same day as start)
startTimeEl.addEventListener('change', function() {
  const sameDay = startDateEl.value && endDateEl.value && startDateEl.value === endDateEl.value;
  if (sameDay && endTimeEl.value && endTimeEl.value <= this.value) {
    endTimeEl.value = '';
    showTimeError('Event end time must be after start time.');
  }
  if (setupTimeEl.value && setupTimeEl.value >= this.value) {
    setupTimeEl.value = '';
    showTimeError('Setup time must be before event start time.');
  }
});

endTimeEl.addEventListener('change', function() {
  const sameDay = startDateEl.value && endDateEl.value && startDateEl.value === endDateEl.value;
  if (sameDay && startTimeEl.value && this.value <= startTimeEl.value) {
    this.value = '';
    showTimeError('Event end time must be after the start time.');
  }
  updateTotals();
});

setupTimeEl.addEventListener('change', function() {
  const setupDate = document.getElementById('setupDateEl');
  const sameDay = setupDate && setupDate.value && startDateEl.value && setupDate.value === startDateEl.value;
  if (sameDay && startTimeEl.value && this.value >= startTimeEl.value) {
    this.value = '';
    showTimeError('Setup / delivery time must be before the event start time when delivery is on the same day.');
  }
});

function showTimeError(msg) {
  let el = document.getElementById('time_error');
  if (!el) {
    el = document.createElement('div');
    el.id = 'time_error';
    el.style.cssText = 'background:#fff5f5;border:1px solid #feb2b2;color:#c53030;padding:.75rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem';
    document.getElementById('submitBtn').before(el);
  }
  el.textContent = msg;
  setTimeout(() => { if (el) el.textContent = ''; }, 4000);
}

// ── Form submit validation ────────────────────────────────────────────────
document.getElementById('bookingForm').addEventListener('submit', function(e) {
  const errors = [];
  const sd = startDateEl.value, ed = endDateEl.value;
  const st = startTimeEl.value, et = endTimeEl.value, sut = setupTimeEl.value;
  const sameDay = sd && ed && sd === ed;

  if (sd && ed && ed < sd)   errors.push('End date cannot be before start date.');
  // End time vs start time only matters when event starts and ends the same day
  if (sameDay && st && et && et <= st)  errors.push('Event end time must be after the start time.');
  // Delivery time vs start time only matters when setup date == event start date
  const setupDateEl = document.getElementById('setupDateEl');
  const setupSameDay = setupDateEl && setupDateEl.value && sd && setupDateEl.value === sd;
  if (setupSameDay && sut && st && sut >= st) errors.push('Setup / delivery time must be before the event start time when delivery is on the same day.');

  if (errors.length) {
    e.preventDefault();
    showTimeError(errors.join(' '));
    return;
  }
  const btn = document.getElementById('submitBtn');
  btn.disabled = true;
  btn.textContent = 'Submitting...';
});
</script>
<script>
function initPublicEventAutocomplete() {
  var streetEl = document.getElementById('event_street');
  if (!streetEl || !window.google) return;
  var ac = new google.maps.places.Autocomplete(streetEl, {
    types: ['address'],
    componentRestrictions: { country: 'us' },
    fields: ['address_components']
  });
  streetEl.addEventListener('keydown', function(e) {
    if (e.key === 'Enter') e.preventDefault();
  });
  ac.addListener('place_changed', function() {
    var place = ac.getPlace();
    if (!place.address_components) return;
    var streetNum = '', route = '', city = '', state = '', zip = '';
    place.address_components.forEach(function(comp) {
      var t = comp.types;
      if (t.includes('street_number'))                    streetNum = comp.long_name;
      else if (t.includes('route'))                       route     = comp.long_name;
      else if (t.includes('locality'))                    city      = comp.long_name;
      else if (t.includes('administrative_area_level_1')) state     = comp.short_name;
      else if (t.includes('postal_code'))                 zip       = comp.long_name;
    });
    streetEl.value = [streetNum, route].filter(Boolean).join(' ');
    var cityEl  = document.getElementById('event_city');
    var stateEl = document.getElementById('event_state');
    var zipEl   = document.getElementById('event_zip');
    if (cityEl)  cityEl.value  = city;
    if (stateEl) stateEl.value = state;
    if (zipEl)   zipEl.value   = zip;
    scheduleDistanceCalc();
  });
}
</script>
{% if google_maps_key %}
<script src="https://maps.googleapis.com/maps/api/js?key={{ google_maps_key }}&libraries=places&callback=initPublicEventAutocomplete" async defer></script>
{% endif %}
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


SUCCESS_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Request Received — {{ business_name }}</title>
  <style>
    body{font-family:-apple-system,sans-serif;background:#f0f4f8;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}
    .box{background:white;border-radius:16px;padding:3rem 2.5rem;text-align:center;max-width:480px;width:90%;box-shadow:0 4px 24px rgba(0,0,0,.1)}
    .icon{font-size:3.5rem;margin-bottom:1rem}
    h1{color:#1a365d;font-size:1.6rem;margin-bottom:.75rem}
    p{color:#4a5568;line-height:1.6;margin-bottom:.75rem}
    .ref{background:#ebf4ff;border-radius:8px;padding:.75rem 1.25rem;display:inline-block;margin:.5rem 0;color:#1a365d;font-weight:700;font-size:1.3rem;letter-spacing:1px}
    a{display:inline-block;margin-top:1.5rem;padding:.7rem 1.5rem;background:#2b6cb0;color:white;border-radius:8px;text-decoration:none;font-weight:600}
  </style>
</head>
<body>
  <div class="box">
    <div class="icon">&#10003;</div>
    <h1>Request Received!</h1>
    <p>Thanks, <strong>{{ name }}</strong>! Your rental request is in.</p>
    <p>Your booking reference:</p>
    <div class="ref">#{{ booking_id }}</div>
    <p>We'll review your request and send a quote to <strong>{{ email }}</strong> shortly.</p>
    {% if business_phone %}<p>Questions? Call <strong>{{ business_phone }}</strong></p>{% endif %}
    <a href="/">Submit Another Request</a>
  </div>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


PAYMENT_SUCCESS_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Payment Received — {{ business_name }}</title>
  <style>
    body{font-family:-apple-system,sans-serif;background:#f0f4f8;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}
    .box{background:white;border-radius:16px;padding:3rem 2.5rem;text-align:center;max-width:480px;width:90%;box-shadow:0 4px 24px rgba(0,0,0,.1)}
    .icon{font-size:3.5rem;margin-bottom:1rem}
    h1{color:#276749;font-size:1.6rem;margin-bottom:.75rem}
    p{color:#4a5568;line-height:1.6;margin-bottom:.75rem}
  </style>
</head>
<body>
  <div class="box">
    <div class="icon">&#127881;</div>
    <h1>Payment Received!</h1>
    <p>Thank you! Your deposit for <strong>Booking #{{ booking_id }}</strong> has been received.</p>
    <p>Your reservation is now <strong>confirmed</strong>. We look forward to serving you!</p>
    <p style="font-weight:600;color:#2d3748">&#8212; {{ business_name }}</p>
    {% if business_phone %}<p style="font-size:.9rem;color:#718096">Questions? Call {{ business_phone }}</p>{% endif %}
  </div>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


ADMIN_LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#1e1e2e">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Admin Login — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;background:linear-gradient(160deg,#1e1e2e 0%,#1e3a5f 100%);padding:1.5rem}
    .card{background:rgba(255,255,255,.06);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);border:1px solid rgba(255,255,255,.12);border-radius:20px;padding:2.25rem 2rem;width:100%;max-width:340px;text-align:center}
    .logo{width:62px;height:62px;background:linear-gradient(135deg,#6366f1,#8b5cf6);border-radius:16px;margin:0 auto 1rem;display:flex;align-items:center;justify-content:center;font-size:1.6rem}
    .brand{font-size:1.2rem;font-weight:700;color:#fff;margin-bottom:.2rem}
    .subtitle{font-size:.78rem;color:rgba(255,255,255,.4);margin-bottom:2rem}
    .bio-btn{width:100%;padding:.95rem;background:linear-gradient(135deg,#6366f1,#8b5cf6);color:#fff;border:none;border-radius:12px;font-size:.93rem;font-weight:700;cursor:pointer;margin-bottom:.85rem;display:flex;align-items:center;justify-content:center;gap:.65rem;transition:opacity .15s}
    .bio-btn:hover{opacity:.9}
    .bio-btn:disabled{opacity:.6;cursor:default}
    .divider{display:flex;align-items:center;gap:.75rem;margin-bottom:.85rem}
    .divider-line{flex:1;height:1px;background:rgba(255,255,255,.1)}
    .divider-text{font-size:.7rem;color:rgba(255,255,255,.3)}
    .field{width:100%;padding:.7rem 1rem;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.15);border-radius:10px;color:#fff;font-size:.9rem;margin-bottom:.65rem;outline:none;transition:border .15s}
    .field:focus{border-color:rgba(99,102,241,.7)}
    .field::placeholder{color:rgba(255,255,255,.3)}
    .remember{display:flex;align-items:center;gap:.55rem;cursor:pointer;margin-bottom:.85rem;text-align:left}
    .chk{width:18px;height:18px;border-radius:5px;border:1.5px solid rgba(255,255,255,.3);background:transparent;display:flex;align-items:center;justify-content:center;flex-shrink:0;transition:all .15s}
    .chk.on{background:rgba(99,102,241,.8);border-color:#6366f1}
    .remember-lbl{font-size:.8rem;color:rgba(255,255,255,.55)}
    .pw-btn{width:100%;padding:.75rem;background:rgba(255,255,255,.1);color:#fff;border:1px solid rgba(255,255,255,.15);border-radius:10px;font-size:.88rem;font-weight:600;cursor:pointer;transition:background .15s}
    .pw-btn:hover{background:rgba(255,255,255,.16)}
    .msg{border-radius:10px;padding:.75rem;margin-bottom:.85rem;font-size:.85rem;font-weight:600}
    .msg-err{background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.35);color:#fca5a5}
    .msg-ok{background:rgba(22,163,74,.15);border:1px solid rgba(22,163,74,.4);color:#4ade80}
    .footer{margin-top:1.25rem;font-size:.68rem;color:rgba(255,255,255,.2)}
    .setup-link{display:inline-block;margin-top:.85rem;font-size:.75rem;color:rgba(255,255,255,.35);text-decoration:none}
    .setup-link:hover{color:rgba(255,255,255,.65)}
  </style>
</head>
<body>
  <div class="card">
    <div class="logo">🎪</div>
    <div class="brand">{{ business_name }}</div>
    <div class="subtitle">Admin Portal</div>

    <div id="msg-area">
      {% if error %}<div class="msg msg-err">{{ error }}</div>{% endif %}
    </div>

    {% if has_biometric %}
    <button class="bio-btn" id="bio-btn" onclick="doBiometric()">
      <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2z"/><circle cx="12" cy="10" r="3"/><path d="M7 20.662A8 8 0 0 1 12 18c1.75 0 3.37.56 4.69 1.5"/></svg>
      Sign in with Face ID / Touch ID
    </button>
    <div class="divider"><div class="divider-line"></div><span class="divider-text">or use password</span><div class="divider-line"></div></div>
    {% endif %}

    <form method="POST" id="pw-form">
      <input class="field" type="password" name="password" placeholder="Password"
             {% if not has_biometric %}autofocus{% endif %} required>
      <label class="remember" onclick="toggleRemember()">
        <div class="chk" id="chk-box"></div>
        <span class="remember-lbl">Remember this device</span>
      </label>
      <button class="pw-btn" type="submit">Sign In</button>
    </form>

    {% if not has_biometric %}
    <a href="#" class="setup-link" onclick="showSetupHint();return false">
      Set up Face ID / Touch ID →
    </a>
    {% endif %}

    <div class="footer">Secured with WebAuthn · {{ business_name }}</div>
  </div>

  <script>
  var rememberChecked = false;
  function toggleRemember() {
    rememberChecked = !rememberChecked;
    var b = document.getElementById('chk-box');
    if (rememberChecked) {
      b.classList.add('on');
      b.innerHTML = '<svg width="11" height="11" viewBox="0 0 12 12" fill="none" stroke="#fff" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="1,6 5,10 11,2"/></svg>';
    } else {
      b.classList.remove('on');
      b.innerHTML = '';
    }
  }

  function showMsg(text, type) {
    var el = document.getElementById('msg-area');
    el.innerHTML = '<div class="msg msg-' + type + '">' + text + '</div>';
  }

  function showSetupHint() {
    showMsg('Sign in with your password first, then go to Settings to set up biometric login.', 'ok');
  }

  function b64u_to_buf(b64u) {
    var b64 = b64u.replace(/-/g,'+').replace(/_/g,'/');
    var pad = b64.length % 4;
    if (pad) b64 += '===='.slice(pad);
    var bin = atob(b64), arr = new Uint8Array(bin.length);
    for (var i=0; i<bin.length; i++) arr[i] = bin.charCodeAt(i);
    return arr.buffer;
  }

  function buf_to_b64u(buf) {
    var arr = new Uint8Array(buf), str = '';
    for (var i=0; i<arr.length; i++) str += String.fromCharCode(arr[i]);
    return btoa(str).replace(/\+/g,'-').replace(/\//g,'_').replace(/=/g,'');
  }

  async function doBiometric() {
    var btn = document.getElementById('bio-btn');
    btn.disabled = true;
    btn.innerHTML = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="animation:spin 1s linear infinite"><circle cx="12" cy="12" r="10" stroke-dasharray="40 20"/></svg>&nbsp; Verifying…';
    try {
      var r = await fetch('/admin/webauthn/authenticate/begin', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'});
      if (!r.ok) throw new Error('Server error');
      var opts = await r.json();

      opts.challenge = b64u_to_buf(opts.challenge);
      if (opts.allowCredentials) {
        opts.allowCredentials = opts.allowCredentials.map(function(c) {
          return {id: b64u_to_buf(c.id), type: c.type};
        });
      }

      var assertion = await navigator.credentials.get({publicKey: opts});
      var payload = {
        id: assertion.id,
        rawId: buf_to_b64u(assertion.rawId),
        type: assertion.type,
        response: {
          clientDataJSON: buf_to_b64u(assertion.response.clientDataJSON),
          authenticatorData: buf_to_b64u(assertion.response.authenticatorData),
          signature: buf_to_b64u(assertion.response.signature),
          userHandle: assertion.response.userHandle ? buf_to_b64u(assertion.response.userHandle) : null
        }
      };

      var r2 = await fetch('/admin/webauthn/authenticate/complete', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify(payload)
      });
      var result = await r2.json();
      if (result.ok) {
        showMsg('✓ Verified — signing you in…', 'ok');
        setTimeout(function(){ window.location.href = '/admin/dashboard'; }, 600);
      } else {
        throw new Error(result.error || 'Verification failed');
      }
    } catch(e) {
      showMsg(e.message.indexOf('cancelled') > -1 || e.name === 'NotAllowedError' ? 'Biometric cancelled.' : ('Error: ' + e.message), 'err');
      btn.disabled = false;
      btn.innerHTML = '<svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2C6.48 2 2 6.48 2 12s4.48 10 10 10 10-4.48 10-10S17.52 2 12 2z"/><circle cx="12" cy="10" r="3"/><path d="M7 20.662A8 8 0 0 1 12 18c1.75 0 3.37.56 4.69 1.5"/></svg> Sign in with Face ID / Touch ID';
    }
  }
  </script>
  <style>@keyframes spin{from{transform:rotate(0deg)}to{transform:rotate(360deg)}}</style>
  <div class="card" style="display:none"></div>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


ADMIN_DASH_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#1a365d">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Admin">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Dashboard — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,sans-serif;background:#f0f2f5;color:#111827;min-height:100vh;display:flex}
    .sidebar{width:200px;min-height:100vh;background:#1e1e2e;border-right:none;display:flex;flex-direction:column;position:fixed;top:0;left:0;z-index:100;transition:transform .2s}
    .sb-brand{padding:1.1rem 1rem .9rem;display:flex;align-items:center;gap:.55rem;border-bottom:1px solid rgba(255,255,255,.08)}
    .sb-brand img{height:1.8rem;width:auto;object-fit:contain}
    .sb-brand-name{font-size:.82rem;font-weight:700;color:#fff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
    .sb-new-btn{display:block;margin:.85rem .85rem .5rem;padding:.55rem .75rem;background:#16a34a;color:white;border-radius:8px;font-size:.84rem;font-weight:700;text-decoration:none;text-align:center;transition:background .15s}
    .sb-new-btn:hover{background:#15803d}
    .sb-nav{display:flex;flex-direction:column;padding:.25rem 0;flex:1}
    .sb-link{display:flex;align-items:center;gap:.6rem;padding:.6rem 1rem;font-size:.84rem;font-weight:500;color:rgba(255,255,255,.55);text-decoration:none;transition:all .1s;border-left:3px solid transparent}
    .sb-link:hover{background:rgba(255,255,255,.07);color:rgba(255,255,255,.9)}
    .sb-link.active{background:rgba(99,102,241,.25);color:#a5b4fc;font-weight:600;border-left-color:#6366f1}
    .sb-icon{width:1.1rem;text-align:center;font-size:.95rem}
    .sb-divider{height:1px;background:rgba(255,255,255,.07);margin:.4rem 0}
    .sb-bottom{border-top:1px solid rgba(255,255,255,.08);padding:.5rem 0}
    .page-content{margin-left:200px;flex:1;min-height:100vh;display:flex;flex-direction:column}
    .page-header{background:white;border-bottom:1px solid #e5e7eb;padding:.85rem 1.5rem;display:flex;align-items:center;gap:1rem;position:sticky;top:0;z-index:50}
    .page-header h1{font-size:1.3rem;font-weight:700;color:#111827;flex:1}
    .mobile-menu-btn{display:none;background:none;border:none;font-size:1.4rem;cursor:pointer;color:#374151;padding:.25rem;line-height:1}
    .page-body{padding:1.5rem;flex:1}
    .stat-cards{display:grid;grid-template-columns:repeat(4,1fr);gap:.85rem;margin-bottom:1.25rem}
    .stat-card{border-radius:12px;padding:.9rem 1.1rem;color:#fff}
    .stat-card-label{font-size:.65rem;font-weight:600;text-transform:uppercase;letter-spacing:.06em;opacity:.75;margin-bottom:.25rem}
    .stat-card-value{font-size:1.65rem;font-weight:700;line-height:1}
    .stat-card-sub{font-size:.7rem;opacity:.6;margin-top:.25rem}
    @media(max-width:768px){.stat-cards{grid-template-columns:repeat(2,1fr)}}
    .day-panels{display:grid;grid-template-columns:1fr 1fr;gap:1.25rem;margin-bottom:1.5rem}
    .day-panel{background:white;border-radius:12px;overflow:hidden}
    .day-panel-delivery{border:1.5px solid #5DCAA5}
    .day-panel-pickup{border:1.5px solid #AFA9EC}
    .panel-hdr{display:flex;justify-content:space-between;align-items:center;padding:.75rem 1.1rem .6rem;border-bottom:1px solid #f3f4f6}
    .panel-hdr-delivery{background:#E1F5EE;border-bottom-color:#9FE1CB}
    .panel-hdr-pickup{background:#EEEDFE;border-bottom-color:#CECBF6}
    .panel-hdr-title{font-size:.9rem;font-weight:700;color:#111827}
    .panel-hdr-title-delivery{color:#085041}
    .panel-hdr-title-pickup{color:#26215C}
    .panel-hdr-date{color:#fff;padding:.2rem .6rem;border-radius:20px;font-size:.72rem;font-weight:700;border:none}
    .panel-hdr-date-delivery{background:#085041}
    .panel-hdr-date-pickup{background:#3C3489}
    .booking-row{display:flex;align-items:center;gap:.75rem;padding:.75rem 1.1rem;border-bottom:1px solid #f9fafb;text-decoration:none;color:inherit;transition:background .1s}
    .booking-row:hover{background:#fafafa}
    .bk-time{font-size:.78rem;font-weight:600;color:#374151;width:3.5rem;flex-shrink:0}
    .avatar{width:34px;height:34px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:.82rem;font-weight:700;color:white;flex-shrink:0}
    .bk-info{flex:1;min-width:0}
    .bk-name{font-size:.86rem;font-weight:600;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    .bk-email{font-size:.75rem;color:#9ca3af;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    .bk-items{font-size:.74rem;color:#2563eb;font-weight:500;margin-top:.1rem}
    .bk-num{font-size:.8rem;font-weight:700;color:#6b7280;flex-shrink:0}
    .panel-empty{padding:2.5rem;text-align:center;color:#9ca3af;font-size:.88rem}
    .panel-footer{display:flex;justify-content:space-between;align-items:center;padding:.6rem 1.1rem;border-top:1px solid #f3f4f6}
    .panel-footer a{font-size:.78rem;color:#2563eb;text-decoration:none;font-weight:500}
    .panel-footer a:hover{text-decoration:underline}
    .panel-footer-delivery a{color:#0F6E56}
    .panel-footer-pickup a{color:#534AB7}
    .section-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:.75rem;flex-wrap:wrap;gap:.5rem}
    .section-title-lg{font-size:1.05rem;font-weight:700;color:#111827}
    .search-wrap{position:relative}
    .search-input{border:1px solid #d1d5db;border-radius:8px;padding:.4rem 2rem .4rem .85rem;font-size:.85rem;width:220px;outline:none;background:white;transition:border .12s}
    .search-input:focus{border-color:#2563eb}
    .search-count{position:absolute;right:.55rem;top:50%;transform:translateY(-50%);font-size:.72rem;color:#9ca3af}
    .tabs{display:flex;background:white;border:1px solid #e5e7eb;border-bottom:none;border-radius:10px 10px 0 0;overflow-x:auto}
    .tab{padding:.7rem 1rem;font-size:.82rem;font-weight:500;color:#6b7280;text-decoration:none;border-bottom:2px solid transparent;white-space:nowrap;flex-shrink:0;transition:all .12s}
    .tab:hover{color:#111827;background:#f9fafb}
    .tab.active{color:#2563eb;border-bottom-color:#2563eb;font-weight:600}
    .table-card{background:white;border:1px solid #e5e7eb;border-radius:0 0 10px 10px;overflow:hidden}
    .table-scroll{overflow-x:auto;-webkit-overflow-scrolling:touch}
    .table-scroll table{min-width:880px}
    table{width:100%;border-collapse:collapse}
    thead tr{background:#f9fafb}
    th{padding:.65rem .9rem;text-align:left;font-size:.72rem;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid #e5e7eb;white-space:nowrap}
    td{padding:.8rem .9rem;border-bottom:1px solid #f3f4f6;vertical-align:middle;font-size:.85rem;color:#374151}
    tr:last-child td{border-bottom:none}
    tbody tr:hover td{background:#fafafa}
    .client-cell{display:flex;align-items:center;gap:.55rem}
    .client-name{font-weight:600;color:#111827;font-size:.85rem}
    .client-email{font-size:.74rem;color:#9ca3af;margin-top:.05rem}
    .badge{display:inline-flex;align-items:center;padding:.2rem .6rem;border-radius:20px;font-size:.73rem;font-weight:600;white-space:nowrap}
    .badge-pending{background:#fef9c3;color:#854d0e}
    .badge-agree_to_pay{background:#d1fae5;color:#065f46}
    .badge-accepted{background:#dbeafe;color:#1e40af}
    .badge-confirmed{background:#dbeafe;color:#1e40af}
    .badge-partial{background:#dbeafe;color:#1e40af}
    .badge-denied{background:#fee2e2;color:#991b1b}
    .badge-cancelled{background:#f3f4f6;color:#6b7280}
    .badge-concluded{background:#e5e7eb;color:#374151}
    .pay-badge.pay-waiting{background:#fef3c7;color:#b45309}
    .pay-badge{display:inline-flex;align-items:center;padding:.2rem .6rem;border-radius:20px;font-size:.73rem;font-weight:600;white-space:nowrap}
    .pay-paid{background:#dcfce7;color:#166534}
    .pay-due{background:#fef9c3;color:#854d0e}
    .pay-partial{background:#dbeafe;color:#1e40af}
    .pay-none{color:#9ca3af;font-size:.77rem}
    .date-range{display:flex;align-items:center;gap:.3rem;font-size:.82rem;white-space:nowrap}
    .date-arrow{color:#d1d5db;font-size:.7rem}
    .action-btns{display:flex;gap:.3rem;flex-wrap:nowrap;align-items:center}
    .btn{display:inline-block;padding:.28rem .6rem;border-radius:6px;font-size:.75rem;font-weight:600;cursor:pointer;border:1px solid transparent;text-decoration:none;line-height:1.5;white-space:nowrap;transition:all .12s}
    .btn-view{background:#eff6ff;color:#2563eb;border-color:#bfdbfe}
    .btn-view:hover{background:#dbeafe}
    .btn-accept{background:#f0fdf4;color:#166534;border-color:#bbf7d0}
    .btn-accept:hover{background:#dcfce7}
    .btn-deny{background:#fef2f2;color:#991b1b;border-color:#fecaca}
    .btn-deny:hover{background:#fee2e2}
    .btn-confirm{background:#eff6ff;color:#1e40af;border-color:#bfdbfe}
    .btn-confirm:hover{background:#dbeafe}
    .btn-cancel{background:#f9fafb;color:#6b7280;border-color:#d1d5db}
    .empty-state{padding:3rem;text-align:center;color:#9ca3af;font-size:.95rem}
    .row-pending{border-left:4px solid #8b5cf6 !important}
    .row-accepted{border-left:4px solid #0ea5e9 !important}
    .row-delivered{border-left:4px solid #f97316 !important}
    .row-picked-up{border-left:4px solid #16a34a !important}
    .row-denied{border-left:4px solid #ef4444 !important}
    .row-cancelled{border-left:4px solid #9ca3af !important}
    @media(max-width:768px){
      .sidebar{transform:translateX(-100%)}
      .sidebar.open{transform:translateX(0);box-shadow:6px 0 30px rgba(0,0,0,.4)}
      .page-content{margin-left:0}
      .mobile-menu-btn{display:block}
      .day-panels{grid-template-columns:1fr}
      .page-body{padding:1rem}
    }
    .overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.3);z-index:99}
    .overlay.show{display:block}
  </style>
</head>
<body>

<div class="overlay" id="overlay" onclick="closeSidebar()"></div>

<aside class="sidebar" id="sidebar">
  <div class="sb-brand">
    <img src="/logo.png" alt="">
    <span class="sb-brand-name">{{ business_name }}</span>
  </div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link active"><span class="sb-icon">🏠</span> Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link"><span class="sb-icon">👥</span> Clients</a>
    <a href="/admin/inventory" class="sb-link"><span class="sb-icon">📦</span> Inventory</a>
    <a href="/admin/calendar" class="sb-link"><span class="sb-icon">📅</span> Calendar</a>
    <a href="/admin/reports" class="sb-link"><span class="sb-icon">📊</span> Reports</a>
    <a href="/admin/route" class="sb-link"><span class="sb-icon">🗺</span> Route</a>
    <a href="/driver/{{ today }}" class="sb-link"><span class="sb-icon">🚚</span> Driver View</a>
    <a href="/admin/formsite-import" class="sb-link"><span class="sb-icon">📥</span> Import</a>
    <a href="/admin/tax-report" class="sb-link"><span class="sb-icon">💰</span> Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/download-backup" class="sb-link"><span class="sb-icon">💾</span> Backup</a>
    <a href="/admin/logout" class="sb-link"><span class="sb-icon">🚪</span> Sign Out</a>
    <a href="/admin/setup-biometric" style="display:block;padding:.35rem 1rem;font-size:.72rem;color:rgba(255,255,255,.25);text-decoration:none;text-align:center" title="Set up Face ID / Touch ID">🔐 Biometric setup</a>
  </div>
</aside>

<div class="page-content">
  <div class="page-header">
    <button class="mobile-menu-btn" onclick="openSidebar()">☰</button>
    <h1>Dashboard</h1>
    <div class="search-wrap">
      <input type="text" class="search-input" id="dash-search" placeholder="Search bookings…" oninput="filterDash(this.value)">
      <span class="search-count" id="dash-count"></span>
    </div>
  </div>

  <div class="page-body">

    {% if inv_conflicts %}
    <div style="background:#fef2f2;border:2px solid #f87171;border-radius:10px;padding:.9rem 1.1rem;margin-bottom:1.25rem">
      <div style="display:flex;align-items:center;gap:.5rem;font-weight:700;color:#991b1b;font-size:.9rem;margin-bottom:.5rem">
        <span>🚨</span> Inventory Conflict — {{ inv_conflicts|length }} item{{ 's' if inv_conflicts|length != 1 else '' }} over-committed
      </div>
      {% for c in inv_conflicts %}
      <div style="background:white;border:1px solid #fecaca;border-radius:6px;padding:.45rem .8rem;display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:.4rem;font-size:.84rem;margin-top:.3rem">
        <div><a href="/admin/booking/{{ c.booking_id }}" style="font-weight:700;color:#dc2626;text-decoration:none">Booking #{{ c.booking_id }}</a> — {{ c.customer }} <span style="color:#9ca3af;font-size:.78rem">({{ c.event_date }})</span></div>
        <div style="color:#7f1d1d;font-size:.82rem"><strong>{{ c.item }}</strong>: needs <strong>{{ c.needed }}</strong>, only <strong>{{ c.available }}</strong> available <span style="background:#dc2626;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.73rem;font-weight:700;margin-left:.3rem">-{{ c.shortfall }} short</span></div>
      </div>
      {% endfor %}
    </div>
    {% endif %}

    <!-- Stat cards -->
    <div class="stat-cards">
      <a href="/admin/dashboard?tab=all&status=pending" class="stat-card" style="background:linear-gradient(135deg,#6366f1,#8b5cf6);text-decoration:none">
        <div class="stat-card-label">Pending</div>
        <div class="stat-card-value">{{ stats.pending }}</div>
        <div class="stat-card-sub">need review</div>
      </a>
      <a href="/admin/dashboard?tab=upcoming" class="stat-card" style="background:linear-gradient(135deg,#0ea5e9,#06b6d4);text-decoration:none">
        <div class="stat-card-label">Upcoming</div>
        <div class="stat-card-value">{{ stats.upcoming }}</div>
        <div class="stat-card-sub">within 8 days</div>
      </a>
      <a href="/admin/dashboard?tab=delivered" class="stat-card" style="background:linear-gradient(135deg,#f97316,#ef4444);text-decoration:none">
        <div class="stat-card-label">Still Out</div>
        <div class="stat-card-value">{{ stats.delivered }}</div>
        <div class="stat-card-sub">awaiting pickup</div>
      </a>
      <div class="stat-card" style="background:linear-gradient(135deg,#16a34a,#059669)">
        <div class="stat-card-label">Revenue</div>
        <div class="stat-card-value">${{ "{:,.0f}".format(stats.revenue) }}</div>
        <div class="stat-card-sub">paid bookings</div>
      </div>
    </div>

    <div class="day-panels">
      <div class="day-panel day-panel-delivery">
        <div class="panel-hdr panel-hdr-delivery">
          <a href="/admin/dashboard?tab=going_out" class="panel-hdr-title panel-hdr-title-delivery" style="text-decoration:none">🚚 Deliveries Today</a>
          <span class="panel-hdr-date panel-hdr-date-delivery">{{ today_label }}</span>
        </div>
        {% if going_out %}
        {% for b in going_out %}
        <a href="/admin/booking/{{ b.id }}" class="booking-row">
          <div class="bk-time">{{ b.event_start_time or '' }}</div>
          <div class="avatar" style="background:{{ b.avatar_color }}">{{ b.avatar_initials }}</div>
          <div class="bk-info">
            <div class="bk-name">{{ b.full_name }}</div>
            <div class="bk-email">{{ b.email or '' }}</div>
            <div class="bk-items">{{ b.item_count }} reserved</div>
          </div>
          <div class="bk-num">#{{ b.id }}</div>
        </a>
        {% endfor %}
        {% else %}
        <div class="panel-empty">No deliveries scheduled today</div>
        {% endif %}
        <div class="panel-footer panel-footer-delivery">
          <a href="/admin/dashboard?past=1">View late ({{ stats.past }})</a>
          <a href="/admin/dashboard">View all</a>
        </div>
      </div>

      <div class="day-panel day-panel-pickup">
        <div class="panel-hdr panel-hdr-pickup">
          <a href="/admin/dashboard?tab=coming_back" class="panel-hdr-title panel-hdr-title-pickup" style="text-decoration:none">🔄 Pickups Today</a>
          <span class="panel-hdr-date panel-hdr-date-pickup">{{ today_label }}</span>
        </div>
        {% if coming_back %}
        {% for b in coming_back %}
        <a href="/admin/booking/{{ b.id }}" class="booking-row">
          <div class="bk-time">{{ b.event_start_time or '' }}</div>
          <div class="avatar" style="background:{{ b.avatar_color }}">{{ b.avatar_initials }}</div>
          <div class="bk-info">
            <div class="bk-name">{{ b.full_name }}</div>
            <div class="bk-email">{{ b.email or '' }}</div>
            <div class="bk-items">{{ b.item_count }} reserved</div>
          </div>
          <div class="bk-num">#{{ b.id }}</div>
        </a>
        {% endfor %}
        {% else %}
        <div class="panel-empty">No results</div>
        {% endif %}
        <div class="panel-footer panel-footer-pickup">
          <a href="/admin/dashboard">View late (0)</a>
          <a href="/admin/dashboard">View all</a>
        </div>
      </div>
    </div>

    <div class="section-header">
      <span class="section-title-lg">All Bookings</span>
    </div>

    {% set df = ('&date_from=' ~ date_from) if date_from else '' %}
    {% set dt = ('&date_to=' ~ date_to) if date_to else '' %}
    {% set pf = ('&pay_filter=' ~ pay_filter) if pay_filter else '' %}
    {% set sf = ('&sort=' ~ sort_by) if (sort_by and sort_by != 'date') else '' %}
    <div class="tabs">
      <a href="/admin/dashboard?tab=all" class="tab {% if tab=='all' %}active{% endif %}">All ({{ stats.total }})</a>
      <a href="/admin/dashboard?tab=going_out" class="tab {% if tab=='going_out' %}active{% endif %}" style="{% if tab=='going_out' %}color:#dc2626;border-bottom-color:#dc2626;{% endif %}">🚚 Deliveries{% if stats.going_out > 0 %} <span style="background:#dc2626;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.going_out }}</span>{% endif %}</a>
      <a href="/admin/dashboard?tab=coming_back" class="tab {% if tab=='coming_back' %}active{% endif %}" style="{% if tab=='coming_back' %}color:#7c3aed;border-bottom-color:#7c3aed;{% endif %}">🔄 Pickups{% if stats.coming_back > 0 %} <span style="background:#7c3aed;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.coming_back }}</span>{% endif %}</a>
      <a href="/admin/dashboard?tab=upcoming" class="tab {% if tab=='upcoming' %}active{% endif %}" style="{% if tab=='upcoming' %}color:#f97316;border-bottom-color:#f97316;{% endif %}">🔔 Upcoming{% if stats.upcoming > 0 %} <span style="background:#f97316;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.upcoming }}</span>{% endif %}</a>
      <a href="/admin/dashboard?tab=delivered" class="tab {% if tab=='delivered' %}active{% endif %}" style="{% if tab=='delivered' %}color:#d97706;border-bottom-color:#d97706;{% endif %}">📦 Still Out{% if stats.delivered > 0 %} <span style="background:#d97706;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.delivered }}</span>{% endif %}</a>
      <a href="/admin/dashboard?tab=picked_up" class="tab {% if tab=='picked_up' %}active{% endif %}" style="{% if tab=='picked_up' %}color:#2563eb;border-bottom-color:#2563eb;{% endif %}">✅ Picked Up{% if stats.picked_up > 0 %} <span style="background:#2563eb;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.picked_up }}</span>{% endif %}</a>
      <a href="/admin/dashboard?tab=still_waiting" class="tab {% if tab=='still_waiting' %}active{% endif %}" style="{% if tab=='still_waiting' %}color:#dc2626;border-bottom-color:#dc2626;{% endif %}">⚠️ Still Waiting{% if stats.still_waiting > 0 %} <span style="background:#dc2626;color:white;border-radius:99px;padding:.05rem .45rem;font-size:.72rem;font-weight:700;margin-left:.2rem">{{ stats.still_waiting }}</span>{% endif %}</a>
      <a href="/admin/dashboard?archived=1" class="tab {% if archived_filter %}active{% endif %}" style="{% if archived_filter %}color:#9ca3af;border-bottom-color:#9ca3af;{% endif %}">📦 Archived</a>
    </div>

    <form method="GET" action="/admin/dashboard" style="background:white;border:1px solid #e5e7eb;border-bottom:none;padding:.6rem 1rem;display:flex;flex-wrap:wrap;gap:.55rem;align-items:center">
      <input type="hidden" name="status" value="{{ status_filter }}">
      <label style="font-size:.78rem;font-weight:600;color:#6b7280">Event Date:</label>
      <input type="date" name="date_from" value="{{ date_from }}" style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .5rem;font-size:.82rem;color:#374151">
      <span style="font-size:.82rem;color:#9ca3af">to</span>
      <input type="date" name="date_to" value="{{ date_to }}" style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .5rem;font-size:.82rem;color:#374151">
      <label style="font-size:.78rem;font-weight:600;color:#6b7280;margin-left:.4rem">Payment:</label>
      <select name="pay_filter" style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .5rem;font-size:.82rem;color:#374151">
        <option value="" {% if not pay_filter %}selected{% endif %}>All</option>
        <option value="paid"    {% if pay_filter=='paid'    %}selected{% endif %}>Paid in Full</option>
        <option value="partial" {% if pay_filter=='partial' %}selected{% endif %}>Partially Paid</option>
        <option value="due"     {% if pay_filter=='due'     %}selected{% endif %}>Waiting</option>
      </select>
      <label style="font-size:.78rem;font-weight:600;color:#6b7280;margin-left:.4rem">Sort:</label>
      <select name="sort" style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .5rem;font-size:.82rem;color:#374151">
        <option value="date"      {% if sort_by=='date'      %}selected{% endif %}>Event Date ↑</option>
        <option value="date_desc" {% if sort_by=='date_desc' %}selected{% endif %}>Event Date ↓</option>
        <option value="name"      {% if sort_by=='name'      %}selected{% endif %}>Client A→Z</option>
        <option value="name_desc" {% if sort_by=='name_desc' %}selected{% endif %}>Client Z→A</option>
        <option value="id"        {% if sort_by=='id'        %}selected{% endif %}>Booking # ↓</option>
        <option value="id_asc"    {% if sort_by=='id_asc'    %}selected{% endif %}>Booking # ↑</option>
        <option value="total"     {% if sort_by=='total'     %}selected{% endif %}>Total ↓</option>
        <option value="created"   {% if sort_by=='created'   %}selected{% endif %}>Date Added ↓</option>
      </select>
      <button type="submit" style="background:#2563eb;color:white;border:none;border-radius:6px;padding:.32rem .8rem;font-size:.82rem;font-weight:600;cursor:pointer">Apply</button>
      {% if date_from or date_to or pay_filter %}<a href="/admin/dashboard?status={{ status_filter }}" style="font-size:.78rem;color:#6b7280;text-decoration:none">✕ Clear</a>{% endif %}
    </form>

    <div id="bulkBar" style="display:none;background:#1e3a5f;color:white;padding:.55rem 1rem;gap:.75rem;align-items:center;border-radius:0 0 8px 8px;margin-bottom:.4rem">
      <span id="bulkCount" style="font-size:.84rem;font-weight:600"></span>
      <form method="POST" action="/admin/bookings/bulk-archive" id="bulkArchiveForm" style="display:inline">
        <input type="hidden" name="ids" id="bulkArchiveIds">
        <button type="button" onclick="bulkAction('archive')" style="background:#f97316;color:white;border:none;border-radius:6px;padding:.28rem .75rem;font-size:.81rem;font-weight:600;cursor:pointer">📦 Archive Selected</button>
      </form>
      <form method="POST" action="/admin/bookings/bulk-delete" id="bulkDeleteForm" style="display:inline">
        <input type="hidden" name="ids" id="bulkDeleteIds">
        <button type="button" onclick="bulkAction('delete')" style="background:#ef4444;color:white;border:none;border-radius:6px;padding:.28rem .75rem;font-size:.81rem;font-weight:600;cursor:pointer">🗑 Delete Selected</button>
      </form>
      <button type="button" onclick="clearAll()" style="background:transparent;color:#9ca3af;border:1px solid #4b5563;border-radius:6px;padding:.28rem .75rem;font-size:.81rem;cursor:pointer">✕ Clear</button>
    </div>

    <div class="table-card">
      {% if bookings %}
      <div class="table-scroll">
      <table>
        <thead><tr>
          <th style="width:36px;padding-left:.75rem"><input type="checkbox" id="selectAll" onchange="toggleAll(this)" style="cursor:pointer;width:15px;height:15px;accent-color:#2563eb"></th>
          <th>#</th><th>Client</th><th>Status</th><th>Delivery → Pickup</th><th>Items</th><th>Total</th><th>Payment</th><th>Actions</th>
        </tr></thead>
        <tbody>
          {% for b in bookings %}
          {% set _bc = conflict_map.get(b.id, []) %}
          <tr id="row-{{ b.id }}" data-search="{{ (b.full_name or '')|lower }} {{ (b.email or '')|lower }} {{ (b.phone or '')|lower }} {{ (b.event_start_date or '') }} {{ (b.items_summary or '')|lower }}"
            class="{% if _bc %}row-conflict{% elif b.status == 'pending' %}row-pending{% elif b.status == 'agree_to_pay' %}row-accepted{% elif b.delivery_status == 'picked_up' %}row-picked-up{% elif b.delivery_status == 'delivered' %}row-delivered{% elif b.status == 'accepted' %}row-accepted{% elif b.status == 'denied' %}row-denied{% elif b.status == 'cancelled' %}row-cancelled{% endif %}"
            {% if _bc %}style="background:#fff5f5;border-left:4px solid #e53e3e;"{% elif b.red_flag %}style="background:#fffbfb;border-left:4px solid #dc2626;"{% endif %}>
            <td style="padding-left:.75rem"><input type="checkbox" class="row-cb" value="{{ b.id }}" onchange="updateBulkBar()" style="cursor:pointer;width:15px;height:15px;accent-color:#2563eb"></td>
            <td style="font-weight:700;color:#2563eb;font-size:.83rem"><a href="/admin/booking/{{ b.id }}" style="color:#2563eb;text-decoration:none">#{{ b.id }}</a></td>
            <td>
              <div class="client-cell">
                <div class="avatar" style="background:{{ b.avatar_color }}">{{ b.avatar_initials }}</div>
                <div>
                  <div class="client-name" style="display:flex;align-items:center;gap:.35rem">
                    <a href="/admin/booking/{{ b.id }}" style="color:#111827;text-decoration:none;font-weight:600" title="View booking">{{ b.full_name }}</a>
                    {% if b.red_flag %}<span title="Delivery within 5 days — payment not received!" style="display:inline-flex;align-items:center;gap:.2rem;background:#fef2f2;color:#dc2626;border:1px solid #fca5a5;border-radius:4px;padding:.08rem .38rem;font-size:.7rem;font-weight:700;white-space:nowrap">🚩 UNPAID</span>{% endif %}
                  </div>
                  <div class="client-email">{{ b.email }}</div>
                  {% if b.phone %}<div style="font-size:.74rem;color:#6b7280;margin-top:.05rem"><a href="tel:{{ b.phone }}" style="color:#6b7280;text-decoration:none">📞 {{ b.phone }}</a></div>{% endif %}
                </div>
              </div>
              {% if _bc %}
              <div style="margin-top:.35rem;display:flex;flex-wrap:wrap;gap:.3rem">
                {% for _cf in _bc %}
                <span style="display:inline-flex;align-items:center;gap:.25rem;background:#fed7d7;color:#c53030;border-radius:4px;padding:.15rem .45rem;font-size:.72rem;font-weight:700">
                  🚨 {{ _cf.item }}: {{ _cf.shortfall }} short
                </span>
                {% endfor %}
              </div>
              {% endif %}
            </td>
            <td>
              {% if b.status == 'accepted' %}
                <span class="badge badge-accepted">Accepted</span>
              {% elif b.status == 'agree_to_pay' %}
                <span class="badge badge-agree_to_pay">Agree to Pay</span>
              {% elif b.status == 'pending' %}
                <span class="badge badge-pending">Pending</span>
              {% elif b.status == 'denied' %}
                <span class="badge badge-denied">Denied</span>
              {% elif b.status == 'cancelled' %}
                <span class="badge badge-cancelled">Cancelled</span>
              {% elif b.status == 'concluded' %}
                <span class="badge badge-concluded">Concluded</span>
              {% else %}
                <span class="badge badge-{{ b.status }}">{{ b.status|capitalize }}</span>
              {% endif %}
            </td>
            <td>
              <div class="date-range">
                <span>{{ b.setup_date.strftime('%m/%d/%Y') if b.setup_date else (b.event_start_date.strftime('%m/%d/%Y') if b.event_start_date else '—') }}</span>
                <span class="date-arrow">→</span>
                <span>{{ b.event_end_date.strftime('%m/%d/%Y') if b.event_end_date else '—' }}</span>
              </div>
              {% if b.maps_url %}<a href="{{ b.maps_url }}" target="_blank" rel="noopener noreferrer" style="display:inline-flex;align-items:center;gap:.15rem;margin-top:.2rem;font-size:.72rem;color:#2563eb;text-decoration:none;font-weight:500">📍 Map</a>{% endif %}
            </td>
            <td style="max-width:155px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#6b7280;font-size:.79rem">{{ b.items_summary }}</td>
            <td style="font-weight:700;white-space:nowrap">${{ "%.2f"|format(b.grand_total or 0) }}</td>
            <td>{% if b.pay_label != '—' %}<span class="pay-badge {{ b.pay_class }}">{{ b.pay_label }}</span>{% else %}<span class="pay-none">—</span>{% endif %}</td>
            <td>
              <div class="action-btns">
                <a href="/admin/booking/{{ b.id }}" class="btn btn-view">View</a>
                {% if b.status in ('pending', 'agree_to_pay') %}
                <form method="POST" action="/admin/booking/{{ b.id }}/accept" style="display:inline"><button class="btn btn-accept" onclick="return confirm('Accept #{{ b.id }}? This emails {{ b.email }} their invoice + Stripe payment link.')">Accept</button></form>
                <form method="POST" action="/admin/booking/{{ b.id }}/deny" style="display:inline"><button class="btn btn-deny" onclick="return confirm('Deny booking #{{ b.id }}?')">Deny</button></form>
                {% endif %}
                {% if b.status == 'accepted' %}
                <form method="POST" action="/admin/booking/{{ b.id }}/confirm" style="display:inline"><button class="btn btn-confirm" onclick="return confirm('Manually mark #{{ b.id }} as paid?')">Mark Paid</button></form>
                {% endif %}
                {% if b.status not in ('denied', 'cancelled') %}
                <form method="POST" action="/admin/booking/{{ b.id }}/cancel" style="display:inline"><button class="btn btn-cancel" onclick="return confirm('Cancel booking #{{ b.id }}?')">Cancel</button></form>
                {% endif %}
                {% if b.delivery_status != 'picked_up' %}
                <form method="POST" action="/admin/booking/{{ b.id }}/delivery-status" style="display:inline">
                  {% if not b.delivery_status %}<button class="btn" style="background:#fffbeb;color:#92400e;border:1px solid #fcd34d;font-size:.74rem" onclick="return confirm('Mark #{{ b.id }} as DELIVERED?')">🚚 Delivered</button>
                  {% elif b.delivery_status == 'delivered' %}<button class="btn" style="background:#eff6ff;color:#1e40af;border:1px solid #93c5fd;font-size:.74rem" onclick="return confirm('Mark #{{ b.id }} as PICKED UP?')">✅ Picked Up</button>{% endif %}
                </form>
                {% else %}
                <span style="font-size:.74rem;color:#16a34a;font-weight:600;padding:.25rem .5rem;background:#f0fdf4;border:1px solid #86efac;border-radius:6px">✔ Picked Up</span>
                {% endif %}
                <div class="more-wrap" style="position:relative;display:inline-block;margin-left:.2rem">
                  <button type="button" onclick="toggleMore(event,this,'more-{{ b.id }}')"
                    style="border:1px solid #d1d5db;border-radius:6px;padding:.25rem .55rem;font-size:.77rem;color:#374151;cursor:pointer;background:#fff;white-space:nowrap">
                    ⚙ More ▾
                  </button>
                  <div id="more-{{ b.id }}" style="display:none;position:fixed;background:#fff;border:1px solid #e5e7eb;border-radius:8px;box-shadow:0 4px 20px rgba(0,0,0,.18);z-index:9999;min-width:150px;overflow:hidden">
                    {% if b.archived %}
                    <form method="POST" action="/admin/booking/{{ b.id }}/unarchive" style="margin:0">
                      <button type="submit" style="width:100%;text-align:left;padding:.5rem .85rem;font-size:.82rem;border:none;background:none;cursor:pointer;color:#374151;white-space:nowrap" onmouseover="this.style.background='#f9fafb'" onmouseout="this.style.background='none'">↩ Unarchive</button>
                    </form>
                    {% else %}
                    <form method="POST" action="/admin/booking/{{ b.id }}/archive" style="margin:0">
                      <button type="submit" style="width:100%;text-align:left;padding:.5rem .85rem;font-size:.82rem;border:none;background:none;cursor:pointer;color:#374151;white-space:nowrap" onmouseover="this.style.background='#f9fafb'" onmouseout="this.style.background='none'">📦 Archive</button>
                    </form>
                    {% endif %}
                    <form method="POST" action="/admin/booking/{{ b.id }}/delete" style="margin:0" onsubmit="return confirm('Permanently delete booking #{{ b.id }}? Cannot be undone.')">
                      <button type="submit" style="width:100%;text-align:left;padding:.5rem .85rem;font-size:.82rem;border:none;background:none;cursor:pointer;color:#dc2626;white-space:nowrap;border-top:1px solid #f3f4f6" onmouseover="this.style.background='#fff5f5'" onmouseout="this.style.background='none'">🗑 Delete</button>
                    </form>
                  </div>
                </div>
              </div>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      </div>
      {% else %}
      <div class="empty-state">No bookings found.</div>
      {% endif %}
    </div>

  </div>
</div>

<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('overlay').classList.remove('show');}
function toggleMore(e,btn,id){
  e.stopPropagation();
  e.preventDefault();
  var m=document.getElementById(id);
  var isOpen=m.style.display==='block';
  document.querySelectorAll('[id^="more-"]').forEach(function(el){el.style.display='none';});
  if(!isOpen){
    var rect=btn.getBoundingClientRect();
    var menuW=160;
    var left=Math.min(rect.right-menuW, window.innerWidth-menuW-8);
    left=Math.max(8,left);
    var top=rect.bottom+4;
    if(top+160>window.innerHeight) top=rect.top-164;
    m.style.top=top+'px';
    m.style.left=left+'px';
    m.style.display='block';
  }
}
document.addEventListener('mousedown',function(e){
  if(!e.target.closest('[id^="more-"]')&&!e.target.closest('.more-wrap')){
    document.querySelectorAll('[id^="more-"]').forEach(function(el){el.style.display='none';});
  }
});
document.addEventListener('scroll',function(){document.querySelectorAll('[id^="more-"]').forEach(function(el){el.style.display='none';});},true);
function getChecked(){return Array.from(document.querySelectorAll('.row-cb:checked')).map(c=>c.value);}
function updateBulkBar(){var ids=getChecked();var bar=document.getElementById('bulkBar');if(ids.length>0){bar.style.display='flex';document.getElementById('bulkCount').textContent=ids.length+' selected';}else{bar.style.display='none';}var all=document.querySelectorAll('.row-cb');document.getElementById('selectAll').indeterminate=ids.length>0&&ids.length<all.length;document.getElementById('selectAll').checked=ids.length===all.length&&all.length>0;}
function toggleAll(cb){document.querySelectorAll('.row-cb').forEach(c=>c.checked=cb.checked);updateBulkBar();}
function clearAll(){document.querySelectorAll('.row-cb').forEach(c=>c.checked=false);document.getElementById('selectAll').checked=false;document.getElementById('bulkBar').style.display='none';}
function filterDash(q){const term=q.toLowerCase().trim();const rows=document.querySelectorAll('tbody tr[data-search]');let shown=0;rows.forEach(row=>{const match=!term||row.dataset.search.includes(term);row.style.display=match?'':'none';if(match)shown++;});const cnt=document.getElementById('dash-count');if(cnt)cnt.textContent=term?shown+' found':'';}
function bulkAction(type){var ids=getChecked();if(ids.length===0)return;var msg=type==='delete'?'Permanently delete '+ids.length+' booking(s)?':'Archive '+ids.length+' booking(s)?';if(!confirm(msg))return;var idStr=ids.join(',');if(type==='delete'){document.getElementById('bulkDeleteIds').value=idStr;document.getElementById('bulkDeleteForm').submit();}else{document.getElementById('bulkArchiveIds').value=idStr;document.getElementById('bulkArchiveForm').submit();}}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


ADMIN_BOOKING_EDIT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json"><meta name="apple-mobile-web-app-capable" content="yes"><meta name="apple-mobile-web-app-title" content="Admin">
  <title>Edit Booking #{{ b.id }} — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,sans-serif;background:#f0f4f8;color:#1a202c;min-height:100vh}
    header{background:linear-gradient(135deg,#1a365d,#2b6cb0);color:white;padding:1.25rem 2rem;display:flex;justify-content:space-between;align-items:center}
    header h1{font-size:1.2rem}
    .container{max-width:800px;margin:0 auto;padding:1.5rem 1rem}
    .card{background:white;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08);padding:1.5rem;margin-bottom:1.5rem}
    .card h2{font-size:.95rem;font-weight:700;color:#2b6cb0;border-bottom:2px solid #ebf4ff;padding-bottom:.5rem;margin-bottom:1rem;text-transform:uppercase;letter-spacing:.4px}
    .field-grid{display:grid;grid-template-columns:1fr 1fr;gap:.75rem 1rem}
    .field-grid.single{grid-template-columns:1fr}
    .field-grid.triple{grid-template-columns:1fr 1fr 1fr}
    label{display:block;font-size:.78rem;font-weight:600;color:#6b7280;margin-bottom:.25rem;text-transform:uppercase;letter-spacing:.3px}
    input,select,textarea{width:100%;border:1px solid #d1d5db;border-radius:7px;padding:.5rem .75rem;font-size:.92rem;color:#1a202c;background:white}
    input:focus,select:focus,textarea:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 3px rgba(37,99,235,.1)}
    textarea{resize:vertical;min-height:80px}
    .actions{display:flex;gap:.75rem;flex-wrap:wrap;margin-top:1.5rem}
    .btn{padding:.65rem 1.4rem;border-radius:8px;font-size:.9rem;font-weight:600;cursor:pointer;border:none;text-decoration:none;display:inline-block}
    .btn-save{background:#16a34a;color:white}
    .btn-cancel{background:#f0f4f8;color:#4a5568;text-decoration:none}
    a{color:#2b6cb0}
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link active">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>✏️ Edit Booking #{{ b.id }}</h1>
  <a href="/admin/booking/{{ b.id }}" class="pg-back">← Cancel</a>
</div>
<div class="container">
<form method="POST" action="/admin/booking/{{ b.id }}/edit">

  <div class="card">
    <h2>Customer</h2>
    <div class="field-grid">
      <div style="position:relative">
        <label>Full Name</label>
        <input name="full_name" id="eb_name" value="{{ b.full_name or '' }}" autocomplete="off">
        <ul id="eb-suggestions" style="display:none;position:absolute;top:100%;left:0;right:0;z-index:999;background:white;border:1px solid #cbd5e1;border-radius:6px;box-shadow:0 4px 12px rgba(0,0,0,.15);margin:2px 0;padding:0;list-style:none;max-height:200px;overflow-y:auto"></ul>
      </div>
      <div><label>Company</label><input name="company_name" id="eb_company" value="{{ b.company_name or '' }}"></div>
      <div><label>Email</label><input name="email" type="email" id="eb_email" value="{{ b.email or '' }}"></div>
      <div><label>Phone</label><input name="phone" id="eb_phone" value="{{ b.phone or '' }}"></div>
      <div><label>Street</label><input name="renter_street" id="eb_street" value="{{ b.renter_street or '' }}"></div>
      <div><label>City</label><input name="renter_city" id="eb_city" value="{{ b.renter_city or '' }}"></div>
      <div><label>State</label><input name="renter_state" id="eb_state" value="{{ b.renter_state or '' }}"></div>
      <div><label>ZIP</label><input name="renter_zip" id="eb_zip" value="{{ b.renter_zip or '' }}"></div>
    </div>
  </div>

  <div class="card">
    <h2>Event</h2>
    <div class="field-grid">
      <div><label>Start Date</label><input name="event_start_date" type="date" value="{{ b.event_start_date or '' }}"></div>
      <div><label>End Date</label><input name="event_end_date" type="date" value="{{ b.event_end_date or '' }}"></div>
      <div><label>Event Start Time</label><input name="event_start_time" type="time" value="{{ b.event_start_time or '' }}"></div>
      <div><label>Pickup Time</label><input name="event_end_time" type="time" value="{{ b.event_end_time or '' }}"></div>
      <div><label>Delivery Date</label><input name="setup_date" type="date" value="{{ b.setup_date.strftime('%Y-%m-%d') if b.setup_date else '' }}"></div>
      <div><label>Delivery Time</label><input name="setup_time" type="time" value="{{ b.setup_time or '' }}"></div>
      <div><label>Venue Type</label>
        <select name="venue_type">
          <option value="venue"        {% if b.venue_type=='venue'        %}selected{% endif %}>Venue</option>
          <option value="residential"  {% if b.venue_type=='residential'  %}selected{% endif %}>Residential</option>
        </select>
      </div>
    </div>
    <div class="field-grid" style="margin-top:.75rem">
      <div><label>Event Street</label><input name="event_street" id="eb_event_street" value="{{ b.event_street or '' }}" autocomplete="off"></div>
      <div><label>Event City</label><input name="event_city" id="eb_event_city" value="{{ b.event_city or '' }}"></div>
      <div><label>Event State</label><input name="event_state" id="eb_event_state" value="{{ b.event_state or '' }}"></div>
      <div><label>Event ZIP</label><input name="event_zip" id="eb_event_zip" value="{{ b.event_zip or '' }}"></div>
      <div class="single" style="grid-column:1/-1"><label>Delivery Location / Notes on Venue</label><input name="delivery_location" value="{{ b.delivery_location or '' }}"></div>
    </div>
  </div>

  <div class="card">
    <h2>Financials & Status</h2>
    <div class="field-grid">
      <div><label>Status</label>
        <select name="status">
          <option value="pending"       {% if b.status=='pending'       %}selected{% endif %}>Pending</option>
          <option value="agree_to_pay" {% if b.status=='agree_to_pay' %}selected{% endif %}>Agree to Pay (Cash/Check at Delivery)</option>
          <option value="accepted"  {% if b.status=='accepted'  %}selected{% endif %}>Accepted (Awaiting Payment)</option>
          <option value="concluded" {% if b.status=='concluded' %}selected{% endif %}>Concluded</option>
          <option value="partial"   {% if b.status=='partial'   %}selected{% endif %}>Confirmed (Partial Payment)</option>
          <option value="denied"    {% if b.status=='denied'    %}selected{% endif %}>Denied</option>
          <option value="cancelled" {% if b.status=='cancelled' %}selected{% endif %}>Cancelled</option>
        </select>
      </div>
      <div><label>Grand Total ($)</label><input name="grand_total" type="number" step="0.01" value="{{ b.grand_total or 0 }}"></div>
      <div><label>Amount Paid ($)</label><input name="amount_paid" type="number" step="0.01" value="{{ b.amount_paid or 0 }}"></div>
      <div><label>Delivery Fee ($)</label><input name="delivery_fee" type="number" step="0.01" value="{{ b.delivery_fee or 0 }}"></div>
      <div><label>Late Night Fee ($)</label><input name="late_night_fee" type="number" step="0.01" value="{{ b.late_night_fee or 0 }}"></div>
      <div><label>Distance (miles)</label><input name="distance_miles" type="number" step="0.1" value="{{ b.distance_miles or '' }}"></div>
    </div>
  </div>

  <div class="card">
    <h2>Notes</h2>
    <div class="field-grid single">
      <div><textarea name="notes">{{ b.notes or '' }}</textarea></div>
    </div>
  </div>

  <div class="actions">
    <button type="submit" class="btn btn-save">💾 Save Changes</button>
    <a href="/admin/booking/{{ b.id }}" class="btn btn-cancel">Cancel</a>
  </div>
</form>
</div>
<script>
(function() {
  const nameInput   = document.getElementById('eb_name');
  const suggestions = document.getElementById('eb-suggestions');
  if (!nameInput) return;
  const LI_STYLE = 'padding:.55rem .85rem;cursor:pointer;font-size:.92rem;border-bottom:1px solid #f1f5f9';
  let debounce;

  nameInput.addEventListener('input', function() {
    clearTimeout(debounce);
    const q = this.value.trim();
    if (q.length < 1) { suggestions.style.display='none'; return; }
    debounce = setTimeout(() => {
      fetch('/admin/customer-search?q=' + encodeURIComponent(q))
        .then(r => r.json())
        .then(data => {
          suggestions.innerHTML = '';
          if (!data.length) { suggestions.style.display='none'; return; }
          data.forEach(c => {
            const li = document.createElement('li');
            li.style.cssText = LI_STYLE;
            li.innerHTML = '<strong>' + c.full_name + '</strong>' +
              (c.email ? ' <span style="color:#64748b;font-size:.82rem">— ' + c.email + '</span>' : '');
            li.addEventListener('mousedown', function(e) {
              e.preventDefault();
              nameInput.value = c.full_name;
              if (c.email)         document.getElementById('eb_email').value   = c.email;
              if (c.phone)         document.getElementById('eb_phone').value   = c.phone;
              if (c.company_name)  document.getElementById('eb_company').value = c.company_name;
              if (c.renter_street) document.getElementById('eb_street').value  = c.renter_street;
              if (c.renter_city)   document.getElementById('eb_city').value    = c.renter_city;
              if (c.renter_state)  document.getElementById('eb_state').value   = c.renter_state;
              if (c.renter_zip)    document.getElementById('eb_zip').value     = c.renter_zip;
              suggestions.style.display = 'none';
            });
            suggestions.appendChild(li);
          });
          suggestions.style.display = 'block';
        })
        .catch(() => { suggestions.style.display='none'; });
    }, 250);
  });

  nameInput.addEventListener('blur', function() {
    setTimeout(() => { suggestions.style.display='none'; }, 150);
  });
})();

// ── Google Maps Places Autocomplete for Edit Booking Event Address ─────────────
function initEditEventAutocomplete() {
  var streetEl = document.getElementById('eb_event_street');
  if (!streetEl || !window.google) return;
  var ac = new google.maps.places.Autocomplete(streetEl, {
    types: ['address'],
    componentRestrictions: { country: 'us' },
    fields: ['address_components']
  });
  streetEl.addEventListener('keydown', function(e) {
    if (e.key === 'Enter') e.preventDefault();
  });
  ac.addListener('place_changed', function() {
    var place = ac.getPlace();
    if (!place.address_components) return;
    var streetNum = '', route = '', city = '', state = '', zip = '';
    place.address_components.forEach(function(comp) {
      var t = comp.types;
      if (t.includes('street_number'))                    streetNum = comp.long_name;
      else if (t.includes('route'))                       route     = comp.long_name;
      else if (t.includes('locality'))                    city      = comp.long_name;
      else if (t.includes('administrative_area_level_1')) state     = comp.short_name;
      else if (t.includes('postal_code'))                 zip       = comp.long_name;
    });
    streetEl.value = [streetNum, route].filter(Boolean).join(' ');
    document.getElementById('eb_event_city').value  = city;
    document.getElementById('eb_event_state').value = state;
    document.getElementById('eb_event_zip').value   = zip;
  });
}
</script>
{% if google_maps_key %}
<script src="https://maps.googleapis.com/maps/api/js?key={{ google_maps_key }}&libraries=places&callback=initEditEventAutocomplete" async defer></script>
{% endif %}
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


ADMIN_NEW_BOOKING_HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Book a Rental — {{ business_name }}</title>
  <style>
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#1a202c;min-height:100vh;display:flex}
    .sidebar{width:200px;min-height:100vh;background:#1e1e2e;display:flex;flex-direction:column;position:fixed;top:0;left:0;z-index:100;transition:transform .2s}
    .sb-brand{padding:1.1rem 1rem .9rem;display:flex;align-items:center;gap:.55rem;border-bottom:1px solid rgba(255,255,255,.08)}
    .sb-brand img{height:1.8rem;width:auto;object-fit:contain}
    .sb-brand-name{font-size:.82rem;font-weight:700;color:#fff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
    .sb-new-btn{display:block;margin:.85rem .85rem .5rem;padding:.55rem .75rem;background:#16a34a;color:white;border-radius:8px;font-size:.84rem;font-weight:700;text-decoration:none;text-align:center}
    .sb-new-btn:hover{background:#15803d}
    .sb-nav{display:flex;flex-direction:column;padding:.25rem 0;flex:1}
    .sb-link{display:flex;align-items:center;gap:.6rem;padding:.6rem 1rem;font-size:.84rem;font-weight:500;color:rgba(255,255,255,.55);text-decoration:none;transition:all .1s;border-left:3px solid transparent}
    .sb-link:hover{background:rgba(255,255,255,.07);color:rgba(255,255,255,.9)}
    .sb-icon{width:1.1rem;text-align:center;font-size:.95rem}
    .sb-divider{height:1px;background:rgba(255,255,255,.07);margin:.4rem 0}
    .sb-bottom{border-top:1px solid rgba(255,255,255,.08);padding:.5rem 0}
    .page-content{margin-left:200px;flex:1;min-height:100vh;display:flex;flex-direction:column}
    .page-header{background:white;border-bottom:1px solid #e5e7eb;padding:.85rem 1.5rem;display:flex;align-items:center;gap:1rem;position:sticky;top:0;z-index:50}
    .page-header h1{font-size:1.3rem;font-weight:700;color:#111827;flex:1}
    .mobile-menu-btn{display:none;background:none;border:none;font-size:1.4rem;cursor:pointer;color:#374151;padding:.25rem}
    .page-body{padding:1.5rem;flex:1}
    @media(max-width:768px){.sidebar{transform:translateX(-100%)}.sidebar.open{transform:translateX(0);box-shadow:6px 0 30px rgba(0,0,0,.4)}.page-content{margin-left:0}.mobile-menu-btn{display:block}}
    .container{max-width:720px;margin:0 auto;padding:0 0 4rem}
    .card{background:white;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.08);padding:1.75rem;margin-bottom:1.5rem}
    .card h2{font-size:1rem;font-weight:700;color:#2b6cb0;border-bottom:2px solid #ebf4ff;padding-bottom:.6rem;margin-bottom:1.25rem;text-transform:uppercase;letter-spacing:.5px}
    .field{margin-bottom:1rem}
    .field label{display:block;font-size:.85rem;font-weight:600;color:#4a5568;margin-bottom:.3rem}
    .field input,.field select,.field textarea{width:100%;padding:.65rem .85rem;border:1.5px solid #cbd5e0;border-radius:8px;font-size:.97rem;color:#1a202c;background:#fff;transition:border-color .15s}
    .field input:focus,.field select:focus,.field textarea:focus{outline:none;border-color:#2b6cb0;box-shadow:0 0 0 3px rgba(43,108,176,.12)}
    .field textarea{resize:vertical;min-height:80px}
    .row{display:grid;gap:1rem;grid-template-columns:1fr 1fr}
    .row3{display:grid;gap:1rem;grid-template-columns:2fr 1fr 1fr}
    @media(max-width:560px){.row,.row3{grid-template-columns:1fr}}
    .required{color:#e53e3e}
    .section-note{font-size:.82rem;color:#718096;margin-bottom:1rem;font-style:italic}
    .type-toggle{display:flex;gap:.75rem;margin-bottom:1rem}
    .type-btn{flex:1;padding:.65rem;border:2px solid #cbd5e0;border-radius:8px;background:white;font-size:.9rem;font-weight:600;color:#718096;cursor:pointer;text-align:center;transition:all .15s}
    .type-btn.active{border-color:#2b6cb0;background:#ebf4ff;color:#2b6cb0}
    .exact-toggle{display:flex;align-items:center;gap:.75rem;padding:1rem;background:#fffaf0;border:2px solid #ed8936;border-radius:10px;cursor:pointer;margin-bottom:.75rem}
    .exact-toggle input[type=checkbox]{width:20px;height:20px;cursor:pointer;accent-color:#2b6cb0}
    .exact-label{flex:1}
    .exact-label strong{display:block;font-size:.97rem;color:#1a202c}
    .exact-label span{font-size:.82rem;color:#718096}
    .exact-badge{background:#ed8936;color:white;padding:.2rem .6rem;border-radius:20px;font-size:.8rem;font-weight:700}
    .product-row{display:grid;grid-template-columns:1fr auto auto;align-items:center;gap:.75rem;padding:.85rem 0;border-bottom:1px solid #f0f4f8}
    .product-row:last-child{border-bottom:none}
    .product-name{font-weight:600;font-size:.95rem}
    .product-meta{display:flex;gap:1rem;font-size:.8rem;margin-top:.15rem}
    .product-price{color:#718096}
    .avail-badge{font-weight:600}
    .avail-badge.ok{color:#38a169}
    .avail-badge.low{color:#d69e2e}
    .avail-badge.out{color:#e53e3e}
    .qty-control{display:flex;align-items:center;border:1.5px solid #cbd5e0;border-radius:8px;overflow:hidden}
    .qty-btn{background:#f7fafc;border:none;width:34px;height:36px;font-size:1.1rem;color:#2b6cb0;cursor:pointer;transition:background .12s}
    .qty-btn:hover{background:#ebf4ff}
    .qty-input{width:52px;border:none;border-left:1.5px solid #cbd5e0;border-right:1.5px solid #cbd5e0;text-align:center;font-size:.95rem;font-weight:600;padding:.4rem .2rem;outline:none}
    .product-sub{text-align:right;min-width:70px;font-weight:600;color:#718096;font-size:.95rem}
    .product-sub.has-val{color:#2b6cb0}
    .total-bar{background:linear-gradient(135deg,#1a365d,#2b6cb0);border-radius:12px;padding:1.25rem 1.75rem;color:white;margin-bottom:1.5rem}
    .total-row{display:flex;justify-content:space-between;padding:.2rem 0;font-size:.95rem;opacity:.85}
    .total-row.grand{font-size:1.35rem;font-weight:700;opacity:1;border-top:1px solid rgba(255,255,255,.25);margin-top:.5rem;padding-top:.6rem}
    .total-note{font-size:.78rem;opacity:.7;margin-top:.5rem;font-style:italic}
    .alert{background:#fff5f5;border:1px solid #feb2b2;color:#c53030;padding:.85rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem}
    .submit-btn{width:100%;padding:1rem;background:linear-gradient(135deg,#1a365d,#2b6cb0);color:white;border:none;border-radius:10px;font-size:1.1rem;font-weight:700;cursor:pointer;transition:opacity .15s}
    .submit-btn:hover{opacity:.92}
    .submit-btn:disabled{opacity:.6;cursor:not-allowed}
  </style>
</head>
<body>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand">
    <img src="/logo.png" alt="">
    <span class="sb-brand-name">{{ business_name }}</span>
  </div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link"><span class="sb-icon">🏠</span> Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link"><span class="sb-icon">👥</span> Clients</a>
    <a href="/admin/inventory" class="sb-link"><span class="sb-icon">📦</span> Inventory</a>
    <a href="/admin/calendar" class="sb-link"><span class="sb-icon">📅</span> Calendar</a>
    <a href="/admin/reports" class="sb-link"><span class="sb-icon">📊</span> Reports</a>
    <a href="/admin/route" class="sb-link"><span class="sb-icon">🗺</span> Route</a>
    <a href="/driver/{{ today }}" class="sb-link"><span class="sb-icon">🚚</span> Driver View</a>
    <a href="/admin/formsite-import" class="sb-link"><span class="sb-icon">📥</span> Import</a>
    <a href="/admin/tax-report" class="sb-link"><span class="sb-icon">💰</span> Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/download-backup" class="sb-link"><span class="sb-icon">💾</span> Backup</a>
    <a href="/admin/logout" class="sb-link"><span class="sb-icon">🚪</span> Sign Out</a>
  </div>
</aside>
<div class="page-content">
  <div class="page-header">
    <button class="mobile-menu-btn" onclick="document.getElementById('sidebar').classList.toggle('open')">☰</button>
    <h1>New Booking</h1>
    <a href="/admin/dashboard" style="font-size:.85rem;color:#2563eb;text-decoration:none">← Dashboard</a>
  </div>
  <div class="page-body">
<div class="container">
{% if error %}<div class="alert">{{ error }}</div>{% endif %}
<form method="POST" action="/admin/booking/new" id="bookingForm">
  <input type="hidden" name="admin_create" value="1">
  <div class="card">
    <h2>Your Information</h2>
    <div class="row">
      <div class="field"><label>Full Name <span class="required">*</span></label><input name="full_name" required placeholder="Jane Smith" value="{{ form.full_name or '' }}"></div>
      <div class="field"><label>Company Name <span style="color:#718096;font-weight:400">(if applicable)</span></label><input name="company_name" placeholder="ABC Events LLC" value="{{ form.company_name or '' }}"></div>
    </div>
    <div class="field"><label>Street Address <span class="required">*</span></label><input name="renter_street" required placeholder="123 Main Street" value="{{ form.renter_street or '' }}"></div>
    <div class="row3">
      <div class="field"><label>City <span class="required">*</span></label><input name="renter_city" required placeholder="Hartford" value="{{ form.renter_city or '' }}"></div>
      <div class="field"><label>State <span class="required">*</span></label><input name="renter_state" required placeholder="CT" maxlength="2" value="{{ form.renter_state or '' }}"></div>
      <div class="field"><label>Zip <span class="required">*</span></label><input name="renter_zip" required placeholder="06101" value="{{ form.renter_zip or '' }}"></div>
    </div>
    <div class="row">
      <div class="field"><label>Phone <span class="required">*</span></label><input name="phone" type="tel" required placeholder="(555) 000-0000" value="{{ form.phone or '' }}"></div>
      <div class="field"><label>Email <span class="required">*</span></label><input name="email" type="email" required placeholder="jane@email.com" value="{{ form.email or '' }}"></div>
    </div>
    <div style="margin-top:1rem;padding:.75rem 1rem;background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px;display:flex;align-items:flex-start;gap:.75rem">
      <input type="checkbox" name="tax_exempt_request" id="tax_exempt_request" value="1" onchange="updateTotals()" style="width:18px;height:18px;margin-top:.15rem;accent-color:#16a34a;cursor:pointer;flex-shrink:0">
      <label for="tax_exempt_request" style="cursor:pointer;font-size:.9rem;color:#166534;font-weight:600;line-height:1.4">
        I have a Connecticut Tax-Exempt Certificate
        <span style="display:block;font-weight:400;font-size:.8rem;color:#4b7c5a;margin-top:.1rem">Check this if your organization is tax-exempt. You will need to provide your certificate number to the rental office.</span>
      </label>
    </div>
  </div>

  <div class="card">
    <h2>Event Details</h2>

    {%- set time_opts -%}
    <option value="">-- Select --</option>
    <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
    <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
    <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
    <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
    <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
    <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
    <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
    <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
    <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
    <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
    <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
    <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
    <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
    <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
    <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
    <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
    <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
    <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
    {%- endset -%}

    <div class="row">
      <div class="field"><label>Event Start Date <span class="required">*</span></label><input id="event_start_date" name="event_start_date" type="date" required onchange="onDateChange()" value="{{ form.event_start_date or '' }}"></div>
      <div class="field"><label>Event Start Time <span class="required">*</span></label>
        <select name="event_start_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field"><label>Event End Date <span class="required">*</span></label><input id="event_end_date" name="event_end_date" type="date" required onchange="onDateChange()" value="{{ form.event_end_date or '' }}"></div>
      <div class="field"><label>Event End Time <span class="required">*</span></label>
        <select name="event_end_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field"><label>Setup Date <span class="required">*</span></label><input name="setup_date" type="date" required value="{{ form.setup_date or '' }}" id="setupDateEl" onchange="checkDeliveryBeforeEvent()"></div>
      <div class="field"><label>Setup Time <span class="required">*</span></label>
        <select name="setup_time" required style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field">
        <label>🚚 Delivery Date</label>
        <input name="delivery_date" id="deliveryDateEl" type="date" value="{{ form.delivery_date or '' }}"
               style="width:100%;border:1px solid #93c5fd;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#eff6ff">
      </div>
      <div class="field">
        <label>🚚 Delivery Time</label>
        <select name="delivery_time" id="deliveryTimeEl" style="width:100%;border:1px solid #93c5fd;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#eff6ff;color:#1a202c">
          <option value="">-- Select --</option>
          {{ time_opts }}
        </select>
      </div>
    </div>
    <div class="row">
      <div class="field">
        <label>📦 Pickup Date</label>
        <input name="pickup_date" id="pickupDateEl" type="date" value="{{ form.pickup_date or '' }}"
               style="width:100%;border:1px solid #86efac;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#f0fdf4">
      </div>
      <div class="field">
        <label>📦 Pickup Time</label>
        <select name="pickup_time" id="pickupTimeEl" style="width:100%;border:1px solid #86efac;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#f0fdf4;color:#1a202c">
          <option value="">-- Select --</option>
          {{ time_opts }}
        </select>
      </div>
    </div>
    <script>
    function doWeekendSchedule(){
      var s=document.getElementById('event_start_date');
      if(!s||!s.value){alert('Enter Event Start Date first.');return;}
      var d=new Date(s.value+'T00:00:00');
      var w=d.getDay();
      var dd,pu,label;
      if(w===6){
        dd=new Date(d);dd.setDate(d.getDate()-1);
        pu=new Date(d);pu.setDate(d.getDate()+1);
        label='Saturday event: Deliver Friday, Pickup Sunday';
      }else if(w===0){
        dd=new Date(d);dd.setDate(d.getDate()-2);
        pu=new Date(d);pu.setDate(d.getDate()+1);
        label='Sunday event: Deliver Friday, Pickup Monday';
      }else{alert('Event start is not Saturday or Sunday.');return;}
      var fmt=function(x){return x.toISOString().split('T')[0];};
      var ddEl=document.getElementById('deliveryDateEl');
      if(ddEl)ddEl.value=fmt(dd);
      var dtEl=document.getElementById('deliveryTimeEl');
      if(dtEl){for(var i=0;i<dtEl.options.length;i++){if(dtEl.options[i].value==='16:00'){dtEl.value='16:00';break;}}}
      var puEl=document.getElementById('pickupDateEl');
      if(puEl)puEl.value=fmt(pu);
      var ptEl=document.getElementById('pickupTimeEl');
      if(ptEl){for(var i=0;i<ptEl.options.length;i++){if(ptEl.options[i].value==='10:00'){ptEl.value='10:00';break;}}}
      var edEl=document.getElementById('event_end_date');
      if(edEl)edEl.value=fmt(pu);
      var msg=document.getElementById('weekend-msg');
      if(msg){msg.textContent='✓ '+label;msg.style.display='inline';}
    }
    </script>
    <div style="margin-bottom:1.25rem">
      <button type="button" onclick="doWeekendSchedule()"
              style="background:#1d4ed8;color:white;border:none;border-radius:7px;padding:.5rem 1.1rem;font-size:.85rem;font-weight:600;cursor:pointer">
        📅 Apply Weekend Schedule
      </button>
      <span id="weekend-msg" style="margin-left:.75rem;font-size:.8rem;color:#059669;display:none"></span>
    </div>

    <!-- Early delivery acknowledgment — shown only when delivery date is before event date -->
    <div id="early-delivery-notice" style="display:none;background:#fff7ed;border:2px solid #f97316;border-radius:10px;padding:1rem 1.15rem;margin-bottom:1rem">
      <div style="display:flex;align-items:flex-start;gap:.75rem">
        <span style="font-size:1.5rem;line-height:1">📦</span>
        <div style="flex:1">
          <div style="font-weight:700;color:#c2410c;font-size:1rem;margin-bottom:.35rem">Your delivery is scheduled BEFORE your event date</div>
          <div style="font-size:.9rem;color:#7c2d12;line-height:1.5;margin-bottom:.75rem">
            Your event starts on <strong id="notice-event-date"></strong>, but you requested delivery on <strong id="notice-delivery-date"></strong>.
          </div>
          <label style="display:flex;align-items:flex-start;gap:.6rem;cursor:pointer;background:#ffedd5;border:1px solid #fb923c;border-radius:7px;padding:.65rem .85rem">
            <input type="checkbox" id="early_delivery_ack" name="early_delivery_ack" value="1"
              onchange="checkDeliveryAck()"
              style="width:20px;height:20px;margin-top:.1rem;accent-color:#ea580c;flex-shrink:0;cursor:pointer">
            <span style="font-size:.88rem;font-weight:600;color:#9a3412;line-height:1.45">
              I understand that my rental items will be delivered on <strong id="notice-ack-date"></strong> — before my event — and I approve this early delivery.
            </span>
          </label>
        </div>
      </div>
    </div>

    <div class="field">
      <label>Venue Type <span class="required">*</span></label>
      <div class="type-toggle">
        <div class="type-btn active" id="btn_venue" onclick="setVenue('venue')">Venue</div>
        <div class="type-btn" id="btn_residential" onclick="setVenue('residential')">Residential</div>
      </div>
      <input type="hidden" name="venue_type" id="venue_type_input" value="venue">
    </div>
    <div id="venue_pickup_row" class="field"><label>Latest Pickup Time at Venue <span class="required">*</span></label>
      <select id="venue_latest_pickup" name="venue_latest_pickup" style="width:100%;border:1px solid #d1d5db;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
        {{ time_opts }}
      </select>
    </div>
  </div>

  <div class="card">
    <h2>Event Address</h2>
    <p class="section-note">Where will we deliver your rental items?</p>
    <div class="field"><label>Street Address <span class="required">*</span></label><input id="event_street" name="event_street" required placeholder="456 Venue Blvd" value="{{ form.event_street or '' }}" oninput="scheduleDistanceCalc()"></div>
    <div class="row3">
      <div class="field"><label>City <span class="required">*</span></label><input id="event_city" name="event_city" required placeholder="Hartford" value="{{ form.event_city or '' }}" oninput="scheduleDistanceCalc()"></div>
      <div class="field"><label>State <span class="required">*</span></label><input id="event_state" name="event_state" required placeholder="CT" maxlength="2" value="{{ form.event_state or '' }}" oninput="scheduleDistanceCalc()"></div>
      <div class="field"><label>Zip <span class="required">*</span></label><input id="event_zip" name="event_zip" required placeholder="06101" value="{{ form.event_zip or '' }}" oninput="scheduleDistanceCalc()"></div>
    </div>
  </div>

  <div class="card">
    <h2>Delivery Options</h2>
    <label class="exact-toggle">
      <input type="checkbox" id="exact_time_cb" name="exact_time_delivery" value="yes" onchange="updateTotals();toggleExactTimes(this.checked)">
      <div class="exact-label"><strong>Exact Time Delivery</strong><span>Guaranteed delivery at your specified setup time</span></div>
      <span class="exact-badge">+$175</span>
    </label>

    <div id="exact_time_fields" style="display:none;margin-top:1rem;background:#fff8f0;border:1px solid #f6ad55;border-radius:10px;padding:1rem 1.1rem">
      <p style="margin:0 0 .75rem;font-size:.88rem;color:#92400e;font-weight:600">⏰ Exact Time Details</p>
      <div class="row">
        <div class="field">
          <label>Exact Delivery Time <span class="required">*</span></label>
          <select name="exact_delivery_time" id="exact_delivery_time" style="width:100%;border:1px solid #f6ad55;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
            <option value="">-- Select delivery time --</option>
            <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
            <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
            <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
            <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
            <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
            <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
            <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
            <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
            <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
            <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
            <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
            <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
            <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
            <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
            <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
            <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
            <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
            <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
          </select>
        </div>
        <div class="field">
          <label>Exact Pickup Time <span class="required">*</span></label>
          <select name="exact_pickup_time" id="exact_pickup_time" style="width:100%;border:1px solid #f6ad55;border-radius:8px;padding:.55rem .75rem;font-size:1rem;background:#fff;color:#1a202c">
            <option value="">-- Select pickup time --</option>
            <option value="06:00">6:00 AM</option><option value="06:30">6:30 AM</option>
            <option value="07:00">7:00 AM</option><option value="07:30">7:30 AM</option>
            <option value="08:00">8:00 AM</option><option value="08:30">8:30 AM</option>
            <option value="09:00">9:00 AM</option><option value="09:30">9:30 AM</option>
            <option value="10:00">10:00 AM</option><option value="10:30">10:30 AM</option>
            <option value="11:00">11:00 AM</option><option value="11:30">11:30 AM</option>
            <option value="12:00">12:00 PM</option><option value="12:30">12:30 PM</option>
            <option value="13:00">1:00 PM</option><option value="13:30">1:30 PM</option>
            <option value="14:00">2:00 PM</option><option value="14:30">2:30 PM</option>
            <option value="15:00">3:00 PM</option><option value="15:30">3:30 PM</option>
            <option value="16:00">4:00 PM</option><option value="16:30">4:30 PM</option>
            <option value="17:00">5:00 PM</option><option value="17:30">5:30 PM</option>
            <option value="18:00">6:00 PM</option><option value="18:30">6:30 PM</option>
            <option value="19:00">7:00 PM</option><option value="19:30">7:30 PM</option>
            <option value="20:00">8:00 PM</option><option value="20:30">8:30 PM</option>
            <option value="21:00">9:00 PM</option><option value="21:30">9:30 PM</option>
            <option value="22:00">10:00 PM</option><option value="22:30">10:30 PM</option>
            <option value="23:00">11:00 PM</option><option value="23:30">11:30 PM</option>
          </select>
        </div>
      </div>
    </div>

    <script>
    function toggleExactTimes(checked) {
      var el = document.getElementById('exact_time_fields');
      el.style.display = checked ? 'block' : 'none';
      document.getElementById('exact_delivery_time').required = checked;
      document.getElementById('exact_pickup_time').required = checked;
    }
    </script>

    <div class="field" style="margin-top:1rem"><label>Where on the premises will items be delivered? <span class="required">*</span></label><textarea name="delivery_location" required placeholder="e.g. Through the main entrance, set up in the ballroom on the left side...">{{ form.delivery_location or '' }}</textarea></div>
  </div>

  <!-- Hidden qty inputs — submitted with form -->
  {% for p in products %}
  <input type="hidden" class="qty-input" id="qty_{{ p.id }}" name="qty_{{ p.id }}" value="0" data-price="{{ p.price }}" data-max="{{ p.total }}">
  {% endfor %}

  <!-- Product data for JS -->
  <script>
  const ALL_PRODUCTS = [
    {% for p in products %}
    { id:"{{ p.id }}", name:{{ p.name | tojson }}, price:{{ p.price }}, max:{{ p.total }} },
    {% endfor %}
  ];
  </script>

  <div class="card">
    <h2>Select Your Items</h2>
    <p style="color:#6b7280;font-size:.88rem;margin-bottom:1.25rem">Click a category below to browse items. Select an item to add it to your order.</p>

    <!-- Category accordion dropdowns -->
    <div id="category-dropdowns"></div>

    <!-- Selected items list -->
    <div id="selected-items-wrap" style="display:none;margin-top:1.25rem;border-top:2px solid #e5e7eb;padding-top:1rem">
      <div style="font-weight:700;font-size:.95rem;color:#1a202c;margin-bottom:.75rem">🛒 Your Items</div>
      <div id="marquee-tier-notice" style="display:none;background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.88rem;color:#1e40af"></div>
      <div id="selected-items-list"></div>
    </div>
  </div>

  <!-- Stackable Marquee Option — shown only when marquee items are in cart -->
  <div id="stackable-section" style="display:none;margin:.75rem 0;padding:1.1rem 1.25rem;background:#faf5ff;border:1.5px solid #d8b4fe;border-radius:10px">
    <div style="font-weight:700;font-size:.95rem;color:#6b21a8;margin-bottom:.5rem">🔡 Stackable Marquee</div>
    <p style="font-size:.88rem;color:#4c1d95;margin-bottom:.85rem;line-height:1.6">
      Would you like your marquee letters/numbers to be <strong>stackable</strong> (stacked on top of each other)?
      <span style="background:#ede9fe;color:#7c3aed;border-radius:4px;padding:.1rem .45rem;font-size:.8rem;font-weight:700;margin-left:.3rem">+$75 fee</span>
    </p>
    <div style="display:flex;gap:1rem;margin-bottom:.85rem">
      <label style="display:flex;align-items:center;gap:.4rem;cursor:pointer;font-weight:600;color:#1a202c">
        <input type="radio" name="stackable_choice" value="yes" onchange="onStackableChange()"
          style="accent-color:#7c3aed;width:16px;height:16px"> Yes, I want stackable (+$75)
      </label>
      <label style="display:flex;align-items:center;gap:.4rem;cursor:pointer;font-weight:600;color:#1a202c">
        <input type="radio" name="stackable_choice" value="no" onchange="onStackableChange()" checked
          style="accent-color:#7c3aed;width:16px;height:16px"> No thanks
      </label>
    </div>
    <div id="stackable-top-wrap" style="display:none">
      <label style="font-size:.88rem;font-weight:600;color:#1a202c;display:block;margin-bottom:.35rem">
        Which letters or numbers go on top? <span style="color:#dc2626">*</span>
      </label>
      <input type="text" id="stackable_top_display" placeholder="e.g. A, E, O — or 2, 5"
        oninput="document.getElementById('stackable_top_input').value=this.value"
        style="width:100%;border:1.5px solid #c4b5fd;border-radius:8px;padding:.5rem .75rem;font-size:.9rem;box-sizing:border-box">
    </div>
  </div>
  <input type="hidden" name="stackable" id="stackable_input" value="no">
  <input type="hidden" name="stackable_top" id="stackable_top_input" value="">

  <input type="hidden" name="late_night_fee" id="late_night_fee_input" value="0">
  <div id="late_night_notice" style="display:none;margin:.75rem 0;padding:.75rem 1rem;background:#fef3c7;border:1.5px solid #fcd34d;border-radius:8px;font-size:.88rem;color:#92400e">
    <strong>⏰ Late Night / Early Morning Fee: $125.00</strong><br>
    Your pickup or dropoff time falls between 11:30 PM – 7:00 AM. A $125 fee applies for pickups or deliveries outside of standard hours.
  </div>
  <div class="total-bar">
    <div class="total-row"><span>Items Subtotal</span><span id="t_items">$0.00</span></div>
    <div class="total-row"><span>Exact Time Delivery</span><span id="t_exact">-</span></div>
    <div class="total-row" id="t_stackable_row" style="display:none"><span>Stackable Marquee</span><span id="t_stackable">$75.00</span></div>
    <div class="total-row" id="t_latenight_row" style="display:none"><span>Late Night / Early Morning Fee</span><span id="t_latenight">$125.00</span></div>
    <div class="total-row"><span>Delivery Fee</span><span id="t_delivery">Calculated after review</span></div>
    <div class="total-row"><span>CT Sales Tax (6.35%)</span><span id="t_tax">$0.00</span></div>
    <div class="total-row grand"><span>Estimated Total</span><span id="t_grand">$0.00</span></div>
    <p class="total-note">Live estimate — delivery fee calculated from event address.</p>
  </div>



    <div class="card" style="margin-bottom:1.25rem">
    <h2 style="margin-bottom:.75rem;color:#1e40af">⚙️ Booking Status</h2>
    <div class="row">
      <div class="field">
        <label>Status</label>
        <select name="status" style="width:100%;padding:.6rem .75rem;border:1px solid #e2e8f0;border-radius:8px;font-size:.95rem">
          <option value="accepted">Accepted — Awaiting Payment</option>
          <option value="accepted_paid">Accepted — Paid in Full</option>
          <option value="partial">Accepted — Partial Payment</option>
          <option value="agree_to_pay">Agree to Pay (Cash/Check at Delivery)</option>
          <option value="pending">Pending Review</option>
        </select>
      </div>
      <div class="field">
        <label>Amount Already Paid ($)</label>
        <input type="number" name="amount_paid" step="0.01" min="0" placeholder="0.00" value="0"
               style="width:100%;padding:.6rem .75rem;border:1px solid #e2e8f0;border-radius:8px;font-size:.95rem">
      </div>
    </div>
    <div class="field" style="margin-top:.5rem">
      <label>Admin Notes (private)</label>
      <textarea name="notes" rows="2" placeholder="Phone call notes, special instructions…"
        style="width:100%;padding:.6rem .75rem;border:1px solid #e2e8f0;border-radius:8px;font-size:.95rem;resize:vertical"></textarea>
    </div>
  </div>
  <button type="submit" class="submit-btn" id="submitBtn" style="background:linear-gradient(135deg,#1e40af,#2563eb)">➕ Create Booking</button>
</form>
</div>
<script>
const EXACT_FEE = {{ exact_time_fee }};
const LATE_NIGHT_FEE = 125;
const CT_TAX_RATE = 0.0635;
function isLateNight(timeStr){
  if(!timeStr) return false;
  const [h,m]=timeStr.split(':').map(Number);
  const mins=h*60+m;
  return mins>=1410||mins<420;
}
function checkLateNightFee(){
  const endTime=document.querySelector('[name="event_end_time"]').value;
  const late=isLateNight(endTime);
  document.getElementById('late_night_notice').style.display=late?'block':'none';
  document.getElementById('t_latenight_row').style.display=late?'flex':'none';
  document.getElementById('late_night_fee_input').value=late?LATE_NIGHT_FEE:0;
  return late?LATE_NIGHT_FEE:0;
}

// ── Marquee Tier Pricing ──────────────────────────────────────────
const MARQUEE_NUMBER_TIERS = [
  { count:1, total:80 },
  { count:2, total:150 },
  { count:3, total:215 },
  { count:4, total:275 },
];
function getMarqueeNumberTotal(n){
  if(n<=0) return 0;
  const t=MARQUEE_NUMBER_TIERS.find(x=>x.count===n);
  if(t) return t.total;
  return 275+(n-4)*55;
}
const MARQUEE_LETTER_TIERS = [
  { count:1, total:85 },
  { count:2, total:160 },
  { count:3, total:225 },
  { count:4, total:285 },
];
function getMarqueeLetterTotal(n){
  if(n<=0) return 0;
  const t=MARQUEE_LETTER_TIERS.find(x=>x.count===n);
  if(t) return t.total;
  return 285+(n-4)*55;
}
function isMarqueeNumber(name){ return /^marquee\s+#?\d/i.test(name); }
function isMarqueeLetter(name){ return /^marquee\s+[a-z]$/i.test(name); }

// ── Item Categories ───────────────────────────────────────────────
const ITEM_CATEGORIES = [
  { label:"🪑 Chairs",           keywords:["chair","stool","bench","seat","chiavari"] },
  { label:"🪣 Tables",           keywords:["table","tablecloth","linen","cloth","runner","overlay","skirt"] },
  { label:"🔢 Marquee Numbers",  keywords:["marquee number"] },
  { label:"🔤 Marquee Letters",  keywords:["marquee letter"] },
  { label:"💡 Lighting",         keywords:["light","lamp","led","glow","neon","bulb","lantern","fairy","chandelier","uplighting"] },
  { label:"🎭 Backdrops & Décor",keywords:["backdrop","banner","balloon","arch","flower","floral","decor","sign","drape","curtain","pillar","column","centerpiece","vase","frame","wall"] },
  { label:"🎪 Entertainment",    keywords:["bounce","slide","game","popcorn","cotton candy","machine","photo booth","casino","carnival","inflatable"] },
  { label:"⛺ Tents & Canopies", keywords:["tent","canopy","pergola","gazebo","umbrella"] },
];
function getCat(name){
  const n=name.trim();
  // Marquee Letter: "Marquee A", "Marquee B", etc. — marquee followed by a single letter
  if(/^marquee\s+[a-zA-Z]$/i.test(n) || /marquee\s+[a-zA-Z]\s*$/i.test(n)) return "🔤 Marquee Letters";
  // Marquee Number: "Marquee #5", "Marquee 3", etc. — marquee followed by # or digit
  if(/marquee\s+#?\d/i.test(n)) return "🔢 Marquee Numbers";
  const nl=n.toLowerCase();
  for(const c of ITEM_CATEGORIES){ if(c.keywords.some(k=>nl.includes(k))) return c.label; }
  return "📦 Other";
}

// ── Build Category Accordion Dropdowns ───────────────────────────
function buildDropdowns(){
  // Group products by category
  const groups={};
  ALL_PRODUCTS.forEach(p=>{
    const cat=getCat(p.name);
    if(!groups[cat]) groups[cat]=[];
    groups[cat].push(p);
  });
  const wrap=document.getElementById('category-dropdowns');
  wrap.innerHTML='';
  const CAT_ORDER=['🪑 Chairs','🪣 Tables','🔤 Marquee Letters','🔢 Marquee Numbers'];
  const sorted=Object.entries(groups).sort(([a],[b])=>{
    if(a==='📦 Other') return 1;
    if(b==='📦 Other') return -1;
    const ai=CAT_ORDER.indexOf(a), bi=CAT_ORDER.indexOf(b);
    if(ai>=0&&bi>=0) return ai-bi;
    if(ai>=0) return -1;
    if(bi>=0) return 1;
    return a.localeCompare(b);
  });
  sorted.forEach(([cat,items])=>{
    const sec=document.createElement('div');
    sec.style.cssText='border:1px solid #e5e7eb;border-radius:8px;margin-bottom:.5rem;overflow:hidden';
    sec.innerHTML=`
      <button type="button" onclick="toggleCat(this)"
        style="width:100%;text-align:left;background:#f9fafb;border:none;padding:.75rem 1rem;font-size:.95rem;font-weight:700;color:#1a202c;cursor:pointer;display:flex;justify-content:space-between;align-items:center">
        <span>${cat} <span style="font-size:.8rem;font-weight:400;color:#9ca3af">(${items.length} item${items.length!==1?'s':''})</span></span>
        <span class="chev" style="transition:transform .2s;font-size:.8rem">▼</span>
      </button>
      <div class="cat-body" style="display:none;padding:.75rem 1rem;background:white">
        ${cat==='🔢 Marquee Numbers'?`<div style="background:#fefce8;border:1px solid #fde047;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.83rem;color:#713f12">
          <strong>📋 Tier Pricing:</strong> 1 for $80 &nbsp;·&nbsp; 2 for $150 &nbsp;·&nbsp; 3 for $215 &nbsp;·&nbsp; 4 for $275 &nbsp;·&nbsp; 5+ = $275 + $55 each additional
        </div>`:''}
        ${cat==='🔤 Marquee Letters'?`<div style="background:#fefce8;border:1px solid #fde047;border-radius:8px;padding:.6rem .9rem;margin-bottom:.75rem;font-size:.83rem;color:#713f12">
          <strong>📋 Tier Pricing:</strong> 1 for $85 &nbsp;·&nbsp; 2 for $160 &nbsp;·&nbsp; 3 for $225 &nbsp;·&nbsp; 4 for $285 &nbsp;·&nbsp; 5+ = $285 + $55 each additional
        </div>`:''}
        <div style="display:flex;flex-wrap:wrap;gap:.5rem">
          ${items.map(p=>`
            <button type="button" onclick="addToCart('${p.id}')"
              data-id="${p.id}"
              style="background:#eff6ff;color:#1d4ed8;border:1.5px solid #bfdbfe;border-radius:20px;padding:.35rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer;transition:background .15s"
              onmouseover="this.style.background='#dbeafe'" onmouseout="this.style.background='#eff6ff'">
              ${p.name} — $${p.price.toFixed(2)}
            </button>`).join('')}
        </div>
      </div>`;
    wrap.appendChild(sec);
  });
}
function toggleCat(btn){
  const body=btn.nextElementSibling;
  const open=body.style.display==='block';
  body.style.display=open?'none':'block';
  btn.querySelector('.chev').style.transform=open?'':'rotate(180deg)';
  btn.style.background=open?'#f9fafb':'#eff6ff';
}

// ── Cart ──────────────────────────────────────────────────────────
const cart={};  // id -> qty
function addToCart(id){
  const p=ALL_PRODUCTS.find(x=>x.id===id);
  if(!p) return;
  if(cart[id]){ setQty(id, cart[id]+1); }
  else { cart[id]=1; renderCart(); }
  setHiddenQty(id, cart[id]);
  updateTotals();
  // Highlight the button
  const btn=document.querySelector(`button[data-id="${id}"]`);
  if(btn){ btn.style.background='#bbf7d0'; btn.style.borderColor='#4ade80'; btn.style.color='#166534';
    setTimeout(()=>{ btn.style.background='#eff6ff'; btn.style.borderColor='#bfdbfe'; btn.style.color='#1d4ed8'; },600); }
}
function setQty(id, val){
  const p=ALL_PRODUCTS.find(x=>x.id===id);
  if(!p) return;
  const v=Math.max(0, val);
  if(v===0){ delete cart[id]; }
  else { cart[id]=v; }
  const inp=document.getElementById('cart-qty-'+id);
  if(inp) inp.value=v===0?'':v;
  setHiddenQty(id, v);
  renderCart();
  updateTotals();
}
function setHiddenQty(id, val){
  const h=document.getElementById('qty_'+id);
  if(h) h.value=val;
}
function removeFromCart(id){
  delete cart[id];
  setHiddenQty(id,0);
  renderCart();
  updateTotals();
}
const STACKABLE_FEE = 75;
function hasMarqueeInCart(){
  return Object.keys(cart).some(id=>{
    const p=ALL_PRODUCTS.find(x=>x.id===id);
    return p && (isMarqueeLetter(p.name)||isMarqueeNumber(p.name));
  });
}
function onStackableChange(){
  const sel=document.querySelector('input[name="stackable_choice"]:checked');
  const yes=sel&&sel.value==='yes';
  document.getElementById('stackable_input').value=yes?'yes':'no';
  document.getElementById('stackable-top-wrap').style.display=yes?'block':'none';
  if(!yes){ document.getElementById('stackable_top_input').value=''; const d=document.getElementById('stackable_top_display'); if(d) d.value=''; }
  updateTotals();
}
function renderCart(){
  const list=document.getElementById('selected-items-list');
  const wrap=document.getElementById('selected-items-wrap');
  const ids=Object.keys(cart);
  if(ids.length===0){
    wrap.style.display='none'; list.innerHTML='';
    // Hide stackable section and reset when cart is empty
    const ss=document.getElementById('stackable-section');
    if(ss) ss.style.display='none';
    document.getElementById('stackable_input').value='no';
    const noRadio=document.querySelector('input[name="stackable_choice"][value="no"]');
    if(noRadio){ noRadio.checked=true; }
    document.getElementById('stackable-top-wrap').style.display='none';
    return;
  }
  wrap.style.display='block';
  // Show stackable section only when marquee items are in cart
  const ss=document.getElementById('stackable-section');
  if(ss) ss.style.display=hasMarqueeInCart()?'block':'none';
  if(!hasMarqueeInCart()){
    document.getElementById('stackable_input').value='no';
    const noRadio=document.querySelector('input[name="stackable_choice"][value="no"]');
    if(noRadio){ noRadio.checked=true; }
    document.getElementById('stackable-top-wrap').style.display='none';
  }
  // Calculate total marquee numbers for proration
  let mnCount=0, mlCount=0;
  ids.forEach(id=>{ const p=ALL_PRODUCTS.find(x=>x.id===id); if(!p) return;
    if(isMarqueeNumber(p.name)) mnCount+=cart[id]||0;
    else if(isMarqueeLetter(p.name)) mlCount+=cart[id]||0;
  });
  const mnTierTotal=getMarqueeNumberTotal(mnCount);
  const mlTierTotal=getMarqueeLetterTotal(mlCount);
  const mnUnitPrice=mnCount>0?(mnTierTotal/mnCount):0;
  const mlUnitPrice=mlCount>0?(mlTierTotal/mlCount):0;
  list.innerHTML=ids.map(id=>{
    const p=ALL_PRODUCTS.find(x=>x.id===id);
    const q=cart[id]||1;
    const isMN=isMarqueeNumber(p.name);
    const isML=isMarqueeLetter(p.name);
    const unitPrice=isMN?mnUnitPrice:isML?mlUnitPrice:p.price;
    const lineTotal=(unitPrice*q).toFixed(2);
    const tierTag=(isMN||isML)?` <span style="font-size:.75rem;color:#2563eb;font-weight:600">(tier)</span>`:'';
    const unitLabel=`$${unitPrice.toFixed(2)} ea${tierTag}`;
    return `<div style="display:flex;align-items:center;gap:.75rem;padding:.6rem .5rem;border-bottom:1px solid #f3f4f6">
      <span style="flex:1;font-size:.92rem;font-weight:600;color:#1a202c">${p.name}</span>
      <span style="font-size:.82rem;color:#6b7280;white-space:nowrap">${unitLabel}</span>
      <div style="display:flex;align-items:center;gap:.3rem">
        <button type="button" onclick="setQty('${id}',${q-1})"
          style="width:28px;height:28px;border:1px solid #d1d5db;border-radius:6px;background:white;font-size:1rem;cursor:pointer;line-height:1">−</button>
        <input id="cart-qty-${id}" type="number" value="${q}" min="1" max="9999"
          onchange="setQty('${id}',parseInt(this.value)||1)"
          style="width:44px;text-align:center;border:1px solid #d1d5db;border-radius:6px;padding:.2rem;font-size:.9rem;font-weight:700">
        <button type="button" onclick="setQty('${id}',${q+1})"
          style="width:28px;height:28px;border:1px solid #d1d5db;border-radius:6px;background:white;font-size:1rem;cursor:pointer;line-height:1">+</button>
      </div>
      <span style="font-size:.9rem;font-weight:700;color:#2563eb;min-width:52px;text-align:right">$${lineTotal}</span>
      <button type="button" onclick="removeFromCart('${id}')"
        style="background:none;border:none;color:#9ca3af;font-size:1.1rem;cursor:pointer;padding:.1rem .3rem" title="Remove">✕</button>
    </div>`;
  }).join('');
}
function updateTotals(){
  let regularSub=0, mnCount=0, mlCount=0;
  ALL_PRODUCTS.forEach(p=>{
    const qty=cart[p.id]||0;
    if(!qty) return;
    if(isMarqueeNumber(p.name)) mnCount+=qty;
    else if(isMarqueeLetter(p.name)) mlCount+=qty;
    else regularSub+=qty*p.price;
  });
  const mnSub=getMarqueeNumberTotal(mnCount);
  const mlSub=getMarqueeLetterTotal(mlCount);
  const sub=regularSub+mnSub+mlSub;
  // Update tier notices in cart
  const tierEl=document.getElementById('marquee-tier-notice');
  if(tierEl){
    let html='';
    if(mnCount>0){
      const nextN=MARQUEE_NUMBER_TIERS.find(t=>t.count===mnCount+1);
      const savN=nextN?` <span style="color:#16a34a;font-size:.82rem">· Add 1 more for $${nextN.total} total</span>`:'';
      html+=`🔢 <strong>${mnCount} Marquee Number${mnCount!==1?'s':''}</strong> — Tier Price: <strong>$${mnSub.toFixed(2)}</strong>${savN}<br>`;
    }
    if(mlCount>0){
      const nextL=MARQUEE_LETTER_TIERS.find(t=>t.count===mlCount+1);
      const savL=nextL?` <span style="color:#16a34a;font-size:.82rem">· Add 1 more for $${nextL.total} total</span>`:'';
      html+=`🔤 <strong>${mlCount} Marquee Letter${mlCount!==1?'s':''}</strong> — Tier Price: <strong>$${mlSub.toFixed(2)}</strong>${savL}`;
    }
    tierEl.innerHTML=html;
    tierEl.style.display=(mnCount>0||mlCount>0)?'block':'none';
  }
  const exactCb=document.getElementById('exact_time_cb');
  const ef=exactCb&&exactCb.checked?EXACT_FEE:0;
  const lf=checkLateNightFee();
  const stackableSel=document.querySelector('input[name="stackable_choice"]:checked');
  const sf=(stackableSel&&stackableSel.value==='yes'&&hasMarqueeInCart())?STACKABLE_FEE:0;
  const stackRow=document.getElementById('t_stackable_row');
  if(stackRow) stackRow.style.display=sf>0?'flex':'none';
  const exemptCb=document.getElementById('tax_exempt_request');
  const exempt=exemptCb&&exemptCb.checked;
  const df=typeof _calcDeliveryFee!=='undefined'?_calcDeliveryFee:0;
  const tax=exempt?0:(sub+ef+sf+lf+df)*CT_TAX_RATE;
  const taxEl=document.getElementById('t_tax');
  if(taxEl){ taxEl.textContent='$'+tax.toFixed(2); taxEl.style.color=exempt?'#16a34a':'';
    const lbl=taxEl.previousElementSibling; if(lbl) lbl.textContent=exempt?'CT Sales Tax (EXEMPT)':'CT Sales Tax (6.35%)'; }
  document.getElementById('t_items').textContent='$'+sub.toFixed(2);
  document.getElementById('t_exact').textContent=ef>0?'$'+ef.toFixed(2):'-';
  document.getElementById('t_grand').textContent='$'+(sub+ef+sf+lf+df+tax).toFixed(2)+'+';
}
document.addEventListener('DOMContentLoaded', buildDropdowns);
function setVenue(type){document.getElementById('venue_type_input').value=type;document.getElementById('btn_venue').classList.toggle('active',type==='venue');document.getElementById('btn_residential').classList.toggle('active',type==='residential');const row=document.getElementById('venue_pickup_row');const inp=document.getElementById('venue_latest_pickup');row.style.display=type==='venue'?'block':'none';inp.required=type==='venue';}
setVenue('venue');
function onDateChange(){const start=document.getElementById('event_start_date').value;const end=document.getElementById('event_end_date').value;if(!start||!end||end<start)return;fetch('/availability?start='+start+'&end='+end).then(r=>r.json()).then(data=>{ALL_PRODUCTS.forEach(p=>{if(data[p.id]!==undefined){p.max=data[p.id];}});updateTotals();}).catch(()=>{});checkDeliveryBeforeEvent();}
function fmtDateNice(d){if(!d)return'';const p=d.split('-');if(p.length!==3)return d;const months=['January','February','March','April','May','June','July','August','September','October','November','December'];return months[parseInt(p[1],10)-1]+' '+parseInt(p[2],10)+', '+p[0];}
function checkDeliveryBeforeEvent(){
  const ed=document.getElementById('event_start_date').value;
  const sd=document.getElementById('setupDateEl').value;
  const notice=document.getElementById('early-delivery-notice');
  const ack=document.getElementById('early_delivery_ack');
  if(!notice)return;
  if(sd&&ed&&sd<ed){
    document.getElementById('notice-event-date').textContent=fmtDateNice(ed);
    document.getElementById('notice-delivery-date').textContent=fmtDateNice(sd);
    document.getElementById('notice-ack-date').textContent=fmtDateNice(sd);
    // Calculate days before
    const diff=Math.round((new Date(ed)-new Date(sd))/(1000*60*60*24));
    document.getElementById('notice-days-before').textContent=diff===1?'1 day':diff+' days';
    notice.style.display='';
  } else {
    notice.style.display='none';
    if(ack)ack.checked=false;
  }
}
function checkDeliveryAck(){
  // no-op, submit handler checks state
}
function applyWeekendDelivery(){
  const startDate = document.getElementById('event_start_date');
  if (!startDate || !startDate.value) {
    alert('Please enter the Event Start Date first.');
    return;
  }
  const esd = new Date(startDate.value + 'T00:00:00');
  const wd = esd.getDay(); // 0=Sun, 6=Sat
  let deliveryDate, pickupDate, label;
  if (wd === 6) { // Saturday
    deliveryDate = new Date(esd); deliveryDate.setDate(esd.getDate() - 1);
    pickupDate   = new Date(esd); pickupDate.setDate(esd.getDate() + 1);
    label = 'Saturday event → Deliver Friday, Pickup Sunday';
  } else if (wd === 0) { // Sunday
    deliveryDate = new Date(esd); deliveryDate.setDate(esd.getDate() - 2);
    pickupDate   = new Date(esd); pickupDate.setDate(esd.getDate() + 1);
    label = 'Sunday event → Deliver Friday, Pickup Monday';
  } else {
    alert('Event start date is not a Saturday or Sunday. Weekend schedule only applies to weekend events.');
    return;
  }
  const fmt = d => d.toISOString().split('T')[0];
  const ddEl = document.getElementById('deliveryDateEl');
  if (ddEl) ddEl.value = fmt(deliveryDate);
  const dtEl = document.getElementById('deliveryTimeEl');
  if (dtEl) { for (let o of dtEl.options) { if (o.value === '16:00') { dtEl.value = '16:00'; break; } } }
  const endDateEl = document.getElementById('event_end_date');
  if (endDateEl && !endDateEl.value) endDateEl.value = fmt(pickupDate);
  const msg = document.getElementById('weekend-msg');
  if (msg) { msg.textContent = '✓ ' + label; msg.style.display = 'inline'; }
}
let distTimer;
let _calcDeliveryFee=0;
function scheduleDistanceCalc(){clearTimeout(distTimer);distTimer=setTimeout(()=>{const street=document.getElementById('event_street').value;const city=document.getElementById('event_city').value;const state=document.getElementById('event_state').value;const zip=document.getElementById('event_zip').value;if(street&&city&&state&&zip){const addr=street+', '+city+', '+state+' '+zip;fetch('/delivery_fee?address='+encodeURIComponent(addr)).then(r=>r.json()).then(d=>{document.getElementById('t_delivery').textContent='$'+d.fee.toFixed(2)+' ('+d.note+')';_calcDeliveryFee=d.fee;updateTotals();}).catch(()=>{});}},800);}
// ── Date validation ───────────────────────────────────────────────────────
const today=new Date().toISOString().split('T')[0];
const startDateEl = document.getElementById('event_start_date');
const endDateEl   = document.getElementById('event_end_date');
const startTimeEl = document.querySelector('[name="event_start_time"]');
const endTimeEl   = document.querySelector('[name="event_end_time"]');
const setupTimeEl = document.querySelector('[name="setup_time"]');

startDateEl.min = today;
endDateEl.min   = today;

// When start date changes, end date must be >= start date
startDateEl.addEventListener('change', function() {
  endDateEl.min = this.value;
  if (endDateEl.value && endDateEl.value < this.value) {
    endDateEl.value = this.value;
  }
  onDateChange();
});

endDateEl.addEventListener('change', function() {
  if (this.value < startDateEl.value) {
    this.value = startDateEl.value;
    showTimeError('End date cannot be before start date.');
  }
  onDateChange();
});

// When start time changes:
//   - end time rule only applies if start date == end date (same-day event)
//   - delivery (setup) time must always be before start time (delivery is always same day as start)
startTimeEl.addEventListener('change', function() {
  const sameDay = startDateEl.value && endDateEl.value && startDateEl.value === endDateEl.value;
  if (sameDay && endTimeEl.value && endTimeEl.value <= this.value) {
    endTimeEl.value = '';
    showTimeError('Event end time must be after start time.');
  }
  if (setupTimeEl.value && setupTimeEl.value >= this.value) {
    setupTimeEl.value = '';
    showTimeError('Setup time must be before event start time.');
  }
});

endTimeEl.addEventListener('change', function() {
  const sameDay = startDateEl.value && endDateEl.value && startDateEl.value === endDateEl.value;
  if (sameDay && startTimeEl.value && this.value <= startTimeEl.value) {
    this.value = '';
    showTimeError('Event end time must be after the start time.');
  }
  updateTotals();
});

setupTimeEl.addEventListener('change', function() {
  const setupDate = document.getElementById('setupDateEl');
  const sameDay = setupDate && setupDate.value && startDateEl.value && setupDate.value === startDateEl.value;
  if (sameDay && startTimeEl.value && this.value >= startTimeEl.value) {
    this.value = '';
    showTimeError('Setup / delivery time must be before the event start time when delivery is on the same day.');
  }
});

function showTimeError(msg) {
  let el = document.getElementById('time_error');
  if (!el) {
    el = document.createElement('div');
    el.id = 'time_error';
    el.style.cssText = 'background:#fff5f5;border:1px solid #feb2b2;color:#c53030;padding:.75rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem';
    document.getElementById('submitBtn').before(el);
  }
  el.textContent = msg;
  setTimeout(() => { if (el) el.textContent = ''; }, 4000);
}

// ── Form submit validation ────────────────────────────────────────────────
document.getElementById('bookingForm').addEventListener('submit', function(e) {
  const errors = [];
  const sd = startDateEl.value, ed = endDateEl.value;
  const st = startTimeEl.value, et = endTimeEl.value, sut = setupTimeEl.value;
  const sameDay = sd && ed && sd === ed;

  if (sd && ed && ed < sd)   errors.push('End date cannot be before start date.');
  // End time vs start time only matters when event starts and ends the same day
  if (sameDay && st && et && et <= st)  errors.push('Event end time must be after the start time.');
  // Delivery time vs start time only matters when setup date == event start date
  const setupDateEl = document.getElementById('setupDateEl');
  const setupSameDay = setupDateEl && setupDateEl.value && sd && setupDateEl.value === sd;
  if (setupSameDay && sut && st && sut >= st) errors.push('Setup / delivery time must be before the event start time when delivery is on the same day.');
  // Block if early delivery not acknowledged
  const earlyNotice = document.getElementById('early-delivery-notice');
  const earlyAck = document.getElementById('early_delivery_ack');
  if (earlyNotice && earlyNotice.style.display !== 'none' && earlyAck && !earlyAck.checked) {
    errors.push('Please check the box acknowledging your early delivery date before submitting.');
  }

  if (errors.length) {
    e.preventDefault();
    showTimeError(errors.join(' '));
    return;
  }
  const btn = document.getElementById('submitBtn');
  btn.disabled = true;
  btn.textContent = 'Submitting...';
});
</script>
<script>
function initPublicEventAutocomplete() {
  var streetEl = document.getElementById('event_street');
  if (!streetEl || !window.google) return;
  var ac = new google.maps.places.Autocomplete(streetEl, {
    types: ['address'],
    componentRestrictions: { country: 'us' },
    fields: ['address_components']
  });
  streetEl.addEventListener('keydown', function(e) {
    if (e.key === 'Enter') e.preventDefault();
  });
  ac.addListener('place_changed', function() {
    var place = ac.getPlace();
    if (!place.address_components) return;
    var streetNum = '', route = '', city = '', state = '', zip = '';
    place.address_components.forEach(function(comp) {
      var t = comp.types;
      if (t.includes('street_number'))                    streetNum = comp.long_name;
      else if (t.includes('route'))                       route     = comp.long_name;
      else if (t.includes('locality'))                    city      = comp.long_name;
      else if (t.includes('administrative_area_level_1')) state     = comp.short_name;
      else if (t.includes('postal_code'))                 zip       = comp.long_name;
    });
    streetEl.value = [streetNum, route].filter(Boolean).join(' ');
    var cityEl  = document.getElementById('event_city');
    var stateEl = document.getElementById('event_state');
    var zipEl   = document.getElementById('event_zip');
    if (cityEl)  cityEl.value  = city;
    if (stateEl) stateEl.value = state;
    if (zipEl)   zipEl.value   = zip;
    scheduleDistanceCalc();
  });
}
</script>
{% if google_maps_key %}
<script src="https://maps.googleapis.com/maps/api/js?key={{ google_maps_key }}&libraries=places&callback=initPublicEventAutocomplete" async defer></script>
{% endif %}
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</div><!-- /container -->
  </div><!-- /page-body -->
</div><!-- /page-content -->
</body></html>
"""


ADMIN_BOOKING_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Booking #{{ b.id }} — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,sans-serif;background:#f0f4f8;color:#1a202c;min-height:100vh}
    header{background:linear-gradient(135deg,#1a365d,#2b6cb0);color:white;padding:1.25rem 2rem;display:flex;justify-content:space-between;align-items:center}
    header h1{font-size:1.2rem}
    .container{max-width:1100px;margin:0 auto;padding:.85rem 1rem .5rem}
    .card{background:white;border-radius:12px;box-shadow:0 1px 8px rgba(0,0,0,.06);padding:1.25rem 1.5rem;margin-bottom:0}
    .card h2{font-size:.82rem;font-weight:700;color:#6b7280;border-bottom:1px solid #f1f5f9;padding-bottom:.45rem;margin-bottom:.85rem;text-transform:uppercase;letter-spacing:.05em}
    .row{display:grid;grid-template-columns:140px 1fr;gap:.4rem .65rem;font-size:.88rem}
    .row .k{color:#9ca3af;font-size:.82rem}
    .row .v{font-weight:500;color:#111827}
    .badge{display:inline-block;padding:.2rem .7rem;border-radius:20px;font-size:.75rem;font-weight:700;text-transform:uppercase;margin-bottom:0}
    @media(max-width:768px){[style*="grid-template-columns:1fr 360px"]{grid-template-columns:1fr!important}[style*="position:sticky"]{position:relative!important}}
    .badge-pending{background:#fefcbf;color:#975a16}
    .badge-accepted{background:#bee3f8;color:#2c5282}
    .badge-confirmed{background:#bee3f8;color:#2c5282}
    .badge-partial{background:#bee3f8;color:#2c5282}
    .badge-denied{background:#fbd38d;color:#744210}
    .badge-cancelled{background:#fed7d7;color:#9b2c2c}
    .badge-concluded{background:#e2e8f0;color:#475569}
    table{width:100%;border-collapse:collapse;font-size:.9rem}
    th{padding:8px 10px;text-align:left;color:#718096;font-size:.78rem;text-transform:uppercase;border-bottom:1px solid #e2e8f0}
    td{padding:8px 10px;border-bottom:1px solid #f0f4f8}
    .total-row{font-weight:700;background:#1a365d;color:white}
    .total-row td{padding:10px}
    .actions{display:flex;gap:.75rem;flex-wrap:wrap;margin-top:1.5rem}
    .btn{padding:.65rem 1.25rem;border-radius:8px;font-size:.9rem;font-weight:600;cursor:pointer;border:none;text-decoration:none;display:inline-block}
    .btn-back{background:#f0f4f8;color:#4a5568}
    .btn-accept{background:#38a169;color:white}
    .btn-deny{background:#e53e3e;color:white}
    .btn-confirm{background:#3182ce;color:white}
    .btn-cancel{background:#718096;color:white}
    .btn-reminder{background:linear-gradient(135deg,#c05621,#dd6b20);color:white}
    .payment-link-box{background:#f0fff4;border:2px solid #38a169;border-radius:10px;padding:1.1rem 1.25rem;margin-bottom:1rem}
    .alert{background:#fffaf0;border-left:4px solid #ed8936;padding:.85rem 1rem;border-radius:0 8px 8px 0;font-size:.9rem;margin-bottom:1rem}
    a{color:#2b6cb0}
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link active">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr" style="flex-wrap:wrap;gap:.5rem;padding:.6rem 1rem">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <!-- ── Sticky page header ── -->
  <div style="display:flex;align-items:center;gap:.75rem;flex-wrap:wrap;flex:1">
    <div>
      <h1 style="font-size:1.1rem;font-weight:700;color:#111827;margin:0">Booking #{{ b.id }}</h1>
      <div style="font-size:.75rem;color:#9ca3af;margin-top:.1rem">{{ b.created_at|string|truncate(19,True,'') }}</div>
    </div>
    <div style="display:flex;gap:.4rem;align-items:center;flex-wrap:wrap">
      {% if b.status == "agree_to_pay" %}
        <span class="badge badge-agree_to_pay" style="margin:0">AGREE TO PAY</span>
      {% elif b.status == "accepted" and b.payment_status == "paid" %}
        <span class="badge badge-accepted" style="margin:0;background:#dcfce7;color:#166534">✅ PAID IN FULL</span>
      {% elif b.status == "accepted" and b.payment_status == "partial" %}
        <span class="badge badge-accepted" style="margin:0;background:#fef9c3;color:#854d0e">💳 DEPOSIT PAID</span>
      {% else %}
        <span class="badge badge-{{ b.status }}" style="margin:0">{{ b.status|upper }}</span>
      {% endif %}
      {% if b.delivery_status == 'delivered' %}
      <span style="background:#fffbeb;color:#92400e;border:1.5px solid #fcd34d;border-radius:20px;padding:.2rem .75rem;font-size:.75rem;font-weight:700">🚚 DELIVERED</span>
      {% elif b.delivery_status == 'picked_up' %}
      <span style="background:#f0fdf4;color:#15803d;border:1.5px solid #86efac;border-radius:20px;padding:.2rem .75rem;font-size:.75rem;font-weight:700">✅ PICKED UP</span>
      {% endif %}
    </div>
    {% if b.status == 'agree_to_pay' %}
    <div style="margin-left:auto;display:flex;gap:.5rem;flex-wrap:wrap;align-items:center">
      <span style="background:#d1fae5;color:#065f46;border:1.5px solid #6ee7b7;border-radius:20px;padding:.25rem .9rem;font-size:.8rem;font-weight:700">💵 Agree to Pay at Delivery</span>
      <form method="POST" action="/admin/booking/{{ b.id }}/accept">
        <button type="submit"
          style="background:#2563eb;color:#fff;border:none;border-radius:8px;padding:.45rem 1rem;font-size:.84rem;font-weight:700;cursor:pointer;white-space:nowrap">
          ✅ Convert to Stripe
        </button>
      </form>
      <form method="POST" action="/admin/booking/{{ b.id }}/deny">
        <button onclick="return confirm('Deny this booking?')"
          style="background:#fff;color:#dc2626;border:1.5px solid #dc2626;border-radius:8px;padding:.45rem .85rem;font-size:.84rem;font-weight:700;cursor:pointer;white-space:nowrap">
          ✕ Deny
        </button>
      </form>
    </div>
    {% endif %}
    {% if b.status == 'pending' %}
    <div style="margin-left:auto;display:flex;gap:.5rem;flex-wrap:wrap">
      <form id="accept-form" method="POST" action="/admin/booking/{{ b.id }}/accept">
        <input type="hidden" name="custom_amount" id="accept-amount-input">
        <button type="button" id="accept-btn"
          style="background:#16a34a;color:#fff;border:none;border-radius:8px;padding:.45rem 1.1rem;font-size:.88rem;font-weight:700;cursor:pointer;white-space:nowrap">
          ✅ Accept — Send Invoice
        </button>
      </form>
      <form method="POST" action="/admin/booking/{{ b.id }}/deny">
        <button onclick="return confirm('Deny this booking? A rejection email will be sent to {{ b.email }}.')"
          style="background:#fff;color:#dc2626;border:1.5px solid #dc2626;border-radius:8px;padding:.45rem 1rem;font-size:.88rem;font-weight:700;cursor:pointer;white-space:nowrap">
          ✕ Deny
        </button>
      </form>
    </div>
    {% endif %}
    {% if b.status in ('accepted', 'confirmed') and b.payment_status in ('waiting', None, '') %}
    <div style="margin-left:auto;display:flex;gap:.5rem;flex-wrap:wrap;align-items:center">
      <form id="accept-form" method="POST" action="/admin/booking/{{ b.id }}/accept">
        <input type="hidden" name="custom_amount" id="accept-amount-input">
        <button type="button" id="accept-btn"
          style="background:#16a34a;color:#fff;border:none;border-radius:8px;padding:.45rem 1.1rem;font-size:.88rem;font-weight:700;cursor:pointer;white-space:nowrap">
          📧 Send Invoice + Payment Link
        </button>
      </form>
    </div>
    {% endif %}
    <a href="/admin/booking/{{ b.id }}/edit" style="margin-left:{% if b.status not in ('pending', 'accepted', 'confirmed') or b.payment_status not in ('waiting', None, '') %}auto{% else %}0{% endif %};font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap;border:1px solid #e5e7eb;border-radius:6px;padding:.3rem .75rem">✏️ Edit</a>
  </div>
</div>

<!-- ── Two-column layout ── -->
<div style="max-width:1100px;margin:0 auto;padding:1rem 1rem 2rem;display:grid;grid-template-columns:1fr 360px;gap:1.25rem;align-items:start">

<!-- ══ LEFT COLUMN ══ -->
<div style="display:flex;flex-direction:column;gap:1.1rem">

  {% if b.status == 'accepted' and b.payment_status in ('waiting', None, '') %}
  <div class="payment-link-box">
    <div style="font-weight:700;color:#276749;margin-bottom:.4rem">⏳ Awaiting Deposit Payment</div>
    {% if b.stripe_payment_link %}
    <p style="font-size:.9rem;color:#4a5568;margin-bottom:.5rem">Payment link sent to {{ b.email }}:</p>
    <a href="{{ b.stripe_payment_link }}" target="_blank" style="word-break:break-all;font-size:.85rem">{{ b.stripe_payment_link }}</a>
    {% else %}
    <p style="font-size:.9rem;color:#744210">No payment link generated yet. Use "Send Invoice + Payment Link" to email one.</p>
    {% endif %}
  </div>
  {% endif %}

  {% if booking_inv_issues %}
  <div style="background:#fef2f2;border:2px solid #f87171;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1rem">
    <div style="font-weight:700;color:#dc2626;font-size:1rem;margin-bottom:.5rem">🚨 Inventory Shortage — Review Before Accepting</div>
    {% for c in booking_inv_issues %}
    <div style="margin-bottom:.75rem">
      <div style="font-size:.9rem;color:#7f1d1d;margin-bottom:.4rem">
        <strong>{{ c.item }}</strong>: customer needs <strong>{{ c.needed }}</strong>,
        {% if c.get('absolute') %}only <strong>{{ c.available }}</strong> in stock total
        <span style="background:#7f1d1d;color:white;border-radius:4px;padding:.1rem .45rem;font-size:.78rem;font-weight:700;margin-left:.4rem">EXCEEDS TOTAL STOCK by {{ c.shortfall }}</span>
        {% else %}only <strong>{{ c.available }}</strong> available after other bookings
        <span style="background:#fee2e2;color:#b91c1c;border-radius:4px;padding:.1rem .45rem;font-size:.78rem;font-weight:700;margin-left:.4rem">{{ c.shortfall }} short</span>
        {% endif %}
      </div>
      {% if c.conflicting_bookings %}
      <div style="font-size:.8rem;color:#991b1b;margin-bottom:.3rem;font-weight:600">Bookings using this item on overlapping dates:</div>
      <div style="display:flex;flex-wrap:wrap;gap:.4rem">
        {% for cb in c.conflicting_bookings %}
        <a href="/admin/booking/{{ cb.id }}"
           style="display:inline-flex;align-items:center;gap:.35rem;background:#fff;border:1px solid #f87171;border-radius:6px;padding:.3rem .65rem;font-size:.8rem;color:#b91c1c;text-decoration:none;font-weight:600"
           title="{{ cb.name }} — {{ cb.start }} to {{ cb.end }} ({{ cb.qty }} units)">
          #{{ cb.id }} {{ cb.name }}
          <span style="color:#6b7280;font-weight:400">{{ cb.start }}{% if cb.end and cb.end != cb.start %} → {{ cb.end }}{% endif %}</span>
          <span style="background:#fecaca;border-radius:3px;padding:.05rem .3rem">&times;{{ cb.qty }}</span>
          <span style="font-size:.75rem;background:#dc2626;color:white;border-radius:4px;padding:.05rem .35rem;margin-left:.1rem">View →</span>
        </a>
        {% endfor %}
      </div>
      {% endif %}
    </div>
    {% endfor %}
    <div style="font-size:.82rem;color:#991b1b;margin-top:.4rem;border-top:1px solid #fca5a5;padding-top:.6rem">You may need to source more inventory or contact the customer to adjust quantities.</div>
  </div>
  {% endif %}

  {% if booking_inv_status %}
  <div style="background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:1rem 1.25rem;margin-bottom:1rem">
    <div style="font-weight:700;font-size:.88rem;color:#374151;margin-bottom:.65rem">📦 Inventory Status for This Booking</div>
    <div style="display:flex;flex-direction:column;gap:.4rem">
    {% for s in booking_inv_status %}
    <div style="display:flex;align-items:center;justify-content:space-between;padding:.45rem .75rem;border-radius:7px;background:{% if s.ok %}#f0fdf4{% else %}#fef2f2{% endif %};border:1px solid {% if s.ok %}#bbf7d0{% else %}#fca5a5{% endif %}">
      <span style="font-weight:600;font-size:.88rem;color:{% if s.ok %}#166534{% else %}#991b1b{% endif %}">
        {% if not s.ok %}🚨{% else %}✅{% endif %} {{ s.item }}
      </span>
      <span style="font-size:.82rem;color:#6b7280">
        Needs <strong style="color:{% if s.ok %}#166534{% else %}#dc2626{% endif %}">{{ s.needed }}</strong>
        &nbsp;·&nbsp; {{ s.reserved }} reserved by others
        &nbsp;·&nbsp; <strong>{{ s.available }}</strong> of {{ s.total }} available
        {% if not s.ok %}<span style="background:#dc2626;color:white;border-radius:4px;padding:.1rem .4rem;font-size:.75rem;font-weight:700;margin-left:.4rem">{{ s.shortfall }} SHORT</span>{% endif %}
      </span>
    </div>
    {% endfor %}
    </div>
  </div>
  {% endif %}

  {% if b.status == 'pending' %}
  <div class="alert">
    This booking is waiting for your review. Click Accept to send the customer their invoice, contract, and Stripe payment link.
    Click Deny to send a polite rejection.
  </div>
  {% endif %}

  {% if matched_customer %}
  {% set mc = matched_customer %}
  <div style="background:#fffbeb;border:1px solid #fcd34d;border-radius:10px;padding:.85rem 1.1rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;flex-wrap:wrap">
    <div style="flex:1;min-width:200px">
      <div style="font-weight:700;color:#92400e;font-size:.88rem;margin-bottom:.15rem">⚠️ Existing account found for <strong>{{ mc.full_name }}</strong></div>
      <div style="font-size:.82rem;color:#78350f">{{ mc.email or '—' }} &nbsp;·&nbsp; {{ mc.phone or '—' }}</div>
    </div>
    <form method="POST" action="/admin/booking/{{ b.id }}/sync-customer-profile">
      <input type="hidden" name="action" value="link">
      <button type="submit"
        style="background:#d97706;color:white;border:none;border-radius:7px;padding:.45rem 1rem;font-size:.84rem;font-weight:700;cursor:pointer;white-space:nowrap"
        onclick="return confirm('Link this booking to {{ mc.full_name }}'s existing account?')">
        🔗 Link to Existing Account
      </button>
    </form>
    <a href="/admin/customers/{{ mc.id }}"
      style="font-size:.82rem;color:#92400e;text-decoration:underline;white-space:nowrap">
      View account →
    </a>
  </div>
  {% endif %}

  <div class="card" style="background:#EEEDFE;border:1.5px solid #AFA9EC">
    <h2 style="display:flex;align-items:center;justify-content:space-between;margin-bottom:.75rem;color:#26215C">
      👤 Customer
      <button id="cust-edit-btn" onclick="custEditToggle(true)"
              style="background:#CECBF6;color:#3C3489;border:1px solid #AFA9EC;border-radius:6px;padding:.3rem .8rem;font-size:.8rem;font-weight:600;cursor:pointer">
        ✏️ Edit
      </button>
    </h2>

    <!-- READ-ONLY VIEW -->
    <div id="cust-view" class="row">
      <span class="k">Name</span><span class="v">{{ b.full_name or '—' }}</span>
      {% if b.company_name and b.company_name != 'None' %}<span class="k">Company</span><span class="v">{{ b.company_name }}</span>{% endif %}
      <span class="k">Address</span>
      <span class="v">
        {% set _addr_parts = [] %}
        {% if b.renter_street and b.renter_street != 'None' %}{% set _ = _addr_parts.append(b.renter_street) %}{% endif %}
        {% if b.renter_city and b.renter_city != 'None' %}{% set _ = _addr_parts.append(b.renter_city) %}{% endif %}
        {% if b.renter_state and b.renter_state != 'None' %}{% set _ = _addr_parts.append(b.renter_state) %}{% endif %}
        {% if b.renter_zip and b.renter_zip != 'None' %}{% set _ = _addr_parts.append(b.renter_zip) %}{% endif %}
        {{ _addr_parts | join(', ') or '—' }}
      </span>
      {% set _phone = b.phone if (b.phone and b.phone != 'None') else '' %}
      <span class="k">Phone</span>
      <span class="v">{% if _phone %}<a href="tel:{{ _phone }}">{{ _phone }}</a>{% else %}—{% endif %}</span>
      <span class="k">Email</span>
      <span class="v">{% if b.email %}<a href="mailto:{{ b.email }}">{{ b.email }}</a>{% else %}—{% endif %}</span>
    </div>

    <!-- EDIT VIEW (hidden by default) -->
    <div id="cust-edit-view" style="display:none">
      <form method="POST" action="/admin/booking/{{ b.id }}/update-address" id="cust-addr-form" style="margin:0">
        <div class="row" style="margin-bottom:.5rem">
          <span class="k">Name</span><span class="v" style="font-weight:600">{{ b.full_name or '—' }}</span>
          <span class="k">Address</span>
          <span class="v" style="display:flex;flex-wrap:wrap;gap:.4rem;align-items:center">
            <input id="bk_renter_street" name="renter_street" value="{{ b.renter_street if (b.renter_street and b.renter_street != 'None') else '' }}" placeholder="Street"
                   style="border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.88rem;width:180px">
            <input id="bk_renter_city" name="renter_city" value="{{ b.renter_city if (b.renter_city and b.renter_city != 'None') else '' }}" placeholder="City"
                   style="border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.88rem;width:120px">
            <input id="bk_renter_state" name="renter_state" value="{{ b.renter_state if (b.renter_state and b.renter_state != 'None') else '' }}" placeholder="ST"
                   style="border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.88rem;width:50px">
            <input id="bk_renter_zip" name="renter_zip" value="{{ b.renter_zip if (b.renter_zip and b.renter_zip != 'None') else '' }}" placeholder="ZIP"
                   style="border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.88rem;width:75px">
          </span>
        </div>
      </form>
      <form method="POST" action="/admin/booking/{{ b.id }}/update-phone" id="cust-phone-form" style="margin:0">
        <div class="row" style="margin-bottom:.5rem">
          <span class="k">Phone</span>
          <span class="v">
            <input type="tel" name="phone" id="cust-phone-input" value="{{ b.phone if (b.phone and b.phone != 'None') else '' }}" placeholder="(555) 000-0000"
                   style="border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.88rem;width:175px">
          </span>
          <span class="k">Email</span>
          <span class="v" style="font-size:.88rem;color:#374151">{{ b.email or '—' }}</span>
        </div>
      </form>
      <div style="display:flex;gap:.5rem;margin-top:.75rem">
        <button onclick="custSaveAll()" style="background:#16a34a;color:white;border:none;border-radius:6px;padding:.4rem 1.1rem;font-size:.85rem;font-weight:700;cursor:pointer">
          💾 Save Changes
        </button>
        <button onclick="custEditToggle(false)" style="background:#f3f4f6;color:#374151;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer">
          Cancel
        </button>
      </div>
    </div>
  </div>
  <script>
  function custEditToggle(on) {
    document.getElementById('cust-view').style.display = on ? 'none' : '';
    document.getElementById('cust-edit-view').style.display = on ? '' : 'none';
    document.getElementById('cust-edit-btn').style.display = on ? 'none' : '';
  }
  function custSaveAll() {
    // Submit address form first, phone form will chain via hidden iframe trick
    // Simpler: submit address, then immediately submit phone in background
    var addrForm = document.getElementById('cust-addr-form');
    var phoneForm = document.getElementById('cust-phone-form');
    // Put phone value into address form as hidden field so we only need one POST
    var ph = document.getElementById('cust-phone-input').value;
    var hidden = document.createElement('input');
    hidden.type = 'hidden'; hidden.name = '_phone_val'; hidden.value = ph;
    addrForm.appendChild(hidden);
    // Submit phone form silently via fetch, then submit address form normally
    var fd = new FormData(phoneForm);
    fetch(phoneForm.action, {method:'POST', body:fd, credentials:'same-origin'})
      .then(function(){ addrForm.submit(); })
      .catch(function(){ addrForm.submit(); });
  }
  </script>

  <div class="card" style="background:#f8fbff;border:1.5px solid #bfdbfe">
    <h2 style="display:flex;align-items:center;justify-content:space-between;margin-bottom:.75rem;color:#1e3a5f">
      📅 Event
      <button id="evt-edit-btn" onclick="evtEditToggle(true)"
              style="background:#dbeafe;color:#1e40af;border:1px solid #93c5fd;border-radius:6px;padding:.3rem .8rem;font-size:.8rem;font-weight:600;cursor:pointer">
        ✏️ Edit
      </button>
    </h2>


    <!-- READ-ONLY VIEW -->
    <div id="evt-view">

      <!-- Customer's event info — never changed by weekend schedule -->
      <div style="background:#E6F1FB;border:1.5px solid #85B7EB;border-radius:8px;padding:.85rem 1rem;margin-bottom:.85rem">
        <div style="font-size:.7rem;font-weight:700;color:#0C447C;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">🗓 Customer Event Info</div>
        <div class="row" style="margin:0">
          <span class="k">Event Start</span>
          <span class="v">
            {% if b.event_start_date %}{{ b.event_start_date.strftime('%m/%d/%Y') }}{% if b.event_start_time %} &nbsp;{{ b.event_start_time }}{% endif %}{% else %}—{% endif %}
          </span>
          <span class="k">Event End</span>
          <span class="v">
            {% if b.event_end_date %}{{ b.event_end_date.strftime('%m/%d/%Y') }}{% if b.event_end_time %} &nbsp;{{ b.event_end_time }}{% endif %}{% else %}—{% endif %}
          </span>
          <span class="k">Setup</span><span class="v">{{ b.setup_date.strftime('%m/%d/%Y') if b.setup_date else '—' }} &nbsp;{{ b.setup_time or '' }}</span>
          <span class="k">Venue Type</span><span class="v" style="text-transform:capitalize">{{ b.venue_type or '—' }}</span>
          <span class="k">Event Address</span><span class="v">{{ b.event_street or '' }}{% if b.event_city %}, {{ b.event_city }}{% endif %}{% if b.event_state %}, {{ b.event_state }}{% endif %} {{ b.event_zip or '' }}</span>
          <span class="k">Deliver To</span><span class="v">{{ b.delivery_location or '—' }}</span>
          {% if b.venue_latest_pickup %}<span class="k">Latest Pickup</span><span class="v">{{ b.venue_latest_pickup }}</span>{% endif %}
        </div>
      </div>

      <!-- Delivery schedule — updated by weekend schedule -->
      <div id="delivery-sched-card" style="background:#E1F5EE;border:1.5px solid #5DCAA5;border-radius:8px;padding:.85rem 1rem;transition:all .4s ease">
        <div style="font-size:.7rem;font-weight:700;color:#085041;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">🚚 Delivery Schedule</div>

        {% if weekend_residential %}
        <div id="weekend-banner" style="background:#fef3c7;border:1px solid #f59e0b;border-radius:7px;padding:.7rem .9rem;margin-bottom:.75rem">
          <div style="display:flex;flex-wrap:wrap;align-items:center;gap:.6rem">
            <div style="flex:1;min-width:180px">
              <div style="font-weight:700;color:#92400e;font-size:.88rem;margin-bottom:.15rem">🏠 Weekend Residential Event</div>
              <div style="font-size:.82rem;color:#78350f">
                {{ weekend_residential.day_label }} detected &nbsp;·&nbsp;
                <strong>Delivery:</strong> {{ weekend_residential.delivery_label }} &nbsp;|&nbsp; <strong>Pickup:</strong> {{ weekend_residential.pickup_label }}
              </div>
            </div>
            <button type="button" id="ws-apply-btn-{{ b.id }}"
              onclick="showWeekendConfirm('{{ b.id }}')"
              style="background:#d97706;color:white;border:none;border-radius:6px;padding:.45rem 1rem;font-size:.84rem;font-weight:700;cursor:pointer;white-space:nowrap">
              📅 Apply Weekend Schedule
            </button>
          </div>
          <!-- inline confirm row, hidden until button clicked -->
          <div id="ws-confirm-{{ b.id }}" style="display:none;margin-top:.6rem;padding:.55rem .75rem;background:#fff8e1;border:1px solid #f59e0b;border-radius:6px;display:none;align-items:center;gap:.75rem;flex-wrap:wrap">
            <span style="font-size:.85rem;font-weight:600;color:#92400e;flex:1">
              ⚠️ This will set Delivery to <strong>{{ weekend_residential.delivery_label }}</strong> and Pickup to <strong>{{ weekend_residential.pickup_label }}</strong>. Confirm?
            </span>
            <div style="display:flex;gap:.5rem">
              <form method="POST" action="/admin/booking/{{ b.id }}/apply-weekend-schedule" style="margin:0">
                <button type="submit" style="background:#15803d;color:white;border:none;border-radius:6px;padding:.4rem .9rem;font-size:.83rem;font-weight:700;cursor:pointer">
                  ✓ Yes, Apply
                </button>
              </form>
              <button type="button" onclick="hideWeekendConfirm('{{ b.id }}')"
                style="background:#f3f4f6;color:#374151;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .9rem;font-size:.83rem;font-weight:600;cursor:pointer">
                Cancel
              </button>
            </div>
          </div>
        </div>
        {% endif %}

        <div class="row" style="margin:0">
          <span class="k">Est. Delivery</span>
          <span class="v">{{ b.setup_date.strftime('%m/%d/%Y') if b.setup_date else '—' }}{% if b.setup_time %} &nbsp;{{ b.setup_time }}{% endif %}</span>
          <span class="k">Est. Pickup</span>
          <span class="v">{{ b.event_end_date.strftime('%m/%d/%Y') if b.event_end_date else '—' }}{% if b.event_end_time %} &nbsp;{{ b.event_end_time }}{% endif %}</span>
        </div>
      </div>

    </div>

    <!-- EDIT VIEW (hidden by default) -->
    <div id="evt-edit-view" style="display:none">

      <!-- Customer Event Info — edit section -->
      <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:.85rem 1rem;margin-bottom:.85rem">
        <div style="font-size:.7rem;font-weight:700;color:#0369a1;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">🗓 Customer Event Info</div>
        <form method="POST" action="/admin/booking/{{ b.id }}/update-event-dates" style="margin:0">
          {%- set tsel -%}
          <option value="">-- Select --</option>
          <option value="06:00" {% if b.event_start_time=="06:00" %}selected{% endif %}>6:00 AM</option><option value="06:30" {% if b.event_start_time=="06:30" %}selected{% endif %}>6:30 AM</option>
          <option value="07:00" {% if b.event_start_time=="07:00" %}selected{% endif %}>7:00 AM</option><option value="07:30" {% if b.event_start_time=="07:30" %}selected{% endif %}>7:30 AM</option>
          <option value="08:00" {% if b.event_start_time=="08:00" %}selected{% endif %}>8:00 AM</option><option value="08:30" {% if b.event_start_time=="08:30" %}selected{% endif %}>8:30 AM</option>
          <option value="09:00" {% if b.event_start_time=="09:00" %}selected{% endif %}>9:00 AM</option><option value="09:30" {% if b.event_start_time=="09:30" %}selected{% endif %}>9:30 AM</option>
          <option value="10:00" {% if b.event_start_time=="10:00" %}selected{% endif %}>10:00 AM</option><option value="10:30" {% if b.event_start_time=="10:30" %}selected{% endif %}>10:30 AM</option>
          <option value="11:00" {% if b.event_start_time=="11:00" %}selected{% endif %}>11:00 AM</option><option value="11:30" {% if b.event_start_time=="11:30" %}selected{% endif %}>11:30 AM</option>
          <option value="12:00" {% if b.event_start_time=="12:00" %}selected{% endif %}>12:00 PM</option><option value="12:30" {% if b.event_start_time=="12:30" %}selected{% endif %}>12:30 PM</option>
          <option value="13:00" {% if b.event_start_time=="13:00" %}selected{% endif %}>1:00 PM</option><option value="13:30" {% if b.event_start_time=="13:30" %}selected{% endif %}>1:30 PM</option>
          <option value="14:00" {% if b.event_start_time=="14:00" %}selected{% endif %}>2:00 PM</option><option value="14:30" {% if b.event_start_time=="14:30" %}selected{% endif %}>2:30 PM</option>
          <option value="15:00" {% if b.event_start_time=="15:00" %}selected{% endif %}>3:00 PM</option><option value="15:30" {% if b.event_start_time=="15:30" %}selected{% endif %}>3:30 PM</option>
          <option value="16:00" {% if b.event_start_time=="16:00" %}selected{% endif %}>4:00 PM</option><option value="16:30" {% if b.event_start_time=="16:30" %}selected{% endif %}>4:30 PM</option>
          <option value="17:00" {% if b.event_start_time=="17:00" %}selected{% endif %}>5:00 PM</option><option value="17:30" {% if b.event_start_time=="17:30" %}selected{% endif %}>5:30 PM</option>
          <option value="18:00" {% if b.event_start_time=="18:00" %}selected{% endif %}>6:00 PM</option><option value="18:30" {% if b.event_start_time=="18:30" %}selected{% endif %}>6:30 PM</option>
          <option value="19:00" {% if b.event_start_time=="19:00" %}selected{% endif %}>7:00 PM</option><option value="19:30" {% if b.event_start_time=="19:30" %}selected{% endif %}>7:30 PM</option>
          <option value="20:00" {% if b.event_start_time=="20:00" %}selected{% endif %}>8:00 PM</option><option value="20:30" {% if b.event_start_time=="20:30" %}selected{% endif %}>8:30 PM</option>
          <option value="21:00" {% if b.event_start_time=="21:00" %}selected{% endif %}>9:00 PM</option><option value="21:30" {% if b.event_start_time=="21:30" %}selected{% endif %}>9:30 PM</option>
          <option value="22:00" {% if b.event_start_time=="22:00" %}selected{% endif %}>10:00 PM</option><option value="22:30" {% if b.event_start_time=="22:30" %}selected{% endif %}>10:30 PM</option>
          <option value="23:00" {% if b.event_start_time=="23:00" %}selected{% endif %}>11:00 PM</option><option value="23:30" {% if b.event_start_time=="23:30" %}selected{% endif %}>11:30 PM</option>
          {%- endset -%}
          {%- set tsel2 -%}
          <option value="">-- Select --</option>
          <option value="06:00" {% if b.event_end_time=="06:00" %}selected{% endif %}>6:00 AM</option><option value="06:30" {% if b.event_end_time=="06:30" %}selected{% endif %}>6:30 AM</option>
          <option value="07:00" {% if b.event_end_time=="07:00" %}selected{% endif %}>7:00 AM</option><option value="07:30" {% if b.event_end_time=="07:30" %}selected{% endif %}>7:30 AM</option>
          <option value="08:00" {% if b.event_end_time=="08:00" %}selected{% endif %}>8:00 AM</option><option value="08:30" {% if b.event_end_time=="08:30" %}selected{% endif %}>8:30 AM</option>
          <option value="09:00" {% if b.event_end_time=="09:00" %}selected{% endif %}>9:00 AM</option><option value="09:30" {% if b.event_end_time=="09:30" %}selected{% endif %}>9:30 AM</option>
          <option value="10:00" {% if b.event_end_time=="10:00" %}selected{% endif %}>10:00 AM</option><option value="10:30" {% if b.event_end_time=="10:30" %}selected{% endif %}>10:30 AM</option>
          <option value="11:00" {% if b.event_end_time=="11:00" %}selected{% endif %}>11:00 AM</option><option value="11:30" {% if b.event_end_time=="11:30" %}selected{% endif %}>11:30 AM</option>
          <option value="12:00" {% if b.event_end_time=="12:00" %}selected{% endif %}>12:00 PM</option><option value="12:30" {% if b.event_end_time=="12:30" %}selected{% endif %}>12:30 PM</option>
          <option value="13:00" {% if b.event_end_time=="13:00" %}selected{% endif %}>1:00 PM</option><option value="13:30" {% if b.event_end_time=="13:30" %}selected{% endif %}>1:30 PM</option>
          <option value="14:00" {% if b.event_end_time=="14:00" %}selected{% endif %}>2:00 PM</option><option value="14:30" {% if b.event_end_time=="14:30" %}selected{% endif %}>2:30 PM</option>
          <option value="15:00" {% if b.event_end_time=="15:00" %}selected{% endif %}>3:00 PM</option><option value="15:30" {% if b.event_end_time=="15:30" %}selected{% endif %}>3:30 PM</option>
          <option value="16:00" {% if b.event_end_time=="16:00" %}selected{% endif %}>4:00 PM</option><option value="16:30" {% if b.event_end_time=="16:30" %}selected{% endif %}>4:30 PM</option>
          <option value="17:00" {% if b.event_end_time=="17:00" %}selected{% endif %}>5:00 PM</option><option value="17:30" {% if b.event_end_time=="17:30" %}selected{% endif %}>5:30 PM</option>
          <option value="18:00" {% if b.event_end_time=="18:00" %}selected{% endif %}>6:00 PM</option><option value="18:30" {% if b.event_end_time=="18:30" %}selected{% endif %}>6:30 PM</option>
          <option value="19:00" {% if b.event_end_time=="19:00" %}selected{% endif %}>7:00 PM</option><option value="19:30" {% if b.event_end_time=="19:30" %}selected{% endif %}>7:30 PM</option>
          <option value="20:00" {% if b.event_end_time=="20:00" %}selected{% endif %}>8:00 PM</option><option value="20:30" {% if b.event_end_time=="20:30" %}selected{% endif %}>8:30 PM</option>
          <option value="21:00" {% if b.event_end_time=="21:00" %}selected{% endif %}>9:00 PM</option><option value="21:30" {% if b.event_end_time=="21:30" %}selected{% endif %}>9:30 PM</option>
          <option value="22:00" {% if b.event_end_time=="22:00" %}selected{% endif %}>10:00 PM</option><option value="22:30" {% if b.event_end_time=="22:30" %}selected{% endif %}>10:30 PM</option>
          <option value="23:00" {% if b.event_end_time=="23:00" %}selected{% endif %}>11:00 PM</option><option value="23:30" {% if b.event_end_time=="23:30" %}selected{% endif %}>11:30 PM</option>
          {%- endset -%}
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:.6rem .75rem;margin-bottom:.6rem">
            <div>
              <div style="font-size:.72rem;font-weight:600;color:#0369a1;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.25rem">Event Start</div>
              <div style="display:flex;gap:.35rem">
                <input type="date" name="event_start_date" value="{{ b.event_start_date.strftime('%Y-%m-%d') if b.event_start_date else '' }}"
                  style="flex:1;border:1px solid #bae6fd;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;min-width:0">
                <select name="event_start_time" style="border:1px solid #bae6fd;border-radius:6px;padding:.3rem .35rem;font-size:.82rem;color:#111827;width:105px;background:#fff">
                  {{ tsel }}
                </select>
              </div>
            </div>
            <div>
              <div style="font-size:.72rem;font-weight:600;color:#0369a1;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.25rem">Event End</div>
              <div style="display:flex;gap:.35rem">
                <input type="date" name="event_end_date" value="{{ b.event_end_date.strftime('%Y-%m-%d') if b.event_end_date else '' }}"
                  style="flex:1;border:1px solid #bae6fd;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;min-width:0">
                <select name="event_end_time" style="border:1px solid #bae6fd;border-radius:6px;padding:.3rem .35rem;font-size:.82rem;color:#111827;width:105px;background:#fff">
                  {{ tsel2 }}
                </select>
              </div>
            </div>
          </div>
          <button type="submit" style="background:#0369a1;color:white;border:none;border-radius:6px;padding:.35rem 1rem;font-size:.82rem;font-weight:600;cursor:pointer">Save Event Dates</button>
        </form>
        <div class="row" style="margin:.65rem 0 0">
          <span class="k">Venue Type</span><span class="v" style="text-transform:capitalize">{{ b.venue_type or '—' }}</span>
          <span class="k">Event Address</span><span class="v">{{ b.event_street or '' }}{% if b.event_city %}, {{ b.event_city }}{% endif %}{% if b.event_state %}, {{ b.event_state }}{% endif %} {{ b.event_zip or '' }}</span>
          <span class="k">Deliver To</span><span class="v">{{ b.delivery_location or '—' }}</span>
          {% if b.venue_latest_pickup %}<span class="k">Latest Pickup</span><span class="v">{{ b.venue_latest_pickup }}</span>{% endif %}
        </div>
      </div>

      <!-- Delivery Schedule — edit section -->
      <div style="background:#fafafa;border:1px solid #e5e7eb;border-radius:8px;padding:.85rem 1rem;margin-bottom:.75rem">
        <div style="font-size:.7rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">🚚 Delivery Schedule</div>
        <form method="POST" action="/admin/booking/{{ b.id }}/update-times" style="margin:0">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:.6rem .75rem;margin-bottom:.6rem">
            <div>
              <div style="font-size:.72rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.25rem">Est. Delivery</div>
              <div style="display:flex;gap:.35rem">
                <input type="date" name="setup_date" value="{{ b.setup_date.strftime('%Y-%m-%d') if b.setup_date else '' }}"
                  style="flex:1;border:1px solid #d1d5db;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;min-width:0">
                <input type="time" name="setup_time" value="{{ b.setup_time or '' }}"
                  style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;width:90px">
              </div>
            </div>
            <div>
              <div style="font-size:.72rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.25rem">Est. Pickup</div>
              <div style="display:flex;gap:.35rem">
                <input type="date" name="event_end_date" value="{{ b.event_end_date.strftime('%Y-%m-%d') if b.event_end_date else '' }}"
                  style="flex:1;border:1px solid #d1d5db;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;min-width:0">
                <input type="time" name="event_end_time" value="{{ b.event_end_time or '' }}"
                  style="border:1px solid #d1d5db;border-radius:6px;padding:.3rem .45rem;font-size:.85rem;color:#111827;width:90px">
              </div>
            </div>
          </div>
          <button type="submit" style="background:#374151;color:white;border:none;border-radius:6px;padding:.35rem 1rem;font-size:.82rem;font-weight:600;cursor:pointer">Save Schedule</button>
        </form>
      </div>

      <div>
        <button onclick="evtEditToggle(false)" style="background:#f3f4f6;color:#374151;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer">
          Cancel
        </button>
      </div>
    </div>
    <script>
    function evtEditToggle(on) {
      document.getElementById('evt-view').style.display = on ? 'none' : '';
      document.getElementById('evt-edit-view').style.display = on ? '' : 'none';
      document.getElementById('evt-edit-btn').style.display = on ? 'none' : '';
    }
    function showWeekendConfirm(bookingId) {
      document.getElementById('ws-apply-btn-' + bookingId).style.display = 'none';
      var c = document.getElementById('ws-confirm-' + bookingId);
      c.style.display = 'flex';
    }
    function hideWeekendConfirm(bookingId) {
      document.getElementById('ws-apply-btn-' + bookingId).style.display = '';
      var c = document.getElementById('ws-confirm-' + bookingId);
      c.style.display = 'none';
    }
    </script>
    <div style="background:#f0f9ff;border-left:3px solid #38bdf8;padding:.6rem .9rem;margin-top:.9rem;border-radius:0 6px 6px 0;font-size:.82rem;color:#0c4a6e">
      ℹ️ <strong>Delivery times are approximate</strong> unless the customer has opted for exact-time delivery/pickup.
    </div>
    <!-- Booking Calendar -->
    <div style="margin-top:1.25rem">
      <div style="font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#6b7280;margin-bottom:.75rem">📅 Confirmed Bookings Calendar</div>
      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:.5rem">
        <button onclick="calPrev()" style="background:none;border:1px solid #d1d5db;border-radius:6px;padding:.25rem .65rem;cursor:pointer;font-size:.9rem;color:#374151">‹</button>
        <span id="cal-title" style="font-weight:700;font-size:.9rem;color:#1a202c"></span>
        <button onclick="calNext()" style="background:none;border:1px solid #d1d5db;border-radius:6px;padding:.25rem .65rem;cursor:pointer;font-size:.9rem;color:#374151">›</button>
      </div>
      <div id="cal-grid" style="display:grid;grid-template-columns:repeat(7,1fr);gap:2px;font-size:.78rem"></div>
      <div style="display:flex;flex-wrap:wrap;gap:.6rem;margin-top:.75rem;font-size:.75rem">
        <span style="display:flex;align-items:center;gap:.3rem"><span style="width:12px;height:12px;border-radius:3px;background:#fef3c7;border:1px solid #f59e0b;display:inline-block"></span>This booking</span>
        <span style="display:flex;align-items:center;gap:.3rem"><span style="width:12px;height:12px;border-radius:3px;background:#dcfce7;border:1px solid #86efac;display:inline-block"></span>Confirmed</span>
        <span style="display:flex;align-items:center;gap:.3rem"><span style="width:12px;height:12px;border-radius:3px;background:#dbeafe;border:1px solid #93c5fd;display:inline-block"></span>Accepted</span>
      </div>
    </div>

    <script>
    (function(){
      const CAL_BOOKINGS = {{ cal_bookings | tojson }};
      const THIS_START   = "{{ b.event_start_date or '' }}".slice(0,10);
      const THIS_END     = "{{ b.event_end_date or '' }}".slice(0,10);
      const THIS_ID      = {{ b.id }};

      // Start calendar at this booking's month
      let calYear, calMonth;
      if(THIS_START){
        const d=new Date(THIS_START+'T00:00:00');
        calYear=d.getFullYear(); calMonth=d.getMonth();
      } else {
        const now=new Date();
        calYear=now.getFullYear(); calMonth=now.getMonth();
      }

      const MONTHS=['January','February','March','April','May','June',
                    'July','August','September','October','November','December'];
      const DAYS=['Su','Mo','Tu','We','Th','Fr','Sa'];

      function dateStr(y,m,d){ return y+'-'+(String(m+1).padStart(2,'0'))+'-'+(String(d).padStart(2,'0')); }
      function inRange(ds, start, end){ return ds>=start && ds<=end; }

      function render(){
        document.getElementById('cal-title').textContent=MONTHS[calMonth]+' '+calYear;
        const grid=document.getElementById('cal-grid');
        grid.innerHTML='';
        // Day headers
        DAYS.forEach(d=>{
          const h=document.createElement('div');
          h.textContent=d;
          h.style.cssText='text-align:center;font-weight:700;color:#9ca3af;padding:4px 2px;font-size:.7rem';
          grid.appendChild(h);
        });
        const firstDay=new Date(calYear,calMonth,1).getDay();
        const daysInMonth=new Date(calYear,calMonth+1,0).getDate();
        // Empty cells before first day
        for(let i=0;i<firstDay;i++){
          const e=document.createElement('div'); grid.appendChild(e);
        }
        for(let d=1;d<=daysInMonth;d++){
          const ds=dateStr(calYear,calMonth,d);
          const cell=document.createElement('div');
          cell.textContent=d;
          let bg='transparent', border='transparent', color='#374151', fw='400';

          // Check if this day falls in THIS booking's range
          const isThis=THIS_START&&THIS_END&&inRange(ds,THIS_START,THIS_END);
          if(isThis){ bg='#fef3c7'; border='#f59e0b'; fw='700'; }

          // Check confirmed/accepted bookings (others)
          let tip='';
          CAL_BOOKINGS.forEach(bk=>{
            if(bk.id===THIS_ID) return;
            if(bk.start&&bk.end&&inRange(ds,bk.start,bk.end)){
              if(!isThis){
                bg=bk.payment_status==='paid'?'#dcfce7':bk.payment_status==='partial'?'#ede9fe':'#dbeafe';
                border=bk.payment_status==='paid'?'#86efac':bk.payment_status==='partial'?'#c4b5fd':'#93c5fd';
              } else {
                // Overlap — show red ring
                border='#f87171';
              }
              fw='700';
              tip+=(tip?', ':'')+bk.name+' #'+bk.id;
            }
          });

          cell.style.cssText=`text-align:center;padding:5px 2px;border-radius:5px;cursor:default;font-weight:${fw};color:${color};background:${bg};border:1.5px solid ${border};font-size:.78rem;line-height:1.4;min-height:24px`;
          if(tip) cell.title=tip;
          grid.appendChild(cell);
        }
      }

      window.calPrev=function(){ calMonth--; if(calMonth<0){calMonth=11;calYear--;} render(); };
      window.calNext=function(){ calMonth++; if(calMonth>11){calMonth=0;calYear++;} render(); };
      render();
    })();
    </script>
  </div>

  <div class="card">
    <h2>Items & Totals</h2>
    <table>
      <thead><tr><th>Item</th><th style="text-align:center">Qty</th><th style="text-align:right">Unit</th><th style="text-align:right">Total</th></tr></thead>
      <tbody>
        {% for item in items %}
        <tr><td>{{ item.name }}</td><td style="text-align:center">{{ item.qty }}</td><td style="text-align:right">${{ "%.2f"|format((item.unit_price or 0)|float) }}</td><td style="text-align:right;font-weight:600">${{ "%.2f"|format((item.total or 0)|float) }}</td></tr>
        {% endfor %}
        {% if b.exact_time_delivery %}
        <tr><td colspan="3">Exact Time Delivery</td><td style="text-align:right;font-weight:600">$175.00</td></tr>
        {% endif %}
        <tr>
          <td colspan="3">
            Delivery Fee ({{ b.distance_miles or '?' }} mi)
            <form method="POST" action="/admin/booking/{{ b.id }}/recalc-delivery" style="display:inline;margin-left:.5rem">
              <button type="submit" style="font-size:.7rem;padding:.15rem .5rem;background:#e0f2fe;color:#0369a1;border:1px solid #7dd3fc;border-radius:4px;cursor:pointer">🔄 Recalc</button>
            </form>
          </td>
          <td style="text-align:right;font-weight:600">${{ "%.2f"|format(b.delivery_fee or 0) }}</td>
        </tr>
        {% if (b.discount_amount or 0)|float > 0 %}
        <tr style="background:#f0fdf4">
          <td colspan="3" style="color:#16a34a;font-weight:600">
            🏷️ Discount
            {% if b.discount_type == 'percent' %}({{ b.discount_value|float|round(1) }}% off)
            {% else %}(${{ "%.2f"|format(b.discount_value|float) }} off){% endif %}
          </td>
          <td style="text-align:right;font-weight:600;color:#16a34a">- ${{ "%.2f"|format(b.discount_amount|float) }}</td>
        </tr>
        {% endif %}
        {% if b.tax_exempt %}
        <tr style="background:#f0fff4"><td colspan="3" style="color:#276749">CT Sales Tax <span style="font-size:.78rem;background:#c6f6d5;color:#276749;border-radius:4px;padding:.1rem .4rem;margin-left:.4rem">TAX EXEMPT</span></td><td style="text-align:right;color:#276749">$0.00</td></tr>
        {% elif b.tax_amount %}
        <tr><td colspan="3" style="color:#718096">CT Sales Tax ({{ "%.2f"|format((b.tax_rate or 0)*100) }}%)</td><td style="text-align:right;font-weight:600">${{ "%.2f"|format(b.tax_amount or 0) }}</td></tr>
        {% endif %}
        <tr class="total-row"><td colspan="3">GRAND TOTAL</td><td style="text-align:right">${{ "%.2f"|format(b.grand_total or 0) }}</td></tr>
        {% set paid = (b.amount_paid or 0)|float %}
        {% set total = (b.grand_total or 0)|float %}
        {% set balance = [total - paid, 0]|max %}
        {% if paid > 0 %}
        <tr style="background:#f0fff4">
          <td colspan="3" style="color:#276749;font-weight:700">✅ Amount Paid</td>
          <td style="text-align:right;font-weight:700;color:#276749">- ${{ "%.2f"|format(paid) }}</td>
        </tr>
        {% if balance > 0.01 %}
        <tr style="background:#fff5f5">
          <td colspan="3" style="color:#c53030;font-weight:700">⚠️ Balance Due</td>
          <td style="text-align:right;font-weight:700;color:#c53030">${{ "%.2f"|format(balance) }}</td>
        </tr>
        {% else %}
        <tr style="background:#f0fff4">
          <td colspan="4" style="color:#276749;font-weight:700;text-align:center">✅ Paid In Full</td>
        </tr>
        {% endif %}
        {% elif b.status == 'accepted' and b.payment_status == 'paid' %}
        <tr style="background:#f0fff4">
          <td colspan="4" style="color:#276749;font-weight:700;text-align:center">✅ Paid In Full</td>
        </tr>
        {% elif b.status == 'accepted' %}
        <tr style="background:#fffbeb">
          <td colspan="3" style="color:#92400e;font-weight:700">⏳ Balance Due</td>
          <td style="text-align:right;font-weight:700;color:#92400e">${{ "%.2f"|format(total) }}</td>
        </tr>
        {% endif %}
      </tbody>
    </table>
  </div>

  <!-- ── Edit Items ── -->
  <datalist id="inv-list">
    {% for p in products %}<option value="{{ p.name }}">{% endfor %}
  </datalist>
  <div class="card">
    <h2>Edit Items</h2>
    <form method="POST" action="/admin/booking/{{ b.id }}/update-items">
      <!-- Header labels -->
      <div style="display:flex;gap:.5rem;margin-bottom:.25rem;font-size:.75rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.04em">
        <span style="flex:1;padding-left:.65rem">Item Name</span>
        <span style="width:62px;text-align:center">Qty</span>
        <span style="width:90px;text-align:center">Unit Price</span>
        <span style="width:30px"></span>
      </div>
      <div id="items-editor" style="display:flex;flex-direction:column;gap:.5rem;margin-bottom:.75rem">
        {% for item in items %}
        <div class="item-row" style="display:flex;gap:.5rem;align-items:center">
          <input type="text" name="item_name" value="{{ item.name }}" list="inv-list"
                 placeholder="Type to search inventory…"
                 style="flex:1;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .65rem;font-size:.9rem">
          <input type="number" name="item_qty" value="{{ item.qty }}" min="1"
                 style="width:62px;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .5rem;font-size:.9rem;text-align:center" title="Qty">
          <input type="number" name="item_price" value="{{ item.unit_price or 0 }}" min="0" step="0.01"
                 style="width:90px;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .5rem;font-size:.9rem;text-align:center" placeholder="Price" title="Unit Price">
          <button type="button" onclick="this.closest('.item-row').remove()"
                  style="background:#fee2e2;color:#dc2626;border:none;border-radius:6px;padding:.4rem .65rem;font-size:.85rem;cursor:pointer;font-weight:700">✕</button>
        </div>
        {% endfor %}
      </div>

      <!-- Exact Time Delivery & Delivery Fee row -->
      <div style="display:flex;gap:1.25rem;flex-wrap:wrap;align-items:center;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:.65rem 1rem;margin-bottom:.75rem">
        <label style="display:flex;align-items:center;gap:.45rem;font-size:.88rem;font-weight:600;color:#374151;cursor:pointer">
          <input type="checkbox" name="exact_time_delivery" value="1"
                 {% if b.exact_time_delivery %}checked{% endif %}
                 style="width:16px;height:16px;accent-color:#2563eb">
          Exact Time Delivery <span style="color:#6b7280;font-weight:400">(+$175.00)</span>
        </label>
        <label style="display:flex;align-items:center;gap:.45rem;font-size:.88rem;font-weight:600;color:#374151">
          Delivery Fee ($)
          <input type="number" name="delivery_fee" value="{{ b.delivery_fee or 0 }}" min="0" step="0.01"
                 style="width:90px;border:1px solid #d1d5db;border-radius:6px;padding:.35rem .5rem;font-size:.9rem;text-align:center">
        </label>
      </div>

      <div style="display:flex;gap:.5rem;flex-wrap:wrap">
        <button type="button" id="add-item-btn"
                style="background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe;border-radius:6px;padding:.45rem 1rem;font-size:.85rem;font-weight:600;cursor:pointer">+ Add Item</button>
        <button type="submit"
                style="background:#16a34a;color:white;border:none;border-radius:6px;padding:.45rem 1.1rem;font-size:.85rem;font-weight:600;cursor:pointer">💾 Save &amp; Recalculate</button>
      </div>
      <p style="font-size:.78rem;color:#6b7280;margin-top:.5rem;margin-bottom:0">Totals, tax, and grand total are automatically recalculated on save.</p>
    </form>
  </div>
  <script>
  // Price lookup map from inventory
  var _invPrices = {};
  {% for p in products %}_invPrices[{{ p.name|tojson }}] = {{ p.price|float }};{% endfor %}

  function makePriceInput(val) {
    var p = document.createElement('input');
    p.type = 'number'; p.name = 'item_price'; p.min = '0'; p.step = '0.01';
    p.value = val || '0'; p.placeholder = 'Price'; p.title = 'Unit Price';
    p.style.cssText = 'width:80px;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .5rem;font-size:.9rem;text-align:center';
    return p;
  }

  document.getElementById('items-editor').addEventListener('change', function(e) {
    if (e.target.name === 'item_name') {
      var row = e.target.closest('.item-row');
      if (!row) return;
      var priceInput = row.querySelector('input[name="item_price"]');
      var price = _invPrices[e.target.value];
      if (priceInput && price !== undefined && parseFloat(priceInput.value) === 0) {
        priceInput.value = price.toFixed(2);
      }
    }
  });

  document.getElementById('add-item-btn').addEventListener('click', function() {
    var editor = document.getElementById('items-editor');
    var row = document.createElement('div');
    row.className = 'item-row';
    row.style.cssText = 'display:flex;gap:.5rem;align-items:center';
    var nameInput = document.createElement('input');
    nameInput.type = 'text'; nameInput.name = 'item_name';
    nameInput.setAttribute('list', 'inv-list');
    nameInput.placeholder = 'Type to search inventory…';
    nameInput.style.cssText = 'flex:1;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .65rem;font-size:.9rem';
    var qtyInput = document.createElement('input');
    qtyInput.type = 'number'; qtyInput.name = 'item_qty'; qtyInput.value = '1'; qtyInput.min = '1';
    qtyInput.style.cssText = 'width:62px;border:1px solid #d1d5db;border-radius:6px;padding:.4rem .5rem;font-size:.9rem;text-align:center';
    qtyInput.title = 'Qty';
    var priceInput = makePriceInput(0);
    nameInput.addEventListener('change', function() {
      var price = _invPrices[nameInput.value];
      if (price !== undefined && parseFloat(priceInput.value) === 0) {
        priceInput.value = price.toFixed(2);
      }
    });
    var removeBtn = document.createElement('button');
    removeBtn.type = 'button'; removeBtn.textContent = '✕';
    removeBtn.style.cssText = 'background:#fee2e2;color:#dc2626;border:none;border-radius:6px;padding:.4rem .65rem;font-size:.85rem;cursor:pointer;font-weight:700';
    removeBtn.addEventListener('click', function() { row.remove(); });
    row.appendChild(nameInput); row.appendChild(qtyInput); row.appendChild(priceInput); row.appendChild(removeBtn);
    editor.appendChild(row);
    nameInput.focus();
  });
  </script>

  {% if b.notes %}
  <div class="card" style="background:#FAEEDA;border:1.5px solid #EF9F27"><h2 style="color:#633806">📝 Notes</h2><p style="color:#412402;line-height:1.6">{{ b.notes }}</p></div>
  {% endif %}

</div><!-- end left column -->

<!-- ══ RIGHT COLUMN ══ -->
<div style="display:flex;flex-direction:column;gap:1rem;position:sticky;top:1rem">



  <!-- ── agree_to_pay summary banner ── -->
  {% if b.status == 'agree_to_pay' %}
  <div style="background:#d1fae5;border:1px solid #6ee7b7;border-radius:7px;padding:.4rem .75rem;font-size:.8rem;color:#065f46;font-weight:600;display:flex;align-items:center;gap:.4rem">
    ✅ Agree to Pay &mdash; {{ (b.payment_method or 'cash')|title }} at delivery &nbsp;·&nbsp; Inventory reserved
  </div>
  {% endif %}

  <!-- ── Payment Summary ── -->
  <div class="card" style="border:none;overflow:hidden">
    {% set paid = (b.amount_paid or 0)|float %}
    {% set total = (b.grand_total or 0)|float %}
    {% set balance = [total - paid, 0]|max %}
    <!-- Hero -->
    <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);padding:1.1rem 1.25rem;margin:-1.5rem -1.5rem 1rem">
      <div style="font-size:.7rem;color:rgba(255,255,255,.65);text-transform:uppercase;letter-spacing:.07em;margin-bottom:.25rem">Grand Total</div>
      <div style="font-size:1.8rem;font-weight:700;color:#fff">${{ "%.2f"|format(total) }}</div>
      {% if b.status == 'accepted' %}
      {% set dep_pct = 0.25 %}
      {% set deposit = (total * dep_pct)|round(2) %}
      <div style="font-size:.8rem;color:rgba(255,255,255,.7);margin-top:.2rem">Deposit due: ${{ "%.2f"|format(deposit) }}</div>
      {% endif %}
    </div>

    <!-- Line items -->
    {% set subtotal  = (b.items_subtotal or 0)|float %}
    {% set del_fee   = (b.delivery_fee or 0)|float %}
    {% set exact_fee = 175.0 if b.exact_time_delivery else 0.0 %}
    {% set disc_amt  = (b.discount_amount or 0)|float %}
    {% set tax_amt   = (b.tax_amount or 0)|float %}
    <div style="font-size:.86rem">
      <div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#6b7280"><span>Items</span><span>${{ "%.2f"|format(subtotal) }}</span></div>
      {% if del_fee > 0 %}<div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#6b7280"><span>Delivery</span><span>${{ "%.2f"|format(del_fee) }}</span></div>{% endif %}
      {% if b.exact_time_delivery %}<div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#6b7280"><span>Exact Time Delivery</span><span>$175.00</span></div>{% endif %}
      {% if disc_amt > 0 %}<div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#16a34a;font-weight:600"><span>Discount</span><span>- ${{ "%.2f"|format(disc_amt) }}</span></div>{% endif %}
      <div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#6b7280"><span>Tax ({{ "%.2f"|format((b.tax_rate or 0)*100) }}%)</span><span>${{ "%.2f"|format(tax_amt) }}</span></div>
      {% if paid > 0 %}<div style="display:flex;justify-content:space-between;padding:.3rem 0;border-bottom:1px solid #f1f5f9;color:#16a34a;font-weight:600"><span>✅ Paid</span><span>- ${{ "%.2f"|format(paid) }}</span></div>{% endif %}
    </div>

    {% if paid > 0 and balance > 0.01 %}
    <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:.6rem .85rem;display:flex;justify-content:space-between;align-items:center;margin-top:.75rem">
      <span style="font-size:.88rem;font-weight:600;color:#c2410c">Balance Due</span>
      <span style="font-size:1.15rem;font-weight:700;color:#c2410c">${{ "%.2f"|format(balance) }}</span>
    </div>
    {% elif b.status == 'accepted' and (b.payment_status == 'paid' or balance <= 0.01) %}
    {% if (b.amount_paid or 0)|float < 0.50 %}
    <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:.55rem .85rem;margin-top:.75rem">
      <div style="font-size:.8rem;color:#c2410c;font-weight:600;margin-bottom:.4rem">⚠️ Marked paid but $0 collected</div>
      <form method="POST" action="/admin/booking/{{ b.id }}/reset-payment-status">
        <button type="submit" style="width:100%;background:#2563eb;color:#fff;border:none;border-radius:6px;padding:.4rem .8rem;font-size:.83rem;font-weight:700;cursor:pointer">
          🔄 Reset to Awaiting Payment
        </button>
      </form>
    </div>
    {% else %}
    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:.55rem .85rem;text-align:center;font-weight:700;color:#16a34a;margin-top:.75rem">✅ Paid in Full</div>
    {% endif %}
    {% elif b.status == 'accepted' %}
    <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:8px;padding:.6rem .85rem;display:flex;justify-content:space-between;align-items:center;margin-top:.75rem">
      <span style="font-size:.88rem;font-weight:600;color:#c2410c">Balance Due</span>
      <span style="font-size:1.15rem;font-weight:700;color:#c2410c">${{ "%.2f"|format(total) }}</span>
    </div>
    {% endif %}

    <!-- Discount form -->
    <div style="margin-top:1rem;padding-top:.85rem;border-top:1px solid #f1f5f9">
      <div style="font-size:.7rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">🏷️ Discount</div>
      {% if disc_amt > 0 %}
      <div style="font-size:.82rem;color:#16a34a;font-weight:600;margin-bottom:.5rem">
        Applied: {% if b.discount_type == 'percent' %}{{ b.discount_value|float|round(1) }}% off{% else %}${{ "%.2f"|format(b.discount_value|float) }} off{% endif %} → saves ${{ "%.2f"|format(disc_amt) }}
        <a href="/admin/booking/{{ b.id }}/remove-discount" style="margin-left:.5rem;font-size:.78rem;color:#dc2626;text-decoration:none;border:1px solid #fca5a5;border-radius:5px;padding:.1rem .45rem">Remove</a>
      </div>
      {% endif %}
      <form method="POST" action="/admin/booking/{{ b.id }}/apply-discount">
        <div style="display:flex;gap:.4rem;flex-wrap:wrap;align-items:flex-end">
          <select name="discount_type" style="border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.84rem;background:#fff;flex:1;min-width:100px">
            <option value="amount" {% if b.discount_type == 'amount' %}selected{% endif %}>$ Fixed</option>
            <option value="percent" {% if b.discount_type == 'percent' %}selected{% endif %}>% Percent</option>
          </select>
          <input type="number" name="discount_value" min="0" step="0.01" value="{{ b.discount_value|float if b.discount_value else '' }}" placeholder="Value"
            style="border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.84rem;width:75px">
          <button type="submit" style="background:#16a34a;color:#fff;border:none;border-radius:6px;padding:.38rem .85rem;font-size:.84rem;font-weight:600;cursor:pointer">Apply</button>
        </div>
      </form>
      <form method="POST" action="/admin/booking/{{ b.id }}/recalc-total" style="margin-top:.5rem">
        <button type="submit" style="font-size:.78rem;color:#1e40af;background:#eff6ff;border:1px solid #bfdbfe;border-radius:5px;padding:.25rem .65rem;cursor:pointer">🔄 Recalc Total</button>
      </form>
    </div>

    <!-- Payment link -->
    <div style="margin-top:.85rem;padding-top:.85rem;border-top:1px solid #f1f5f9">
      <div style="font-size:.7rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem">💳 Payment Links</div>
      {% set balance_due = ((b.grand_total or 0)|float - (b.amount_paid or 0)|float) %}
      {% if balance_due > 0.50 and b.status in ('accepted', 'pending', 'confirmed', 'agree_to_pay') %}
      <form method="POST" action="/admin/booking/{{ b.id }}/custom-stripe-link" style="margin-bottom:.6rem"
            onsubmit="return confirm('Send a ${{ '%.2f'|format(balance_due) }} payment link to {{ b.email }}?')">
        <input type="hidden" name="amount" value="{{ '%.2f'|format(balance_due) }}">
        <input type="hidden" name="label" value="Balance Due — Booking #{{ b.id }}">
        <button type="submit" style="width:100%;background:#15803d;color:#fff;border:none;border-radius:7px;padding:.55rem .9rem;font-size:.86rem;font-weight:700;cursor:pointer">
          💳 Send Balance Link — ${{ '%.2f'|format(balance_due) }}
        </button>
      </form>
      {% endif %}
      {% if b.status == 'accepted' and b.payment_status in ('partial',) %}
      {% set _bal = [((b.grand_total or 0)|float - (b.amount_paid or 0)|float), 0]|max %}
      <form method="POST" action="/admin/booking/{{ b.id }}/send-final-reminder"
            style="margin-bottom:.6rem"
            onsubmit="return confirm('Email final payment link for ${{ "%.2f"|format(_bal) }} to {{ b.email }}?')">
        <input type="hidden" name="custom_amount" value="{{ "%.2f"|format(_bal) }}">
        <button type="submit"
          style="width:100%;background:#dc2626;color:#fff;border:none;border-radius:7px;padding:.55rem .9rem;font-size:.86rem;font-weight:700;cursor:pointer">
          📧 Email Final Payment to Customer — ${{ "%.2f"|format(_bal) }}
        </button>
      </form>
      {% endif %}
      <form method="POST" action="/admin/booking/{{ b.id }}/custom-stripe-link" style="display:flex;gap:.4rem;align-items:center;flex-wrap:wrap">
        <input type="number" name="amount" min="0.50" step="0.01" placeholder="$ Amount" required
               style="flex:1;min-width:80px;border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.9rem">
        <input type="text" name="label" placeholder="Label (opt)"
               style="flex:2;min-width:100px;border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.84rem">
        <button type="submit" style="background:#2563eb;color:#fff;border:none;border-radius:6px;padding:.38rem .85rem;font-size:.84rem;font-weight:600;cursor:pointer;white-space:nowrap">Send</button>
      </form>
      {% if request.args.get('custom_link') %}
      <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:7px;padding:.55rem .75rem;margin-top:.5rem">
        <div style="font-size:.78rem;font-weight:600;color:#15803d;margin-bottom:.3rem">✅ Link sent to {{ b.email }}</div>
        <div style="display:flex;gap:.4rem">
          <input id="cl-val" type="text" value="{{ request.args.get('custom_link') }}" readonly
                 style="flex:1;border:1px solid #d1d5db;border-radius:5px;padding:.3rem .5rem;font-size:.78rem;background:#f9fafb">
          <button onclick="document.getElementById('cl-val').select();document.execCommand('copy');this.textContent='✓';setTimeout(()=>this.textContent='Copy',1500)"
                  style="background:#2563eb;color:#fff;border:none;border-radius:5px;padding:.3rem .65rem;font-size:.78rem;cursor:pointer">Copy</button>
        </div>
      </div>
      {% endif %}
      <!-- Stripe payment link if accepted -->
      {% if b.status == 'accepted' and b.stripe_payment_link %}
      <div style="margin-top:.5rem;font-size:.8rem;color:#6b7280">
        Invoice link:
        {% if b.payment_status in ('paid', 'partial') %}
        <span style="color:#059669;font-weight:600">✓ Used</span>
        <a href="{{ b.stripe_payment_link }}" target="_blank" style="color:#9ca3af;font-size:.75rem;margin-left:.35rem">Open ↗</a>
        {% else %}
        <a href="{{ b.stripe_payment_link }}" target="_blank" style="color:#2563eb">Open ↗</a>
        {% endif %}
      </div>
      {% endif %}
      {% if payment_links %}
      <div style="margin-top:.65rem">
        {% for pl in payment_links %}
        {% set _pl_paid = b.payment_status in ('paid', 'partial') %}
        <div style="display:flex;align-items:center;gap:.4rem;padding:.35rem 0;border-top:1px solid #f1f5f9;font-size:.8rem">
          <span style="flex:1;color:#374151;font-weight:500">{{ pl.label or 'Link' }} — ${{ "%.2f"|format(pl.amount or 0) }}</span>
          {% if _pl_paid %}
          <span style="color:#059669;font-weight:600;font-size:.78rem">✓ Paid</span>
          <a href="{{ pl.url }}" target="_blank" style="color:#9ca3af;font-size:.75rem">Open ↗</a>
          {% elif pl.status == 'active' %}
          <a href="{{ pl.url }}" target="_blank" style="color:#2563eb;font-size:.78rem">Open ↗</a>
          <form method="POST" action="/admin/payment-link/{{ pl.id }}/cancel" style="margin:0">
            <button type="submit" onclick="return confirm('Cancel this link?')"
                    style="background:#fee2e2;color:#dc2626;border:none;border-radius:4px;padding:.15rem .45rem;font-size:.75rem;cursor:pointer">✕</button>
          </form>
          {% else %}
          <span style="color:#9ca3af;font-size:.75rem">Cancelled</span>
          {% endif %}
        </div>
        {% endfor %}
      </div>
      {% endif %}
    </div>
  </div>

  <!-- ── Actions card ── -->
  <div class="card" style="border:none;padding:1.1rem 1.25rem">

    {% if b.status in ('accepted', 'pending', 'agree_to_pay', 'confirmed') %}
    <!-- Record Payment -->
    <div style="margin-bottom:1rem">
      <div style="font-size:.7rem;font-weight:700;color:#0C447C;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.55rem;padding:.35rem .6rem;background:#E6F1FB;border-radius:6px">💳 Record Payment</div>
      <form method="POST" action="/admin/booking/{{ b.id }}/record-payment"
            onsubmit="return confirm('Record this payment?')"
            style="display:flex;gap:.4rem;align-items:center;flex-wrap:wrap">
        <input type="number" name="amount" step="0.01" min="0.01" placeholder="Amount $" required
               style="width:105px;border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.88rem">
        <select name="method" style="flex:1;border:1px solid #d1d5db;border-radius:6px;padding:.38rem .5rem;font-size:.84rem;background:#fff">
          <option value="stripe">Stripe</option>
          <option value="cash">Cash</option>
          <option value="check">Check</option>
          <option value="other">Other</option>
        </select>
        <button type="submit" style="background:#2563eb;color:#fff;border:none;border-radius:6px;padding:.38rem .85rem;font-size:.84rem;font-weight:600;cursor:pointer">Record</button>
      </form>
      <div style="display:flex;gap:.4rem;flex-wrap:wrap;margin-top:.5rem">
        {% if b.status not in ('denied','cancelled','concluded') %}
        <form method="POST" action="/admin/booking/{{ b.id }}/cash-payment" onsubmit="return confirm('Mark as paid in full with cash?')">
          <button style="background:#E1F5EE;color:#085041;border:1px solid #5DCAA5;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:600;cursor:pointer">💵 Cash Full</button>
        </form>
        {% endif %}
        {% if b.status not in ('denied','cancelled','concluded','agree_to_pay') %}
        <form method="POST" action="/admin/booking/{{ b.id }}/agree-to-pay" onsubmit="return confirm('Mark as Agree to Pay at delivery?')" style="display:flex;gap:.3rem;align-items:center">
          <select name="pay_method" style="border:1px solid #6ee7b7;border-radius:6px;padding:.28rem .4rem;font-size:.78rem;background:#f0fdf4;color:#065f46;height:28px">
            <option value="cash">💵 Cash</option>
            <option value="check">📝 Check</option>
          </select>
          <button style="background:#d1fae5;color:#065f46;border:1.5px solid #6ee7b7;border-radius:6px;padding:.3rem .65rem;font-size:.8rem;font-weight:700;cursor:pointer;white-space:nowrap">🤝 Agree to Pay</button>
        </form>
        {% endif %}
        {% if b.status == 'agree_to_pay' %}
        <form method="POST" action="/admin/booking/{{ b.id }}/revert-agree-to-pay" onsubmit="return confirm('Revert Agree to Pay back to Accepted?')">
          <button style="background:#d1fae5;color:#065f46;border:1.5px solid #6ee7b7;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:700;cursor:pointer">✅ Agree to Pay ↩</button>
        </form>
        {% endif %}
        {% if b.status not in ('denied','cancelled') %}
        <form method="POST" action="/admin/booking/{{ b.id }}/no-charge" onsubmit="return confirm('Mark as No Charge?')">
          <button style="background:#f8fafc;color:#475569;border:1px solid #cbd5e1;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:600;cursor:pointer">🚫 No Charge</button>
        </form>
        {% endif %}
        {% if b.status == 'accepted' %}
        <form method="POST" action="/admin/booking/{{ b.id }}/send-receipt" onsubmit="return confirm('Send receipt to {{ b.email }}?')">
          <button style="background:#f8fafc;color:#374151;border:1px solid #d1d5db;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:600;cursor:pointer">📄 Receipt</button>
        </form>
        {% endif %}
        <!-- Send Agreement -->
        {% if b.status not in ('denied','cancelled') %}
        <form method="POST" action="/admin/booking/{{ b.id }}/send-agreement"
              onsubmit="return confirm('Send rental agreement link to {{ b.customer_name }}?')">
          <button style="background:{% if b.agreement_signed %}#dcfce7;color:#15803d;border:1px solid #86efac{% else %}#fef3c7;color:#92400e;border:1px solid #fcd34d{% endif %};border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:600;cursor:pointer;white-space:nowrap">
            {% if b.agreement_signed %}✅ Agreement Signed{% else %}✍️ Send Agreement{% endif %}
          </button>
        </form>
        {% endif %}
        <!-- Rebook -->
        <a href="/admin/booking/rebook/{{ b.id }}"
           style="background:#eff6ff;color:#1d4ed8;border:1px solid #93c5fd;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:600;cursor:pointer;text-decoration:none;white-space:nowrap;display:inline-flex;align-items:center">
          🔁 Rebook
        </a>
        {% if b.status == 'accepted' and b.payment_status in ('paid','partial') %}
        <form id="final-form" method="POST" action="/admin/booking/{{ b.id }}/send-final-reminder"
              onsubmit="return confirm('Send final payment link for ${{ "%.2f"|format([((b.grand_total or 0)|float - (b.amount_paid or 0)|float), 0]|max) }} to {{ b.email }}?')">
          <input type="hidden" name="custom_amount" id="final-amount-input">
          <button type="submit" id="final-btn" style="background:#fff7ed;color:#c2410c;border:1.5px solid #fdba74;border-radius:6px;padding:.3rem .75rem;font-size:.8rem;font-weight:700;cursor:pointer;white-space:nowrap">🔔 Remind</button>
        </form>
        {% endif %}
      </div>
    </div>
    {% endif %}

    <!-- Payment History Log -->
    {% set _paid_amt = (b.amount_paid or 0)|float %}
    {% if payment_history or _paid_amt > 0 %}
    <div style="padding-top:.85rem;border-top:1px solid #f1f5f9;margin-bottom:.85rem">
      <div style="font-size:.7rem;font-weight:700;color:#1e40af;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.55rem;padding:.35rem .6rem;background:#eff6ff;border-radius:6px">📋 Payment History</div>

      {% if payment_history %}
        {% for p in payment_history %}
        {% if p.amount|float > 0 %}
        <div style="display:flex;justify-content:space-between;align-items:flex-start;padding:.45rem .55rem;border-radius:7px;background:{{ '#f8faff' if loop.index is odd else '#fff' }};margin-bottom:.2rem;gap:.5rem;border:1px solid #e5e7eb">
          <div style="flex:1;min-width:0">
            <span style="font-weight:700;color:#15803d;font-size:.9rem">${{ "%.2f"|format(p.amount|float) }}</span>
            <span style="color:#6b7280;font-size:.75rem;margin-left:.35rem;text-transform:capitalize">via {{ p.method }}</span>
            {% if p.note %}<div style="color:#9ca3af;font-size:.7rem;margin-top:.1rem">{{ p.note }}</div>{% endif %}
          </div>
          <div style="text-align:right;white-space:nowrap;flex-shrink:0;font-size:.75rem">
            <div style="font-weight:600;color:#374151">{{ p.paid_at.strftime('%b %-d, %Y') if p.paid_at else '' }}</div>
            <div style="color:#9ca3af">{{ p.paid_at.strftime('%-I:%M %p') if p.paid_at else '' }}</div>
          </div>
        </div>
        {% endif %}
        {% endfor %}
        {% set ns = namespace(tot=0) %}{% for p in payment_history %}{% if p.amount|float > 0 %}{% set ns.tot = ns.tot + p.amount|float %}{% endif %}{% endfor %}
        {% if ns.tot > 0 %}
        <div style="display:flex;justify-content:space-between;padding:.4rem .55rem;font-size:.8rem;font-weight:700;color:#059669;border-top:1px solid #e5e7eb;margin-top:.3rem">
          <span>Total Logged</span><span>${{ "%.2f"|format(ns.tot) }}</span>
        </div>
        {% endif %}
      {% elif _paid_amt > 0 %}
        {# Legacy booking — amount_paid exists but no log entries yet #}
        <div style="display:flex;justify-content:space-between;align-items:center;padding:.45rem .55rem;border-radius:7px;background:#f8faff;border:1px solid #e5e7eb;margin-bottom:.2rem;gap:.5rem">
          <div>
            <span style="font-weight:700;color:#15803d;font-size:.9rem">${{ "%.2f"|format(_paid_amt) }}</span>
            <span style="color:#6b7280;font-size:.75rem;margin-left:.35rem">recorded</span>
            <div style="color:#9ca3af;font-size:.7rem;margin-top:.1rem">Payment received before detailed logging was enabled</div>
          </div>
          <div style="font-size:.72rem;color:#d1d5db;white-space:nowrap">date unknown</div>
        </div>
        <div style="display:flex;justify-content:space-between;padding:.4rem .55rem;font-size:.8rem;font-weight:700;color:#059669;border-top:1px solid #e5e7eb;margin-top:.3rem">
          <span>Total Logged</span><span>${{ "%.2f"|format(_paid_amt) }}</span>
        </div>
      {% else %}
        <div style="color:#9ca3af;font-size:.8rem;padding:.4rem .55rem">No payments recorded yet.</div>
      {% endif %}
    </div>
    {% endif %}

    <!-- Delivery -->
    {% if b.status not in ('denied','cancelled') %}
    <div style="padding-top:.85rem;border-top:1px solid #f1f5f9;margin-bottom:.85rem">
      <div style="font-size:.7rem;font-weight:700;color:#085041;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.55rem;padding:.35rem .6rem;background:#E1F5EE;border-radius:6px">🚚 Delivery</div>
      {% if b.delivery_status == 'picked_up' %}
        <span style="background:#f0fdf4;color:#16a34a;border:1px solid #86efac;border-radius:7px;padding:.35rem .85rem;font-size:.84rem;font-weight:600">✔ Picked Up</span>
      {% elif b.delivery_status == 'delivered' %}
        <form method="POST" action="/admin/booking/{{ b.id }}/delivery-status">
          <button onclick="return confirm('Mark as Picked Up?')"
            style="background:#eff6ff;color:#1e40af;border:1px solid #93c5fd;border-radius:7px;padding:.38rem .9rem;font-size:.84rem;font-weight:600;cursor:pointer">✅ Mark Picked Up</button>
        </form>
      {% else %}
        <form method="POST" action="/admin/booking/{{ b.id }}/delivery-status">
          <button onclick="return confirm('Mark as Delivered?')"
            style="background:#fffbeb;color:#92400e;border:1px solid #fcd34d;border-radius:7px;padding:.38rem .9rem;font-size:.84rem;font-weight:600;cursor:pointer">🚚 Mark Delivered</button>
        </form>
      {% endif %}
    </div>
    {% endif %}

    <!-- Admin Notes -->
    <div style="padding-top:.85rem;border-top:1px solid #f1f5f9;margin-bottom:.85rem">
      <div style="font-size:.7rem;font-weight:700;color:#633806;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.55rem;padding:.35rem .6rem;background:#FAEEDA;border-radius:6px">🔒 Private Notes</div>
      <form method="POST" action="/admin/booking/{{ b.id }}/admin-notes">
        <textarea name="admin_notes" rows="3" placeholder="Follow-up reminders, payment notes…"
          style="width:100%;border:1px solid #d1d5db;border-radius:7px;padding:.5rem .65rem;font-size:.84rem;color:#1a202c;background:#fff;resize:vertical;line-height:1.5">{{ b.admin_notes or '' }}</textarea>
        <button type="submit" style="margin-top:.35rem;background:#BA7517;color:#fff;border:none;border-radius:6px;padding:.35rem .9rem;font-size:.82rem;font-weight:600;cursor:pointer">Save Notes</button>
      </form>
    </div>

    <!-- Cancel / Delete / Back -->
    <div style="padding-top:.85rem;border-top:1px solid #f1f5f9;display:flex;flex-direction:column;gap:.4rem">
      <a href="/admin/dashboard" style="display:block;text-align:center;background:#f9fafb;color:#374151;border:1px solid #e5e7eb;border-radius:7px;padding:.42rem .9rem;font-size:.85rem;font-weight:600;text-decoration:none">← Dashboard</a>
      {% if b.status not in ('denied','cancelled') %}
      <form method="POST" action="/admin/booking/{{ b.id }}/cancel">
        <button onclick="return confirm('Cancel booking #{{ b.id }}?')" style="width:100%;background:#FAECE7;color:#993C1D;border:1px solid #F09575;border-radius:7px;padding:.42rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer">Cancel Booking</button>
      </form>
      {% endif %}
      <form method="POST" action="/admin/booking/{{ b.id }}/delete">
        <button onclick="return confirm('Permanently DELETE booking #{{ b.id }}?')" style="width:100%;background:#fff;color:#dc2626;border:1px solid #fca5a5;border-radius:7px;padding:.42rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer">Delete Booking</button>
      </form>
    </div>

  </div>

</div><!-- end right column -->
</div><!-- end two-column grid -->
</div>

<!-- ── Payment Amount Modal ── -->
<div id="payment-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);z-index:9999;align-items:center;justify-content:center">
  <div style="background:white;border-radius:14px;padding:2rem;width:min(420px,92vw);box-shadow:0 20px 60px rgba(0,0,0,.3)">
    <h2 id="modal-title" style="margin:0 0 .35rem;font-size:1.2rem;color:#1a202c">Send Stripe Payment Link</h2>
    <p id="modal-subtitle" style="margin:0 0 1.25rem;font-size:.9rem;color:#718096"></p>
    <label style="display:block;font-size:.85rem;font-weight:600;color:#374151;margin-bottom:.4rem">Amount to charge ($)</label>
    <input id="modal-amount" type="number" min="0.01" step="0.01"
           style="width:100%;box-sizing:border-box;border:2px solid #d1d5db;border-radius:8px;padding:.6rem .85rem;font-size:1.25rem;font-weight:700;color:#1a202c;margin-bottom:.3rem">
    <p style="font-size:.78rem;color:#9ca3af;margin:0 0 1.5rem">You can change this to any amount — the customer will be charged exactly what you enter.</p>
    <div style="display:flex;gap:.75rem">
      <button id="modal-confirm" style="flex:1;background:#16a34a;color:white;border:none;border-radius:8px;padding:.7rem;font-size:.95rem;font-weight:700;cursor:pointer">
        ✓ Send Payment Link
      </button>
      <button onclick="document.getElementById('payment-modal').style.display='none'"
              style="background:#f3f4f6;color:#374151;border:none;border-radius:8px;padding:.7rem 1.25rem;font-size:.95rem;cursor:pointer">
        Cancel
      </button>
    </div>
  </div>
</div>

<script>
(function() {
  var grandTotal = parseFloat('{{ b.grand_total or 0 }}') || 0;
  var eventDate  = '{{ b.event_start_date or "" }}';
  var modal      = document.getElementById('payment-modal');
  var amountIn   = document.getElementById('modal-amount');
  var confirmBtn = document.getElementById('modal-confirm');
  var activeForm = null;
  var activeHidden = null;

  function daysUntil(dateStr) {
    if (!dateStr) return 999;
    var d = new Date(dateStr + 'T00:00:00');
    var today = new Date(); today.setHours(0,0,0,0);
    return Math.round((d - today) / 86400000);
  }

  function openModal(title, subtitle, suggestedAmount, form, hiddenInput) {
    document.getElementById('modal-title').textContent = title;
    document.getElementById('modal-subtitle').textContent = subtitle;
    amountIn.value = suggestedAmount.toFixed(2);
    activeForm = form;
    activeHidden = hiddenInput;
    modal.style.display = 'flex';
    amountIn.focus();
    amountIn.select();
  }

  // Accept button
  var acceptBtn = document.getElementById('accept-btn');
  if (acceptBtn) {
    acceptBtn.addEventListener('click', function() {
      var days = daysUntil(eventDate);
      var suggested = days <= 7 ? grandTotal : Math.round(grandTotal * 0.25 * 100) / 100;
      var label = days <= 7 ? 'Full payment (event within 7 days)' : '25% deposit suggested';
      openModal(
        'Accept Booking — Set Payment Amount',
        'Grand total: $' + grandTotal.toFixed(2) + ' | ' + label,
        suggested,
        document.getElementById('accept-form'),
        document.getElementById('accept-amount-input')
      );
    });
  }

  // Final payment button
  var finalBtn = document.getElementById('final-btn');
  if (finalBtn) {
    finalBtn.addEventListener('click', function() {
      var suggested = Math.round(grandTotal * 0.75 * 100) / 100;
      openModal(
        'Send Final Payment Reminder',
        'Grand total: $' + grandTotal.toFixed(2) + ' | Remaining 75% suggested',
        suggested,
        document.getElementById('final-form'),
        document.getElementById('final-amount-input')
      );
    });
  }

  // Confirm button submits the active form
  confirmBtn.addEventListener('click', function() {
    var amt = parseFloat(amountIn.value);
    if (!amt || amt <= 0) { amountIn.style.borderColor='#ef4444'; return; }
    activeHidden.value = amt.toFixed(2);
    modal.style.display = 'none';
    activeForm.submit();
  });

  // Close on backdrop click
  modal.addEventListener('click', function(e) {
    if (e.target === modal) modal.style.display = 'none';
  });

  // Enter key confirms
  amountIn.addEventListener('keydown', function(e) {
    if (e.key === 'Enter') confirmBtn.click();
  });
})();
</script>

</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — BOOKING FORM
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/manifest.json")
def serve_manifest():
    """PWA manifest for the customer booking form."""
    import json as _json
    manifest = {
        "name": BUSINESS_NAME,
        "short_name": "Book Now",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#2563eb",
        "icons": [{"src": "/logo.png", "sizes": "192x192", "type": "image/png"},
                  {"src": "/logo.png", "sizes": "512x512", "type": "image/png"}]
    }
    return Response(_json.dumps(manifest), mimetype="application/manifest+json")

@app.route("/admin-manifest.json")
def serve_admin_manifest():
    """PWA manifest for the admin panel — starts at /admin/dashboard."""
    import json as _json
    manifest = {
        "name": f"{BUSINESS_NAME} Admin",
        "short_name": "Admin",
        "start_url": "/admin/dashboard",
        "display": "standalone",
        "background_color": "#1a365d",
        "theme_color": "#1a365d",
        "icons": [{"src": "/logo.png", "sizes": "192x192", "type": "image/png"},
                  {"src": "/logo.png", "sizes": "512x512", "type": "image/png"}]
    }
    return Response(_json.dumps(manifest), mimetype="application/manifest+json")

@app.route("/logo.png")
def serve_logo():
    """Serve logo from the repo root — tries several common filenames."""
    base = os.path.dirname(__file__)
    for name in ("logo.png", "Rent a Party Logo.png", "rent a party logo.png",
                 "Logo.png", "RentAPartyLogo.png"):
        p = os.path.join(base, name)
        if os.path.exists(p):
            return send_file(p, mimetype="image/png")
    # last resort: any .png in the repo root
    try:
        for f in os.listdir(base):
            if f.lower().endswith(".png"):
                return send_file(os.path.join(base, f), mimetype="image/png")
    except Exception:
        pass
    return "", 404

@app.route("/", methods=["GET"])
def index():
    return render_template_string(FORM_HTML,
        business_name=BUSINESS_NAME,
        products=get_products(),
        exact_time_fee=EXACT_TIME_FEE,
        google_maps_key=GOOGLE_MAPS_KEY,
        error=None,
        form={},
    )


@app.route("/availability")
def availability():
    start = request.args.get("start", "")
    end   = request.args.get("end",   "")
    if not start or not end:
        return jsonify({p["id"]: p["total"] for p in get_products()})
    avail = get_available(start, end)
    return jsonify(avail)


@app.route("/delivery_fee")
def delivery_fee_check():
    address = request.args.get("address", "")
    miles   = get_distance_miles(address) if address else None
    fee, note = calc_delivery_fee(miles)
    return jsonify({"fee": fee, "note": note, "miles": miles})


@app.route("/submit", methods=["POST"])
def submit():
    import traceback as _tb_sub
    try:
        return _submit_inner()
    except Exception as _sub_err:
        _trace = _tb_sub.format_exc()
        log.error(f"SUBMIT UNHANDLED ERROR: {_sub_err}\n{_trace}")
        return f"<pre style='color:red;padding:2rem'>Booking submission error (please screenshot and report):\n\n{_trace}</pre>", 500

def _submit_inner():
    f = request.form

    full_name        = f.get("full_name",        "").strip()
    company_name     = f.get("company_name",     "").strip()
    renter_street    = f.get("renter_street",    "").strip()
    renter_city      = f.get("renter_city",      "").strip()
    renter_state     = f.get("renter_state",     "").strip()
    renter_zip       = f.get("renter_zip",       "").strip()
    phone            = f.get("phone",            "").strip()
    email            = f.get("email",            "").strip()
    event_start_date = f.get("event_start_date", "").strip()
    event_end_date   = f.get("event_end_date",   "").strip()
    event_start_time = f.get("event_start_time", "").strip()
    event_end_time   = f.get("event_end_time",   "").strip()
    setup_time       = f.get("setup_time",       "").strip()
    setup_date       = f.get("setup_date",       "").strip()
    venue_type       = f.get("venue_type",       "venue").strip()
    venue_latest     = f.get("venue_latest_pickup","").strip()
    delivery_date    = f.get("delivery_date",    "").strip()
    delivery_time    = f.get("delivery_time",    "").strip()
    pickup_date      = f.get("pickup_date",      "").strip()
    pickup_time      = f.get("pickup_time",      "").strip()

    # Auto-apply weekend residential schedule if delivery_date not provided
    # Saturday event → deliver Friday, pickup Sunday
    # Sunday event   → deliver Friday, pickup Monday
    if not delivery_date and event_start_date and venue_type.lower() == "residential":
        try:
            _esd = datetime.strptime(event_start_date[:10], "%Y-%m-%d").date()
            _wd  = _esd.weekday()  # 5=Saturday, 6=Sunday
            if _wd == 5:  # Saturday
                delivery_date = (_esd - timedelta(days=1)).strftime("%Y-%m-%d")
                if not delivery_time:
                    delivery_time = "16:00"
                if not event_end_date:
                    event_end_date = (_esd + timedelta(days=1)).strftime("%Y-%m-%d")
            elif _wd == 6:  # Sunday
                delivery_date = (_esd - timedelta(days=2)).strftime("%Y-%m-%d")
                if not delivery_time:
                    delivery_time = "16:00"
                if not event_end_date:
                    event_end_date = (_esd + timedelta(days=1)).strftime("%Y-%m-%d")
        except Exception:
            pass
    event_street     = f.get("event_street",     "").strip()
    event_city       = f.get("event_city",       "").strip()
    event_state      = f.get("event_state",      "").strip()
    event_zip        = f.get("event_zip",        "").strip()
    exact_delivery   = f.get("exact_time_delivery","") == "yes"
    delivery_location= f.get("delivery_location","").strip()
    notes            = f.get("notes",            "").strip()

    _products = get_products()
    if not email or not full_name:
        return render_template_string(FORM_HTML, business_name=BUSINESS_NAME,
            google_maps_key=GOOGLE_MAPS_KEY,
            products=_products, exact_time_fee=EXACT_TIME_FEE,
            error="Name and email are required.", form=f), 400

    # Accept any quantity — admin is alerted of conflicts on the dashboard
    # Collect items first so we can apply marquee tier pricing
    _raw_items = []
    for p in _products:
        qty = int(f.get(f"qty_{p['id']}", 0) or 0)
        qty = max(0, qty)
        if qty > 0:
            _raw_items.append({"id": p["id"], "name": p["name"], "qty": qty, "base_price": float(p["price"])})

    # Marquee tier pricing (must match JS logic exactly)
    def _is_ml(name): return bool(re.match(r'^marquee\s+[a-z]$', name.strip(), re.IGNORECASE))
    def _is_mn(name): return bool(re.match(r'^marquee\s+#?\d', name.strip(), re.IGNORECASE))
    _ML_TIERS = [(1,85),(2,160),(3,225),(4,285)]
    _MN_TIERS = [(1,80),(2,150),(3,215),(4,275)]
    def _ml_total(n):
        for c,t in _ML_TIERS:
            if c==n: return float(t)
        return 285.0+(n-4)*55.0
    def _mn_total(n):
        for c,t in _MN_TIERS:
            if c==n: return float(t)
        return 275.0+(n-4)*55.0
    ml_count = sum(i["qty"] for i in _raw_items if _is_ml(i["name"]))
    mn_count = sum(i["qty"] for i in _raw_items if _is_mn(i["name"]))
    ml_unit = (_ml_total(ml_count)/ml_count) if ml_count>0 else 0.0
    mn_unit = (_mn_total(mn_count)/mn_count) if mn_count>0 else 0.0

    order_items, subtotal = [], 0.0
    for item in _raw_items:
        name = item["name"]
        qty  = item["qty"]
        if _is_ml(name):
            unit_price = ml_unit
        elif _is_mn(name):
            unit_price = mn_unit
        else:
            unit_price = item["base_price"]
        line = round(qty * unit_price, 2)
        subtotal += line
        order_items.append({"id": item["id"], "name": name, "qty": qty,
                             "unit_price": round(unit_price, 2), "total": line})

    # Stackable marquee fee
    stackable     = f.get("stackable", "no").strip().lower() == "yes"
    stackable_top = f.get("stackable_top", "").strip()
    has_marquee   = any(_is_ml(i["name"]) or _is_mn(i["name"]) for i in order_items)
    if stackable and has_marquee:
        subtotal += 75.0
        stack_label = f"Stackable Marquee (top: {stackable_top})" if stackable_top else "Stackable Marquee"
        order_items.append({"name": stack_label, "qty": 1, "unit_price": 75.0, "total": 75.0})

    # Delivery
    event_address = f"{event_street}, {event_city}, {event_state} {event_zip}"
    miles = get_distance_miles(event_address)
    delivery_fee, delivery_note = calc_delivery_fee(miles)

    exact_fee   = EXACT_TIME_FEE if exact_delivery else 0.0
    late_night_fee = float(f.get("late_night_fee") or 0)
    # Validate server-side: check end time actually falls in late night window
    def _is_late_night(t):
        if not t: return False
        try:
            h, m = map(int, t.strip().split(':')[:2])
            mins = h * 60 + m
            return mins >= 23*60+30 or mins < 7*60
        except: return False
    if late_night_fee and not _is_late_night(event_end_time):
        late_night_fee = 0.0
    if _is_late_night(event_end_time):
        late_night_fee = 125.0
    pre_tax_total = round(subtotal + exact_fee + late_night_fee + delivery_fee, 2)

    # Check if customer is tax exempt (either from DB record or self-reported on form)
    is_tax_exempt = f.get("tax_exempt_request") == "1"
    if not is_tax_exempt:
        tax_check_conn = get_db()
        if tax_check_conn and email:
            try:
                tc = tax_check_conn.cursor()
                tc.execute("SELECT tax_exempt FROM customers WHERE email=%s", (email,))
                row = tc.fetchone()
                if row and row[0]:
                    is_tax_exempt = True
                tc.close()
                tax_check_conn.close()
            except Exception:
                pass

    applied_tax_rate = 0.0 if is_tax_exempt else CT_TAX_RATE
    tax_amount  = round(pre_tax_total * applied_tax_rate, 2)
    grand_total = round(pre_tax_total + tax_amount, 2)

    # Save to DB — guard against duplicate submissions (same email+date within 3 min)
    booking_id = None
    conn = get_db()
    if conn:
        try:
            dup_cur = conn.cursor()
            dup_cur.execute("""
                SELECT id FROM bookings
                WHERE email=%s
                  AND event_start_date=%s
                  AND created_at >= NOW() - INTERVAL '3 minutes'
                ORDER BY id DESC LIMIT 1
            """, (email, event_start_date or None))
            dup_row = dup_cur.fetchone()
            dup_cur.close()
            if dup_row:
                log.warning(f"Duplicate submission blocked for {email} on {event_start_date} — returning existing #{dup_row[0]}")
                conn.close()
                return redirect(url_for("booking_success", booking_id=dup_row[0]))
        except Exception as _dup_err:
            log.error(f"Duplicate check error: {_dup_err}")

        try:
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO bookings (
                    full_name, company_name,
                    renter_street, renter_city, renter_state, renter_zip,
                    phone, email,
                    event_start_date, event_end_date,
                    event_start_time, event_end_time, setup_time, setup_date,
                    delivery_date, delivery_time,
                    pickup_date, pickup_time,
                    venue_type, venue_latest_pickup,
                    event_street, event_city, event_state, event_zip,
                    exact_time_delivery, delivery_location,
                    delivery_fee, distance_miles,
                    items_json, items_subtotal, exact_time_fee, late_night_fee, tax_rate, tax_amount, tax_exempt, grand_total,
                    notes
                ) VALUES (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
                ) RETURNING id
            """, (
                full_name, company_name,
                renter_street, renter_city, renter_state, renter_zip,
                phone, email,
                event_start_date or None, event_end_date or None,
                event_start_time, event_end_time, setup_time,
                setup_date or None,
                delivery_date or None, delivery_time or None,
                pickup_date or None, pickup_time or None,
                venue_type, venue_latest or None,
                event_street, event_city, event_state, event_zip,
                exact_delivery, delivery_location,
                delivery_fee, miles,
                json.dumps(order_items), subtotal, exact_fee, late_night_fee, applied_tax_rate, tax_amount, is_tax_exempt, grand_total,
                notes,
            ))
            booking_id = cur.fetchone()[0]
            view_token = secrets.token_urlsafe(24)
            cur.execute("UPDATE bookings SET view_token=%s WHERE id=%s", (view_token, booking_id))
            conn.commit()
            cur.close()
            conn.close()
            log.info(f"Booking #{booking_id} saved for {full_name}")
        except Exception as e:
            log.error(f"DB insert error: {e}")

    booking_data = {
        "id": booking_id, "view_token": view_token if booking_id else None,
        "full_name": full_name, "company_name": company_name,
        "renter_street": renter_street, "renter_city": renter_city,
        "renter_state": renter_state, "renter_zip": renter_zip,
        "phone": phone, "email": email,
        "event_start_date": event_start_date, "event_end_date": event_end_date,
        "event_start_time": event_start_time, "event_end_time": event_end_time,
        "setup_time": setup_time, "venue_type": venue_type,
        "venue_latest_pickup": venue_latest,
        "event_street": event_street, "event_city": event_city,
        "event_state": event_state, "event_zip": event_zip,
        "exact_time_delivery": exact_delivery,
        "delivery_location": delivery_location,
        "delivery_fee": delivery_fee, "delivery_fee_note": delivery_note,
        "distance_miles": miles,
        "items_json": json.dumps(order_items),
        "items_subtotal": subtotal, "exact_time_fee": exact_fee,
        "tax_rate": applied_tax_rate, "tax_amount": tax_amount, "tax_exempt": is_tax_exempt,
        "grand_total": grand_total, "notes": notes,
    }
    try:
        send_owner_email(booking_data)
    except Exception as _oe:
        import traceback as _tb2
        log.error(f"send_owner_email error: {_oe}\n{_tb2.format_exc()}")
    try:
        send_customer_email(booking_data)
    except Exception as _ce:
        import traceback as _tb3
        log.error(f"send_customer_email error: {_ce}\n{_tb3.format_exc()}")

    # ── Push notification to owner ───────────────────────────────────────────
    try:
        _base = os.environ.get("APP_BASE_URL", "").rstrip("/")
        _ntfy = os.environ.get("NTFY_TOPIC", "")
        log.info(f"Push attempt — booking_id={booking_id}, NTFY_TOPIC={_ntfy!r}, APP_BASE_URL={_base!r}")
        _admin_link = f"{_base}/admin/booking/{booking_id}" if _base else f"https://rental-booking-biv0.onrender.com/admin/booking/{booking_id}"
        send_push(
            title=f"New Booking #{booking_id} - {full_name}",
            body=f"{_fmt_date(event_start_date)} | Total: ${grand_total:.2f}\n{_admin_link}",
            click_url=_admin_link
        )
        log.info("Push notification dispatched successfully")
    except Exception as _pe:
        log.error(f"Push notification failed in submit: {_pe}")

    # Upsert customer record — if email already exists update their info, else insert
    cust_conn = get_db()
    if cust_conn and email:
        try:
            cust_cur = cust_conn.cursor()
            cust_cur.execute("""
                INSERT INTO customers (full_name, company_name, email, phone, street, city, state, zip)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (email) WHERE email IS NOT NULL DO UPDATE SET
                    full_name    = EXCLUDED.full_name,
                    company_name = EXCLUDED.company_name,
                    phone        = EXCLUDED.phone,
                    street       = EXCLUDED.street,
                    city         = EXCLUDED.city,
                    state        = EXCLUDED.state,
                    zip          = EXCLUDED.zip
            """, (
                full_name,
                company_name or None,
                email,
                phone or None,
                renter_street or None,
                renter_city or None,
                renter_state or None,
                renter_zip or None,
            ))
            cust_conn.commit()
            cust_cur.close()
            cust_conn.close()
            log.info(f"Customer record upserted for {email}")
        except Exception as e:
            log.error(f"Customer upsert error: {e}")

    # PRG: redirect so browser refresh doesn't re-POST
    if booking_id:
        if request.form.get("admin_create") == "1":
            # Admin created — set status/payment_status from form
            _st = request.form.get("status", "accepted")
            _ap = round(float(request.form.get("amount_paid") or 0), 2)
            _notes = request.form.get("notes", "").strip()
            _pst = "paid" if _st == "accepted_paid" else ("partial" if _ap > 0 else "waiting")
            _db_st = "accepted" if _st in ("accepted_paid", "partial") else _st
            conn2 = get_db()
            if conn2:
                try:
                    cur2 = conn2.cursor()
                    cur2.execute(
                        "UPDATE bookings SET status=%s, payment_status=%s, amount_paid=%s, admin_notes=%s WHERE id=%s",
                        (_db_st, _pst, _ap, _notes or None, booking_id)
                    )
                    conn2.commit(); cur2.close(); conn2.close()
                except Exception as _e:
                    log.error(f"Admin new booking status update error: {_e}")
            return redirect(url_for("admin_booking", booking_id=booking_id))
        return redirect(url_for("booking_success", booking_id=booking_id))
    return render_template_string(SUCCESS_HTML,
        business_name=BUSINESS_NAME,
        business_phone=BUSINESS_PHONE,
        name=full_name.split()[0],
        email=email,
        booking_id=booking_id,
    )


@app.route("/booking/success/<int:booking_id>")
def booking_success(booking_id):
    """Success page after booking — GET-only so refresh is safe."""
    b = None
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT full_name, email FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                b = dict(row)
        except Exception as e:
            log.error(f"booking_success fetch error: {e}")
    if not b:
        return redirect(url_for("index"))
    return render_template_string(SUCCESS_HTML,
        business_name=BUSINESS_NAME,
        business_phone=BUSINESS_PHONE,
        name=(b["full_name"] or "").split()[0],
        email=b["email"] or "",
        booking_id=booking_id,
    )


ADMIN_INVENTORY_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Inventory — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Helvetica,sans-serif;background:#f5f6fa;color:#111827;min-height:100vh}
    .topbar{background:white;border-bottom:1px solid #e5e7eb;padding:.9rem 1.25rem;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:50;flex-wrap:wrap;gap:.4rem}
    .topbar-brand{font-size:1rem;font-weight:700;color:#111827}
    .topbar-nav{display:flex;gap:.25rem;align-items:center;overflow-x:auto;-webkit-overflow-scrolling:touch;flex-wrap:nowrap;max-width:100%}
    .nav-link{color:#6b7280;text-decoration:none;font-size:.85rem;font-weight:500;padding:.38rem .65rem;border-radius:6px;transition:all .12s;white-space:nowrap}
    .nav-link:hover{background:#f3f4f6;color:#111827}
    .nav-link.active{background:#eff6ff;color:#2563eb;font-weight:600}
    .logout-btn{background:white;border:1px solid #d1d5db;color:#6b7280;padding:.38rem .65rem;border-radius:6px;cursor:pointer;font-size:.82rem;font-weight:500;text-decoration:none;white-space:nowrap}
    .main{max-width:900px;margin:0 auto;padding:1.75rem}
    .page-title{font-size:1.4rem;font-weight:700;color:#111827;margin-bottom:.35rem}
    .page-sub{font-size:.88rem;color:#6b7280;margin-bottom:1.5rem}
    .flash{padding:.75rem 1rem;border-radius:8px;margin-bottom:1.25rem;font-size:.9rem;font-weight:500}
    .flash-ok{background:#dcfce7;color:#166534;border:1px solid #bbf7d0}
    .flash-err{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}
    .card{background:white;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:1.5rem}
    .card-header{padding:.85rem 1.25rem;background:#f9fafb;border-bottom:1px solid #e5e7eb;font-weight:700;font-size:.88rem;color:#374151;display:flex;justify-content:space-between;align-items:center}
    table{width:100%;border-collapse:collapse}
    th{padding:.65rem 1rem;text-align:left;font-size:.72rem;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid #e5e7eb;background:#f9fafb}
    td{padding:.6rem 1rem;border-bottom:1px solid #f3f4f6;vertical-align:middle}
    tr:last-child td{border-bottom:none}
    tbody tr:hover td{background:#fafafa}
    input[type=text],input[type=number]{width:100%;padding:.4rem .6rem;border:1px solid #d1d5db;border-radius:6px;font-size:.86rem;color:#111827;background:white;transition:border .12s}
    input[type=text]:focus,input[type=number]:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,.1)}
    input[type=number]{max-width:90px}
    .btn{display:inline-block;padding:.4rem .85rem;border-radius:6px;font-size:.82rem;font-weight:600;cursor:pointer;border:none;text-decoration:none;transition:all .12s;line-height:1.5}
    .btn-primary{background:#2563eb;color:white}
    .btn-primary:hover{background:#1d4ed8}
    .btn-danger{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}
    .btn-danger:hover{background:#fecaca}
    .btn-outline{background:white;color:#374151;border:1px solid #d1d5db}
    .btn-outline:hover{background:#f3f4f6}
    .add-form{display:grid;grid-template-columns:1fr auto auto auto;gap:.6rem;align-items:center;padding:1rem 1.25rem;background:#f9fafb;border-top:1px solid #e5e7eb}
    .add-form input{width:100%}
    .save-bar{display:flex;justify-content:flex-end;gap:.75rem;padding:1rem 1.25rem;border-top:1px solid #e5e7eb;background:#f9fafb}
    .item-id{font-size:.75rem;color:#9ca3af;font-family:monospace}
    @media(max-width:600px){
      .add-form{grid-template-columns:1fr;gap:.5rem}
      .main{padding:1rem}
    }
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link active">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>Inventory</h1>
</div>
<div class="main">
  <div class="page-title">Inventory</div>
  <div class="page-sub">Edit item names, rental prices, and total quantities. Changes take effect immediately on the booking form.</div>

  {% if flash_ok %}<div class="flash flash-ok">✓ {{ flash_ok }}</div>{% endif %}
  {% if flash_err %}<div class="flash flash-err">⚠ {{ flash_err }}</div>{% endif %}

  <!-- ── Availability Checker ── -->
  <div class="card" style="margin-bottom:1.5rem">
    <div class="card-header">📅 Check Availability</div>
    <div style="padding:1.1rem 1.25rem">
      <div style="display:flex;gap:.75rem;margin-bottom:.9rem">
        <button type="button" id="btn_single" onclick="setMode('single')"
          style="padding:.35rem .9rem;border-radius:6px;font-size:.83rem;font-weight:600;cursor:pointer;border:1.5px solid #2563eb;background:#2563eb;color:white">Single Day</button>
        <button type="button" id="btn_range" onclick="setMode('range')"
          style="padding:.35rem .9rem;border-radius:6px;font-size:.83rem;font-weight:600;cursor:pointer;border:1.5px solid #d1d5db;background:white;color:#374151">Date Range</button>
      </div>
      <form method="GET" action="/admin/inventory" style="display:flex;flex-wrap:wrap;gap:.65rem;align-items:flex-end">
        <div>
          <label style="display:block;font-size:.75rem;font-weight:600;color:#6b7280;margin-bottom:.25rem">Date</label>
          <input type="date" name="check_from" id="check_from" value="{{ check_from }}"
            style="border:1px solid #d1d5db;border-radius:6px;padding:.38rem .6rem;font-size:.86rem">
        </div>
        <div id="to_wrapper" style="display:none">
          <label style="display:block;font-size:.75rem;font-weight:600;color:#6b7280;margin-bottom:.25rem">To</label>
          <input type="date" name="check_to" id="check_to" value="{{ check_to }}"
            style="border:1px solid #d1d5db;border-radius:6px;padding:.38rem .6rem;font-size:.86rem">
        </div>
        <button type="submit" style="background:#2563eb;color:white;border:none;border-radius:6px;padding:.42rem 1rem;font-size:.85rem;font-weight:600;cursor:pointer;align-self:flex-end">Check</button>
        {% if check_from %}<a href="/admin/inventory" style="align-self:flex-end;font-size:.82rem;color:#6b7280;text-decoration:none;padding:.42rem .5rem">✕ Clear</a>{% endif %}
      </form>
    </div>
    {% if avail_data %}
    <div style="border-top:1px solid #e5e7eb;padding:.75rem 1.25rem .5rem">
      <div style="font-size:.8rem;font-weight:600;color:#374151;margin-bottom:.6rem">
        Availability for {{ check_from }}{% if check_to and check_to != check_from %} → {{ check_to }}{% endif %}
      </div>
      <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.6rem">
        {% for item in avail_data %}
        {% set pct = ((item.reserved / item.total * 100)|int) if item.total > 0 else 0 %}
        <div style="border:1px solid #e5e7eb;border-radius:8px;padding:.65rem .85rem;background:{% if item.available == 0 %}#fef2f2{% elif item.available <= 2 %}#fffbeb{% else %}#f0fdf4{% endif %}">
          <div style="font-size:.84rem;font-weight:600;color:#111827;margin-bottom:.3rem">{{ item.name }}</div>
          <div style="display:flex;justify-content:space-between;align-items:center">
            <span style="font-size:.78rem;color:#6b7280">{{ item.reserved }} reserved</span>
            <span style="font-size:.88rem;font-weight:700;color:{% if item.available == 0 %}#dc2626{% elif item.available <= 2 %}#d97706{% else %}#16a34a{% endif %}">
              {{ item.available }}/{{ item.total }} avail{% if item.available == 0 %} — SOLD OUT{% endif %}
            </span>
          </div>
          <div style="height:4px;background:#e5e7eb;border-radius:2px;margin-top:.4rem">
            <div style="height:4px;border-radius:2px;width:{{ pct }}%;background:{% if pct>=100 %}#ef4444{% elif pct>=70 %}#f59e0b{% else %}#10b981{% endif %}"></div>
          </div>
        </div>
        {% endfor %}
      </div>
    </div>
    {% endif %}
  </div>

  <div class="card">
    <div class="card-header">
      <span>Rental Items ({{ products|length }})</span>
      <input type="text" id="inv-search" placeholder="🔍 Search items…" oninput="filterInv(this.value)"
        style="border:1px solid #d1d5db;border-radius:7px;padding:.35rem .75rem;font-size:.85rem;width:220px;outline:none;transition:border .12s"
        onfocus="this.style.borderColor='#2563eb'" onblur="this.style.borderColor='#d1d5db'">
    </div>
    <!-- Bulk action bar -->
    <div id="inv-bulk-bar" style="display:none;background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;padding:.65rem 1.1rem;margin:0 1.25rem .75rem;display:none;align-items:center;gap:.75rem;flex-wrap:wrap">
      <span id="inv-sel-count" style="font-size:.88rem;font-weight:600;color:#92400e"></span>
      <button onclick="invDeleteSelected()" style="background:#dc2626;color:white;border:none;border-radius:6px;padding:.35rem .9rem;font-size:.83rem;font-weight:700;cursor:pointer">🗑 Delete Selected</button>
      <button onclick="invClearAll()" style="background:white;color:#374151;border:1px solid #d1d5db;border-radius:6px;padding:.35rem .8rem;font-size:.83rem;font-weight:600;cursor:pointer">✕ Clear</button>
    </div>

    <form method="POST" action="/admin/inventory/save" id="inv-save-form">
      <table>
        <thead>
          <tr>
            <th style="width:36px;padding:.65rem .5rem .65rem 1rem">
              <input type="checkbox" id="inv-select-all" onchange="invToggleAll(this.checked)"
                     style="width:15px;height:15px;accent-color:#2563eb;cursor:pointer">
            </th>
            <th>Item Name</th>
            <th>Price / Unit</th>
            <th>Total Qty</th>
            <th>Remove</th>
          </tr>
        </thead>
        <tbody>
          {% for p in products %}
          <tr id="inv-row-{{ p.id }}">
            <td style="padding:.6rem .5rem .6rem 1rem">
              <input type="checkbox" class="inv-cb" value="{{ p.id }}"
                     onchange="invUpdateBulk()"
                     style="width:15px;height:15px;accent-color:#2563eb;cursor:pointer">
            </td>
            <td>
              <input type="hidden" name="id_{{ loop.index0 }}" value="{{ p.id }}">
              <input type="text" name="name_{{ loop.index0 }}" value="{{ p.name }}" required>
            </td>
            <td>
              <div style="display:flex;align-items:center;gap:.3rem">
                <span style="color:#9ca3af;font-size:.9rem">$</span>
                <input type="number" name="price_{{ loop.index0 }}" value="{{ '%.2f'|format(p.price|float) }}" min="0" step="0.01" required>
              </div>
            </td>
            <td>
              <input type="number" name="total_{{ loop.index0 }}" value="{{ p.total }}" min="0" step="1" required>
            </td>
            <td>
              <button type="submit" form="del_{{ p.id }}" class="btn btn-danger"
                      onclick="return confirm('Remove {{ p.name }}?')">✕</button>
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table>
      <input type="hidden" name="count" value="{{ products|length }}">
      <div class="save-bar">
        <a href="/admin/dashboard" class="btn btn-outline">Cancel</a>
        <button type="submit" class="btn btn-primary">Save Changes</button>
      </div>
    </form>

    <!-- Delete forms (one per item, outside main form) -->
    {% for p in products %}
    <form id="del_{{ p.id }}" method="POST" action="/admin/inventory/delete/{{ p.id }}" style="display:none"></form>
    {% endfor %}

    <script>
    function invToggleAll(checked) {
      document.querySelectorAll('.inv-cb').forEach(cb => cb.checked = checked);
      invUpdateBulk();
    }
    function invUpdateBulk() {
      const checked = [...document.querySelectorAll('.inv-cb:checked')];
      const bar = document.getElementById('inv-bulk-bar');
      const allCb = document.getElementById('inv-select-all');
      const total = document.querySelectorAll('.inv-cb').length;
      bar.style.display = checked.length > 0 ? 'flex' : 'none';
      document.getElementById('inv-sel-count').textContent = checked.length + ' item' + (checked.length > 1 ? 's' : '') + ' selected';
      allCb.indeterminate = checked.length > 0 && checked.length < total;
      allCb.checked = checked.length === total;
    }
    function invClearAll() {
      document.querySelectorAll('.inv-cb').forEach(cb => cb.checked = false);
      document.getElementById('inv-select-all').checked = false;
      invUpdateBulk();
    }
    async function invDeleteSelected() {
      const ids = [...document.querySelectorAll('.inv-cb:checked')].map(cb => cb.value);
      if (!ids.length) return;
      if (!confirm('Delete ' + ids.length + ' item' + (ids.length > 1 ? 's' : '') + '? This cannot be undone.')) return;
      for (const id of ids) {
        const form = document.getElementById('del_' + id);
        if (form) {
          const fd = new FormData(form);
          await fetch(form.action, {method:'POST', body:fd, credentials:'same-origin'});
          const row = document.getElementById('inv-row-' + id);
          if (row) row.remove();
        }
      }
      location.reload();
    }
    </script>

    <!-- Add new item -->
    <form method="POST" action="/admin/inventory/add">
      <div class="add-form">
        <input type="text" name="name" placeholder="New item name (e.g. Tents 20x20)" required>
        <div style="display:flex;align-items:center;gap:.3rem">
          <span style="color:#9ca3af">$</span>
          <input type="number" name="price" placeholder="Price" min="0" step="0.01" style="max-width:90px" required>
        </div>
        <input type="number" name="total" placeholder="Qty" min="1" step="1" style="max-width:70px" required>
        <button type="submit" class="btn btn-primary">+ Add Item</button>
      </div>
    </form>
  </div>
</div>
<script>
function filterInv(q){
  const term=q.toLowerCase();
  document.querySelectorAll('tbody tr').forEach(row=>{
    const name=(row.querySelector('input[type=text]')?.value||'').toLowerCase();
    row.style.display=name.includes(term)?'':'none';
  });
}
function setMode(m){
  const isRange=m==='range';
  document.getElementById('to_wrapper').style.display=isRange?'block':'none';
  const bs=document.getElementById('btn_single').style;
  const br=document.getElementById('btn_range').style;
  bs.background=isRange?'white':'#2563eb'; bs.color=isRange?'#374151':'white'; bs.borderColor=isRange?'#d1d5db':'#2563eb';
  br.background=isRange?'#2563eb':'white'; br.color=isRange?'white':'#374151'; br.borderColor=isRange?'#2563eb':'#d1d5db';
  if(!isRange) document.getElementById('check_to').value='';
}
{% if check_to and check_to != check_from %}setMode('range');{% else %}setMode('single');{% endif %}
</script>
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — CUSTOMER ORDER VIEW (read-only, token-gated)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/booking/view/<token>")
def customer_booking_view(token):
    if not token:
        return "Invalid link.", 404
    conn = get_db()
    if not conn:
        return "Unable to load booking. Please try again later.", 503
    b, items = None, []
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE view_token=%s", (token,))
        row = cur.fetchone()
        if row:
            b = _row(row)
            try: items = json.loads(b.get("items_json") or "[]")
            except Exception: items = []
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"Customer view error: {e}")
    if not b:
        return "Booking not found or link has expired.", 404

    STATUS_LABELS = {
        "pending":   ("🕐 Under Review",      "#92400e", "#fef3c7"),
        "accepted":  ("✅ Accepted",           "#1e40af", "#dbeafe"),
        "confirmed": ("✅ Paid in Full",       "#166534", "#dcfce7"),
        "partial":   ("💳 Partial Payment",   "#7c3aed", "#ede9fe"),
        "denied":    ("❌ Denied",             "#991b1b", "#fee2e2"),
        "cancelled": ("🚫 Cancelled",          "#6b7280", "#f3f4f6"),
    }
    status_key = (b.get("status") or "pending").lower()
    status_label, status_color, status_bg = STATUS_LABELS.get(status_key, ("Unknown", "#6b7280", "#f3f4f6"))

    disc_amount = float(b.get("discount_amount") or 0)
    disc_type   = b.get("discount_type") or ""
    disc_value  = float(b.get("discount_value") or 0)

    items_html = ""
    for it in items:
        up  = float(it.get("unit_price") or 0)
        tot = float(it.get("total") or round(up * int(it.get("qty",1)), 2))
        items_html += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #e5e7eb">{it.get('name','')}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:center">{it.get('qty','')}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right">${up:.2f}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600">${tot:.2f}</td>
        </tr>"""

    exact_fee = float(b.get("exact_time_fee") or 0)
    late_fee  = float(b.get("late_night_fee") or 0)
    if exact_fee:
        items_html += f'<tr><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280">Exact Time Delivery</td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600">${exact_fee:.2f}</td></tr>'
    if late_fee:
        items_html += f'<tr><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280">Late Night Fee</td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600">${late_fee:.2f}</td></tr>'
    delivery_fee = float(b.get("delivery_fee") or 0)
    if delivery_fee:
        miles_str = f"{b.get('distance_miles','?')} mi" if b.get("distance_miles") else ""
        items_html += f'<tr><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280">Delivery Fee {miles_str}</td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600">${delivery_fee:.2f}</td></tr>'
    if disc_amount > 0:
        disc_lbl = f"{disc_value:.1f}% Discount" if disc_type == "percent" else f"${disc_value:.2f} Discount"
        items_html += f'<tr style="background:#f0fdf4"><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#166534;font-weight:600">🏷️ {disc_lbl}</td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:700;color:#16a34a">- ${disc_amount:.2f}</td></tr>'
    tax_amount = float(b.get("tax_amount") or 0)
    tax_rate   = float(b.get("tax_rate") or 0)
    if b.get("tax_exempt"):
        items_html += '<tr><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280">CT Sales Tax <span style="font-size:.75rem;background:#dcfce7;color:#166534;border-radius:4px;padding:.1rem .35rem;margin-left:.3rem">TAX EXEMPT</span></td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;color:#166534">$0.00</td></tr>'
    elif tax_amount:
        items_html += f'<tr><td colspan="3" style="padding:10px 14px;border-bottom:1px solid #e5e7eb;color:#6b7280">CT Sales Tax ({tax_rate*100:.2f}%)</td><td style="padding:10px 14px;border-bottom:1px solid #e5e7eb;text-align:right;font-weight:600">${tax_amount:.2f}</td></tr>'

    grand_total = float(b.get("grand_total") or 0)
    amount_paid = float(b.get("amount_paid") or 0)
    balance_due = max(round(grand_total - amount_paid, 2), 0)

    venue_map = {"venue": "Event Venue", "backyard": "Backyard", "residential": "Residential", "park": "Park", "other": "Other"}

    html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Your Order — {BUSINESS_NAME}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f4f8;color:#1a202c;padding:1.5rem 1rem}}
  .wrap{{max-width:680px;margin:0 auto}}
  .card{{background:white;border-radius:12px;box-shadow:0 2px 12px rgba(0,0,0,.07);padding:1.5rem;margin-bottom:1.25rem}}
  .section-title{{font-size:.72rem;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#6b7280;margin-bottom:.9rem}}
  .row{{display:grid;grid-template-columns:1fr 1fr;gap:.5rem 1.5rem}}
  .field label{{font-size:.75rem;color:#6b7280;font-weight:600;display:block;margin-bottom:.2rem}}
  .field span{{font-size:.95rem;color:#1a202c}}
  table{{width:100%;border-collapse:collapse;font-size:.9rem}}
  th{{padding:9px 14px;background:#f8fafc;color:#6b7280;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;text-align:left}}
  th:last-child{{text-align:right}}
  .grand-row td{{padding:13px 14px;background:#1a365d;color:white;font-weight:700;font-size:1rem}}
  .grand-row td:last-child{{text-align:right;font-size:1.15rem}}
  @media(max-width:500px){{.row{{grid-template-columns:1fr}}}}
</style>
</head><body>
<div class="wrap">

  <div style="text-align:center;margin-bottom:1.5rem">
    <div style="font-size:1.5rem;font-weight:800;color:#1a365d">{BUSINESS_NAME}</div>
    <div style="font-size:.88rem;color:#6b7280;margin-top:.25rem">Order Confirmation</div>
  </div>

  <!-- Status banner -->
  <div style="background:{status_bg};border-radius:10px;padding:.85rem 1.25rem;margin-bottom:1.25rem;display:flex;align-items:center;justify-content:space-between">
    <span style="font-weight:700;color:{status_color};font-size:.95rem">{status_label}</span>
    <span style="font-size:.8rem;color:{status_color};opacity:.75">Booking #{b.get('id')}</span>
  </div>

  <!-- Your Information -->
  <div class="card">
    <div class="section-title">Your Information</div>
    <div class="row">
      <div class="field"><label>Name</label><span>{b.get('full_name') or '—'}</span></div>
      <div class="field"><label>Email</label><span>{b.get('email') or '—'}</span></div>
      <div class="field"><label>Phone</label><span>{b.get('phone') or '—'}</span></div>
      {'<div class="field"><label>Company</label><span>' + str(b.get('company_name')) + '</span></div>' if b.get('company_name') else ''}
      <div class="field" style="grid-column:1/-1"><label>Address</label>
        <span>{b.get('renter_street') or ''}{', ' + str(b.get('renter_city')) if b.get('renter_city') else ''}{', ' + str(b.get('renter_state')) if b.get('renter_state') else ''} {b.get('renter_zip') or ''}</span>
      </div>
    </div>
  </div>

  <!-- Event Details -->
  <div class="card">
    <div class="section-title">Event Details</div>
    <div class="row">
      <div class="field"><label>Delivery Date</label><span>{_fmt_date(b.get('event_start_date')) or '—'}</span></div>
      <div class="field"><label>Pickup Date</label><span>{_fmt_date(b.get('event_end_date')) or '—'}</span></div>
      <div class="field"><label>Delivery / Setup Time</label><span>{b.get('setup_time') or '—'}</span></div>
      <div class="field"><label>Pickup Time</label><span>{b.get('event_end_time') or '—'}</span></div>
      <div class="field"><label>Venue Type</label><span>{venue_map.get(b.get('venue_type',''), b.get('venue_type','') or '—')}</span></div>
      <div class="field" style="grid-column:1/-1"><label>Event Address</label>
        <span>{b.get('event_street') or ''}{', ' + str(b.get('event_city')) if b.get('event_city') else ''}{', ' + str(b.get('event_state')) if b.get('event_state') else ''} {b.get('event_zip') or ''}</span>
      </div>
      <div class="field" style="grid-column:1/-1"><label>Deliver Items To</label><span>{b.get('delivery_location') or '—'}</span></div>
      {'<div class="field" style="grid-column:1/-1"><label>Notes</label><span>' + str(b.get('notes')) + '</span></div>' if b.get('notes') else ''}
    </div>
  </div>

  <!-- Items & Totals -->
  <div class="card">
    <div class="section-title">Items &amp; Totals</div>
    <table>
      <thead><tr>
        <th>Item</th><th style="text-align:center">Qty</th><th style="text-align:right">Unit</th><th style="text-align:right">Total</th>
      </tr></thead>
      <tbody>
        {items_html}
        <tr class="grand-row"><td colspan="3">Grand Total</td><td>${grand_total:.2f}</td></tr>
      </tbody>
    </table>
    {f'<div style="margin-top:.85rem;padding:.75rem 1rem;background:#f0fdf4;border-radius:8px;display:flex;justify-content:space-between;font-size:.9rem"><span style="color:#166534;font-weight:600">✅ Amount Paid</span><span style="color:#166534;font-weight:700">${amount_paid:.2f}</span></div>' if amount_paid > 0 else ''}
    {f'<div style="margin-top:.5rem;padding:.75rem 1rem;background:#fff5f5;border-radius:8px;display:flex;justify-content:space-between;font-size:.9rem"><span style="color:#991b1b;font-weight:600">⚠️ Balance Due</span><span style="color:#991b1b;font-weight:700">${balance_due:.2f}</span></div>' if balance_due > 0.01 else ''}
  </div>

  <div style="text-align:center;color:#9ca3af;font-size:.8rem;padding:.5rem 0 1.5rem">
    {f'Questions? Call <strong style="color:#374151">{BUSINESS_PHONE}</strong> or reply to your confirmation email.' if BUSINESS_PHONE else 'Reply to your confirmation email with any questions.'}
    <br>This page is view-only. To make changes, please contact us directly.
  </div>

</div></body></html>"""
    return html


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — ADMIN
# ══════════════════════════════════════════════════════════════════════════════

# ── WebAuthn helpers ─────────────────────────────────────────────────────────

def _wau_rp_id():
    if WEBAUTHN_RP_ID:
        return WEBAUTHN_RP_ID
    if BASE_URL:
        from urllib.parse import urlparse
        return urlparse(BASE_URL).hostname or "localhost"
    return request.host.split(":")[0]

def _wau_origin():
    if BASE_URL:
        return BASE_URL.rstrip("/")
    return request.url_root.rstrip("/")

def _wau_creds():
    """Return list of stored webauthn credentials."""
    creds = []
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT credential_id, public_key, sign_count FROM webauthn_credentials")
            creds = [dict(r) for r in cur.fetchall()]
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"wau_creds error: {e}")
    return creds

def _wau_has_creds():
    return len(_wau_creds()) > 0


# ── WebAuthn registration (requires existing password login) ─────────────────

@app.route("/admin/webauthn/register/begin", methods=["POST"])
@admin_required
def webauthn_register_begin():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"error": "WebAuthn library not installed"}), 503
    try:
        opts = generate_registration_options(
            rp_id=_wau_rp_id(),
            rp_name=BUSINESS_NAME,
            user_id=b"admin",
            user_name="admin",
            user_display_name="Admin",
            authenticator_selection=AuthenticatorSelectionCriteria(
                resident_key=ResidentKeyRequirement.PREFERRED,
                user_verification=UserVerificationRequirement.REQUIRED,
            ),
        )
        session["wau_reg_challenge"] = opts.challenge
        return options_to_json(opts), 200, {"Content-Type": "application/json"}
    except Exception as e:
        log.error(f"webauthn_register_begin: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/admin/webauthn/register/complete", methods=["POST"])
@admin_required
def webauthn_register_complete():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"error": "WebAuthn library not installed"}), 503
    challenge = session.pop("wau_reg_challenge", None)
    if not challenge:
        return jsonify({"error": "No challenge in session"}), 400
    try:
        credential = parse_registration_credential_json(request.get_data(as_text=True))
        verified = verify_registration_response(
            credential=credential,
            expected_challenge=challenge,
            expected_rp_id=_wau_rp_id(),
            expected_origin=_wau_origin(),
        )
        import base64
        cred_id_b64 = base64.urlsafe_b64encode(verified.credential_id).rstrip(b"=").decode()
        conn = get_db()
        if conn:
            cur = conn.cursor()
            cur.execute(
                """INSERT INTO webauthn_credentials (credential_id, public_key, sign_count)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (credential_id) DO UPDATE SET sign_count=%s""",
                (cred_id_b64, bytes(verified.credential_public_key),
                 verified.sign_count, verified.sign_count)
            )
            conn.commit(); cur.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        log.error(f"webauthn_register_complete: {e}")
        return jsonify({"error": str(e)}), 400


# ── WebAuthn authentication ───────────────────────────────────────────────────

@app.route("/admin/webauthn/authenticate/begin", methods=["POST"])
def webauthn_authenticate_begin():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"error": "WebAuthn not available"}), 503
    try:
        import base64
        creds = _wau_creds()
        if not creds:
            return jsonify({"error": "No biometric registered"}), 400
        allow_creds = [
            PublicKeyCredentialDescriptor(
                id=base64.urlsafe_b64decode(row["credential_id"] + "==")
            )
            for row in creds
        ]
        opts = generate_authentication_options(
            rp_id=_wau_rp_id(),
            allow_credentials=allow_creds,
            user_verification=UserVerificationRequirement.REQUIRED,
        )
        session["wau_auth_challenge"] = opts.challenge
        return options_to_json(opts), 200, {"Content-Type": "application/json"}
    except Exception as e:
        log.error(f"webauthn_authenticate_begin: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/admin/webauthn/authenticate/complete", methods=["POST"])
def webauthn_authenticate_complete():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"error": "WebAuthn not available"}), 503
    challenge = session.pop("wau_auth_challenge", None)
    if not challenge:
        return jsonify({"error": "No challenge"}), 400
    try:
        import base64
        credential = parse_authentication_credential_json(request.get_data(as_text=True))
        cred_id_b64 = base64.urlsafe_b64encode(credential.raw_id).rstrip(b"=").decode()
        # Look up stored credential
        conn = get_db(); row = None
        if conn:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM webauthn_credentials WHERE credential_id=%s", (cred_id_b64,))
            row = cur.fetchone()
            cur.close(); conn.close()
        if not row:
            return jsonify({"error": "Credential not recognised"}), 400
        verified = verify_authentication_response(
            credential=credential,
            expected_challenge=challenge,
            expected_rp_id=_wau_rp_id(),
            expected_origin=_wau_origin(),
            credential_public_key=bytes(row["public_key"]),
            credential_current_sign_count=row["sign_count"],
        )
        # Update sign count
        conn2 = get_db()
        if conn2:
            cur2 = conn2.cursor()
            cur2.execute("UPDATE webauthn_credentials SET sign_count=%s WHERE credential_id=%s",
                         (verified.new_sign_count, cred_id_b64))
            conn2.commit(); cur2.close(); conn2.close()
        session.permanent = True
        session["admin_logged_in"] = True
        return jsonify({"ok": True})
    except Exception as e:
        log.error(f"webauthn_authenticate_complete: {e}")
        return jsonify({"error": str(e)}), 400


# ── Setup biometric page (requires password login) ───────────────────────────

BIOMETRIC_SETUP_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Set Up Biometric — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;background:linear-gradient(160deg,#1e1e2e 0%,#1e3a5f 100%);padding:1.5rem}
    .card{background:rgba(255,255,255,.06);border:1px solid rgba(255,255,255,.12);border-radius:20px;padding:2.25rem 2rem;width:100%;max-width:360px;text-align:center;color:#fff}
    h1{font-size:1.15rem;font-weight:700;margin-bottom:.4rem}
    p{font-size:.84rem;color:rgba(255,255,255,.55);line-height:1.6;margin-bottom:1.5rem}
    .btn{width:100%;padding:.9rem;border:none;border-radius:12px;font-size:.93rem;font-weight:700;cursor:pointer;margin-bottom:.65rem}
    .btn-bio{background:linear-gradient(135deg,#6366f1,#8b5cf6);color:#fff}
    .btn-back{background:rgba(255,255,255,.08);color:rgba(255,255,255,.7);border:1px solid rgba(255,255,255,.12)}
    .msg{border-radius:10px;padding:.75rem;margin-bottom:.85rem;font-size:.85rem;font-weight:600}
    .msg-err{background:rgba(239,68,68,.15);border:1px solid rgba(239,68,68,.35);color:#fca5a5}
    .msg-ok{background:rgba(22,163,74,.15);border:1px solid rgba(22,163,74,.4);color:#4ade80}
    .existing{background:rgba(99,102,241,.12);border:1px solid rgba(99,102,241,.3);border-radius:10px;padding:.75rem;margin-bottom:1rem;font-size:.82rem;color:rgba(255,255,255,.65)}
    @keyframes spin{from{transform:rotate(0deg)}to{transform:rotate(360deg)}}
  </style>
</head>
<body>
<div class="card">
  <div style="font-size:2.5rem;margin-bottom:.75rem">🔐</div>
  <h1>Set Up Biometric Login</h1>
  <p>Register your Face ID or fingerprint so you can sign in without a password next time.</p>

  {% if cred_count > 0 %}
  <div class="existing">✓ {{ cred_count }} device{{ 's' if cred_count > 1 else '' }} already registered. You can add another below.</div>
  {% endif %}

  <div id="msg-area"></div>

  <button class="btn btn-bio" id="reg-btn" onclick="doRegister()">
    📲 Register This Device
  </button>
  {% if cred_count > 0 %}
  <form method="POST" action="/admin/webauthn/delete-all" onsubmit="return confirm('Remove all biometric credentials?')">
    <button class="btn" style="background:rgba(239,68,68,.15);color:#fca5a5;border:1px solid rgba(239,68,68,.3)">🗑 Remove All Biometrics</button>
  </form>
  {% endif %}
  <br>
  <a href="/admin/dashboard" class="btn btn-back" style="display:block;text-decoration:none;padding:.7rem">← Back to Dashboard</a>
</div>
<script>
function showMsg(text, type) {
  document.getElementById('msg-area').innerHTML = '<div class="msg msg-' + type + '">' + text + '</div>';
}
function b64u_to_buf(b64u) {
  var b64 = b64u.replace(/-/g,'+').replace(/_/g,'/');
  var pad = b64.length % 4; if (pad) b64 += '===='.slice(pad);
  var bin = atob(b64), arr = new Uint8Array(bin.length);
  for (var i=0; i<bin.length; i++) arr[i] = bin.charCodeAt(i);
  return arr.buffer;
}
function buf_to_b64u(buf) {
  var arr = new Uint8Array(buf), str = '';
  for (var i=0; i<arr.length; i++) str += String.fromCharCode(arr[i]);
  return btoa(str).replace(/\+/g,'-').replace(/\//g,'_').replace(/=/g,'');
}
async function doRegister() {
  var btn = document.getElementById('reg-btn');
  btn.disabled = true;
  btn.innerHTML = '<svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" style="animation:spin 1s linear infinite;vertical-align:middle"><circle cx="12" cy="12" r="10" stroke-dasharray="40 20"/></svg>  Waiting for biometric…';
  try {
    var r = await fetch('/admin/webauthn/register/begin', {method:'POST', headers:{'Content-Type':'application/json'}, body:'{}'});
    if (!r.ok) { var e=await r.json(); throw new Error(e.error||'Server error'); }
    var opts = await r.json();
    opts.challenge = b64u_to_buf(opts.challenge);
    opts.user.id = b64u_to_buf(opts.user.id);
    if (opts.excludeCredentials) opts.excludeCredentials = opts.excludeCredentials.map(function(c){ return {id:b64u_to_buf(c.id),type:c.type}; });
    var cred = await navigator.credentials.create({publicKey: opts});
    var payload = {
      id: cred.id, rawId: buf_to_b64u(cred.rawId), type: cred.type,
      response: {
        attestationObject: buf_to_b64u(cred.response.attestationObject),
        clientDataJSON: buf_to_b64u(cred.response.clientDataJSON)
      }
    };
    var r2 = await fetch('/admin/webauthn/register/complete', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
    var result = await r2.json();
    if (result.ok) {
      showMsg('✓ Biometric registered! You can now sign in with Face ID / Touch ID.', 'ok');
      btn.innerHTML = '✓ Registered';
    } else { throw new Error(result.error || 'Registration failed'); }
  } catch(e) {
    showMsg(e.name === 'NotAllowedError' ? 'Registration cancelled.' : ('Error: ' + e.message), 'err');
    btn.disabled = false;
    btn.innerHTML = '📲 Register This Device';
  }
}
</script>
</body></html>
"""


@app.route("/admin/setup-biometric")
@admin_required
def setup_biometric():
    cred_count = len(_wau_creds())
    return render_template_string(BIOMETRIC_SETUP_HTML, business_name=BUSINESS_NAME, cred_count=cred_count)


@app.route("/admin/webauthn/delete-all", methods=["POST"])
@admin_required
def webauthn_delete_all():
    conn = get_db()
    if conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM webauthn_credentials")
        conn.commit(); cur.close(); conn.close()
    return redirect(url_for("setup_biometric"))


# ── Login / Logout ────────────────────────────────────────────────────────────

@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    if session.get("admin_logged_in"):
        return redirect(url_for("admin_dashboard"))
    error = None
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session.permanent = True
            session["admin_logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        error = "Incorrect password."
    has_biometric = WEBAUTHN_AVAILABLE and _wau_has_creds()
    return render_template_string(ADMIN_LOGIN_HTML,
                                  business_name=BUSINESS_NAME,
                                  error=error,
                                  has_biometric=has_biometric)


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    status_filter   = request.args.get("status", "")
    date_from       = request.args.get("date_from", "")
    date_to         = request.args.get("date_to", "")
    pay_filter      = request.args.get("pay_filter", "")   # paid | partial | due | ""
    upcoming_filter = bool(request.args.get("upcoming", ""))
    archived_filter = bool(request.args.get("archived", ""))
    past_filter     = bool(request.args.get("past", ""))
    sort_by         = request.args.get("sort", "")           # date | name | id | created | ""=auto
    tab             = request.args.get("tab", "all")        # all | going_out | upcoming | delivered | picked_up
    conn = get_db()
    bookings = []
    stats = {"total": 0, "pending": 0, "accepted": 0, "confirmed": 0, "partial": 0, "revenue": 0, "amount_due": 0, "upcoming": 0, "past": 0, "going_out": 0, "coming_back": 0, "delivered": 0, "picked_up": 0, "still_waiting": 0}
    inventory_status = []

    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT COUNT(*) FROM bookings"); stats["total"] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM bookings WHERE status IN ('pending','agree_to_pay')"); stats["pending"] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM bookings WHERE status='accepted'"); stats["accepted"] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM bookings WHERE status='accepted' AND payment_status='paid'"); stats["confirmed"] = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM bookings WHERE status='accepted' AND payment_status='partial'"); stats["partial"] = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(grand_total),0) FROM bookings WHERE status='accepted' AND payment_status IN ('paid','partial')")
            stats["revenue"] = float(cur.fetchone()[0])
            cur.execute("SELECT COALESCE(SUM(grand_total),0) FROM bookings WHERE status='accepted'")
            stats["amount_due"] = float(cur.fetchone()[0])
            today_dt = date.today()
            seven_days_ago = (today_dt - timedelta(days=7)).isoformat()
            in_8_days = (today_dt + timedelta(days=8)).isoformat()

            # Tab counts
            cur.execute("""
                SELECT COUNT(*) FROM bookings
                WHERE COALESCE(delivery_date, setup_date, event_start_date) = %s
                  AND status NOT IN ('cancelled','denied')
                  AND (archived IS NULL OR archived = FALSE)
            """, (today_dt.isoformat(),))
            stats["going_out"] = cur.fetchone()[0]

            cur.execute("""
                SELECT COUNT(*) FROM bookings
                WHERE event_end_date = %s
                  AND status NOT IN ('cancelled','denied')
                  AND (archived IS NULL OR archived = FALSE)
            """, (today_dt.isoformat(),))
            stats["coming_back"] = cur.fetchone()[0]

            cur.execute("""
                SELECT COUNT(*) FROM bookings
                WHERE COALESCE(delivery_date, setup_date, event_start_date) > %s
                  AND COALESCE(delivery_date, setup_date, event_start_date) <= %s
                  AND status NOT IN ('cancelled','denied')
                  AND (archived IS NULL OR archived = FALSE)
            """, (today_dt.isoformat(), in_8_days))
            stats["upcoming"] = cur.fetchone()[0]

            cur.execute("""
                SELECT COUNT(*) FROM bookings
                WHERE COALESCE(delivery_date, setup_date, event_start_date) >= %s
                  AND COALESCE(delivery_date, setup_date, event_start_date) <= %s
                  AND (archived IS NULL OR archived = FALSE)
                  AND (
                    status = 'pending'
                    OR (status = 'accepted' AND (payment_status IS NULL OR payment_status = 'waiting'))
                  )
            """, (today_dt.isoformat(), in_8_days))
            stats["still_waiting"] = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM bookings WHERE delivery_status = 'delivered'")
            stats["delivered"] = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM bookings WHERE delivery_status = 'picked_up' AND picked_up_at >= NOW() - INTERVAL '48 hours'")
            stats["picked_up"] = cur.fetchone()[0]

            # Count past (kept for legacy)
            cur.execute("""
                SELECT COUNT(*) FROM bookings
                WHERE (event_start_date < %s OR event_start_date IS NULL)
                  AND (archived IS NULL OR archived = FALSE)
            """, (seven_days_ago,))
            stats["past"] = cur.fetchone()[0]

            # Build filtered query
            wheres = []
            params = []
            if tab == "going_out":
                wheres.append("COALESCE(delivery_date, setup_date, event_start_date) = %s"); params.append(today_dt.isoformat())
                wheres.append("status NOT IN ('cancelled','denied')")
                wheres.append("(archived IS NULL OR archived = FALSE)")
            elif tab == "coming_back":
                wheres.append("event_end_date = %s"); params.append(today_dt.isoformat())
                wheres.append("status NOT IN ('cancelled','denied')")
                wheres.append("(archived IS NULL OR archived = FALSE)")
            elif tab == "upcoming":
                wheres.append("COALESCE(delivery_date, setup_date, event_start_date) > %s"); params.append(today_dt.isoformat())
                wheres.append("COALESCE(delivery_date, setup_date, event_start_date) <= %s"); params.append(in_8_days)
                wheres.append("status NOT IN ('cancelled','denied')")
                wheres.append("(archived IS NULL OR archived = FALSE)")
            elif tab == "still_waiting":
                wheres.append("COALESCE(delivery_date, setup_date, event_start_date) >= %s"); params.append(today_dt.isoformat())
                wheres.append("COALESCE(delivery_date, setup_date, event_start_date) <= %s"); params.append(in_8_days)
                wheres.append("(archived IS NULL OR archived = FALSE)")
                wheres.append("(status = 'pending' OR (status = 'accepted' AND (payment_status IS NULL OR payment_status = 'waiting')))")
            elif tab == "delivered":
                wheres.append("delivery_status = 'delivered'")
            elif tab == "picked_up":
                wheres.append("delivery_status = 'picked_up'")
                wheres.append("picked_up_at >= NOW() - INTERVAL '48 hours'")
            elif archived_filter:
                wheres.append("archived = TRUE")
            elif upcoming_filter:
                # legacy support
                wheres.append("event_start_date >= %s"); params.append(today_dt.isoformat())
                wheres.append("event_start_date <= %s"); params.append(in_8_days)
                wheres.append("status NOT IN ('cancelled','denied')")
                wheres.append("(archived IS NULL OR archived = FALSE)")
            elif past_filter:
                wheres.append("(event_start_date < %s OR event_start_date IS NULL)")
                params.append(seven_days_ago)
                wheres.append("(archived IS NULL OR archived = FALSE)")
                if status_filter:
                    wheres.append("status=%s"); params.append(status_filter)
            else:
                # All tab — no date restriction, newest first
                wheres.append("(archived IS NULL OR archived = FALSE)")
                if status_filter:
                    if status_filter == "pending":
                        wheres.append("status IN ('pending','agree_to_pay')")
                    else:
                        wheres.append("status=%s"); params.append(status_filter)
                if date_from:
                    wheres.append("event_start_date >= %s"); params.append(date_from)
                if date_to:
                    wheres.append("event_start_date <= %s"); params.append(date_to)
            q = "SELECT * FROM bookings"
            if wheres:
                q += " WHERE " + " AND ".join(wheres)
            # Sort order — auto-default based on tab
            sort_map = {
                "date":        "event_start_date ASC NULLS LAST, created_at DESC",
                "date_desc":   "event_start_date DESC NULLS LAST, created_at DESC",
                "name":        "full_name ASC",
                "name_desc":   "full_name DESC",
                "id":          "id DESC",
                "id_asc":      "id ASC",
                "total":       "grand_total DESC NULLS LAST",
                "created":     "created_at DESC",
                "created_asc": "created_at ASC",
            }
            tab_default_sort = {
                "all":           "created_at DESC",
                "going_out":     "COALESCE(delivery_date, setup_date, event_start_date) ASC NULLS LAST, created_at DESC",
                "coming_back":   "event_end_date ASC NULLS LAST, created_at DESC",
                "upcoming":      "COALESCE(delivery_date, setup_date, event_start_date) ASC NULLS LAST, created_at DESC",
                "delivered":     "delivered_at DESC NULLS LAST",
                "picked_up":     "picked_up_at DESC NULLS LAST",
                "still_waiting": "COALESCE(delivery_date, setup_date, event_start_date) ASC NULLS LAST, created_at DESC",
            }
            if sort_by and sort_by in sort_map:
                order_clause = sort_map[sort_by]
            else:
                order_clause = tab_default_sort.get(tab, "created_at DESC")
            q += " ORDER BY " + order_clause
            q += " LIMIT 1000"
            cur.execute(q, params)
            rows = cur.fetchall()
            _avatar_colors = ['#ef4444','#f97316','#eab308','#22c55e','#14b8a6',
                               '#3b82f6','#8b5cf6','#ec4899','#06b6d4','#84cc16']
            for row in rows:
                b = _row(row)
                items = json.loads(b.get("items_json") or "[]")
                b["items_summary"] = "  ·  ".join(f"{i['qty']}x" for i in items)
                # Payment label + class — driven by payment_status field
                pmt_st  = b.get("payment_status") or ""
                paid    = float(b.get("amount_paid") or 0)
                total   = float(b.get("grand_total") or 0)
                owed    = round(total - paid, 2) if total > 0 else 0

                if b["status"] == "agree_to_pay":
                    b["pay_label"], b["pay_class"] = "Cash/Check at Delivery", "pay-paid"
                elif b["status"] not in ("accepted",):
                    b["pay_label"], b["pay_class"] = "—", "pay-none"
                elif pmt_st == "paid":
                    b["pay_label"], b["pay_class"] = "Paid in Full", "pay-paid"
                elif pmt_st == "partial":
                    b["pay_label"] = f"Partial — ${owed:,.2f} owed"
                    b["pay_class"] = "pay-partial"
                elif pmt_st == "waiting":
                    b["pay_label"], b["pay_class"] = "Waiting", "pay-due"
                else:
                    # Legacy fallback: infer from amount_paid
                    if paid > 0 and owed > 0.01:
                        b["pay_label"] = f"Partial — ${owed:,.2f} owed"
                        b["pay_class"] = "pay-partial"
                    elif paid >= total - 0.50 and total > 0:
                        b["pay_label"], b["pay_class"] = "Paid in Full", "pay-paid"
                    elif b["status"] == "accepted":
                        b["pay_label"], b["pay_class"] = "Waiting", "pay-due"
                    else:
                        b["pay_label"], b["pay_class"] = "—", "pay-none"
                # Apply pay_filter AFTER labelling
                if pay_filter:
                    if pay_filter == "paid" and b["pay_class"] != "pay-paid": continue
                    if pay_filter == "partial" and b["pay_class"] != "pay-partial": continue
                    if pay_filter == "due" and b["pay_class"] != "pay-due": continue
                # Google Maps URL for delivery address
                addr_parts = [p for p in [
                    (b.get("event_street") or "").strip(),
                    (b.get("event_city") or "").strip(),
                    (b.get("event_state") or "").strip(),
                    (b.get("event_zip") or "").strip()
                ] if p]
                if addr_parts:
                    b["maps_url"] = "https://www.google.com/maps/search/?api=1&query=" + urllib.parse.quote_plus(", ".join(addr_parts))
                else:
                    b["maps_url"] = ""
                # Avatar
                name = b.get("full_name") or "?"
                b["avatar_color"]    = _avatar_colors[ord(name[0].lower()) % len(_avatar_colors)]
                b["avatar_initials"] = name[0].upper()
                # Red flag: delivery within 5 days and not paid/agreed
                _setup = b.get("delivery_date") or b.get("setup_date")
                _pst   = b.get("payment_status") or ""
                _st    = b.get("status") or ""
                if _setup:
                    try:
                        _days_out = (_setup - today_dt).days
                        b["red_flag"] = (
                            0 <= _days_out <= 5
                            and _st not in ("denied", "cancelled", "concluded")
                            and _st != "agree_to_pay"
                            and _pst != "paid"
                        )
                    except Exception:
                        b["red_flag"] = False
                else:
                    b["red_flag"] = False
                bookings.append(b)

            today_str = date.today().isoformat()
            avail = get_available(today_str, "2099-12-31")
            for p2 in get_products():
                reserved = p2["total"] - avail.get(p2["id"], p2["total"])
                inventory_status.append({
                    "name": p2["name"], "total": p2["total"],
                    "reserved": reserved, "available": avail.get(p2["id"], p2["total"])
                })

            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"Admin dashboard error: {e}")

    # ── Auto-archive: older than 2 weeks + picked-up past 48 hours ──────────
    try:
        conn_aa = get_db()
        if conn_aa:
            cur_aa = conn_aa.cursor()
            two_weeks_ago = (date.today() - timedelta(days=14)).isoformat()
            # Archive bookings whose event ended more than 2 weeks ago
            cur_aa.execute("""
                UPDATE bookings
                SET archived = TRUE
                WHERE (archived IS NULL OR archived = FALSE)
                  AND event_start_date < %s
                  AND status NOT IN ('denied','cancelled','concluded')
            """, (two_weeks_ago,))
            # Archive picked-up orders older than 48 hours
            cur_aa.execute("""
                UPDATE bookings
                SET archived = TRUE
                WHERE (archived IS NULL OR archived = FALSE)
                  AND delivery_status = 'picked_up'
                  AND picked_up_at IS NOT NULL
                  AND picked_up_at < NOW() - INTERVAL '48 hours'
            """)
            conn_aa.commit()
            cur_aa.close()
            conn_aa.close()
    except Exception as e:
        log.error(f"Auto-archive error: {e}")

    inv_conflicts = get_inventory_conflicts()
    conflict_map = {}
    for _c in inv_conflicts:
        conflict_map.setdefault(_c["booking_id"], []).append(_c)

    # Going out today / Coming back today
    going_out   = []
    coming_back = []
    try:
        conn2 = get_db()
        if conn2:
            cur2 = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
            today_iso = date.today().isoformat()
            cur2.execute("""
                SELECT id, full_name, email,
                       COALESCE(delivery_time, setup_time) AS event_start_time,
                       items_json, status
                FROM bookings
                WHERE COALESCE(delivery_date, setup_date) = %s
                  AND status NOT IN ('denied','cancelled','concluded')
                  AND (archived IS NULL OR archived = FALSE)
                ORDER BY COALESCE(delivery_time, setup_time) ASC NULLS LAST, id ASC
            """, (today_iso,))
            for row in cur2.fetchall():
                b2 = dict(row)
                name = b2.get('full_name') or '?'
                b2['avatar_initials'] = name[0].upper()
                b2['avatar_color'] = _avatar_colors[ord(name[0].lower()) % len(_avatar_colors)]
                try:
                    b2['item_count'] = sum(int(i.get('qty', 1)) for i in json.loads(b2.get('items_json') or '[]'))
                except Exception:
                    b2['item_count'] = 0
                going_out.append(b2)
            cur2.execute("""
                SELECT id, full_name, email, event_start_time, items_json, status
                FROM bookings
                WHERE event_end_date = %s
                  AND status NOT IN ('denied','cancelled','concluded')
                  AND (archived IS NULL OR archived = FALSE)
                ORDER BY event_start_time ASC NULLS LAST, id ASC
            """, (today_iso,))
            for row in cur2.fetchall():
                b2 = dict(row)
                name = b2.get('full_name') or '?'
                b2['avatar_initials'] = name[0].upper()
                b2['avatar_color'] = _avatar_colors[ord(name[0].lower()) % len(_avatar_colors)]
                try:
                    b2['item_count'] = sum(int(i.get('qty', 1)) for i in json.loads(b2.get('items_json') or '[]'))
                except Exception:
                    b2['item_count'] = 0
                coming_back.append(b2)
            cur2.close()
            conn2.close()
    except Exception as e:
        log.error(f"going_out/coming_back error: {e}")

    return render_template_string(ADMIN_DASH_HTML,
        business_name=BUSINESS_NAME,
        bookings=bookings,
        stats=stats,
        inventory=inventory_status,
        status_filter=status_filter,
        date_from=date_from,
        date_to=date_to,
        pay_filter=pay_filter,
        upcoming_filter=upcoming_filter,
        archived_filter=archived_filter,
        past_filter=past_filter,
        sort_by=sort_by,
        tab=tab,
        inv_conflicts=inv_conflicts,
        conflict_map=conflict_map,
        going_out=going_out,
        coming_back=coming_back,
        today_label=date.today().strftime('%A, %B %-d %Y'),
    )


@app.route("/admin/booking/<int:booking_id>")
@admin_required
def admin_booking(booking_id):
    conn = get_db()
    b, items = None, []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                b = _row(row)
                try:
                    items_raw = json.loads(b.get("items_json") or "[]")
                    items = items_raw if isinstance(items_raw, list) else []
                except Exception:
                    items = []
                # Auto-fill prices from inventory for items missing unit_price
                _prod_map = {p["name"].lower(): float(p.get("price") or 0) for p in get_products()}
                for item in items:
                    if not item.get("unit_price"):
                        price = _prod_map.get((item.get("name") or "").lower(), 0)
                        item["unit_price"] = price
                        item["total"] = round(price * int(item.get("qty") or 1), 2)
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"Booking fetch error: {e}")

    if not b:
        return "Booking not found", 404

    # Calculate days until event for payment label logic
    days_until = 999
    try:
        event_dt = datetime.strptime(str(b.get("event_start_date", ""))[:10], "%Y-%m-%d").date()
        days_until = (event_dt - date.today()).days
    except Exception:
        pass

    # Check for matching customer profile by name
    matched_customer = None
    try:
        conn2 = get_db()
        if conn2:
            cur2 = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur2.execute("""
                SELECT * FROM customers
                WHERE LOWER(TRIM(full_name)) = LOWER(TRIM(%s))
                LIMIT 1
            """, (b.get("full_name") or "",))
            crow = cur2.fetchone()
            cur2.close(); conn2.close()
            if crow:
                matched_customer = _row(crow)
    except Exception as e:
        log.error(f"Customer match lookup error: {e}")

    # Weekend residential schedule detection
    weekend_residential = None
    try:
        if (b.get("venue_type") or "").lower() == "residential":
            esd = str(b.get("event_start_date", ""))[:10]
            event_date = datetime.strptime(esd, "%Y-%m-%d").date()
            weekday = event_date.weekday()  # 5=Saturday, 6=Sunday
            if weekday == 5:  # Saturday
                friday = event_date - timedelta(days=1)
                sunday = event_date + timedelta(days=1)
                weekend_residential = {
                    "day_label": "Saturday",
                    "delivery_date": friday,
                    "delivery_label": f"Friday {friday.strftime('%b %-d')} at 4:00 PM",
                    "pickup_date": sunday,
                    "pickup_label": f"Sunday {sunday.strftime('%b %-d')} at 10:00 AM",
                    "pickup_weekday": "sunday",
                }
            elif weekday == 6:  # Sunday
                friday = event_date - timedelta(days=2)
                monday = event_date + timedelta(days=1)
                weekend_residential = {
                    "day_label": "Sunday",
                    "delivery_date": friday,
                    "delivery_label": f"Friday {friday.strftime('%b %-d')} at 4:00 PM",
                    "pickup_date": monday,
                    "pickup_label": f"Monday {monday.strftime('%b %-d')} at 10:00 AM",
                    "pickup_weekday": "monday",
                }
    except Exception:
        pass

    # Build full inventory status (always, not just on shortage)
    booking_inv_issues = []
    booking_inv_status = []   # [{item, needed, available, total, ok}]
    try:
        _products_inv   = get_products()
        _prod_totals_inv = {p["id"]: int(p["total"]) for p in _products_inv}
        _name_to_pid    = {p["name"].lower(): p["id"] for p in _products_inv}
        _b_items = json.loads(b.get("items_json") or "[]")

        # Inventory window = est. delivery (setup_date) → est. pickup (event_end_date).
        # Only bookings with BOTH dates set count against inventory.
        _b_out  = str(b.get("setup_date")    or "")[:10]
        _b_back = str(b.get("event_end_date") or "")[:10]

        _others_reserved = {}
        if _b_out and _b_back:
            _conn_inv = get_db()
            if _conn_inv:
                _cur_inv = _conn_inv.cursor(cursor_factory=psycopg2.extras.DictCursor)
                _cur_inv.execute("""
                    SELECT items_json FROM bookings
                    WHERE status = 'accepted' AND payment_status IN ('paid','partial')
                      AND (delivery_status IS NULL OR delivery_status != 'picked_up')
                      AND id != %s
                      AND setup_date     IS NOT NULL
                      AND event_end_date IS NOT NULL
                      AND setup_date     <= %s
                      AND event_end_date >= %s
                """, (booking_id, _b_back, _b_out))
                for _orow in _cur_inv.fetchall():
                    for _oit in json.loads(_orow["items_json"] or "[]"):
                        _opid = _oit.get("id") or _name_to_pid.get((_oit.get("name") or "").lower())
                        _oqty = int(_oit.get("qty") or 0)
                        if _opid and _oqty > 0:
                            _others_reserved[_opid] = _others_reserved.get(_opid, 0) + _oqty
                _cur_inv.close(); _conn_inv.close()

        for _it in _b_items:
            _iname = (_it.get("name") or "").strip()
            _pid   = _it.get("id") or _name_to_pid.get(_iname.lower())
            _qty   = int(_it.get("qty") or 0)
            if _pid and _qty > 0 and _pid in _prod_totals_inv:
                _total  = _prod_totals_inv[_pid]
                _reserved = _others_reserved.get(_pid, 0)
                _avail  = max(0, _total - _reserved)
                _ok     = _qty <= _avail
                booking_inv_status.append({
                    "item":      _iname,
                    "needed":    _qty,
                    "reserved":  _reserved,
                    "available": _avail,
                    "total":     _total,
                    "shortfall": max(0, _qty - _avail),
                    "ok":        _ok,
                })
                if not _ok:
                    booking_inv_issues.append({
                        "item":                 _iname,
                        "needed":               _qty,
                        "available":            _avail,
                        "shortfall":            _qty - _avail,
                        "conflicting_bookings": [],
                    })
        log.info(f"[INV] booking #{booking_id}: status={booking_inv_status} issues={booking_inv_issues}")

        # Paid/partial bookings own their inventory — no shortage alert for them.
        # Only flag shortages for Waiting payment or Pending bookings.
        _pmt = b.get("payment_status") or ""
        if _pmt in ("paid", "partial") or b.get("status") in ("concluded", "delivered"):
            booking_inv_issues = []
            for _s in booking_inv_status:
                _s["ok"] = True
                _s["shortfall"] = 0

    except Exception as _inv_err:
        log.error(f"Inventory check error for #{booking_id}: {_inv_err}")
        import traceback as _inv_tb; log.error(_inv_tb.format_exc())

    # Fetch all confirmed/accepted booking date ranges for the calendar
    cal_bookings = []
    try:
        conn3 = get_db()
        if conn3:
            cur3 = conn3.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur3.execute("""
                SELECT id, full_name, event_start_date, event_end_date, status
                FROM bookings
                WHERE status IN ('confirmed','partial','accepted')
                  AND event_start_date IS NOT NULL
                ORDER BY event_start_date
            """)
            for row in cur3.fetchall():
                r = _row(row)
                cal_bookings.append({
                    "id":    r["id"],
                    "name":  r.get("full_name",""),
                    "start": str(r.get("event_start_date",""))[:10],
                    "end":   str(r.get("event_end_date",""))[:10],
                    "status": r.get("status",""),
                })
            cur3.close(); conn3.close()
    except Exception as e:
        log.error(f"Calendar bookings fetch error: {e}")

    payment_history = []
    _ph_conn = get_db()
    if _ph_conn:
        try:
            _ph_cur = _ph_conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            _ph_cur.execute("SELECT * FROM payment_logs WHERE booking_id=%s ORDER BY paid_at ASC", (booking_id,))
            payment_history = [dict(r) for r in _ph_cur.fetchall()]
            _ph_cur.close(); _ph_conn.close()
        except Exception as _phe:
            log.error(f"payment_history error: {_phe}")
    try:
        return render_template_string(ADMIN_BOOKING_HTML,
            business_name=BUSINESS_NAME, b=b, items=items, days_until=days_until,
            products=get_products(), payment_links=get_payment_links(booking_id),
            matched_customer=matched_customer, weekend_residential=weekend_residential,
            booking_inv_issues=booking_inv_issues,
            booking_inv_status=booking_inv_status,
            cal_bookings=cal_bookings,
            payment_history=payment_history)
    except Exception as e:
        log.error(f"Booking {booking_id} render error: {e}")
        return "Error rendering booking — please contact support.", 500


@app.route("/admin/booking/<int:booking_id>/accept", methods=["POST"])
@admin_required
def accept_booking(booking_id):
    """Accept booking: create Stripe payment link, email invoice + contract + link to customer."""
    import traceback as _tb
    try:
        conn = get_db()
        if not conn:
            return redirect(url_for("admin_dashboard"))

        b = None
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            cur.close()
            conn.close()
            if row:
                b = _row(row)
        except Exception as e:
            log.error(f"Accept fetch error: {e}")
            return redirect(url_for("admin_dashboard"))

        if not b:
            return "Booking not found", 404

        grand_total = float(b.get("grand_total") or 0)
        try:
            items = json.loads(b.get("items_json") or "[]")
        except Exception:
            items = []
        items_desc = ", ".join(f"{i.get('qty','')}x {i.get('name','')}" for i in items if isinstance(i, dict))

        # ── Determine payment type based on days until delivery (setup_date) ──
        delivery_date_raw = b.get("setup_date") or b.get("event_start_date")
        days_until = 999
        try:
            event_dt = datetime.strptime(str(delivery_date_raw)[:10], "%Y-%m-%d").date()
            days_until = (event_dt - date.today()).days
            log.info(f"Booking #{booking_id}: delivery={event_dt}, today={date.today()}, days_until={days_until}")
        except Exception as e:
            log.error(f"Date calc error for booking #{booking_id}: {e} (raw={delivery_date_raw!r})")

        if days_until <= 7:
            # Event within 7 days — full payment required
            charge_amount  = round(grand_total, 2)
            payment_type   = "full"
            product_name   = f"Full Payment — Booking #{booking_id}"
            stripe_desc    = items_desc
            log.info(f"Booking #{booking_id}: {days_until} days away — requiring FULL payment ${charge_amount:.2f}")
        else:
            # More than 7 days away — 25% deposit
            charge_amount  = round(grand_total * DEPOSIT_PERCENT, 2)
            payment_type   = "deposit"
            product_name   = f"25% Deposit — Booking #{booking_id}"
            stripe_desc    = items_desc
            log.info(f"Booking #{booking_id}: {days_until} days away — requiring 25% deposit ${charge_amount:.2f}")

        # Override with admin-specified amount if provided
        try:
            custom = float(request.form.get("custom_amount") or 0)
            if custom > 0:
                charge_amount = round(custom, 2)
                product_name  = f"Payment — Booking #{booking_id}"
                log.info(f"Booking #{booking_id}: admin overrode charge to ${charge_amount:.2f}")
        except Exception:
            pass

        # Create Stripe Payment Link
        payment_link, plink_id, stripe_error = create_stripe_payment_link(
            booking_id, charge_amount, b.get("email"), stripe_desc, product_name
        )
        if stripe_error:
            log.warning(f"Stripe error for #{booking_id}: {stripe_error}")
        if payment_link:
            save_payment_link(booking_id, product_name, charge_amount, payment_link, plink_id)

        # Update DB: status=accepted, payment_status=waiting, store payment link
        conn2 = get_db()
        if conn2:
            try:
                cur2 = conn2.cursor()
                cur2.execute(
                    "UPDATE bookings SET status='accepted', payment_status='waiting', stripe_payment_link=%s WHERE id=%s",
                    (payment_link, booking_id)
                )
                conn2.commit()
                cur2.close()
                conn2.close()
                log.info(f"Booking #{booking_id} accepted → Waiting Payment ({payment_type})")
            except Exception as e:
                log.error(f"Accept DB update error: {e}")

        # Send acceptance email with invoice + contract + payment link
        b["stripe_payment_link"] = payment_link
        try:
            send_accepted_email(b, charge_amount, payment_type)
        except Exception as e:
            log.error(f"send_accepted_email error for #{booking_id}: {e}")

        return redirect(url_for("admin_booking", booking_id=booking_id))

    except Exception as _top_err:
        _trace = _tb.format_exc()
        log.error(f"ACCEPT BOOKING #{booking_id} UNHANDLED ERROR: {_top_err}\n{_trace}")
        return f"<pre style='color:red;padding:2rem'>Accept booking error — please report this:\n\n{_trace}</pre>", 500


@app.route("/admin/booking/<int:booking_id>/agree-to-pay", methods=["POST"])
@admin_required
def agree_to_pay_booking(booking_id):
    """Mark booking as Agree to Pay (cash/check at delivery). Reserves inventory immediately."""
    pay_method = request.form.get("pay_method", "cash")
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE bookings SET status='agree_to_pay', payment_method=%s WHERE id=%s",
                (pay_method, booking_id)
            )
            conn.commit()
            cur.close()
            conn.close()
            log.info(f"Booking #{booking_id} set to agree_to_pay ({pay_method})")
            _log_payment(booking_id, 0, method=pay_method, note="Agreed to pay at delivery — inventory reserved", recorded_by="admin")
        except Exception as e:
            log.error(f"agree_to_pay error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/reset-payment-status", methods=["POST"])
@admin_required
def reset_payment_status(booking_id):
    """Reset payment_status to waiting and amount_paid to 0 for incorrectly marked bookings."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET payment_status='waiting', amount_paid=0 WHERE id=%s", (booking_id,))
            conn.commit(); cur.close(); conn.close()
            log.info(f"Booking #{booking_id} payment status reset to waiting")
        except Exception as e:
            log.error(f"reset_payment_status error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/revert-agree-to-pay", methods=["POST"])
@admin_required
def revert_agree_to_pay(booking_id):
    """Revert agree_to_pay back to accepted."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET status='accepted', payment_method=NULL WHERE id=%s", (booking_id,))
            conn.commit(); cur.close(); conn.close()
        except Exception as e:
            log.error(f"revert_agree_to_pay error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/deny", methods=["POST"])
@admin_required
def deny_booking(booking_id):
    """Deny booking: send polite rejection email."""
    conn = get_db()
    if not conn:
        return redirect(url_for("admin_dashboard"))
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        row = cur.fetchone()
        if row:
            b = _row(row)
            cur2 = conn.cursor()
            cur2.execute("UPDATE bookings SET status='denied' WHERE id=%s", (booking_id,))
            conn.commit()
            cur2.close()
            # Check if denial is due to inventory shortage — if so, send specific email
            try:
                inv_status = get_booking_inventory_check(booking_id)
                short_items = [s["item"] for s in inv_status if not s.get("ok")]
            except Exception:
                short_items = []
            if short_items:
                send_denied_inventory_email(b, short_items)
            else:
                send_denied_email(b)
            log.info(f"Booking #{booking_id} denied (short_items={short_items})")
        cur.close()
        conn.close()
    except Exception as e:
        log.error(f"Deny error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/confirm", methods=["POST"])
@admin_required
def confirm_booking(booking_id):
    """Manually confirm a booking — full or partial payment."""
    req_partial = request.form.get("partial") == "1"
    new_pmt_status_manual = "partial" if req_partial else "paid"
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE bookings SET status='accepted', payment_status=%s WHERE id=%s",
                (new_pmt_status_manual, booking_id,)
            )
            conn.commit()
            cur2 = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur2.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur2.fetchone()
            cur2.close(); cur.close(); conn.close()
            if row:
                send_receipt_email(_row(row))
            log.info(f"Booking #{booking_id} manually confirmed — {new_pmt_status_manual}")
        except Exception as e:
            log.error(f"Confirm error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/cash-payment", methods=["POST"])
@admin_required
def cash_payment(booking_id):
    """Mark a booking as paid in full with cash — no Stripe involved."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                b = _row(row)
                grand_total = float(b.get("grand_total") or 0)
                cur.execute(
                    "UPDATE bookings SET status='accepted', payment_status='paid', amount_paid=%s WHERE id=%s",
                    (grand_total, booking_id)
                )
                conn.commit()
                b["payment_status"] = new_pmt_status_manual
                b["amount_paid"] = grand_total
                send_receipt_email(b)
                log.info(f"Booking #{booking_id} marked as cash payment — Paid in Full")
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Cash payment error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/record-payment", methods=["POST"])
@admin_required
def record_payment(booking_id):
    """Manually record any payment (Stripe missed webhook, cash, check, etc.)."""
    try:
        amount = round(float(request.form.get("amount") or 0), 2)
    except Exception:
        amount = 0
    if amount <= 0:
        return redirect(url_for("admin_booking", booking_id=booking_id))
    method = request.form.get("method", "other")
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                b = _row(row)
                current_paid = round(float(b.get("amount_paid") or 0), 2)
                grand_total  = round(float(b.get("grand_total") or 0), 2)
                new_paid     = round(current_paid + amount, 2)
                balance      = round(grand_total - new_paid, 2)
                new_pmt_status_manual = "paid" if balance <= 0.50 else "partial"
                cur.execute(
                    "UPDATE bookings SET amount_paid=%s, payment_status=%s, status='accepted' WHERE id=%s",
                    (new_paid, new_pmt_status_manual, booking_id)
                )
                conn.commit()
                b["amount_paid"]    = new_paid
                b["payment_status"] = new_pmt_status_manual
                b["status"]         = "accepted"
                send_receipt_email(b)
                # Notify admin
                _send_email(
                    OWNER_BCC,
                    f"Payment Recorded — Booking #{booking_id}",
                    f"<html><body style='font-family:sans-serif;padding:1rem'>"
                    f"<p><strong>{b.get('full_name')}</strong> — ${amount:.2f} recorded via {method}.</p>"
                    f"<p>Total paid: <strong>${new_paid:.2f}</strong> / ${grand_total:.2f} | "
                    f"Balance: <strong>${max(balance,0):.2f}</strong> | Status: <strong>{new_pmt_status_manual.upper()}</strong></p>"
                    f"<p><a href='{BASE_URL}/admin/booking/{booking_id}'>View Booking</a></p>"
                    f"</body></html>",
                    f"Payment Recorded — Booking #{booking_id}: ${amount:.2f} via {method}",
                )
                log.info(f"Booking #{booking_id}: ${amount:.2f} recorded via {method}, payment_status={new_pmt_status_manual}")
                _log_payment(booking_id, amount, method=method, note=f"Manually recorded. Balance after: ${max(balance,0):.2f}", recorded_by="admin")
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Record payment error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/no-charge", methods=["POST"])
@admin_required
def no_charge(booking_id):
    """Zero out all fees and mark booking as confirmed — no payment collected."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE bookings
                SET payment_status='paid',
                    grand_total=0, amount_paid=0,
                    items_subtotal=0, delivery_fee=0,
                    late_night_fee=0, tax_amount=0,
                    discount_amount=0
                WHERE id=%s
            """, (booking_id,))
            conn.commit()
            cur.close(); conn.close()
            log.info(f"Booking #{booking_id} marked No Charge — all fees zeroed")
        except Exception as e:
            log.error(f"No charge error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/cancel", methods=["POST"])
@admin_required
def cancel_booking(booking_id):
    """Cancel a booking — releases inventory."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET status='cancelled' WHERE id=%s", (booking_id,))
            conn.commit()
            cur.close()
            conn.close()
            log.info(f"Booking #{booking_id} cancelled")
        except Exception as e:
            log.error(f"Cancel error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_booking(booking_id):
    conn = get_db()
    if not conn:
        return redirect(url_for("admin_dashboard"))
    if request.method == "GET":
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if not row:
                return "Booking not found", 404
            b = _row(row)
        except Exception as e:
            log.error(f"Edit booking fetch: {e}")
            return redirect(url_for("admin_dashboard"))
        return render_template_string(ADMIN_BOOKING_EDIT_HTML, business_name=BUSINESS_NAME, b=b, google_maps_key=GOOGLE_MAPS_KEY)
    # POST — save changes
    f = request.form
    fields = {
        "full_name":        f.get("full_name","").strip(),
        "company_name":     f.get("company_name","").strip(),
        "email":            f.get("email","").strip(),
        "phone":            f.get("phone","").strip(),
        "renter_street":    f.get("renter_street","").strip(),
        "renter_city":      f.get("renter_city","").strip(),
        "renter_state":     f.get("renter_state","").strip(),
        "renter_zip":       f.get("renter_zip","").strip(),
        "event_start_date": f.get("event_start_date","").strip() or None,
        "event_end_date":   f.get("event_end_date","").strip() or None,
        "event_start_time": f.get("event_start_time","").strip(),
        "event_end_time":   f.get("event_end_time","").strip(),
        "setup_time":       f.get("setup_time","").strip(),
        "setup_date":       f.get("setup_date","").strip() or None,
        "venue_type":       f.get("venue_type","venue"),
        "event_street":     f.get("event_street","").strip(),
        "event_city":       f.get("event_city","").strip(),
        "event_state":      f.get("event_state","").strip(),
        "event_zip":        f.get("event_zip","").strip(),
        "delivery_location":f.get("delivery_location","").strip(),
        "status":           f.get("status","pending"),
        "grand_total":      float(f.get("grand_total") or 0),
        "amount_paid":      float(f.get("amount_paid") or 0),
        "delivery_fee":     float(f.get("delivery_fee") or 0),
        "late_night_fee":   float(f.get("late_night_fee") or 0),
        "distance_miles":   float(f.get("distance_miles") or 0) or None,
        "notes":            f.get("notes","").strip(),
    }
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE bookings SET
              full_name=%(full_name)s, company_name=%(company_name)s,
              email=%(email)s, phone=%(phone)s,
              renter_street=%(renter_street)s, renter_city=%(renter_city)s,
              renter_state=%(renter_state)s, renter_zip=%(renter_zip)s,
              event_start_date=%(event_start_date)s, event_end_date=%(event_end_date)s,
              event_start_time=%(event_start_time)s, event_end_time=%(event_end_time)s,
              setup_time=%(setup_time)s, setup_date=%(setup_date)s, venue_type=%(venue_type)s,
              event_street=%(event_street)s, event_city=%(event_city)s,
              event_state=%(event_state)s, event_zip=%(event_zip)s,
              delivery_location=%(delivery_location)s,
              status=%(status)s, grand_total=%(grand_total)s,
              amount_paid=%(amount_paid)s, delivery_fee=%(delivery_fee)s,
              late_night_fee=%(late_night_fee)s, distance_miles=%(distance_miles)s,
              notes=%(notes)s
            WHERE id=%(id)s
        """, {**fields, "id": booking_id})
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"Edit booking save: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/customer-search")
@admin_required
def customer_search():
    """Search customers + bookings tables, merge in Python, return up to 10 matches."""
    q = request.args.get("q", "").strip()
    if len(q) < 1:
        return jsonify([])
    pat = f"%{q}%"
    try:
        conn = get_db()
        if not conn:
            return jsonify([])
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # 1. Pull from customers table
        cur.execute("""
            SELECT full_name, email, phone, company_name,
                   street AS renter_street, city AS renter_city,
                   state AS renter_state, zip AS renter_zip
            FROM customers
            WHERE full_name ILIKE %s
              AND full_name IS NOT NULL AND TRIM(full_name) != ''
            ORDER BY full_name LIMIT 20
        """, (pat,))
        cust_rows = {r["full_name"].strip().lower(): dict(r) for r in cur.fetchall()}

        # 2. Pull from bookings table (most recent per name)
        cur.execute("""
            SELECT DISTINCT ON (LOWER(TRIM(full_name)))
                full_name, email, phone, company_name,
                COALESCE(NULLIF(renter_street,''), event_street) AS renter_street,
                COALESCE(NULLIF(renter_city,''),   event_city)   AS renter_city,
                COALESCE(NULLIF(renter_state,''),  event_state)  AS renter_state,
                COALESCE(NULLIF(renter_zip,''),    event_zip)    AS renter_zip
            FROM bookings
            WHERE full_name ILIKE %s
              AND full_name IS NOT NULL AND TRIM(full_name) != ''
            ORDER BY LOWER(TRIM(full_name)), id DESC
            LIMIT 20
        """, (pat,))
        book_rows = {r["full_name"].strip().lower(): dict(r) for r in cur.fetchall()}
        cur.close(); conn.close()

        # 3. Merge: customers table wins; bookings fills gaps
        merged = {}
        for key, c in cust_rows.items():
            merged[key] = c
        for key, b in book_rows.items():
            if key not in merged:
                merged[key] = b
            else:
                # fill blank address fields from booking record
                for f in ("renter_street","renter_city","renter_state","renter_zip"):
                    if not merged[key].get(f):
                        merged[key][f] = b.get(f) or ""

        results = sorted(merged.values(), key=lambda x: x.get("full_name",""))[:10]
        return jsonify(results)
    except Exception as e:
        log.error(f"customer_search error: {e}")
        return jsonify([])


@app.route("/admin/booking/rebook/<int:booking_id>")
@admin_required
def admin_rebook(booking_id):
    """Pre-fill the new booking form with an existing customer's data."""
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
    src = cur.fetchone()
    cur.close()
    if not src:
        return redirect(url_for("admin_dashboard"))
    # Pass relevant fields as query params so the form auto-fills
    import urllib.parse as _up
    fields = {
        "prefill": "1",
        "customer_name": src.get("full_name") or "",
        "email": src.get("email") or "",
        "phone": src.get("phone") or "",
        "setup_address": src.get("delivery_location") or "",
        "notes": f"[Rebook from #{booking_id}] ",
    }
    return redirect(url_for("new_booking") + "?" + _up.urlencode(fields))


@app.route("/admin/booking/new", methods=["GET", "POST"])
@admin_required
def new_booking():
    if request.method == "GET":
        # Support rebook prefill via query params
        prefill_form = {}
        if request.args.get("prefill"):
            prefill_form = {k: request.args.get(k, "") for k in
                ("customer_name","email","phone","setup_address","notes")}
        return render_template_string(ADMIN_NEW_BOOKING_HTML,
            business_name=BUSINESS_NAME, products=get_products(),
            google_maps_key=GOOGLE_MAPS_KEY,
            exact_time_fee=EXACT_TIME_FEE,
            error=None, form=prefill_form)
    # POST — delegate to _submit_inner() which handles all pricing + DB insert
    # admin_create=1 is set in the form so _submit_inner redirects to admin booking page
    try:
        return _submit_inner()
    except Exception as e:
        log.error(f"Admin new booking error: {e}")
        return f"Error creating booking: {e}", 500


@app.route("/admin/booking/<int:booking_id>/update-items", methods=["POST"])
@admin_required
def update_booking_items(booking_id):
    names  = request.form.getlist("item_name")
    qtys   = request.form.getlist("item_qty")
    prices = request.form.getlist("item_price")
    while len(prices) < len(names):
        prices.append("0")

    # Build product price map as fallback
    _prod_map = {p["name"].lower(): float(p.get("price") or 0) for p in get_products()}
    items = []
    for name, qty, price in zip(names, qtys, prices):
        name = name.strip()
        if name:
            try:
                q = max(1, int(qty or 1))
            except Exception:
                q = 1
            try:
                up = float(price or 0)
            except Exception:
                up = 0
            if up == 0:
                up = _prod_map.get(name.lower(), 0)
            items.append({"name": name, "qty": q, "unit_price": up, "total": round(up * q, 2)})

    # Exact time delivery & delivery fee
    exact_time = bool(request.form.get("exact_time_delivery"))
    try:
        delivery_fee = round(float(request.form.get("delivery_fee") or 0), 2)
    except Exception:
        delivery_fee = 0.0

    # Recalculate all totals
    items_subtotal = round(sum(i["total"] for i in items), 2)
    exact_fee      = 175.0 if exact_time else 0.0
    pre_tax        = round(items_subtotal + delivery_fee + exact_fee, 2)

    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            # Fetch current discount info so we can reapply it
            cur.execute("SELECT discount_type, discount_value, tax_rate FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            disc_type  = row["discount_type"]  if row else None
            disc_value = float(row["discount_value"] or 0) if row else 0.0
            tax_rate   = float(row["tax_rate"]  or 0.0635) if row else 0.0635

            # Reapply discount to full pre-tax total (items + fees)
            disc_amount = 0.0
            if disc_type == "percent" and disc_value > 0:
                disc_amount = round(pre_tax * disc_value / 100.0, 2)
            elif disc_type == "amount" and disc_value > 0:
                disc_amount = round(min(disc_value, pre_tax), 2)

            taxable    = round(pre_tax - disc_amount, 2)
            tax_amount = round(taxable * tax_rate, 2)
            grand_total = round(taxable + tax_amount, 2)

            cur2 = conn.cursor()
            cur2.execute("""
                UPDATE bookings SET
                    items_json=%s,
                    items_subtotal=%s,
                    exact_time_delivery=%s,
                    delivery_fee=%s,
                    discount_amount=%s,
                    tax_amount=%s,
                    grand_total=%s
                WHERE id=%s
            """, (json.dumps(items), items_subtotal, exact_time, delivery_fee,
                  disc_amount, tax_amount, grand_total, booking_id))
            conn.commit()
            cur.close(); cur2.close(); conn.close()
            log.info(f"Booking #{booking_id} items updated — subtotal ${items_subtotal}, total ${grand_total}")
        except Exception as e:
            log.error(f"Update items error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/sync-customer-profile", methods=["POST"])
@admin_required
def sync_customer_profile(booking_id):
    """Sync between booking info and matching customer profile."""
    action = request.form.get("action", "")
    conn = get_db()
    if not conn:
        return redirect(url_for("admin_booking", booking_id=booking_id))
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        brow = cur.fetchone()
        if not brow:
            cur.close(); conn.close()
            return redirect(url_for("admin_booking", booking_id=booking_id))
        b = _row(brow)
        cur.execute("SELECT * FROM customers WHERE LOWER(TRIM(full_name))=LOWER(TRIM(%s)) LIMIT 1",
                    (b.get("full_name") or "",))
        crow = cur.fetchone()
        if crow:
            mc = _row(crow)
            if action == "update_profile":
                # Push booking phone/email/address → customer profile
                cur2 = conn.cursor()
                cur2.execute("""
                    UPDATE customers SET
                      email   = COALESCE(NULLIF(%s,''), email),
                      phone   = COALESCE(NULLIF(%s,''), phone),
                      street  = COALESCE(NULLIF(%s,''), street),
                      city    = COALESCE(NULLIF(%s,''), city),
                      state   = COALESCE(NULLIF(%s,''), state),
                      zip     = COALESCE(NULLIF(%s,''), zip)
                    WHERE id=%s
                """, (b.get("email") or "", b.get("phone") or "",
                      b.get("renter_street") or "", b.get("renter_city") or "",
                      b.get("renter_state") or "", b.get("renter_zip") or "",
                      mc["id"]))
                conn.commit(); cur2.close()
                log.info(f"Customer profile {mc['id']} updated from booking #{booking_id}")
            elif action == "update_booking":
                # Pull customer profile phone/email/address → booking (force overwrite non-null values)
                cur2 = conn.cursor()
                cur2.execute("""
                    UPDATE bookings SET
                      email         = CASE WHEN %s != '' THEN %s ELSE email END,
                      phone         = CASE WHEN %s != '' THEN %s ELSE phone END,
                      renter_street = CASE WHEN %s != '' THEN %s ELSE renter_street END,
                      renter_city   = CASE WHEN %s != '' THEN %s ELSE renter_city END,
                      renter_state  = CASE WHEN %s != '' THEN %s ELSE renter_state END,
                      renter_zip    = CASE WHEN %s != '' THEN %s ELSE renter_zip END
                    WHERE id=%s
                """, (
                    mc.get("email") or "", mc.get("email") or "",
                    mc.get("phone") or "", mc.get("phone") or "",
                    mc.get("street") or "", mc.get("street") or "",
                    mc.get("city") or "", mc.get("city") or "",
                    mc.get("state") or "", mc.get("state") or "",
                    mc.get("zip") or "", mc.get("zip") or "",
                    booking_id))
                conn.commit(); cur2.close()
                log.info(f"Booking #{booking_id} updated from customer profile {mc['id']}")
            elif action == "link":
                # Link booking to existing customer: copy profile info into booking
                cur2 = conn.cursor()
                cur2.execute("""
                    UPDATE bookings SET
                      email         = COALESCE(NULLIF(%s,''), email),
                      phone         = COALESCE(NULLIF(%s,''), phone),
                      renter_street = COALESCE(NULLIF(%s,''), renter_street),
                      renter_city   = COALESCE(NULLIF(%s,''), renter_city),
                      renter_state  = COALESCE(NULLIF(%s,''), renter_state),
                      renter_zip    = COALESCE(NULLIF(%s,''), renter_zip)
                    WHERE id=%s
                """, (
                    mc.get("email") or "", mc.get("phone") or "",
                    mc.get("street") or "", mc.get("city") or "",
                    mc.get("state") or "", mc.get("zip") or "",
                    booking_id))
                conn.commit(); cur2.close()
                log.info(f"Booking #{booking_id} linked to customer profile {mc['id']}")
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"Sync customer profile error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/send-receipt", methods=["POST"])
@admin_required
def admin_send_receipt(booking_id):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                send_receipt_email(_row(row))
                log.info(f"Manual receipt sent for #{booking_id}")
        except Exception as e:
            log.error(f"Send receipt error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/update-address", methods=["POST"])
@admin_required
def update_booking_address(booking_id):
    f = request.form
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE bookings SET
                  renter_street=%s, renter_city=%s, renter_state=%s, renter_zip=%s
                WHERE id=%s
            """, (
                f.get("renter_street","").strip() or None,
                f.get("renter_city","").strip() or None,
                f.get("renter_state","").strip() or None,
                f.get("renter_zip","").strip() or None,
                booking_id
            ))
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Update address error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/admin-notes", methods=["POST"])
@admin_required
def update_admin_notes(booking_id):
    notes = (request.form.get("admin_notes") or "").strip()
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET admin_notes=%s WHERE id=%s", (notes or None, booking_id))
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Admin notes update error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/apply-discount", methods=["POST"])
@admin_required
def apply_discount(booking_id):
    disc_type  = request.form.get("discount_type", "amount").strip()
    disc_value = float(request.form.get("discount_value") or 0)
    if disc_type not in ("amount", "percent") or disc_value <= 0:
        return redirect(url_for("admin_booking", booking_id=booking_id))

    conn = get_db()
    if not conn:
        return redirect(url_for("admin_booking", booking_id=booking_id))
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return redirect(url_for("admin_booking", booking_id=booking_id))
        b = _row(row)

        items_subtotal = float(b.get("items_subtotal") or 0)
        delivery_fee   = float(b.get("delivery_fee")   or 0)
        exact_fee      = 175.0 if b.get("exact_time_delivery") else 0.0
        tax_rate       = float(b.get("tax_rate")       or 0.0635)
        tax_exempt     = bool(b.get("tax_exempt"))

        # Discount applies to full pre-tax total (items + fees)
        pre_tax_full = round(items_subtotal + delivery_fee + exact_fee, 2)
        if disc_type == "percent":
            disc_amount = round(pre_tax_full * (disc_value / 100.0), 2)
        else:
            disc_amount = round(min(disc_value, pre_tax_full), 2)

        taxable     = round(pre_tax_full - disc_amount, 2)
        tax_amount  = 0.0 if tax_exempt else round(taxable * tax_rate, 2)
        grand_total = round(taxable + tax_amount, 2)

        cur.execute("""
            UPDATE bookings
            SET discount_type=%s, discount_value=%s, discount_amount=%s,
                tax_amount=%s, grand_total=%s
            WHERE id=%s
        """, (disc_type, disc_value, disc_amount, tax_amount, grand_total, booking_id))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"Apply discount error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/remove-discount")
@admin_required
def remove_discount(booking_id):
    conn = get_db()
    if not conn:
        return redirect(url_for("admin_booking", booking_id=booking_id))
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return redirect(url_for("admin_booking", booking_id=booking_id))
        b = _row(row)

        # Restore grand total without discount
        items_subtotal = float(b.get("items_subtotal") or 0)
        delivery_fee   = float(b.get("delivery_fee")   or 0)
        exact_fee      = float(b.get("exact_time_fee") or 0)
        late_night_fee = float(b.get("late_night_fee") or 0)
        tax_rate       = float(b.get("tax_rate")       or 0.0635)
        tax_exempt     = bool(b.get("tax_exempt"))

        pre_tax    = round(items_subtotal + delivery_fee + exact_fee + late_night_fee, 2)
        tax_amount = 0.0 if tax_exempt else round(pre_tax * tax_rate, 2)
        grand_total = round(pre_tax + tax_amount, 2)

        cur.execute("""
            UPDATE bookings
            SET discount_type=NULL, discount_value=0, discount_amount=0,
                tax_amount=%s, grand_total=%s
            WHERE id=%s
        """, (tax_amount, grand_total, booking_id))
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"Remove discount error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/recalc-delivery", methods=["POST"])
@admin_required
def recalc_delivery(booking_id):
    """Recalculate delivery fee from stored event address and update grand total."""
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        b = dict(cur.fetchone())
        # Build address from stored fields
        addr_parts = [b.get("event_street",""), b.get("event_city",""),
                      b.get("event_state",""), b.get("event_zip","")]
        address = ", ".join(p for p in addr_parts if p)
        if not address:
            return redirect(url_for("admin_booking", booking_id=booking_id))
        miles = get_distance_miles(address)
        fee, _ = calc_delivery_fee(miles)
        # Recalculate grand total
        subtotal   = float(b.get("items_subtotal") or 0)
        exact_fee  = 175.0 if b.get("exact_time_delivery") else 0.0
        tax_rate   = float(b.get("tax_rate") or 0.0635)
        tax_exempt = bool(b.get("tax_exempt"))
        pre_tax    = subtotal + exact_fee + fee
        tax        = 0.0 if tax_exempt else round(pre_tax * tax_rate, 2)
        grand      = round(pre_tax + tax, 2)
        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE bookings
            SET delivery_fee=%s, distance_miles=%s, tax_amount=%s, grand_total=%s
            WHERE id=%s
        """, (fee, miles, tax, grand, booking_id))
        conn.commit()
        cur.close(); cur2.close(); conn.close()
    except Exception as e:
        log.error(f"recalc_delivery: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/recalc-total", methods=["POST"])
@admin_required
def recalc_total(booking_id):
    """Recalculate grand total from current items_json + delivery fee + tax."""
    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        b = dict(cur.fetchone())

        items        = json.loads(b.get("items_json") or "[]")
        items_sub    = round(sum(float(i.get("total") or 0) for i in items), 2)
        delivery_fee = float(b.get("delivery_fee") or 0)
        exact_fee    = 175.0 if b.get("exact_time_delivery") else 0.0
        disc_type    = b.get("discount_type")
        disc_value   = float(b.get("discount_value") or 0)
        tax_rate     = float(b.get("tax_rate") or 0.0635)
        tax_exempt   = bool(b.get("tax_exempt"))

        pre_tax = round(items_sub + delivery_fee + exact_fee, 2)

        # Reapply existing discount
        disc_amount = 0.0
        if disc_type == "percent" and disc_value > 0:
            disc_amount = round(pre_tax * disc_value / 100.0, 2)
        elif disc_type == "amount" and disc_value > 0:
            disc_amount = round(min(disc_value, pre_tax), 2)

        taxable    = round(pre_tax - disc_amount, 2)
        tax_amount = 0.0 if tax_exempt else round(taxable * tax_rate, 2)
        grand      = round(taxable + tax_amount, 2)

        cur2 = conn.cursor()
        cur2.execute("""
            UPDATE bookings
            SET items_subtotal=%s, discount_amount=%s, tax_amount=%s, grand_total=%s
            WHERE id=%s
        """, (items_sub, disc_amount, tax_amount, grand, booking_id))
        conn.commit()
        cur.close(); cur2.close(); conn.close()
        log.info(f"Booking #{booking_id} grand total recalculated → ${grand}")
    except Exception as e:
        log.error(f"recalc_total error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/update-event-dates", methods=["POST"])
@admin_required
def update_event_dates(booking_id):
    event_start_date = request.form.get("event_start_date", "").strip()
    event_start_time = request.form.get("event_start_time", "").strip()
    event_end_date   = request.form.get("event_end_date", "").strip()
    event_end_time   = request.form.get("event_end_time", "").strip()
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            "UPDATE bookings SET event_start_date=%s, event_start_time=%s, event_end_date=%s, event_end_time=%s WHERE id=%s",
            (event_start_date or None, event_start_time or None, event_end_date or None, event_end_time or None, booking_id)
        )
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"update_event_dates: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/update-times", methods=["POST"])
@admin_required
def update_booking_times(booking_id):
    setup_date      = request.form.get("setup_date", "").strip()
    setup_time      = request.form.get("setup_time", "").strip()
    event_end_date  = request.form.get("event_end_date", "").strip()
    event_end_time  = request.form.get("event_end_time", "").strip()
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            "UPDATE bookings SET setup_date=%s, setup_time=%s, event_end_date=%s, event_end_time=%s WHERE id=%s",
            (setup_date or None, setup_time or None, event_end_date or None, event_end_time or None, booking_id)
        )
        conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        log.error(f"update_booking_times: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/apply-weekend-schedule", methods=["POST"])
@admin_required
def apply_weekend_schedule(booking_id):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT event_start_date, event_start_time, event_end_date, event_end_time, venue_type, notes FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                venue_type = (row["venue_type"] or "").lower()
                event_date = row["event_start_date"]
                if not hasattr(event_date, "strftime"):
                    event_date = datetime.strptime(str(event_date)[:10], "%Y-%m-%d").date()
                weekday = event_date.weekday()
                if venue_type == "residential" and weekday in (5, 6):
                    friday = event_date - timedelta(days=1 if weekday == 5 else 2)
                    pickup_date = event_date + timedelta(days=1)  # Sunday or Monday
                    # Only update Est. Delivery (setup_date/setup_time) and Est. Pickup (event_end_date/event_end_time)
                    # Event Start, Event End, and Setup fields are left untouched
                    cur.execute("""
                        UPDATE bookings SET
                          delivery_date=%s, delivery_time=%s,
                          event_end_date=%s, event_end_time=%s
                        WHERE id=%s
                    """, (friday, "16:00", pickup_date, "10:00", booking_id))
                    conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Apply weekend schedule error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/update-phone", methods=["POST"])
@admin_required
def update_booking_phone(booking_id):
    phone = (request.form.get("phone") or "").strip()
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET phone=%s WHERE id=%s", (phone or None, booking_id))
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Update phone error: {e}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


@app.route("/admin/booking/<int:booking_id>/custom-stripe-link", methods=["POST"])
@admin_required
def custom_stripe_link(booking_id):
    """Create a Stripe payment link for a custom amount and email it to the customer."""
    try:
        amount = float(request.form.get("amount") or 0)
    except Exception:
        amount = 0
    if amount <= 0:
        return redirect(url_for("admin_booking", booking_id=booking_id))

    label = (request.form.get("label") or "").strip() or f"Payment — Booking #{booking_id}"

    # Fetch booking for customer email / name
    b = None
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                b = _row(row)
        except Exception as e:
            log.error(f"Custom link fetch error: {e}")

    if not b:
        return redirect(url_for("admin_booking", booking_id=booking_id))

    items_desc = ", ".join(f"{i['qty']}x {i['name']}" for i in json.loads(b.get("items_json") or "[]"))
    payment_link, plink_id, err = create_stripe_payment_link(booking_id, amount, b.get("email"), items_desc, label)

    if err:
        log.warning(f"Custom Stripe link error for #{booking_id}: {err}")
    if payment_link:
        save_payment_link(booking_id, label, amount, payment_link, plink_id)

    # Email the link to the customer
    if payment_link and b.get("email"):
        first = (b.get("full_name") or "").split()[0] or "there"
        subject = f"Payment Link — {label} | {BUSINESS_NAME}"
        html = f"""
<html><body style="font-family:-apple-system,sans-serif;background:#f0f4f8;padding:2rem 1rem">
<div style="max-width:560px;margin:0 auto;background:white;border-radius:12px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,.1)">
  <div style="background:linear-gradient(135deg,#1e40af,#2563eb);padding:1.5rem 2rem;color:white;text-align:center">
    <h2 style="margin:0;font-size:1.2rem">{BUSINESS_NAME}</h2>
    <p style="margin:.4rem 0 0;opacity:.88;font-size:.95rem">Payment Request</p>
  </div>
  <div style="padding:1.75rem 2rem">
    <p style="color:#2d3748;font-size:1.05rem;margin-bottom:.75rem">Hi <strong>{first}</strong>,</p>
    <p style="color:#4a5568;line-height:1.7;margin-bottom:1.5rem">
      A payment link has been created for your booking. Please use the button below to complete your payment.
    </p>
    <div style="background:#eff6ff;border:2px solid #bfdbfe;border-radius:12px;padding:1.5rem;text-align:center;margin-bottom:1.5rem">
      <p style="font-size:.85rem;color:#1d4ed8;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin:0 0 .3rem">{label}</p>
      <p style="font-size:2.75rem;font-weight:800;color:#1e40af;margin:.2rem 0 1rem;line-height:1">${amount:.2f}</p>
      <a href="{payment_link}"
         style="display:inline-block;background:linear-gradient(135deg,#1e40af,#2563eb);color:white;padding:1rem 2.5rem;border-radius:10px;font-weight:700;font-size:1.1rem;text-decoration:none">
        Pay ${amount:.2f} Now
      </a>
      <p style="margin:.6rem 0 0;font-size:.8rem;color:#6b7280">Secure payment powered by Stripe</p>
    </div>
    <p style="font-size:.85rem;color:#718096">Booking #{booking_id} &bull; {BUSINESS_NAME}</p>
  </div>
</div>
</body></html>"""
        plain = f"Hi {first},\n\nPayment of ${amount:.2f} is due for your booking.\n\nPay here: {payment_link}\n\n— {BUSINESS_NAME}"
        _send_email(b.get("email"), subject, html, plain)
        log.info(f"Custom payment link sent for #{booking_id}: ${amount:.2f}")

    link_param = urllib.parse.quote(payment_link or "", safe="")
    return redirect(url_for("admin_booking", booking_id=booking_id) + f"?custom_link={link_param}")


@app.route("/admin/payment-link/<int:link_id>/cancel", methods=["POST"])
@admin_required
def cancel_payment_link(link_id):
    """Deactivate a Stripe Payment Link and mark it cancelled in DB."""
    # Fetch the link record so we know booking_id and stripe_link_id
    conn = get_db()
    booking_id = None
    stripe_link_id = None
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT booking_id, stripe_link_id FROM payment_links WHERE id=%s", (link_id,))
            row = cur.fetchone()
            cur.close(); conn.close()
            if row:
                booking_id     = row["booking_id"]
                stripe_link_id = row["stripe_link_id"]
        except Exception as e:
            log.error(f"Cancel link fetch error: {e}")

    # Deactivate in Stripe
    if stripe_link_id and STRIPE_SECRET_KEY:
        try:
            stripe.PaymentLink.modify(stripe_link_id, active=False)
            log.info(f"Stripe link {stripe_link_id} deactivated")
        except Exception as e:
            log.warning(f"Stripe deactivate error for {stripe_link_id}: {e}")

    # Mark cancelled in DB
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE payment_links SET status='cancelled' WHERE id=%s", (link_id,))
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Cancel link DB error: {e}")

    if booking_id:
        return redirect(url_for("admin_booking", booking_id=booking_id))
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/delete", methods=["POST"])
@admin_required
def delete_booking(booking_id):
    """Permanently delete a booking. Customer record in customers table is untouched."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM bookings WHERE id=%s", (booking_id,))
            conn.commit()
            cur.close()
            conn.close()
            log.info(f"Booking #{booking_id} permanently deleted")
        except Exception as e:
            log.error(f"Delete booking error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/bookings/bulk-archive", methods=["POST"])
@admin_required
def bulk_archive_bookings():
    ids_raw = request.form.get("ids", "")
    ids = [int(i) for i in ids_raw.split(",") if i.strip().isdigit()]
    if ids:
        conn = get_db()
        if conn:
            try:
                cur = conn.cursor()
                cur.execute("UPDATE bookings SET archived=TRUE WHERE id = ANY(%s)", (ids,))
                conn.commit(); cur.close(); conn.close()
            except Exception as e:
                log.error(f"Bulk archive error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/bookings/bulk-delete", methods=["POST"])
@admin_required
def bulk_delete_bookings():
    ids_raw = request.form.get("ids", "")
    ids = [int(i) for i in ids_raw.split(",") if i.strip().isdigit()]
    if ids:
        conn = get_db()
        if conn:
            try:
                cur = conn.cursor()
                cur.execute("DELETE FROM bookings WHERE id = ANY(%s)", (ids,))
                conn.commit(); cur.close(); conn.close()
            except Exception as e:
                log.error(f"Bulk delete error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/archive", methods=["POST"])
@admin_required
def archive_booking(booking_id):
    """Hide a booking from the main list (soft delete). Can be viewed in Archived tab."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET archived=TRUE WHERE id=%s", (booking_id,))
            conn.commit()
            cur.close()
            conn.close()
            log.info(f"Booking #{booking_id} archived")
        except Exception as e:
            log.error(f"Archive booking error: {e}")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/booking/<int:booking_id>/unarchive", methods=["POST"])
@admin_required
def unarchive_booking(booking_id):
    """Restore an archived booking back to the main list."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("UPDATE bookings SET archived=FALSE WHERE id=%s", (booking_id,))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"Unarchive booking error: {e}")
    return redirect(url_for("admin_dashboard", archived=1))


@app.route("/admin/booking/<int:booking_id>/delivery-status", methods=["POST"])
@admin_required
def booking_delivery_status(booking_id):
    """Advance delivery_status: None → delivered → picked_up."""
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT delivery_status FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            current = row[0] if row else None
            now = datetime.now(timezone.utc)
            if current is None:
                new_status = "delivered"
                cur.execute("UPDATE bookings SET delivery_status=%s, delivered_at=%s WHERE id=%s", (new_status, now, booking_id))
            elif current == "delivered":
                new_status = "picked_up"
                cur.execute("UPDATE bookings SET delivery_status=%s, picked_up_at=%s WHERE id=%s", (new_status, now, booking_id))
            else:
                new_status = current  # already picked_up, no change
                cur.execute("UPDATE bookings SET delivery_status=%s WHERE id=%s", (new_status, booking_id))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"Delivery status error: {e}")
    # Return to wherever admin came from
    ref = request.referrer or url_for("admin_dashboard")
    return redirect(ref)


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — STRIPE WEBHOOK
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/payment/success/<int:booking_id>")
def payment_success(booking_id):
    """Thank-you page shown after Stripe payment completes."""
    conn = get_db()
    b = {}
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                b = _row(row)
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"payment_success error: {e}")
    name = b.get("full_name", "")
    grand = b.get("grand_total", 0)
    paid  = b.get("amount_paid", 0)
    bal   = round(float(grand or 0) - float(paid or 0), 2)
    html = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Payment Received — {BUSINESS_NAME}</title>
<style>
  body{{font-family:sans-serif;background:#f4f6f8;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}}
  .card{{background:#fff;border-radius:12px;padding:2.5rem 2rem;max-width:480px;width:100%;text-align:center;box-shadow:0 4px 24px rgba(0,0,0,.08)}}
  .check{{font-size:3.5rem;margin-bottom:.5rem}}
  h1{{color:#1a7a4a;margin:.25rem 0 1rem}}
  p{{color:#444;line-height:1.6;margin:.4rem 0}}
  .bal{{margin-top:1.2rem;background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:.75rem 1rem;color:#166534;font-weight:600}}
  .btn{{display:inline-block;margin-top:1.5rem;padding:.65rem 1.5rem;background:#2563eb;color:#fff;text-decoration:none;border-radius:8px;font-size:.95rem}}
</style></head><body>
<div class="card">
  <div class="check">✅</div>
  <h1>Payment Received!</h1>
  <p>Thank you{', ' + name.split()[0] if name else ''}! Your payment has been successfully processed.</p>
  {'<p>Booking #' + str(booking_id) + '</p>' if booking_id else ''}
  {'<div class="bal">Remaining balance: $' + f'{bal:.2f}' + '</div>' if bal > 0.50 else '<div class="bal">Paid in full — thank you!</div>'}
  <p style="margin-top:1.2rem;font-size:.9rem;color:#666">You will receive a confirmation email shortly.</p>
  <a href="/" class="btn">Back to Home</a>
</div>
</body></html>"""
    return html


@app.route("/stripe/webhook", methods=["POST"])
def stripe_webhook():
    """
    Stripe sends this webhook when a payment is completed.
    Auto-confirms the booking so inventory gets locked.
    """
    payload    = request.get_data(as_text=False)
    sig_header = request.headers.get("Stripe-Signature", "")

    if STRIPE_WEBHOOK_SECRET:
        try:
            event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
        except ValueError:
            log.error("Invalid Stripe webhook payload")
            return jsonify({"error": "Invalid payload"}), 400
        except stripe.error.SignatureVerificationError:
            log.error("Invalid Stripe webhook signature")
            return jsonify({"error": "Invalid signature"}), 400
    else:
        try:
            event = json.loads(payload)
        except Exception:
            return jsonify({"error": "Invalid JSON"}), 400

    # Support both StripeObject (attribute access) and plain dict (from json.loads)
    def _ev(obj, key, default=None):
        try:
            v = obj[key]
            return v if v is not None else default
        except Exception:
            return default

    event_type = _ev(event, "type", "")
    if event_type == "checkout.session.completed":
        sess = _ev(event, "data", {})
        if hasattr(sess, "__getitem__"):
            sess = _ev(sess, "object", sess)
        # payment_status
        try:
            pmt_status = sess.payment_status
        except Exception:
            pmt_status = _ev(sess, "payment_status", "")
        if pmt_status == "paid":
            # metadata
            try:
                meta = sess.metadata
                booking_id = meta["booking_id"] if meta and "booking_id" in meta else None
            except Exception:
                meta = _ev(sess, "metadata", {}) or {}
                booking_id = meta.get("booking_id")
            # amount
            try:
                amount_paid_cents = sess.amount_total or 0
            except Exception:
                amount_paid_cents = _ev(sess, "amount_total", 0) or 0
            # session id
            try:
                sess_id = sess.id
            except Exception:
                sess_id = _ev(sess, "id", "")
            amount_paid_dollars = round(amount_paid_cents / 100, 2)
            if booking_id:
                conn = get_db()
                if conn:
                    try:
                        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                        # Deduplication: skip if this Stripe session was already recorded
                        if sess_id:
                            cur.execute("SELECT id FROM bookings WHERE stripe_session_id=%s", (sess_id,))
                            if cur.fetchone():
                                log.info(f"Webhook duplicate: session {sess_id} already processed. Skipping.")
                                cur.close(); conn.close()
                                return jsonify({"status": "already_processed"}), 200
                        # Fetch the booking
                        cur.execute("SELECT * FROM bookings WHERE id=%s", (int(booking_id),))
                        row = cur.fetchone()
                        if row:
                            b = _row(row)
                            current_status  = b.get("status", "")
                            current_pmt     = b.get("payment_status") or ""
                            current_paid    = float(b.get("amount_paid") or 0)
                            grand_total     = float(b.get("grand_total") or 0)
                            new_paid        = round(current_paid + amount_paid_dollars, 2)
                            balance         = round(grand_total - new_paid, 2)
                            new_pmt_status  = "paid" if balance <= 0.50 else "partial"

                            if current_status == "accepted":
                                # Payment received — update payment_status, keep status=accepted
                                cur.execute(
                                    "UPDATE bookings SET payment_status=%s, amount_paid=%s, stripe_session_id=%s WHERE id=%s",
                                    (new_pmt_status, new_paid, sess_id, int(booking_id))
                                )
                                conn.commit()
                                b["amount_paid"]     = new_paid
                                b["payment_status"]  = new_pmt_status
                                send_receipt_email(b)
                                log.info(f"Booking #{booking_id} payment received → {new_pmt_status} (${amount_paid_dollars:.2f})")
                                _log_payment(booking_id, amount_paid_dollars, method="stripe", note=f"Stripe checkout. Session: {sess_id}", recorded_by="stripe")

                            elif current_status in ("pending", "accepted"):
                                # Additional payment
                                cur.execute(
                                    "UPDATE bookings SET payment_status=%s, amount_paid=%s, stripe_session_id=%s WHERE id=%s",
                                    (new_pmt_status, new_paid, sess_id, int(booking_id))
                                )
                                conn.commit()
                                b["amount_paid"]    = new_paid
                                b["payment_status"] = new_pmt_status
                                send_receipt_email(b)
                                _notify_subject = f"{'PAID IN FULL' if balance <= 0.50 else 'Payment Received'} — Booking #{booking_id}"
                                _notify_body = (
                                    f"<p><strong>{b.get('full_name')}</strong> just paid "
                                    f"<strong>${amount_paid_dollars:.2f}</strong> on Booking #{booking_id}.</p>"
                                    f"<p>Amount paid to date: <strong>${new_paid:.2f}</strong> / ${grand_total:.2f}</p>"
                                    f"<p>Balance remaining: <strong>${max(balance,0):.2f}</strong></p>"
                                    f"<p>Payment status: <strong>{new_pmt_status.upper()}</strong></p>"
                                    f'<p><a href="{BASE_URL}/admin/booking/{booking_id}">View Booking</a></p>'
                                )
                                _send_email(
                                    OWNER_BCC,
                                    _notify_subject,
                                    f"<html><body style='font-family:sans-serif;padding:1rem'>{_notify_body}</body></html>",
                                    _notify_subject,
                                )
                                log.info(f"Booking #{booking_id} payment received ${amount_paid_dollars:.2f}, status={new_status}")

                        cur.close()
                        conn.close()
                    except Exception as e:
                        log.error(f"Webhook DB error: {e}")

    return jsonify({"status": "received"}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — FINAL PAYMENT REMINDER
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/booking/<int:booking_id>/send-final-reminder", methods=["POST"])
@admin_required
def send_final_reminder(booking_id):
    """Manually send final payment reminder from the admin panel."""
    conn = get_db()
    if not conn:
        return redirect(url_for("admin_booking", booking_id=booking_id))

    b = None
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row:
            b = _row(row)
    except Exception as e:
        log.error(f"Final reminder fetch error: {e}")
        return redirect(url_for("admin_booking", booking_id=booking_id))

    if not b:
        return "Booking not found", 404

    grand_total  = float(b.get("grand_total") or 0)
    amount_paid  = float(b.get("amount_paid") or 0)
    remaining    = round(grand_total - amount_paid, 2)
    if remaining <= 0:
        remaining = round(grand_total * 0.75, 2)  # fallback if amount_paid not set
    # Override with admin-specified amount if provided
    try:
        custom = float(request.form.get("custom_amount") or 0)
        if custom > 0:
            remaining = round(custom, 2)
            log.info(f"Booking #{booking_id}: admin overrode final payment to ${remaining:.2f}")
    except Exception:
        pass
    items_list      = ", ".join(f"{i['qty']}x {i['name']}" for i in json.loads(b.get("items_json") or "[]"))
    product_name    = f"Final Payment — Booking #{booking_id}"

    payment_link, plink_id, err = create_stripe_payment_link(
        booking_id, remaining, b.get("email"), items_list, product_name
    )
    if err:
        log.warning(f"Stripe error for final payment #{booking_id}: {err}")
    if payment_link:
        save_payment_link(booking_id, product_name, remaining, payment_link, plink_id)

    # Save final payment link to DB
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE bookings SET final_payment_link=%s, final_reminder_sent=TRUE WHERE id=%s",
                (payment_link, booking_id)
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"Final reminder DB error: {e}")

    b["final_payment_link"] = payment_link
    send_final_payment_email(b, remaining, payment_link)
    log.info(f"Final payment reminder sent for booking #{booking_id}")
    return redirect(url_for("admin_booking", booking_id=booking_id))


ADMIN_CALENDAR_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-title" content="Admin">
  <title>Calendar — {{ business_name }}</title>
<style>
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{display:flex;flex-direction:column;gap:.15rem;padding:.25rem .5rem;flex:1}
.sb-link{display:flex;align-items:center;gap:.5rem;padding:.5rem .65rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:500;color:#374151;transition:background .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#2563eb;font-weight:600}
.sb-divider{height:1px;background:#f3f4f6;margin:.3rem 0}
.sb-bottom{padding:.5rem .5rem 1rem}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.4rem;cursor:pointer;color:#374151;padding:.2rem .4rem}
@media(max-width:640px){
  .sidebar{transform:translateX(-100%)}
  .sidebar.open{transform:translateX(0)}
  .sb-overlay.show{display:block}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.main{padding:1.25rem}
.cal-nav{display:flex;align-items:center;gap:.6rem;margin-bottom:1rem;flex-wrap:wrap}
.cal-nav h2{font-size:1.1rem;font-weight:700;flex:1;margin:0;min-width:140px}
.cal-btn{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:.4rem .85rem;cursor:pointer;font-size:.95rem;font-weight:600;color:#374151;text-decoration:none;display:inline-block;line-height:1.4}
.cal-btn:hover{background:#f3f4f6}
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:3px}
.cal-hdr{text-align:center;font-size:.75rem;font-weight:700;color:#6b7280;padding:.4rem 0;text-transform:uppercase;letter-spacing:.5px}
.cal-cell{min-height:72px;background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:.4rem .5rem;cursor:default;transition:border-color .15s;box-sizing:border-box}
.cal-cell.has-events{cursor:pointer}
.cal-cell.has-events:hover{border-color:#2563eb;background:#f0f9ff}
.cal-cell.today{border-color:#2563eb;background:#eff6ff}
.cal-cell.other-month{background:#f9fafb;opacity:.5}
.cal-date{font-size:.8rem;font-weight:600;margin-bottom:.25rem}
.dot-confirmed{background:#16a34a}
.dot-paid{background:#0284c7}
.dot-pending{background:#d97706}
.dot-denied,.dot-cancelled{background:#6b7280}
.cal-dot{display:inline-block;width:7px;height:7px;border-radius:50%;margin:1px}
.cal-count{font-size:.7rem;color:#6b7280;margin-top:.15rem}
.popup-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:150}
.popup{display:none;position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);background:#fff;border-radius:16px;box-shadow:0 8px 32px rgba(0,0,0,.18);width:min(420px,94vw);max-height:80vh;overflow-y:auto;z-index:160;padding:1.5rem}
.popup-hdr{display:flex;align-items:center;gap:.75rem;margin-bottom:1rem}
.popup-hdr h3{flex:1;margin:0;font-size:1rem;font-weight:700}
.popup-close{background:none;border:none;font-size:1.3rem;cursor:pointer;color:#6b7280;padding:.2rem .5rem}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">&#127968; Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">&#128101; Clients</a>
    <a href="/admin/inventory" class="sb-link">&#128230; Inventory</a>
    <a href="/admin/calendar" class="sb-link active">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">&#128682; Sign Out</a>
  </div>
</aside>
<div class="page-content">
  <div class="pg-hdr">
    <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
    <h1>Calendar</h1>
  </div>
  <div class="main">
    <div class="cal-nav">
      <a href="{{ prev_url }}" class="cal-btn">&#8249;</a>
      <h2>{{ month_name }}</h2>
      <a href="{{ next_url }}" class="cal-btn">&#8250;</a>
      <a href="/admin/calendar" class="cal-btn" style="margin-left:auto;font-size:.82rem">Today</a>
      <details style="position:relative;display:inline-block">
        <summary style="list-style:none;background:#2563eb;color:#fff;border:1px solid #2563eb;border-radius:8px;padding:.4rem .85rem;cursor:pointer;font-size:.82rem;font-weight:600;display:inline-block">&#8595; Export</summary>
        <div style="position:absolute;right:0;top:110%;background:#fff;border:1px solid #e5e7eb;border-radius:10px;box-shadow:0 4px 16px rgba(0,0,0,.12);min-width:165px;z-index:200;overflow:hidden">
          <a href="/admin/calendar.ics" download="bookings.ics" style="display:block;padding:.65rem 1rem;font-size:.82rem;font-weight:600;text-decoration:none;color:#374151">&#128197; Download .ics</a>
          <a href="/admin/calendar.csv" download="bookings.csv" style="display:block;padding:.65rem 1rem;font-size:.82rem;font-weight:600;text-decoration:none;color:#374151;border-top:1px solid #f3f4f6">&#128196; Download CSV</a>
          <a href="{{ gcal_url }}" target="_blank" style="display:block;padding:.65rem 1rem;font-size:.82rem;font-weight:600;text-decoration:none;color:#374151;border-top:1px solid #f3f4f6">&#128279; Add to Google Calendar</a>
          <div style="padding:.65rem 1rem;border-top:1px solid #f3f4f6">
            <div style="font-size:.78rem;font-weight:600;color:#6b7280;margin-bottom:.3rem">Other apps (Outlook, Apple):</div>
            <input id="ics-url-copy" value="{{ ics_url }}" readonly style="width:100%;font-size:.72rem;border:1px solid #e5e7eb;border-radius:6px;padding:.3rem .5rem;color:#374151;box-sizing:border-box">
            <button onclick="var i=document.getElementById('ics-url-copy');i.select();document.execCommand('copy');this.textContent='Copied!';setTimeout(()=>this.textContent='Copy URL',1500)" style="margin-top:.3rem;width:100%;font-size:.75rem;font-weight:600;border:1px solid #e5e7eb;border-radius:6px;padding:.3rem;cursor:pointer;background:#f9fafb;color:#374151">Copy URL</button>
          </div>
        </div>
      </details>
    </div>
    <div class="cal-grid">
      <div class="cal-hdr">Sun</div>
      <div class="cal-hdr">Mon</div>
      <div class="cal-hdr">Tue</div>
      <div class="cal-hdr">Wed</div>
      <div class="cal-hdr">Thu</div>
      <div class="cal-hdr">Fri</div>
      <div class="cal-hdr">Sat</div>
      {% for cell in cells %}
        {% if cell is none %}
          <div class="cal-cell other-month"></div>
        {% else %}
          <div class="cal-cell{% if cell.is_today %} today{% endif %}{% if cell.bookings %} has-events{% endif %}"{% if cell.bookings %} data-date="{{ cell.date_str }}" onclick="showPopupForDate(this)"{% endif %}>
            <div class="cal-date">{{ cell.day }}</div>
            <div>{% for b in cell.bookings[:6] %}<span class="cal-dot dot-{{ b.status }}"></span>{% endfor %}</div>
            {% if cell.bookings %}<div class="cal-count">{{ cell.bookings|length }} booking{{ 's' if cell.bookings|length != 1 else '' }}</div>{% endif %}
          </div>
        {% endif %}
      {% endfor %}
    </div>
  </div>
</div>
<div class="popup-overlay" id="popup-overlay" onclick="closePopup()"></div>
<div class="popup" id="popup">
  <div class="popup-hdr">
    <h3 id="popup-title"></h3>
    <button class="popup-close" onclick="closePopup()">&#10005;</button>
  </div>
  <div id="popup-body"></div>
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>
<script type="application/json" id="booking-data">{{ bookings_json | safe }}</script>
<script>
var BDATA=[];
try{BDATA=JSON.parse(document.getElementById('booking-data').textContent);}catch(e){}

function showPopupForDate(el){
  var ds=el.getAttribute('data-date');
  var bks=BDATA.filter(function(b){return b.start<=ds&&b.end>=ds;});
  if(!bks.length)return;
  var d=new Date(ds+'T12:00:00');
  document.getElementById('popup-title').textContent=d.toLocaleDateString('en-US',{weekday:'long',month:'long',day:'numeric',year:'numeric'});
  var sc={'confirmed':'#16a34a','paid':'#0284c7','pending':'#d97706','denied':'#dc2626','cancelled':'#6b7280'};
  document.getElementById('popup-body').innerHTML=bks.map(function(b){
    var t=b.time?' @ '+b.time:'';
    return '<div style="padding:.6rem 0;border-bottom:1px solid #f3f4f6"><strong>#'+b.id+'</strong> '+b.name+t+' <span style="padding:.15rem .5rem;border-radius:9999px;font-size:.75rem;font-weight:600;color:#fff;background:'+(sc[b.status]||'#6b7280')+'">'+b.status+'</span><br><a href="/admin/booking/'+b.id+'" style="font-size:.8rem;color:#2563eb">View booking</a></div>';
  }).join('');
  document.getElementById('popup-overlay').style.display='block';
  document.getElementById('popup').style.display='block';
}
function closePopup(){
  document.getElementById('popup-overlay').style.display='none';
  document.getElementById('popup').style.display='none';
}


</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""

ADMIN_ROUTE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-title" content="Admin">
  <title>Route — {{ business_name }}</title>
<style>
/* ── Sidebar shared ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,sans-serif;background:#f8fafc;color:#1a202c}
    .main{padding:1.25rem;max-width:760px}
    .date-row{display:flex;align-items:center;gap:.75rem;margin-bottom:1.25rem;flex-wrap:wrap}
    .date-row label{font-size:.85rem;font-weight:600;color:#374151}
    .date-row input[type=date]{border:1px solid #d1d5db;border-radius:8px;padding:.45rem .75rem;font-size:.9rem;color:#111827}
    .btn-today{background:#2563eb;color:#fff;border:none;border-radius:8px;padding:.45rem .9rem;font-size:.85rem;font-weight:600;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center}
    /* launch strip */
    .launch-strip{display:flex;gap:.6rem;flex-wrap:wrap;margin-bottom:1.1rem;align-items:center}
    .launch-strip .stop-count{font-size:.82rem;color:#6b7280;flex-basis:100%;margin-bottom:.25rem}
    .btn-launch{display:inline-flex;align-items:center;gap:.45rem;padding:.5rem 1rem;border-radius:9px;font-size:.82rem;font-weight:700;text-decoration:none;border:none;cursor:pointer}
    .btn-gmap{background:#1a73e8;color:#fff}
    .btn-gmap:hover{background:#1557b0}
    .btn-amap{background:#000;color:#fff}
    .btn-amap:hover{background:#222}
    /* stops */
    .empty{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:2rem;text-align:center;color:#6b7280;font-size:.9rem}
    .stop{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:1rem 1.25rem;display:flex;gap:1rem;align-items:flex-start}
    .stop-num{width:2rem;height:2rem;border-radius:50%;background:#2563eb;color:#fff;display:flex;align-items:center;justify-content:center;font-size:.8rem;font-weight:700;flex-shrink:0;margin-top:.1rem}
    .stop-body{flex:1;min-width:0}
    .stop-name{font-weight:700;font-size:.95rem;color:#111827}
    .stop-addr{font-size:.85rem;color:#374151;margin:.2rem 0}
    .stop-meta{font-size:.78rem;color:#6b7280}
    .stop-actions{display:flex;gap:.5rem;margin-top:.55rem;flex-wrap:wrap}
    .btn-sm{display:inline-flex;align-items:center;gap:.3rem;padding:.3rem .7rem;border-radius:7px;font-size:.78rem;font-weight:600;text-decoration:none;border:1px solid #d1d5db;color:#374151;background:#fff}
    .btn-sm:hover{background:#f3f4f6}
    .btn-gm{border-color:#1a73e8;color:#1a73e8}
    .btn-am{border-color:#555;color:#333}
    .btn-view{border-color:#2563eb;color:#2563eb}
    /* leg connector */
    .leg{display:flex;align-items:center;gap:.6rem;padding:.35rem 0 .35rem 1rem;margin:.15rem 0}
    .leg-line{width:2px;height:28px;background:#d1d5db;margin-left:.95rem;flex-shrink:0}
    .leg-arrow{display:flex;flex-direction:column;align-items:center;gap:2px;flex-shrink:0;margin-left:.8rem}
    .leg-arrow span{display:block;width:2px;background:#94a3b8}
    .leg-link{font-size:.75rem;color:#2563eb;text-decoration:none;background:#eff6ff;border:1px solid #bfdbfe;border-radius:6px;padding:.2rem .55rem;white-space:nowrap}
    .leg-link:hover{background:#dbeafe}
    /* badge */
    .badge{display:inline-block;padding:.15rem .45rem;border-radius:10px;font-size:.7rem;font-weight:600;text-transform:uppercase;margin-left:.4rem}
    .badge-confirmed,.badge-partial{background:#dcfce7;color:#15803d}
    .badge-pending{background:#fef3c7;color:#b45309}
    .badge-accepted{background:#ede9fe;color:#5b21b6}
    .no-addr-note{font-size:.8rem;color:#9ca3af;font-style:italic}
    .stop-eta{font-size:.68rem;font-weight:700;color:#fff;background:#2563eb;border-radius:6px;padding:.15rem .35rem;white-space:nowrap;text-align:center;display:none}
    .stop-eta.show{display:block}
    .stop-eta.late{background:#dc2626}
    @media(max-width:520px){.btn-launch{font-size:.78rem;padding:.45rem .8rem}}
  </style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link active">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>{% if view == 'pickup' %}Pickup Route{% elif view == 'count' %}Delivery Count{% else %}Delivery Route{% endif %}</h1>
</div>
<div class="main">
  <form method="GET" action="/admin/route" class="date-row">
    <label for="rdate">Date:</label>
    <input type="date" id="rdate" name="date" value="{{ route_date }}" onchange="this.form.submit()">
    <input type="hidden" name="view" value="{{ view }}">
    <a href="/admin/route?view={{ view }}" class="btn-today">Today</a>
    {% if view == "delivery" %}
    <a href="/sheet/{{ route_date }}/{{ sheet_token }}"
       target="_blank"
       style="background:#059669;color:white;border:none;border-radius:8px;padding:.45rem 1rem;font-size:.85rem;font-weight:600;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:.4rem">
      👁 Preview Sheet
    </a>
    <button onclick="copySheetLink()" id="copy-link-btn"
       style="background:#0891b2;color:white;border:none;border-radius:8px;padding:.45rem 1rem;font-size:.85rem;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:.4rem">
      🔗 Copy Delivery Link
    </button>
    <input type="hidden" id="sheet-url" value="{{ request.host_url }}sheet/{{ route_date }}/{{ sheet_token }}">
    {% endif %}
  </form>

  <div style="display:flex;gap:.5rem;margin-bottom:1.1rem;border-bottom:2px solid #e5e7eb;padding-bottom:.5rem;flex-wrap:wrap">
    <a href="/admin/route?date={{ route_date }}&view=delivery"
       style="padding:.4rem .9rem;border-radius:8px 8px 0 0;font-size:.85rem;font-weight:600;text-decoration:none;border:1px solid #e5e7eb;border-bottom:none;
       {% if view == 'delivery' %}background:#2563eb;color:#fff;border-color:#2563eb{% else %}background:#f9fafb;color:#374151{% endif %}">
      🚚 Deliveries
    </a>
    <a href="/admin/route?date={{ route_date }}&view=pickup"
       style="padding:.4rem .9rem;border-radius:8px 8px 0 0;font-size:.85rem;font-weight:600;text-decoration:none;border:1px solid #e5e7eb;border-bottom:none;
       {% if view == 'pickup' %}background:#7c3aed;color:#fff;border-color:#7c3aed{% else %}background:#f9fafb;color:#374151{% endif %}">
      🔄 Pickups
    </a>
    <a href="/admin/route?date={{ route_date }}&view=count"
       style="padding:.4rem .9rem;border-radius:8px 8px 0 0;font-size:.85rem;font-weight:600;text-decoration:none;border:1px solid #e5e7eb;border-bottom:none;
       {% if view == 'count' %}background:#059669;color:#fff;border-color:#059669{% else %}background:#f9fafb;color:#374151{% endif %}">
      📦 Delivery Count
    </a>
  </div>

  {# ════════════════ DELIVERY COUNT TAB ════════════════ #}
  {% if view == 'count' %}
    {% if item_totals %}
    <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;overflow:hidden;margin-bottom:1.5rem">
      <div style="background:#059669;color:#fff;padding:.75rem 1.25rem;display:flex;align-items:center;justify-content:space-between">
        <span style="font-weight:700;font-size:1rem">📦 Total Items Going Out — {{ route_date }}</span>
        <span style="font-size:.82rem;opacity:.85">{{ item_totals|length }} item type{{ 's' if item_totals|length != 1 }}</span>
      </div>
      <table style="width:100%;border-collapse:collapse">
        <thead>
          <tr style="background:#f0fdf4;border-bottom:2px solid #bbf7d0">
            <th style="text-align:left;padding:.6rem 1.25rem;font-size:.78rem;font-weight:700;color:#065f46;text-transform:uppercase;letter-spacing:.05em">Item</th>
            <th style="text-align:right;padding:.6rem 1.25rem;font-size:.78rem;font-weight:700;color:#065f46;text-transform:uppercase;letter-spacing:.05em">Total Qty</th>
            <th style="text-align:left;padding:.6rem 1.25rem;font-size:.78rem;font-weight:700;color:#065f46;text-transform:uppercase;letter-spacing:.05em;display:none" class="bk-col">Bookings</th>
          </tr>
        </thead>
        <tbody>
          {% for row in item_totals %}
          <tr style="border-bottom:1px solid #e5e7eb{% if loop.last %};border-bottom:none{% endif %}"
              onclick="toggleBreakdown(this)" style="cursor:pointer">
            <td style="padding:.75rem 1.25rem;font-weight:600;font-size:.92rem;color:#111827;cursor:pointer">
              {{ row.item }}
            </td>
            <td style="padding:.75rem 1.25rem;text-align:right">
              <span style="background:#dcfce7;color:#166534;font-weight:800;font-size:1.05rem;padding:.2rem .7rem;border-radius:8px">
                {{ row.qty }}
              </span>
            </td>
            <td style="padding:.75rem 1.25rem">
              <div style="display:flex;gap:.4rem;flex-wrap:wrap">
                {% for bk in row.bookings %}
                <a href="/admin/booking/{{ bk.id }}"
                   style="font-size:.75rem;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:6px;padding:.15rem .5rem;text-decoration:none;white-space:nowrap"
                   onclick="event.stopPropagation()">
                  #{{ bk.id }} {{ bk.name }} ({{ bk.qty }})
                </a>
                {% endfor %}
              </div>
            </td>
          </tr>
          {% endfor %}
        </tbody>
        <tfoot>
          <tr style="background:#f0fdf4;border-top:2px solid #bbf7d0">
            <td style="padding:.65rem 1.25rem;font-weight:700;color:#065f46;font-size:.88rem">TOTAL PIECES</td>
            <td style="padding:.65rem 1.25rem;text-align:right">
              <span style="background:#059669;color:#fff;font-weight:800;font-size:1.1rem;padding:.25rem .8rem;border-radius:8px">
                {{ item_totals | sum(attribute='qty') }}
              </span>
            </td>
            <td style="padding:.65rem 1.25rem"></td>
          </tr>
        </tfoot>
      </table>
    </div>
    <style>
      @media(min-width:520px){ .bk-col{display:table-cell!important} }
      table tr:hover td{background:#f9fafb;cursor:pointer}
    </style>
    {% else %}
    <div class="empty">No deliveries scheduled for {{ route_date }}.</div>
    {% endif %}
  {% endif %}

  {# Show delivery/pickup content only when NOT on count tab #}
  {% if view != 'count' %}

  {% if route_bookings %}
  {% set addrs = route_bookings|selectattr('nav_address')|map(attribute='nav_address')|list %}

  <div class="launch-strip">
    <div class="stop-count">{{ route_bookings|length }} {% if view == "pickup" %}pickup{% else %}delivery{% endif %}{{ 's' if route_bookings|length != 1 }} · {{ route_date }}</div>
    {% if addrs|length >= 1 %}
    {# Google Maps: depot → stop1 → ... → stopN #}
    {% set gurl = "https://www.google.com/maps/dir/" + depot_address|replace(" ","+") + "/" + (addrs|join("/"))|replace(" ", "+") %}
    <a class="btn-launch btn-gmap" href="{{ gurl }}" target="_blank">
      🗺 Start Route — Google Maps
    </a>
    <a class="btn-launch btn-amap" href="https://maps.apple.com/?saddr={{ depot_address|urlencode }}&daddr={{ addrs[-1]|urlencode }}&dirflg=d" target="_blank">
      🗺 Start Route — Apple Maps
    </a>
    <button id="opt-btn" onclick="optimizeRoute()" class="btn-launch" style="background:#7c3aed;color:white;border:none;cursor:pointer">
      ✨ Optimize Route Order
    </button>
    {% else %}
    <span class="no-addr-note">📍 Add delivery addresses to enable route navigation</span>
    {% endif %}
  </div>

  {# ── Origin section: current location → depot ── #}
  <div id="origin-section">
    <!-- Current location card (shown when GPS available) -->
    <div id="curr-loc-card" style="display:none;flex-direction:row;gap:1rem;align-items:center;padding:.4rem .25rem .1rem">
      <div style="width:2rem;display:flex;justify-content:center">
        <div style="width:2rem;height:2rem;border-radius:50%;background:#0891b2;color:#fff;display:flex;align-items:center;justify-content:center;font-size:.9rem">&#128205;</div>
      </div>
      <div style="font-size:.82rem;font-weight:500;color:#374151">
        <strong>Your Location</strong> <span id="curr-loc-label" style="color:#6b7280;font-size:.75rem">(detecting…)</span>
      </div>
    </div>
    <!-- leg: current → depot -->
    <div id="curr-to-depot-leg" style="display:none;flex-direction:row;align-items:center;gap:.6rem;padding:.3rem 0 .3rem 1rem">
      <div style="width:2rem;flex-shrink:0;display:flex;justify-content:center">
        <div style="width:2px;height:32px;background:#d1d5db"></div>
      </div>
      <a id="curr-to-depot-link" class="leg-link" target="_blank" href="#">
        <span class="leg-spin">&#9696;</span> Calculating…
      </a>
    </div>
    <!-- Depot card -->
    <div style="display:flex;gap:1rem;align-items:center;padding:.4rem .25rem .1rem">
      <div style="width:2rem;display:flex;justify-content:center;flex-direction:column;align-items:center;gap:.25rem">
        <div style="width:2rem;height:2rem;border-radius:50%;background:#6b7280;color:#fff;display:flex;align-items:center;justify-content:center;font-size:.9rem">&#127968;</div>
        <div class="stop-eta" id="eta-depot" style="font-size:.62rem"></div>
      </div>
      <div style="font-size:.82rem;color:#6b7280;font-weight:500">
        <strong style="color:#374151">Depot:</strong> {{ depot_address }}
      </div>
    </div>
  </div>
  <div style="display:flex;margin-left:.95rem"><div style="width:2px;height:24px;background:#d1d5db"></div></div>

  {% for b in route_bookings %}

  {# Leg connector between consecutive stops #}
  {% if not loop.first %}
  {% set prev = route_bookings[loop.index0 - 1] %}
  <div class="leg" id="leg-{{ loop.index0 }}">
    <div style="width:2rem;flex-shrink:0;display:flex;justify-content:center">
      <div style="width:2px;height:36px;background:#d1d5db"></div>
    </div>
    {% if prev.nav_address and b.nav_address %}
    <a class="leg-link" id="leg-link-{{ loop.index0 }}" target="_blank"
       href="https://www.google.com/maps/dir/{{ prev.nav_address|urlencode }}/{{ b.nav_address|urlencode }}">
      <span class="leg-spin">&#9696;</span> Calculating…
    </a>
    {% else %}
    <span style="font-size:.72rem;color:#9ca3af">— add event addresses for distance —</span>
    {% endif %}
  </div>
  {% endif %}

  <div class="stop">
    <div style="display:flex;flex-direction:column;align-items:center;gap:.3rem;flex-shrink:0">
      <div class="stop-num">{{ loop.index }}</div>
      <div class="stop-eta" id="eta-{{ loop.index }}"></div>
    </div>
    <div class="stop-body">
      <div class="stop-name">
        {{ b.full_name }}
        <span class="badge badge-{{ b.status }}">{{ b.status }}</span>
      </div>
      {% if b.nav_address %}
      <div class="stop-addr">📍 {{ b.nav_address }}</div>
      {% else %}
      <div class="stop-addr" style="color:#9ca3af">No address on file</div>
      {% endif %}
      {% if b.delivery_location %}
      <div class="stop-addr" style="font-size:.78rem;color:#6b7280">📋 {{ b.delivery_location }}</div>
      {% endif %}
      <div class="stop-meta">
        {% if b.event_start_time %}🕐 {{ b.event_start_time }}{% endif %}
        {% if b.phone %} &nbsp;📞 <a href="tel:{{ b.phone }}" style="color:#374151">{{ b.phone }}</a>{% endif %}
      </div>
      {% if b.items_summary %}
      <div class="stop-meta" style="margin-top:.25rem">📦 {{ b.items_summary }}</div>
      {% endif %}
      <div class="stop-actions">
        <a href="/admin/booking/{{ b.id }}" class="btn-sm btn-view">👁 #{{ b.id }}</a>
        {% if b.nav_address %}
        <a href="https://www.google.com/maps/dir/?api=1&destination={{ b.nav_address|urlencode }}&travelmode=driving"
           target="_blank" class="btn-sm btn-gm">🗺 Google Maps</a>
        <a href="https://maps.apple.com/?daddr={{ b.nav_address|urlencode }}&dirflg=d"
           target="_blank" class="btn-sm btn-am">🗺 Apple Maps</a>
        {% else %}
        <span class="btn-sm" style="opacity:.45;cursor:default" title="No address on file">🗺 No address</span>
        {% endif %}
        {% if b.phone %}
        <a href="tel:{{ b.phone }}" class="btn-sm">📞 Call</a>
        {% endif %}
        {% if b.route_override %}
        <form method="POST" action="/admin/route/override/{{ b.id }}" style="display:inline">
          <input type="hidden" name="redirect" value="/admin/route?date={{ route_date }}&view={{ view }}">
          <button type="submit" class="btn-sm" style="background:#fee2e2;color:#b91c1c;border:1px solid #fca5a5;cursor:pointer" title="Remove from route">✕ Remove Override</button>
        </form>
        {% endif %}
      </div>
    </div>
  </div>
  {% endfor %}
  {% else %}
  <div class="empty">{% if view == "pickup" %}No pickups scheduled for {{ route_date }}.{% else %}No deliveries scheduled for {{ route_date }}.{% endif %}</div>
  {% endif %}

  {% endif %}{# /if view != count #}

  {% if excluded_bookings and view != 'count' %}
  <div style="margin-top:2rem;padding:1rem 1.25rem;background:#f9fafb;border:1.5px dashed #d1d5db;border-radius:10px">
    <div style="font-size:.82rem;font-weight:700;color:#6b7280;margin-bottom:.75rem;text-transform:uppercase;letter-spacing:.05em">
      📋 Not on route — {{ excluded_bookings|length }} booking{{ 's' if excluded_bookings|length != 1 }} scheduled this day
    </div>
    {% for eb in excluded_bookings %}
    <div style="display:flex;align-items:center;justify-content:space-between;padding:.6rem .75rem;background:white;border:1px solid #e5e7eb;border-radius:8px;margin-bottom:.5rem;gap:1rem">
      <div style="flex:1;min-width:0">
        <span style="font-weight:600;color:#111827">{{ eb.full_name }}</span>
        <span class="badge badge-{{ eb.status }}" style="margin-left:.4rem;font-size:.68rem">{{ eb.status }}</span>
        <div style="font-size:.78rem;color:#6b7280;margin-top:.15rem">
          #{{ eb.id }}{% if eb.items_summary %} · {{ eb.items_summary }}{% endif %}
        </div>
      </div>
      <form method="POST" action="/admin/route/override/{{ eb.id }}" style="flex-shrink:0">
        <input type="hidden" name="redirect" value="/admin/route?date={{ route_date }}&view={{ view }}">
        <button type="submit" style="background:#1d4ed8;color:white;border:none;border-radius:6px;padding:.35rem .85rem;font-size:.78rem;font-weight:600;cursor:pointer">+ Add to Route</button>
      </form>
    </div>
    {% endfor %}
  </div>
  {% endif %}
</div>
</div>
<script>
// ── Google Maps DirectionsService optimizer ───────────────────────────────
// Called automatically on load (when no custom order yet) and by the button
function optimizeRoute(isAuto){
  if(STOPS.length < 2) return;
  const btn = document.getElementById('opt-btn');
  if(btn){ btn.disabled=true; btn.innerHTML='⏳ Optimizing…'; }

  // Show a subtle banner while optimizing
  if(isAuto){
    const banner = document.createElement('div');
    banner.id = 'opt-banner';
    banner.style.cssText = 'position:fixed;top:60px;left:50%;transform:translateX(-50%);background:#1d4ed8;color:#fff;padding:.5rem 1.25rem;border-radius:8px;font-size:.85rem;font-weight:600;z-index:999;box-shadow:0 4px 12px rgba(0,0,0,.25)';
    banner.textContent = '✨ Finding fastest route…';
    document.body.appendChild(banner);
  }

  const svc = new google.maps.DirectionsService();
  const validStops = STOPS.filter(s => s.addr);
  const noAddr     = STOPS.filter(s => !s.addr);

  const waypoints = validStops.slice(0, -1).map(s => ({
    location: s.addr,
    stopover: true,
  }));
  const destination = validStops[validStops.length - 1].addr;

  svc.route({
    origin: DEPOT,
    destination: destination,
    waypoints: waypoints,
    optimizeWaypoints: true,   // ← Google solves TSP for us
    travelMode: google.maps.TravelMode.DRIVING,
  }, function(result, status){
    const banner = document.getElementById('opt-banner');
    if(banner) banner.remove();
    if(btn){ btn.disabled=false; btn.innerHTML='✨ Re-optimize Route'; }

    if(status !== 'OK'){
      if(!isAuto) alert('Could not optimize route: ' + status);
      return;
    }
    // result.routes[0].waypoint_order is the reordered indices into validStops[0..-2]
    const order = result.routes[0].waypoint_order;
    const reordered = order.map(i => validStops[i]).concat([validStops[validStops.length - 1]]);
    // Add stops with no address at the end
    const all = [...reordered, ...noAddr];
    const ids = all.map(s => s.id).filter(Boolean).join(',');
    const url = new URL(window.location.href);
    url.searchParams.set('order', ids);
    window.location.href = url.toString();
  });
}
function copySheetLink(){
  const url = document.getElementById('sheet-url').value;
  navigator.clipboard.writeText(url).then(()=>{
    const btn = document.getElementById('copy-link-btn');
    btn.innerHTML = '✓ Link Copied!';
    btn.style.background = '#059669';
    setTimeout(()=>{ btn.innerHTML = '🔗 Copy Delivery Link'; btn.style.background = '#0891b2'; }, 2500);
  }).catch(()=>{ prompt('Copy this link and send it to your brother:', url); });
}
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}

/* ── Route data ── */
const STOPS = {{ stops_json|safe }};
const DEPOT = {{ depot_address|tojson }};
const HAS_CUSTOM_ORDER = {{ 'true' if has_custom_order else 'false' }};

function fmtTime(d){
  let h=d.getHours(),m=d.getMinutes(),a=h>=12?'PM':'AM';
  h=h%12||12; return `${h}:${String(m).padStart(2,'0')} ${a}`;
}
function setEta(idx,dt){
  const el=document.getElementById('eta-'+idx);
  if(!el) return;
  el.textContent=fmtTime(dt);
  el.classList.add('show');
}

function getCurrentPos(){
  return new Promise((resolve,reject)=>{
    if(!navigator.geolocation){reject(new Error('no geo'));return;}
    navigator.geolocation.getCurrentPosition(
      p=>resolve({lat:p.coords.latitude,lon:p.coords.longitude}),
      e=>reject(e),
      {timeout:8000,maximumAge:60000}
    );
  });
}
{% if google_maps_key %}
</script>
<script src="https://maps.googleapis.com/maps/api/js?key={{ google_maps_key }}&callback=mapsReady&loading=async" defer></script>
<script>
// Fallback: if Google Maps fails to load, still show the optimize button as non-functional
window.mapsReady = window.mapsReady || function(){};
{% else %}
// No Google Maps key — optimize button disabled
document.addEventListener('DOMContentLoaded', function(){
  var btn=document.getElementById('opt-btn');
  if(btn){btn.disabled=true;btn.title='Set GOOGLE_MAPS_KEY to enable route optimization';}
});
{% endif %}

function loadDistances(){
  if(!STOPS.length || !window.google) return;
  const startTime = new Date();
  const svc = new google.maps.DistanceMatrixService();
  const addrs = [DEPOT, ...STOPS.map(s=>s.addr).filter(Boolean)];

  svc.getDistanceMatrix({
    origins: addrs.slice(0,-1),
    destinations: addrs.slice(1),
    travelMode: google.maps.TravelMode.DRIVING,
  }, function(resp, status){
    if(status !== 'OK') return;
    let cumSecs = 0;
    for(let i = 0; i < STOPS.length; i++){
      const row = resp.rows[i];
      if(!row) continue;
      const elem = row.elements[0];
      if(!elem || elem.status !== 'OK') continue;
      const secs = elem.duration.value;
      const mins = Math.round(secs / 60);
      const hrs  = Math.floor(mins / 60), rem = mins % 60;
      const timeStr = hrs > 0 ? `${hrs}h ${rem}m` : `${mins} min`;
      const dist = (elem.distance.value * 0.000621371).toFixed(1);
      const el = document.getElementById('leg-link-' + i);
      if(el) el.innerHTML = `⇕ ${dist} mi &nbsp;·&nbsp; ~${timeStr}`;
      cumSecs += secs;
      setEta(i + 1, new Date(startTime.getTime() + cumSecs * 1000));
    }
  });
}

function mapsReady(){
  // Auto-optimize if no custom order and 2+ stops
  if(STOPS.length >= 2 && !HAS_CUSTOM_ORDER){
    optimizeRoute(true);
  }
  if(STOPS.length > 0) loadDistances();
}
</script>
<style>
@keyframes spin{to{transform:rotate(360deg)}}
.leg-spin{display:inline-block;animation:spin 1s linear infinite}
</style>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES — INVENTORY
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/inventory")
@admin_required
def admin_inventory():
    products   = get_products()
    check_from = request.args.get("check_from", "")
    check_to   = request.args.get("check_to", "")
    avail_data = []
    if check_from:
        end = check_to or check_from
        available = get_available(check_from, end)
        for p in products:
            total    = int(p.get("total", 0))
            avail    = available.get(p["id"], total)
            reserved = max(0, total - avail)
            avail_data.append({
                "name": p["name"],
                "total": total,
                "reserved": reserved,
                "available": avail,
            })
    return render_template_string(ADMIN_INVENTORY_HTML,
        business_name=BUSINESS_NAME,
        products=products,
        avail_data=avail_data,
        check_from=check_from,
        check_to=check_to,
        flash_ok=request.args.get("flash_ok", ""),
        flash_err=request.args.get("flash_err", ""),
    )


@app.route("/admin/inventory/save", methods=["POST"])
@admin_required
def save_inventory():
    conn = get_db()
    if conn:
        try:
            cur   = conn.cursor()
            count = int(request.form.get("count", 0))
            for idx in range(count):
                item_id = request.form.get(f"id_{idx}")
                name    = request.form.get(f"name_{idx}", "").strip()
                price   = request.form.get(f"price_{idx}", "0")
                total   = request.form.get(f"total_{idx}", "0")
                if item_id and name:
                    cur.execute(
                        "UPDATE inventory SET name=%s, price=%s, total=%s WHERE id=%s",
                        (name, float(price or 0), int(total or 0), item_id)
                    )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"save_inventory error: {e}")
            return redirect(url_for("admin_inventory", flash_err="Error saving inventory"))
    return redirect(url_for("admin_inventory", flash_ok="Inventory saved successfully"))


@app.route("/admin/inventory/add", methods=["POST"])
@admin_required
def add_inventory():
    import re as _re
    name  = request.form.get("name", "").strip()
    price = request.form.get("price", "0")
    total = request.form.get("total", "0")
    if name:
        # Generate a slug-based id from the name (e.g. "Round Tables" -> "round_tables")
        base_id = _re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_') or "item"
        conn = get_db()
        if conn:
            try:
                cur = conn.cursor()
                # Ensure uniqueness — append _2, _3, etc. if needed
                new_id = base_id
                suffix = 2
                while True:
                    cur.execute("SELECT 1 FROM inventory WHERE id=%s", (new_id,))
                    if not cur.fetchone():
                        break
                    new_id = f"{base_id}_{suffix}"
                    suffix += 1
                cur.execute(
                    "INSERT INTO inventory (id, name, price, total, sort_order) VALUES (%s, %s, %s, %s, 999)",
                    (new_id, name, float(price or 0), int(total or 0))
                )
                conn.commit()
                cur.close()
                conn.close()
            except Exception as e:
                log.error(f"add_inventory error: {e}")
                return redirect(url_for("admin_inventory", flash_err="Error adding item"))
    return redirect(url_for("admin_inventory", flash_ok=f"'{name}' added to inventory"))


@app.route("/admin/inventory/delete/<item_id>", methods=["POST"])
@admin_required
def delete_inventory(item_id):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM inventory WHERE id=%s", (item_id,))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"delete_inventory error: {e}")
    return redirect(url_for("admin_inventory", flash_ok="Item removed"))


@app.route("/admin/reconcile-stripe")
@admin_required
def reconcile_stripe():
    """One-time route: sync all past Stripe payments to bookings DB."""
    if not STRIPE_SECRET_KEY:
        return "<pre>ERROR: STRIPE_SECRET_KEY not set</pre>", 500
    import stripe as _stripe
    _stripe.api_key = STRIPE_SECRET_KEY
    lines = ["<pre style='font-family:monospace;padding:1rem'>Searching Stripe for payments...\n"]
    try:
        # Collect paid checkout sessions that have a booking_id in metadata
        # These come from Payment Links (which your app creates)
        sessions, params = [], {"limit": 100}
        while True:
            page = _stripe.checkout.Session.list(**params)
            sessions.extend(page.data)
            if not page.has_more:
                break
            params["starting_after"] = page.data[-1].id

        paid = []
        for s in sessions:
            try:
                ps = s.payment_status
                meta = s.metadata
                bid = meta["booking_id"] if meta else None
                if ps == "paid" and bid:
                    paid.append((bid, round((s.amount_total or 0) / 100, 2), s.id))
            except Exception:
                pass

        lines.append(f"Found {len(paid)} paid sessions with booking_id\n\n")
        conn = get_db()
        n = 0
        if conn:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            for bid_str, amt, sess_id in paid:
                bid = int(bid_str)
                cur.execute("SELECT id,full_name,amount_paid,grand_total,status FROM bookings WHERE id=%s", (bid,))
                row = cur.fetchone()
                if not row:
                    lines.append(f"  SKIP  #{bid} not found in DB\n")
                    continue
                b = dict(row)
                db_paid = round(float(b["amount_paid"] or 0), 2)
                gt = round(float(b["grand_total"] or 0), 2)
                st = b["status"]
                if abs(db_paid - amt) < 0.02 and st in ("confirmed", "paid"):
                    lines.append(f"  OK    #{bid} {b['full_name']} ${db_paid}/{st}\n")
                    continue
                balance = round(gt - amt, 2)
                ns = "paid" if balance <= 0.50 else "confirmed"
                cur.execute(
                    "UPDATE bookings SET amount_paid=%s, status=%s, stripe_session_id=%s WHERE id=%s",
                    (amt, ns, sess_id, bid)
                )
                lines.append(f"  FIXED #{bid} {b['full_name']} ${db_paid}/{st} → ${amt}/{ns} (bal ${max(balance,0):.2f})\n")
                n += 1
            conn.commit()
            cur.close()
            conn.close()
        lines.append(f"\nDone. {n} bookings updated.\n</pre>")
    except Exception as e:
        lines.append(f"\nERROR: {e}\n</pre>")
    return "".join(lines)


@app.route("/admin/calendar.ics")
def admin_calendar_ics():
    """Serve iCal feed — accessible via token or active admin session."""
    token = request.args.get("token", "")
    if token != CALENDAR_TOKEN:
        # Fall back to session check
        if not session.get("admin_logged_in"):
            return redirect("/admin")

    conn = get_db()
    bookings = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("""
                SELECT id, full_name, event_start_date, event_end_date, event_start_time, status
                FROM bookings
                WHERE status NOT IN ('denied','cancelled')
                  AND (archived IS NULL OR archived = FALSE)
                  AND event_start_date IS NOT NULL
                ORDER BY event_start_date ASC
            """)
            bookings = [dict(r) for r in cur.fetchall()]
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"calendar_ics error: {e}")
    lines = [
        "BEGIN:VCALENDAR", "VERSION:2.0",
        "PRODID:-//Rent a Party LLC//Admin//EN",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
        f"X-WR-CALNAME:{BUSINESS_NAME} Bookings",
    ]
    for b in bookings:
        sd = str(b["event_start_date"]).replace("-", "")
        ed = str(b.get("event_end_date") or b["event_start_date"]).replace("-", "")
        t  = str(b.get("event_start_time") or "")
        if t:
            hh, mm = (t + ":00").split(":")[:2]
            dtstart = f"DTSTART:{sd}T{hh.zfill(2)}{mm.zfill(2)}00"
        else:
            dtstart = f"DTSTART;VALUE=DATE:{sd}"
        lines += [
            "BEGIN:VEVENT",
            f"UID:booking-{b['id']}@rental-booking",
            dtstart,
            f"DTEND;VALUE=DATE:{ed}",
            f"SUMMARY:Booking #{b['id']} — {b.get('full_name','')}",
            f"DESCRIPTION:Status: {b.get('status','')}",
            f"STATUS:{'CONFIRMED' if b.get('payment_status') in ('paid','partial') else 'TENTATIVE'}",
            "END:VEVENT",
        ]
    lines.append("END:VCALENDAR")
    ics = "\r\n".join(lines)
    from flask import Response
    return Response(ics, mimetype="text/calendar",
                    headers={"Content-Disposition": "attachment; filename=bookings.ics"})


@app.route("/admin/calendar.csv")
@admin_required
def admin_calendar_csv():
    """Serve CSV download of all bookings."""
    conn = get_db()
    bookings = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("""
                SELECT id, full_name, event_start_date, event_end_date, event_start_time, status
                FROM bookings
                WHERE status NOT IN ('denied','cancelled')
                  AND (archived IS NULL OR archived = FALSE)
                  AND event_start_date IS NOT NULL
                ORDER BY event_start_date ASC
            """)
            bookings = [dict(r) for r in cur.fetchall()]
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"calendar_csv error: {e}")
    def q(v):
        return '"' + str(v or '').replace('"','""')+'"'
    rows = ['ID,Name,Start,End,Time,Status']
    for b in bookings:
        rows.append(','.join([
            q(b['id']), q(b.get('full_name','')),
            q(b['event_start_date']), q(b.get('event_end_date','')),
            q(b.get('event_start_time','')), q(b.get('status','')),
        ]))
    csv = '\n'.join(rows)
    from flask import Response
    return Response(csv, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=bookings.csv'})


@app.route("/admin/calendar")
@admin_required
def admin_calendar():
    import calendar as _cal
    import json as _json
    m_param = request.args.get("m", "")
    try:
        year, month = int(m_param[:4]), int(m_param[5:7])
        if not (1 <= month <= 12): raise ValueError
    except Exception:
        _t = date.today()
        year, month = _t.year, _t.month
    conn = get_db()
    bookings = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("""
                SELECT id, full_name, event_start_date, event_end_date, event_start_time, status
                FROM bookings
                WHERE status NOT IN ('denied','cancelled')
                  AND (archived IS NULL OR archived = FALSE)
                  AND event_start_date IS NOT NULL
                ORDER BY event_start_date ASC
            """)
            for row in cur.fetchall():
                b = dict(row)
                bookings.append({
                    "id": b["id"],
                    "name": b.get("full_name") or "",
                    "start": str(b["event_start_date"]),
                    "end": str(b.get("event_end_date") or b["event_start_date"]),
                    "time": str(b.get("event_start_time") or ""),
                    "status": b.get("status") or "pending",
                })
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"admin_calendar error: {e}")
    # Build calendar grid server-side
    first_dow_monday, days_count = _cal.monthrange(year, month)
    first_dow_sunday = (first_dow_monday + 1) % 7  # convert Mon-based to Sun-based
    today_str = date.today().isoformat()
    month_name = date(year, month, 1).strftime("%B %Y")
    prev_y, prev_m = (year - 1, 12) if month == 1 else (year, month - 1)
    next_y, next_m = (year + 1, 1) if month == 12 else (year, month + 1)
    by_date = {}
    for b in bookings:
        by_date.setdefault(b["start"], []).append(b)
    cells = [None] * first_dow_sunday
    for d in range(1, days_count + 1):
        ds = f"{year}-{month:02d}-{d:02d}"
        cells.append({
            "day": d,
            "date_str": ds,
            "bookings": by_date.get(ds, []),
            "is_today": ds == today_str,
        })
    _ics_url = f"https://{request.host}/admin/calendar.ics?token={CALENDAR_TOKEN}"
    from urllib.parse import quote as _quote
    _gcal_url = "https://calendar.google.com/calendar/r/settings/addbyurl?url=" + _quote(_ics_url, safe="")
    return render_template_string(ADMIN_CALENDAR_HTML,
        business_name=BUSINESS_NAME,
        month_name=month_name,
        prev_url=f"/admin/calendar?m={prev_y}-{prev_m:02d}",
        next_url=f"/admin/calendar?m={next_y}-{next_m:02d}",
        cells=cells,
        bookings_json=json.dumps(bookings),
        gcal_url=_gcal_url,
        ics_url=_ics_url,
    )

@app.route("/admin/route")
@admin_required
def admin_route():
    route_date = request.args.get("date", date.today().isoformat())
    view = request.args.get("view", "delivery")  # "delivery" or "pickup"
    conn = get_db()
    route_bookings = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            if view == "pickup":
                cur.execute("""
                    SELECT id, full_name, phone, email, delivery_location,
                           event_street, event_city, event_state, event_zip,
                           event_start_time, items_json, status, grand_total,
                           event_end_date AS route_date_field, route_override
                    FROM bookings
                    WHERE event_end_date = %s
                      AND (
                        (status = 'accepted' AND payment_status IN ('paid','partial'))
                        OR route_override = TRUE
                      )
                      AND status NOT IN ('denied','cancelled','concluded')
                      AND (archived IS NULL OR archived = FALSE)
                    ORDER BY event_start_time ASC NULLS LAST, id ASC
                """, (route_date,))
            else:
                cur.execute("""
                    SELECT id, full_name, phone, email, delivery_location,
                           event_street, event_city, event_state, event_zip,
                           COALESCE(delivery_time, setup_time) AS event_start_time,
                           items_json, status, grand_total,
                           COALESCE(delivery_date, setup_date) AS route_date_field,
                           route_override
                    FROM bookings
                    WHERE COALESCE(delivery_date, setup_date) = %s
                      AND (
                        (status = 'accepted' AND payment_status IN ('paid','partial'))
                        OR route_override = TRUE
                      )
                      AND status NOT IN ('denied','cancelled','concluded')
                      AND (archived IS NULL OR archived = FALSE)
                    ORDER BY COALESCE(delivery_time, setup_time) ASC NULLS LAST, id ASC
                """, (route_date,))
            for row in cur.fetchall():
                b = dict(row)
                parts = [
                    (b.get('event_street') or '').strip(),
                    (b.get('event_city')   or '').strip(),
                    (b.get('event_state')  or '').strip(),
                    (b.get('event_zip')    or '').strip(),
                ]
                b['nav_address'] = ', '.join(p for p in parts if p) or ''
                try:
                    items = json.loads(b.get('items_json') or '[]')
                    b['items_summary'] = ', '.join(f"{i.get('qty',1)}x {i.get('name','')}" for i in items)
                except Exception:
                    b['items_summary'] = ''
                route_bookings.append(b)
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"admin_route error: {e}")
    stops_json = json.dumps([
        {"addr": b["nav_address"], "name": b.get("full_name", ""), "id": b["id"]}
        for b in route_bookings
    ])

    # Aggregate item totals across all delivery bookings for this date
    item_totals_map = {}   # name → {qty, bookings: [{"id":..,"name":..}]}
    conn_ct = get_db()
    if conn_ct:
        try:
            cur_ct = conn_ct.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur_ct.execute("""
                SELECT id, full_name, items_json FROM bookings
                WHERE COALESCE(delivery_date, setup_date) = %s
                  AND status NOT IN ('denied','cancelled','concluded')
                  AND (archived IS NULL OR archived = FALSE)
            """, (route_date,))
            for row in cur_ct.fetchall():
                b2 = dict(row)
                try:
                    for itm in json.loads(b2.get('items_json') or '[]'):
                        nm  = (itm.get('name') or '').strip()
                        qty = int(itm.get('qty') or 1)
                        if not nm:
                            continue
                        if nm not in item_totals_map:
                            item_totals_map[nm] = {'qty': 0, 'bookings': []}
                        item_totals_map[nm]['qty'] += qty
                        item_totals_map[nm]['bookings'].append({
                            'id': b2['id'],
                            'name': b2.get('full_name') or f"#{b2['id']}",
                            'qty': qty,
                        })
                except Exception:
                    pass
            cur_ct.close(); conn_ct.close()
        except Exception as e:
            log.error(f"item_totals error: {e}")
    # Sort by qty descending
    item_totals = sorted(
        [{'item': k, **v} for k, v in item_totals_map.items()],
        key=lambda x: x['qty'], reverse=True
    )

    # Bookings on this date that are NOT on the route (pending / waiting / not overridden)
    # Apply custom stop order if provided in URL
    custom_order = request.args.get("order", "")
    if custom_order:
        try:
            order_ids = [int(x) for x in custom_order.split(",") if x.strip()]
            id_to_rank = {bid: i for i, bid in enumerate(order_ids)}
            route_bookings.sort(key=lambda b: id_to_rank.get(b["id"], 999))
        except Exception as e:
            log.error(f"Custom order error: {e}")

    excluded_bookings = []
    date_col = "event_end_date" if view == "pickup" else "setup_date"
    conn_ex = get_db()
    if conn_ex:
        try:
            cur_ex = conn_ex.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur_ex.execute(f"""
                SELECT id, full_name, status, payment_status, items_json
                FROM bookings
                WHERE {date_col} = %s
                  AND NOT (status = 'accepted' AND payment_status IN ('paid','partial'))
                  AND status NOT IN ('denied','cancelled','concluded')
                  AND (route_override IS NULL OR route_override = FALSE)
                  AND (archived IS NULL OR archived = FALSE)
                ORDER BY id ASC
            """, (route_date,))
            for row in cur_ex.fetchall():
                eb = dict(row)
                try:
                    items = json.loads(eb.get('items_json') or '[]')
                    eb['items_summary'] = ', '.join(f"{i.get('qty',1)}x {i.get('name','')}" for i in items)
                except Exception:
                    eb['items_summary'] = ''
                excluded_bookings.append(eb)
            cur_ex.close(); conn_ex.close()
        except Exception as e:
            log.error(f"excluded_bookings error: {e}")

    return render_template_string(ADMIN_ROUTE_HTML,
        business_name=BUSINESS_NAME,
        depot_address=DEPOT_ADDRESS,
        route_date=route_date,
        view=view,
        route_bookings=route_bookings,
        excluded_bookings=excluded_bookings,
        stops_json=stops_json,
        sheet_token=_sheet_token(route_date),
        google_maps_key=GOOGLE_MAPS_KEY,
        has_custom_order=bool(request.args.get("order", "")),
        item_totals=item_totals,
    )


# ─── Delivery Sheet (public, no login) ───────────────────────────────────────

def _sheet_token(date_str):
    """Deterministic token from date + secret. Consistent for the same date."""
    secret = (app.secret_key or "rentaparty").encode()
    return hmac.new(secret, date_str.encode(), hashlib.sha256).hexdigest()[:16]

DELIVERY_SHEET_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Deliveries – {{ sheet_date }}</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f0f4f8;min-height:100vh}
.header{background:#1a365d;color:white;padding:1.1rem 1.25rem;position:sticky;top:0;z-index:10}
.header h1{font-size:1rem;font-weight:700}
.header p{font-size:.78rem;opacity:.8;margin-top:.15rem}
.stops{padding:.75rem;display:flex;flex-direction:column;gap:.75rem}
.stop{background:white;border-radius:12px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.08)}
.stop.done{opacity:.55}
.stop-header{display:flex;align-items:center;gap:.75rem;padding:.85rem 1rem;border-bottom:1px solid #f1f5f9}
.num{width:32px;height:32px;border-radius:50%;background:#1a365d;color:white;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:.9rem;flex-shrink:0}
.stop.done .num{background:#16a34a}
.cname{font-weight:700;font-size:.95rem;color:#111827}
.order{font-size:.75rem;color:#6b7280;margin-top:.1rem}
.stop-body{padding:.85rem 1rem;display:flex;flex-direction:column;gap:.6rem}
.addr{font-size:.88rem;color:#374151;display:flex;gap:.4rem;align-items:flex-start}
.addr-icon{flex-shrink:0;margin-top:.1rem}
.items{background:#f8fafc;border-radius:8px;padding:.6rem .85rem;font-size:.83rem;color:#374151;line-height:1.65}
.items strong{display:block;font-size:.72rem;text-transform:uppercase;letter-spacing:.04em;color:#6b7280;margin-bottom:.25rem}
.note-box{background:#fffbeb;border:1px solid #fde68a;border-radius:8px;padding:.6rem .85rem;font-size:.83rem;color:#92400e}
.note-box strong{display:block;font-size:.72rem;text-transform:uppercase;letter-spacing:.04em;color:#b45309;margin-bottom:.2rem}
.actions{display:flex;gap:.5rem;flex-wrap:wrap;padding:.75rem 1rem;border-top:1px solid #f1f5f9;background:#fafafa}
.btn{display:inline-flex;align-items:center;gap:.35rem;padding:.5rem .85rem;border-radius:8px;font-size:.8rem;font-weight:600;text-decoration:none;border:none;cursor:pointer;white-space:nowrap}
.btn-map{background:#1d4ed8;color:white}
.btn-apple{background:#111827;color:white}
.btn-call{background:#059669;color:white}
.btn-done{background:#f0fdf4;color:#15803d;border:1.5px solid #bbf7d0;margin-left:auto}
.btn-done.marked{background:#15803d;color:white;border-color:#15803d}
.done-stamp{display:none;text-align:center;padding:.5rem;font-size:.78rem;font-weight:700;color:#15803d;background:#f0fdf4;border-top:1px solid #bbf7d0}
.stop.done .done-stamp{display:block}
.empty{text-align:center;padding:3rem 1rem;color:#6b7280}
.footer{text-align:center;padding:1.5rem;font-size:.75rem;color:#9ca3af}
</style>
</head>
<body>
<div class="header">
  <h1>🚚 {{ business_name }} — Deliveries</h1>
  <p>{{ sheet_date_fmt }} &nbsp;·&nbsp; {{ stops|length }} stop{{ 's' if stops|length != 1 }}</p>
</div>

{% if stops %}
<div class="stops">
{% for s in stops %}
<div class="stop{% if s.marked %} done{% endif %}" id="stop-{{ s.id }}">
  <div class="stop-header">
    <div class="num">{{ loop.index }}</div>
    <div>
      <div class="cname">{{ s.full_name }}</div>
      <div class="order">Order #{{ s.id }}{% if s.event_start_time %} &nbsp;·&nbsp; ⏰ {{ s.event_start_time }}{% endif %}</div>
    </div>
  </div>
  <div class="stop-body">
    <div class="addr">
      <span class="addr-icon">📍</span>
      <span>{{ s.nav_address }}</span>
    </div>
    {% if s.delivery_location %}
    <div class="addr">
      <span class="addr-icon">📋</span>
      <span><strong>Where:</strong> {{ s.delivery_location }}</span>
    </div>
    {% endif %}
    {% if s.items_summary %}
    <div class="items">
      <strong>Items to deliver</strong>
      {{ s.items_summary }}
    </div>
    {% endif %}
    {% if s.notes %}
    <div class="note-box">
      <strong>Note</strong>{{ s.notes }}
    </div>
    {% endif %}
    <div class="items" style="background:#eff6ff;border-left:3px solid #3b82f6;border-radius:6px;padding:.6rem .85rem">
      <strong style="color:#1d4ed8">On arrival</strong>
      1. Call the customer to let them know you're there.<br>
      2. Deliver items to the location noted above.<br>
      3. Take a quick photo of the setup.<br>
      4. Tap <em>Mark Delivered</em> below when done.
    </div>
  </div>
  <div class="actions">
    {% if s.nav_address %}
    <a class="btn btn-map" href="https://www.google.com/maps/dir/?api=1&destination={{ s.nav_address|urlencode }}&travelmode=driving" target="_blank">🗺 Google Maps</a>
    <a class="btn btn-apple" href="https://maps.apple.com/?daddr={{ s.nav_address|urlencode }}&dirflg=d" target="_blank"> Apple Maps</a>
    {% endif %}
    {% if s.phone %}
    <a class="btn btn-call" href="tel:{{ s.phone }}">📞 Call</a>
    {% endif %}
    <form method="POST" action="/sheet/{{ sheet_date }}/{{ token }}/mark/{{ s.id }}" style="margin-left:auto">
      <button type="submit" class="btn btn-done{% if s.marked %} marked{% endif %}">
        {% if s.marked %}✓ Delivered{% else %}Mark Delivered{% endif %}
      </button>
    </form>
  </div>
  <div class="done-stamp">✓ Delivered</div>
</div>
{% endfor %}
</div>
{% else %}
<div class="empty"><p>No deliveries scheduled for this date.</p></div>
{% endif %}
<div class="footer">{{ business_name }} · Delivery Sheet · {{ sheet_date }}</div>
</body></html>"""


@app.route("/sheet/<sheet_date>/<token>")
def public_delivery_sheet(sheet_date, token):
    if token != _sheet_token(sheet_date):
        return "<h2 style='font-family:sans-serif;padding:2rem;color:#c00'>Invalid or expired link.</h2>", 403
    conn = get_db()
    stops = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("""
                SELECT id, full_name, phone, delivery_location,
                       event_street, event_city, event_state, event_zip,
                       COALESCE(delivery_time, setup_time) AS event_start_time,
                       items_json, notes, delivery_status
                FROM bookings
                WHERE COALESCE(delivery_date, setup_date) = %s
                  AND status = 'accepted'
                  AND payment_status IN ('paid','partial')
                  AND (archived IS NULL OR archived = FALSE)
                ORDER BY COALESCE(delivery_time, setup_time) ASC NULLS LAST, id ASC
            """, (sheet_date,))
            for row in cur.fetchall():
                s = dict(row)
                parts = [
                    (s.get("event_street") or "").strip(),
                    (s.get("event_city")   or "").strip(),
                    (s.get("event_state")  or "").strip(),
                    (s.get("event_zip")    or "").strip(),
                ]
                s["nav_address"] = ", ".join(p for p in parts if p)
                try:
                    items = json.loads(s.get("items_json") or "[]")
                    s["items_summary"] = ", ".join(
                        f"{i.get('qty',1)}x {i.get('name','')}" for i in items
                    )
                except Exception:
                    s["items_summary"] = ""
                s["marked"] = (s.get("delivery_status") == "delivered")
                stops.append(s)
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Delivery sheet error: {e}")
    try:
        dt = datetime.strptime(sheet_date, "%Y-%m-%d")
        sheet_date_fmt = dt.strftime("%A, %B %-d %Y")
    except Exception:
        sheet_date_fmt = sheet_date
    return render_template_string(DELIVERY_SHEET_HTML,
        stops=stops,
        sheet_date=sheet_date,
        sheet_date_fmt=sheet_date_fmt,
        token=token,
        business_name=BUSINESS_NAME,
    )


@app.route("/sheet/<sheet_date>/<token>/mark/<int:booking_id>", methods=["POST"])
def public_mark_delivered(sheet_date, token, booking_id):
    if token != _sheet_token(sheet_date):
        return "Invalid link", 403
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE bookings SET delivery_status='delivered' WHERE id=%s",
                (booking_id,)
            )
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Mark delivered error: {e}")
    return redirect(f"/sheet/{sheet_date}/{token}")


@app.route("/admin/route/override/<int:booking_id>", methods=["POST"])
@admin_required
def admin_route_override(booking_id):
    """Toggle route_override for a booking — forces it onto (or off of) the route."""
    redirect_to = request.form.get("redirect", "/admin/route")
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("SELECT route_override FROM bookings WHERE id=%s", (booking_id,))
            row = cur.fetchone()
            if row:
                new_val = not bool(row[0])
                cur.execute("UPDATE bookings SET route_override=%s WHERE id=%s", (new_val, booking_id))
                conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            log.error(f"Route override error: {e}")
    return redirect(redirect_to)


@app.route("/admin/formsite-import", methods=["GET", "POST"])
@admin_required
def admin_formsite_import():
    return redirect(url_for("admin_dashboard"))



# ══════════════════════════════════════════════════════════════════════════════
#  REVENUE REPORTS
# ══════════════════════════════════════════════════════════════════════════════

ADMIN_REPORTS_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <title>Reports — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;background:#f0f2f5;color:#1a202c;min-height:100vh;display:flex}
    .sidebar{width:200px;min-height:100vh;background:#1e1e2e;display:flex;flex-direction:column;position:fixed;top:0;left:0;z-index:100;transition:transform .2s}
    .sb-brand{padding:1.1rem 1rem .9rem;display:flex;align-items:center;gap:.55rem;border-bottom:1px solid rgba(255,255,255,.08)}
    .sb-brand img{height:1.8rem;width:auto;object-fit:contain}
    .sb-brand-name{font-size:.82rem;font-weight:700;color:#fff;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}
    .sb-new-btn{display:block;margin:.85rem .85rem .5rem;padding:.55rem .75rem;background:#16a34a;color:white;border-radius:8px;font-size:.84rem;font-weight:700;text-decoration:none;text-align:center}
    .sb-new-btn:hover{background:#15803d}
    .sb-nav{display:flex;flex-direction:column;padding:.25rem 0;flex:1}
    .sb-link{display:flex;align-items:center;gap:.6rem;padding:.6rem 1rem;font-size:.84rem;font-weight:500;color:rgba(255,255,255,.55);text-decoration:none;transition:all .1s;border-left:3px solid transparent}
    .sb-link:hover,.sb-link.active{background:rgba(255,255,255,.07);color:rgba(255,255,255,.9)}
    .sb-link.active{border-left-color:#3b82f6;color:white}
    .sb-icon{width:1.1rem;text-align:center;font-size:.95rem}
    .sb-divider{height:1px;background:rgba(255,255,255,.07);margin:.4rem 0}
    .sb-bottom{border-top:1px solid rgba(255,255,255,.08);padding:.5rem 0}
    .page-content{margin-left:200px;flex:1;min-height:100vh;display:flex;flex-direction:column}
    .page-header{background:white;border-bottom:1px solid #e5e7eb;padding:.85rem 1.5rem;display:flex;align-items:center;gap:1rem;position:sticky;top:0;z-index:50}
    .page-header h1{font-size:1.3rem;font-weight:700;color:#111827;flex:1}
    .mobile-menu-btn{display:none;background:none;border:none;font-size:1.4rem;cursor:pointer;color:#374151;padding:.25rem}
    .page-body{padding:1.5rem;flex:1}
    @media(max-width:768px){.sidebar{transform:translateX(-100%)}.sidebar.open{transform:translateX(0);box-shadow:6px 0 30px rgba(0,0,0,.4)}.page-content{margin-left:0}.mobile-menu-btn{display:block}}
    .stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1rem;margin-bottom:1.5rem}
    .stat-card{background:white;border-radius:12px;padding:1.25rem 1.5rem;box-shadow:0 1px 3px rgba(0,0,0,.06)}
    .stat-label{font-size:.75rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.05em}
    .stat-value{font-size:1.8rem;font-weight:800;color:#111827;margin-top:.25rem}
    .stat-sub{font-size:.78rem;color:#9ca3af;margin-top:.15rem}
    .card{background:white;border-radius:12px;padding:1.25rem 1.5rem;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:1.25rem}
    .card h2{font-size:1rem;font-weight:700;color:#111827;margin-bottom:1rem;padding-bottom:.6rem;border-bottom:1px solid #f3f4f6}
    table{width:100%;border-collapse:collapse}
    th{text-align:left;font-size:.72rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;padding:.5rem .75rem;border-bottom:2px solid #f3f4f6}
    td{padding:.65rem .75rem;font-size:.88rem;border-bottom:1px solid #f9fafb;color:#374151}
    tr:last-child td{border-bottom:none}
    tr:hover td{background:#f9fafb}
    .bar-wrap{background:#f3f4f6;border-radius:99px;height:8px;margin-top:.3rem;overflow:hidden}
    .bar{background:#3b82f6;height:8px;border-radius:99px;transition:width .4s}
    .pill{display:inline-block;padding:.15rem .55rem;border-radius:99px;font-size:.72rem;font-weight:700}
    .pill-green{background:#dcfce7;color:#16a34a}
    .pill-red{background:#fee2e2;color:#dc2626}
    .pill-yellow{background:#fef9c3;color:#92400e}
    .filter-row{display:flex;gap:.75rem;align-items:center;margin-bottom:1.25rem;flex-wrap:wrap}
    .filter-row select,.filter-row input{padding:.45rem .75rem;border:1px solid #d1d5db;border-radius:8px;font-size:.85rem;background:white}
  </style>
</head>
<body>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand">
    <img src="/logo.png" alt="">
    <span class="sb-brand-name">{{ business_name }}</span>
  </div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link"><span class="sb-icon">🏠</span> Dashboard</a>
    <a href="/admin/calendar" class="sb-link"><span class="sb-icon">📅</span> Calendar</a>
    <a href="/admin/reports" class="sb-link"><span class="sb-icon">📊</span> Reports</a>
    <a href="/admin/reports" class="sb-link active"><span class="sb-icon">📊</span> Reports</a>
    <div class="sb-divider"></div>
    <a href="/admin/inventory" class="sb-link"><span class="sb-icon">📦</span> Inventory</a>
    <a href="/admin/customers" class="sb-link"><span class="sb-icon">👥</span> Customers</a>
    <div class="sb-divider"></div>
    <a href="/driver/{{ today }}" class="sb-link"><span class="sb-icon">🚚</span> Driver View</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link"><span class="sb-icon">🚪</span> Log Out</a>
  </div>
</aside>
<div class="page-content">
  <div class="page-header">
    <button class="mobile-menu-btn" onclick="document.getElementById('sidebar').classList.toggle('open')">☰</button>
    <h1>📊 Revenue Reports</h1>
    <form method="get" style="display:flex;gap:.5rem;align-items:center">
      <select name="year" onchange="this.form.submit()">
        {% for y in years %}<option value="{{ y }}" {% if y==sel_year %}selected{% endif %}>{{ y }}</option>{% endfor %}
      </select>
    </form>
  </div>
  <div class="page-body">

    <!-- Summary cards -->
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-label">Total Revenue {{ sel_year }}</div>
        <div class="stat-value">${{ "{:,.0f}".format(yearly_total) }}</div>
        <div class="stat-sub">{{ yearly_bookings }} bookings</div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Collected</div>
        <div class="stat-value" style="color:#16a34a">${{ "{:,.0f}".format(collected) }}</div>
        <div class="stat-sub">paid in full or deposit</div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Outstanding</div>
        <div class="stat-value" style="color:#dc2626">${{ "{:,.0f}".format(outstanding) }}</div>
        <div class="stat-sub">accepted but unpaid</div>
      </div>
      <div class="stat-card">
        <div class="stat-label">Avg Booking Value</div>
        <div class="stat-value">${{ "{:,.0f}".format(avg_value) }}</div>
        <div class="stat-sub">{{ sel_year }}</div>
      </div>
    </div>

    <!-- Monthly breakdown -->
    <div class="card">
      <h2>Monthly Breakdown — {{ sel_year }}</h2>
      <table>
        <thead>
          <tr>
            <th>Month</th>
            <th>Bookings</th>
            <th>Revenue</th>
            <th>Collected</th>
            <th>Outstanding</th>
            <th style="width:160px">Bar</th>
          </tr>
        </thead>
        <tbody>
          {% set max_rev = monthly|map(attribute='revenue')|max|default(1) %}
          {% for row in monthly %}
          <tr>
            <td style="font-weight:600">{{ row.month_name }}</td>
            <td>{{ row.count }}</td>
            <td>${{ "{:,.0f}".format(row.revenue) }}</td>
            <td><span class="pill pill-green">${{ "{:,.0f}".format(row.collected) }}</span></td>
            <td>{% if row.outstanding > 0 %}<span class="pill pill-red">${{ "{:,.0f}".format(row.outstanding) }}</span>{% else %}<span style="color:#9ca3af">—</span>{% endif %}</td>
            <td>
              <div class="bar-wrap"><div class="bar" style="width:{{ (row.revenue / max_rev * 100)|int if max_rev else 0 }}%"></div></div>
            </td>
          </tr>
          {% else %}
          <tr><td colspan="6" style="text-align:center;color:#9ca3af;padding:2rem">No bookings for {{ sel_year }}</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    <!-- Top customers -->
    <div class="card">
      <h2>Top Customers — {{ sel_year }}</h2>
      <table>
        <thead>
          <tr>
            <th>#</th>
            <th>Customer</th>
            <th>Bookings</th>
            <th>Total Spent</th>
          </tr>
        </thead>
        <tbody>
          {% for row in top_customers %}
          <tr>
            <td style="color:#9ca3af;font-size:.8rem">{{ loop.index }}</td>
            <td style="font-weight:600">{{ row.customer_name }}</td>
            <td>{{ row.count }}</td>
            <td>${{ "{:,.0f}".format(row.total) }}</td>
          </tr>
          {% else %}
          <tr><td colspan="4" style="text-align:center;color:#9ca3af;padding:2rem">No data</td></tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

  </div>
</div>
</body>
</html>
"""


@app.route("/admin/reports")
@admin_required
def admin_reports():
    import calendar as cal_mod
    sel_year = int(request.args.get("year", date.today().year))

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Year range for selector
    cur.execute("""
        SELECT DISTINCT EXTRACT(YEAR FROM COALESCE(delivery_date, setup_date, event_start_date))::int AS y
        FROM bookings WHERE COALESCE(delivery_date, setup_date, event_start_date) IS NOT NULL
        ORDER BY y DESC
    """)
    years = [r["y"] for r in cur.fetchall() if r["y"]]
    if sel_year not in years:
        years = sorted(set(years + [sel_year]), reverse=True)

    # Monthly breakdown
    cur.execute("""
        SELECT
            EXTRACT(MONTH FROM COALESCE(delivery_date, setup_date, event_start_date))::int AS month,
            COUNT(*) AS count,
            COALESCE(SUM(CASE WHEN status NOT IN ('denied','cancelled') THEN grand_total ELSE 0 END), 0) AS revenue,
            COALESCE(SUM(CASE WHEN payment_status = 'paid' THEN grand_total
                              WHEN payment_status = 'partial' THEN COALESCE(amount_paid, 0) ELSE 0 END), 0) AS collected,
            COALESCE(SUM(CASE WHEN status = 'accepted' AND (payment_status IS NULL OR payment_status = 'waiting')
                              THEN grand_total ELSE 0 END), 0) AS outstanding
        FROM bookings
        WHERE EXTRACT(YEAR FROM COALESCE(delivery_date, setup_date, event_start_date)) = %s
          AND status NOT IN ('denied','cancelled')
          AND (archived IS NULL OR archived = FALSE)
        GROUP BY month
        ORDER BY month
    """, (sel_year,))
    rows = cur.fetchall()
    monthly = []
    for r in rows:
        m = int(r["month"])
        monthly.append({
            "month": m,
            "month_name": cal_mod.month_name[m],
            "count": r["count"],
            "revenue": float(r["revenue"]),
            "collected": float(r["collected"]),
            "outstanding": float(r["outstanding"]),
        })

    # Yearly totals
    yearly_total   = sum(r["revenue"] for r in monthly)
    collected      = sum(r["collected"] for r in monthly)
    outstanding    = sum(r["outstanding"] for r in monthly)
    yearly_bookings = sum(r["count"] for r in monthly)
    avg_value      = yearly_total / yearly_bookings if yearly_bookings else 0

    # Top customers
    cur.execute("""
        SELECT full_name, COUNT(*) AS count,
               COALESCE(SUM(grand_total), 0) AS total
        FROM bookings
        WHERE EXTRACT(YEAR FROM COALESCE(delivery_date, setup_date, event_start_date)) = %s
          AND status NOT IN ('denied','cancelled')
          AND (archived IS NULL OR archived = FALSE)
        GROUP BY full_name
        ORDER BY total DESC
        LIMIT 10
    """, (sel_year,))
    top_customers = [{"customer_name": r["full_name"], "count": r["count"], "total": float(r["total"])}
                     for r in cur.fetchall()]

    cur.close()
    return render_template_string(ADMIN_REPORTS_HTML,
        business_name=BUSINESS_NAME,
        sel_year=sel_year,
        years=years,
        monthly=monthly,
        yearly_total=yearly_total,
        collected=collected,
        outstanding=outstanding,
        yearly_bookings=yearly_bookings,
        avg_value=avg_value,
        top_customers=top_customers,
        today=date.today().isoformat(),
    )




# ══════════════════════════════════════════════════════════════════════════════
#  DRIVER VIEW  (mobile-optimised, no admin login required but token-protected)
# ══════════════════════════════════════════════════════════════════════════════

DRIVER_VIEW_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
  <title>🚚 Driver Sheet — {{ date_label }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#0f172a;color:#f1f5f9;min-height:100vh}
    .topbar{background:#1e293b;padding:1rem 1.25rem;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:10;border-bottom:1px solid #334155}
    .topbar h1{font-size:1rem;font-weight:700;color:#f8fafc}
    .topbar span{font-size:.78rem;color:#94a3b8}
    .date-nav{display:flex;gap:.5rem;padding:.75rem 1.25rem;background:#1e293b;border-bottom:1px solid #334155}
    .date-nav a{flex:1;padding:.5rem;text-align:center;background:#334155;color:#cbd5e1;border-radius:8px;font-size:.82rem;font-weight:600;text-decoration:none}
    .date-nav a:hover{background:#475569}
    .date-nav span{flex:2;text-align:center;padding:.5rem;font-size:.88rem;font-weight:700;color:#f1f5f9}
    .empty{text-align:center;padding:3rem 1.5rem;color:#64748b}
    .empty .icon{font-size:3rem;margin-bottom:.75rem}
    .stop-card{margin:.75rem 1.25rem;background:#1e293b;border-radius:14px;overflow:hidden;border:1px solid #334155}
    .stop-card.delivered{opacity:.55;border-color:#1e3a2e}
    .stop-header{padding:.85rem 1rem;background:#263148;display:flex;align-items:center;gap:.75rem}
    .stop-num{width:2rem;height:2rem;border-radius:50%;background:#3b82f6;color:white;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:.88rem;flex-shrink:0}
    .stop-card.delivered .stop-num{background:#16a34a}
    .stop-name{font-size:1rem;font-weight:700;color:#f8fafc;flex:1}
    .stop-time{font-size:.8rem;color:#94a3b8;font-weight:600}
    .stop-body{padding:.85rem 1rem}
    .stop-addr{font-size:.9rem;color:#cbd5e1;margin-bottom:.6rem;line-height:1.4}
    .stop-phone{font-size:.88rem;color:#60a5fa;margin-bottom:.6rem}
    .stop-phone a{color:#60a5fa;text-decoration:none;font-weight:600}
    .stop-items{font-size:.82rem;color:#94a3b8;margin-bottom:.85rem;line-height:1.5}
    .stop-items strong{color:#cbd5e1}
    .stop-actions{display:flex;gap:.6rem}
    .btn-map{flex:1;padding:.65rem;background:#334155;color:#cbd5e1;border-radius:10px;font-size:.82rem;font-weight:700;text-align:center;text-decoration:none;display:block}
    .btn-map:hover{background:#475569}
    .btn-call{padding:.65rem 1rem;background:#1d4ed8;color:white;border-radius:10px;font-size:.82rem;font-weight:700;text-decoration:none}
    .btn-done{flex:1;padding:.65rem;background:#16a34a;color:white;border-radius:10px;font-size:.82rem;font-weight:700;border:none;cursor:pointer;text-align:center}
    .btn-done:active{background:#15803d}
    .btn-undone{flex:1;padding:.65rem;background:#374151;color:#9ca3af;border-radius:10px;font-size:.82rem;font-weight:700;border:none;cursor:pointer}
    .summary{margin:0 1.25rem .75rem;background:#1e293b;border-radius:10px;padding:.75rem 1rem;display:flex;gap:1.5rem;border:1px solid #334155}
    .summary-item{text-align:center}
    .summary-item .val{font-size:1.4rem;font-weight:800;color:#f1f5f9}
    .summary-item .lbl{font-size:.68rem;color:#64748b;text-transform:uppercase;font-weight:600}
  </style>
</head>
<body>
<div class="topbar">
  <h1>🚚 Driver Sheet</h1>
  <span>{{ date_label }}</span>
</div>
<div class="date-nav">
  <a href="/driver/{{ prev_date }}">← {{ prev_label }}</a>
  <span>{{ short_label }}</span>
  <a href="/driver/{{ next_date }}">{{ next_label }} →</a>
</div>

{% if stops %}
<div class="summary" style="margin-top:.75rem">
  <div class="summary-item"><div class="val">{{ stops|length }}</div><div class="lbl">Stops</div></div>
  <div class="summary-item"><div class="val">{{ delivered_count }}</div><div class="lbl">Done</div></div>
  <div class="summary-item"><div class="val">{{ stops|length - delivered_count }}</div><div class="lbl">Left</div></div>
</div>
{% for s in stops %}
<div class="stop-card {% if s.delivered %}delivered{% endif %}" id="card-{{ s.id }}">
  <div class="stop-header">
    <div class="stop-num">{% if s.delivered %}✓{% else %}{{ loop.index }}{% endif %}</div>
    <div class="stop-name">{{ s.customer_name }}</div>
    {% if s.time_display %}<div class="stop-time">{{ s.time_display }}</div>{% endif %}
  </div>
  <div class="stop-body">
    {% if s.address %}
    <div class="stop-addr">📍 {{ s.address }}</div>
    {% endif %}
    {% if s.phone %}
    <div class="stop-phone">📞 <a href="tel:{{ s.phone }}">{{ s.phone }}</a></div>
    {% endif %}
    {% if s.items_summary %}
    <div class="stop-items"><strong>Items:</strong> {{ s.items_summary }}</div>
    {% endif %}
    <div class="stop-actions">
      {% if s.address %}
      {% if s.maps_url %}<a class="btn-map" href="{{ s.maps_url }}" target="_blank">🗺 Maps</a>{% endif %}
      {% endif %}
      {% if s.phone %}
      <a class="btn-call" href="tel:{{ s.phone }}">Call</a>
      {% endif %}
      {% if not s.delivered %}
      <button class="btn-done" onclick="markDelivered({{ s.id }}, this)">✓ Mark Done</button>
      {% else %}
      <button class="btn-undone" onclick="markDelivered({{ s.id }}, this)">Undo</button>
      {% endif %}
    </div>
  </div>
</div>
{% endfor %}
{% else %}
<div class="empty">
  <div class="icon">📭</div>
  <div style="font-size:1rem;font-weight:700;color:#cbd5e1;margin-bottom:.4rem">No deliveries today</div>
  <div>Nothing scheduled for {{ date_label }}</div>
</div>
{% endif %}

<script>
function markDelivered(id, btn){
  fetch('/driver/' + id + '/toggle', {method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({date:'{{ date_str }}'})
  }).then(function(r){return r.json();}).then(function(d){
    location.reload();
  }).catch(function(e){alert('Error: '+e);});
}
</script>
</body>
</html>
"""



@app.route("/driver/")
@app.route("/driver")
def driver_today():
    """Redirect /driver/ to today's date."""
    token = request.args.get("token", "")
    suffix = f"?token={token}" if token else ""
    return redirect(f"/driver/{date.today().isoformat()}{suffix}")


@app.route("/driver/<date_str>")
def driver_view(date_str):
    """Mobile driver sheet — accessible without admin login if CALENDAR_TOKEN matches, or if admin session active."""
    token = request.args.get("token", "")
    if not (session.get("admin_logged_in") or token == CALENDAR_TOKEN):
        # Return a simple token-prompt page
        return render_template_string("""<!DOCTYPE html><html><body style='font-family:sans-serif;background:#0f172a;color:#f1f5f9;display:flex;align-items:center;justify-content:center;min-height:100vh'>
        <div style='text-align:center'>
          <div style='font-size:2rem;margin-bottom:1rem'>🔒</div>
          <p>Enter access token to view driver sheet.</p>
          <form method='get'>
            <input name='token' placeholder='Token' style='padding:.5rem;border-radius:8px;border:1px solid #334155;background:#1e293b;color:#f1f5f9;margin:.5rem'>
            <input type='hidden' name='' value=''>
            <button type='submit' style='padding:.5rem 1rem;background:#2563eb;color:white;border:none;border-radius:8px;cursor:pointer;margin:.5rem'>Go</button>
          </form>
        </div></body></html>"""), 401

    try:
        view_date = date.fromisoformat(date_str)
    except ValueError:
        view_date = date.today()
        date_str = view_date.isoformat()

    prev_d = view_date - timedelta(days=1)
    next_d = view_date + timedelta(days=1)

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id, full_name, phone,
               COALESCE(delivery_date, setup_date, event_start_date) AS del_date,
               COALESCE(delivery_time, setup_time) AS del_time,
               event_street, event_city, event_state, event_zip,
               items_json, status, delivered_at
        FROM bookings
        WHERE COALESCE(delivery_date, setup_date, event_start_date) = %s
          AND status IN ('accepted','concluded')
          AND (archived IS NULL OR archived = FALSE)
        ORDER BY COALESCE(delivery_time, setup_time) ASC NULLS LAST
    """, (date_str,))
    rows = cur.fetchall()
    cur.close()

    stops = []
    for b in rows:
        time_display = ""
        if b["del_time"]:
            try:
                t = datetime.strptime(b["del_time"], "%H:%M")
                time_display = t.strftime("%-I:%M %p")
            except Exception:
                time_display = b["del_time"]
        # Build address from event location fields (delivery_location is a description, not address)
        addr_parts = [
            b.get("event_street",""),
            b.get("event_city",""),
            b.get("event_state",""),
            b.get("event_zip",""),
        ]
        addr = ", ".join(p for p in addr_parts if p)
        # Parse items from JSON
        try:
            import json as _j
            items_list = _j.loads(b.get("items_json") or "[]")
            items_raw = ", ".join(
                f"{it.get('qty','1')}x {it.get('name','')}" for it in items_list
                if it.get('name')
            )
        except Exception:
            items_raw = ""
        if len(items_raw) > 120:
            items_raw = items_raw[:117] + "…"
        maps_url = ""
        if addr:
            maps_url = "https://www.google.com/maps/search/?api=1&query=" + urllib.parse.quote_plus(addr)
        stops.append({
            "id": b["id"],
            "customer_name": b["full_name"],
            "phone": b["phone"] or "",
            "address": addr,
            "maps_url": maps_url,
            "time_display": time_display,
            "items_summary": items_raw,
            "delivered": bool(b.get("delivered_at")),
        })

    delivered_count = sum(1 for s in stops if s["delivered"])

    def fmt_date(d):
        today = date.today()
        if d == today:
            return "Today"
        if d == today + timedelta(days=1):
            return "Tomorrow"
        if d == today - timedelta(days=1):
            return "Yesterday"
        return d.strftime("%b %-d")

    token_param = f"?token={token}" if token else ""
    return render_template_string(DRIVER_VIEW_HTML,
        date_str=date_str,
        date_label=view_date.strftime("%A, %B %-d, %Y"),
        short_label=fmt_date(view_date),
        prev_date=prev_d.isoformat() + token_param,
        next_date=next_d.isoformat() + token_param,
        prev_label=fmt_date(prev_d),
        next_label=fmt_date(next_d),
        stops=stops,
        delivered_count=delivered_count,
    )


@app.route("/driver/<int:booking_id>/toggle", methods=["POST"])
def driver_toggle_delivered(booking_id):
    """Toggle delivered_at for a booking from the driver view."""
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT delivered_at FROM bookings WHERE id=%s", (booking_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        return jsonify({"error": "not found"}), 404
    if row["delivered_at"]:
        cur.execute("UPDATE bookings SET delivered_at=NULL WHERE id=%s", (booking_id,))
    else:
        cur.execute("UPDATE bookings SET delivered_at=NOW() WHERE id=%s", (booking_id,))
    conn.commit()
    cur.close()
    return jsonify({"ok": True})



# ══════════════════════════════════════════════════════════════════════════════
#  RENTAL AGREEMENT / E-SIGNATURE
# ══════════════════════════════════════════════════════════════════════════════

def _sign_token(booking_id):
    """Generate a deterministic token for the signing link."""
    raw = f"{booking_id}:{app.secret_key}:sign"
    return hashlib.sha256(raw.encode()).hexdigest()[:32]


AGREEMENT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Rental Agreement — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f4f8;color:#1a202c;padding:1.5rem 1rem 3rem;max-width:680px;margin:0 auto}
    h1{font-size:1.4rem;font-weight:800;margin-bottom:.25rem;color:#111827}
    .sub{font-size:.9rem;color:#6b7280;margin-bottom:1.75rem}
    .booking-box{background:white;border-radius:12px;padding:1.25rem 1.5rem;margin-bottom:1.5rem;border:1px solid #e5e7eb}
    .booking-box h2{font-size:.88rem;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.75rem}
    .row{display:flex;gap:1rem;margin-bottom:.4rem;font-size:.88rem}
    .row .lbl{width:130px;color:#9ca3af;flex-shrink:0}
    .row .val{color:#111827;font-weight:600}
    .agreement-text{background:white;border-radius:12px;padding:1.25rem 1.5rem;border:1px solid #e5e7eb;margin-bottom:1.5rem;max-height:320px;overflow-y:auto;font-size:.83rem;line-height:1.7;color:#374151}
    .agreement-text h3{font-size:.9rem;font-weight:700;color:#111827;margin-bottom:.5rem}
    .agreement-text p{margin-bottom:.75rem}
    .sign-box{background:white;border-radius:12px;padding:1.5rem;border:1px solid #e5e7eb}
    .sign-box h2{font-size:1rem;font-weight:700;margin-bottom.75rem;color:#111827;margin-bottom:.75rem}
    label{display:block;font-size:.82rem;font-weight:600;color:#374151;margin-bottom:.35rem;margin-top:.85rem}
    input[type=text]{width:100%;padding:.6rem .85rem;border:1px solid #d1d5db;border-radius:8px;font-size:.95rem;color:#111827}
    input[type=text]:focus{outline:2px solid #2563eb;border-color:#2563eb}
    .checkbox-row{display:flex;align-items:flex-start;gap:.75rem;margin-top:1rem}
    .checkbox-row input{width:1.1rem;height:1.1rem;margin-top:.1rem;flex-shrink:0;cursor:pointer}
    .checkbox-row label{font-size:.85rem;color:#374151;font-weight:400;margin:0}
    .btn-sign{width:100%;margin-top:1.25rem;padding:.85rem;background:#16a34a;color:white;font-size:1rem;font-weight:700;border:none;border-radius:10px;cursor:pointer}
    .btn-sign:hover{background:#15803d}
    .btn-sign:disabled{background:#9ca3af;cursor:not-allowed}
    .signed-banner{background:#dcfce7;border:1px solid #bbf7d0;border-radius:12px;padding:1.5rem;text-align:center;margin-bottom:1.5rem}
    .signed-banner .icon{font-size:2.5rem;margin-bottom:.5rem}
    .signed-banner h2{font-size:1.1rem;font-weight:800;color:#15803d}
    .signed-banner p{font-size:.85rem;color:#166534;margin-top:.35rem}
  </style>
</head>
<body>
<h1>📋 Rental Agreement</h1>
<p class="sub">{{ business_name }} · Booking #{{ b.id }}</p>

{% if b.agreement_signed %}
<div class="signed-banner">
  <div class="icon">✅</div>
  <h2>Agreement Signed</h2>
  <p>Signed by <strong>{{ b.agreement_signer_name or '' }}</strong><br>
  {% if b.agreement_signed_at %}on {{ b.agreement_signed_at.strftime('%B %-d, %Y at %-I:%M %p') }}{% endif %}</p>
</div>
{% endif %}

<div class="booking-box">
  <h2>Booking Details</h2>
  <div class="row"><span class="lbl">Customer</span><span class="val">{{ b.full_name }}</span></div>
  <div class="row"><span class="lbl">Event Date</span><span class="val">{{ b.event_start_date.strftime('%B %-d, %Y') if b.event_start_date else '—' }}</span></div>
  <div class="row"><span class="lbl">Location</span><span class="val">{{ b.setup_address or '—' }}</span></div>
  <div class="row"><span class="lbl">Total</span><span class="val">${{ "%.2f"|format(b.total_cost or 0) }}</span></div>
</div>

<div class="agreement-text">
  <h3>Terms & Conditions — {{ business_name }}</h3>
  <p><strong>1. Rental Period.</strong> Equipment is rented for the dates specified in this booking. Late returns may incur additional fees.</p>
  <p><strong>2. Deposit & Payment.</strong> A deposit is required to secure your booking. The remaining balance is due prior to or on the delivery date. Failure to pay may result in cancellation without refund of the deposit.</p>
  <p><strong>3. Damage & Loss.</strong> The renter is responsible for any damage to or loss of equipment during the rental period. Normal wear and tear is excluded. Replacement costs will be charged for lost or severely damaged items.</p>
  <p><strong>4. Setup & Conditions.</strong> The renter must ensure a safe and accessible setup area. {{ business_name }} reserves the right to decline setup if conditions are unsafe.</p>
  <p><strong>5. Cancellation.</strong> Cancellations made less than 7 days before the event are non-refundable. Cancellations 7+ days in advance may receive a credit toward a future booking.</p>
  <p><strong>6. Weather.</strong> Outdoor rentals are subject to weather conditions. {{ business_name }} may cancel or modify delivery in cases of severe weather. No refunds are issued for weather events outside our control.</p>
  <p><strong>7. Indemnification.</strong> The renter agrees to hold {{ business_name }} harmless from any injury, loss, or liability arising from the use of rented equipment.</p>
  <p><strong>8. Governing Law.</strong> This agreement is governed by the laws of the state in which {{ business_name }} operates.</p>
</div>

{% if not b.agreement_signed %}
<div class="sign-box">
  <h2>Sign to Confirm</h2>
  <form method="post">
    <label>Your Full Name</label>
    <input type="text" name="signer_name" placeholder="Type your full legal name" required>
    <div class="checkbox-row">
      <input type="checkbox" id="agree_check" required>
      <label for="agree_check">I have read and agree to the Rental Agreement terms above. I understand this is a legally binding agreement.</label>
    </div>
    <button type="submit" class="btn-sign">✍️ I Agree &amp; Sign</button>
  </form>
</div>
{% endif %}
</body>
</html>
"""


@app.route("/booking/sign/<int:booking_id>/<token>", methods=["GET","POST"])
def booking_sign(booking_id, token):
    if token != _sign_token(booking_id):
        return "Invalid or expired link.", 403

    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
    b = cur.fetchone()
    if not b:
        cur.close()
        return "Booking not found.", 404

    if request.method == "POST" and not b["agreement_signed"]:
        signer_name = request.form.get("signer_name", "").strip()
        if signer_name:
            cur.execute("""
                UPDATE bookings
                SET agreement_signed=TRUE, agreement_signed_at=NOW(), agreement_signer_name=%s
                WHERE id=%s
            """, (signer_name, booking_id))
            conn.commit()
            cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
            b = cur.fetchone()

    cur.close()
    return render_template_string(AGREEMENT_HTML,
        business_name=BUSINESS_NAME,
        b=b,
    )


@app.route("/admin/booking/<int:booking_id>/send-agreement", methods=["POST"])
@admin_required
def admin_send_agreement(booking_id):
    """Send the e-sign link to the customer via email."""
    conn = get_db()
    cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM bookings WHERE id=%s", (booking_id,))
    b = cur.fetchone()
    cur.close()
    if not b:
        return "Not found", 404

    token = _sign_token(booking_id)
    base  = os.environ.get("APP_BASE_URL", BASE_URL).rstrip("/")
    sign_url = f"{base}/booking/sign/{booking_id}/{token}"

    # Send email
    if b.get("email") and GMAIL_USER:
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"Please Sign Your Rental Agreement — {BUSINESS_NAME}"
            msg["From"]    = GMAIL_USER
            msg["To"]      = b["email"]
            html_body = f"""
            <div style="font-family:sans-serif;max-width:600px;margin:0 auto;padding:2rem">
              <h2 style="color:#111827">Rental Agreement Ready to Sign</h2>
              <p style="color:#374151">Hi {b['full_name'].split()[0]},</p>
              <p style="color:#374151;margin-top:.75rem">Your booking with {BUSINESS_NAME} is almost confirmed.
              Please review and sign your rental agreement to complete the process.</p>
              <div style="text-align:center;margin:2rem 0">
                <a href="{sign_url}" style="background:#16a34a;color:white;padding:.85rem 2rem;border-radius:10px;font-weight:700;text-decoration:none;font-size:1rem">
                  ✍️ Review &amp; Sign Agreement
                </a>
              </div>
              <p style="color:#6b7280;font-size:.85rem">This link is unique to your booking. If you have questions, reply to this email or call {BUSINESS_PHONE or BUSINESS_NAME}.</p>
            </div>"""
            msg.attach(MIMEText(html_body, "html"))
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
                s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                s.send_message(msg)
        except Exception as e:
            log.error(f"Agreement email error: {e}")

    # Also send SMS if phone
    if b.get("phone"):
        send_sms(b["phone"], f"Hi {b['full_name'].split()[0]}! Please sign your rental agreement for {BUSINESS_NAME}: {sign_url}")

    return redirect(url_for("admin_booking_detail", booking_id=booking_id))


ADMIN_CUSTOMERS_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Customers — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#111827;min-height:100vh}
    .topbar{background:white;border-bottom:1px solid #e5e7eb;padding:.9rem 1.25rem;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:50;flex-wrap:wrap;gap:.4rem}
    .topbar-brand{font-size:1rem;font-weight:700;color:#111827}
    .topbar-nav{display:flex;gap:.25rem;align-items:center;overflow-x:auto;-webkit-overflow-scrolling:touch;flex-wrap:nowrap;max-width:100%}
    .nav-link{color:#6b7280;font-size:.85rem;font-weight:500;text-decoration:none;padding:.38rem .65rem;border-radius:6px;transition:all .12s;white-space:nowrap}
    .nav-link:hover{background:#f3f4f6;color:#111827}
    .nav-link.active{background:#eff6ff;color:#2563eb;font-weight:600}
    .logout-btn{background:white;border:1px solid #d1d5db;color:#6b7280;padding:.38rem .65rem;border-radius:6px;font-size:.82rem;font-weight:500;text-decoration:none;white-space:nowrap}
    .main{max-width:1100px;margin:0 auto;padding:1.75rem}
    .top-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:1.25rem}
    .page-title{font-size:1.4rem;font-weight:700;color:#111827}
    .flash{padding:.75rem 1rem;border-radius:8px;margin-bottom:1.25rem;font-size:.9rem;font-weight:500}
    .flash-ok{background:#dcfce7;color:#166534;border:1px solid #bbf7d0}
    .flash-err{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}

    /* Add form panel */
    .add-panel{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:1.25rem 1.5rem;margin-bottom:1.5rem;display:none}
    .add-panel.open{display:block}
    .add-panel h3{font-size:.95rem;font-weight:700;color:#374151;margin-bottom:1rem}
    .form-grid{display:grid;grid-template-columns:1fr 1fr;gap:.75rem}
    .form-grid.three{grid-template-columns:1fr 1fr 1fr}
    .form-group{display:flex;flex-direction:column;gap:.3rem}
    .form-group label{font-size:.75rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.4px}
    .form-group input,.form-group textarea{padding:.45rem .65rem;border:1px solid #d1d5db;border-radius:6px;font-size:.86rem;color:#111827;font-family:inherit}
    .form-group input:focus,.form-group textarea:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,.1)}
    .form-group textarea{resize:vertical;min-height:60px}
    .form-actions{display:flex;gap:.6rem;justify-content:flex-end;margin-top:.75rem}

    /* Table */
    .card{background:white;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden}
    .card-header{padding:.85rem 1.25rem;background:#f9fafb;border-bottom:1px solid #e5e7eb;display:flex;justify-content:space-between;align-items:center}
    .card-title{font-weight:700;font-size:.88rem;color:#374151}
    table{width:100%;border-collapse:collapse}
    thead tr{background:#f9fafb}
    th{padding:.65rem 1rem;text-align:left;font-size:.72rem;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid #e5e7eb}
    td{padding:.85rem 1rem;border-bottom:1px solid #f3f4f6;vertical-align:middle;font-size:.86rem}
    tr:last-child td{border-bottom:none}
    tbody tr:hover td{background:#fafafa}
    .avatar{width:34px;height:34px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:.8rem;font-weight:700;color:white;flex-shrink:0}
    .client-cell{display:flex;align-items:center;gap:.65rem}
    .client-name{font-weight:600;color:#111827}
    .client-company{font-size:.78rem;color:#9ca3af}
    .btn{display:inline-block;padding:.32rem .7rem;border-radius:6px;font-size:.78rem;font-weight:600;cursor:pointer;border:1px solid transparent;text-decoration:none;line-height:1.5;transition:all .12s}
    .btn-primary{background:#2563eb;color:white;border:none}
    .btn-primary:hover{background:#1d4ed8}
    .btn-outline{background:white;color:#374151;border-color:#d1d5db}
    .btn-outline:hover{background:#f3f4f6}
    .btn-edit{background:#eff6ff;color:#2563eb;border-color:#bfdbfe}
    .btn-edit:hover{background:#dbeafe}
    .btn-danger{background:#fef2f2;color:#991b1b;border-color:#fecaca}
    .btn-danger:hover{background:#fee2e2}
    .action-btns{display:flex;gap:.4rem}
    .empty-state{padding:3rem;text-align:center;color:#9ca3af}
    .search-box{padding:.45rem .75rem;border:1px solid #d1d5db;border-radius:7px;font-size:.86rem;width:220px}
    .search-box:focus{outline:none;border-color:#2563eb}
    @media(max-width:700px){.form-grid,.form-grid.three{grid-template-columns:1fr}.main{padding:1rem}}
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link active">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>Customers</h1>
</div>
<div class="main">
  <div class="top-row">
    <div style="display:flex;gap:.6rem;align-items:center">
      <a href="/admin/booking/new" class="btn btn-primary" style="background:#16a34a">+ New Booking</a>
      <a href="/admin/customers/import" class="btn btn-outline">⬆ Import CSV</a>
      <button class="btn btn-primary" onclick="toggleAdd()">+ Add Customer</button>
    </div>
  </div>

  {% if flash_ok %}<div class="flash flash-ok">✓ {{ flash_ok }}</div>{% endif %}
  {% if flash_err %}<div class="flash flash-err">⚠ {{ flash_err }}</div>{% endif %}

  <!-- Add Customer Panel -->
  <div class="add-panel" id="addPanel">
    <h3>New Customer</h3>
    <form method="POST" action="/admin/customers/add">
      <div class="form-grid">
        <div class="form-group">
          <label>Full Name *</label>
          <input type="text" name="full_name" required>
        </div>
        <div class="form-group">
          <label>Company Name</label>
          <input type="text" name="company_name">
        </div>
        <div class="form-group">
          <label>Email</label>
          <input type="email" name="email">
        </div>
        <div class="form-group">
          <label>Phone</label>
          <input type="tel" name="phone">
        </div>
      </div>
      <div class="form-grid" style="margin-top:.75rem">
        <div class="form-group" style="grid-column:1/-1">
          <label>Street Address</label>
          <input type="text" name="street">
        </div>
        <div class="form-group">
          <label>City</label>
          <input type="text" name="city">
        </div>
        <div class="form-group three">
          <label>State</label>
          <input type="text" name="state" maxlength="2">
        </div>
        <div class="form-group three">
          <label>Zip</label>
          <input type="text" name="zip" maxlength="10">
        </div>
      </div>
      <div class="form-group" style="margin-top:.75rem">
        <label>Notes</label>
        <textarea name="notes" placeholder="Any notes about this customer..."></textarea>
      </div>
      <div style="margin-top:.75rem;padding:.75rem 1rem;background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px">
        <label style="display:flex;align-items:center;gap:.6rem;cursor:pointer;font-weight:600;color:#166534;font-size:.95rem;text-transform:none;letter-spacing:0">
          <input type="checkbox" name="tax_exempt" value="1"
                 style="width:18px;height:18px;accent-color:#16a34a;cursor:pointer">
          Tax Exempt
        </label>
        <p style="margin:.3rem 0 0;font-size:.8rem;color:#4b7c5a">
          Check if this customer has a valid CT tax-exempt certificate. Tax (6.35%) will be removed from all their bookings.
        </p>
      </div>
      <div class="form-actions">
        <button type="button" class="btn btn-outline" onclick="toggleAdd()">Cancel</button>
        <button type="submit" class="btn btn-primary">Save Customer</button>
      </div>
    </form>
  </div>

  <!-- Customer List -->
  <div class="card">
    <div class="card-header">
      <span class="card-title">All Customers ({{ customers|length }})</span>
      <input class="search-box" type="text" id="searchBox" placeholder="Search name or email..." onkeyup="filterCustomers()">
    </div>
    {% if customers %}
    <table id="customerTable">
      <thead>
        <tr>
          <th>Customer</th>
          <th>Email</th>
          <th>Phone</th>
          <th>Location</th>
          <th>Orders</th>
          <th>Actions</th>
        </tr>
      </thead>
      <tbody>
        {% set avatar_colors = ['#ef4444','#f97316','#eab308','#22c55e','#14b8a6','#3b82f6','#8b5cf6','#ec4899','#06b6d4','#84cc16'] %}
        {% for c in customers %}
        {% set color = avatar_colors[loop.index0 % 10] %}
        <tr class="cust-row">
          <td>
            <div class="client-cell">
              <div class="avatar" style="background:{{ color }}">{{ (c.full_name or '?')[0]|upper }}</div>
              <div>
                <div class="client-name cust-name">{{ c.full_name }}</div>
                {% if c.company_name %}<div class="client-company">{{ c.company_name }}</div>{% endif %}
              </div>
            </div>
          </td>
          <td class="cust-email">
            <a href="mailto:{{ c.email }}" style="color:#2563eb;text-decoration:none">{{ c.email or '—' }}</a>
            {% if c.tax_exempt %}<span style="display:inline-block;margin-left:.4rem;font-size:.7rem;background:#dcfce7;color:#166534;border-radius:4px;padding:.1rem .4rem;font-weight:600">TAX EXEMPT</span>{% endif %}
          </td>
          <td><a href="tel:{{ c.phone }}" style="color:#374151;text-decoration:none">{{ c.phone or '—' }}</a></td>
          <td style="color:#6b7280;font-size:.82rem">
            {% if c.city %}{{ c.city }}{% if c.state %}, {{ c.state }}{% endif %}{% else %}—{% endif %}
          </td>
          <td>
            {% if c.bookings %}
            <div style="display:flex;flex-wrap:wrap;gap:.3rem">
              {% for b in c.bookings %}
              <a href="/admin/booking/{{ b.id }}"
                 style="display:inline-flex;align-items:center;gap:.25rem;padding:.2rem .5rem;border-radius:5px;font-size:.75rem;font-weight:600;text-decoration:none;
                        background:{% if b.status=='accepted' and b.payment_status=='paid' %}#dcfce7;color:#166534{% elif b.status=='accepted' and b.payment_status=='partial' %}#ede9fe;color:#7c3aed{% elif b.status=='accepted' %}#dbeafe;color:#1e40af{% elif b.status=='pending' %}#fef9c3;color:#854d0e{% elif b.status=='concluded' %}#e5e7eb;color:#374151{% else %}#f3f4f6;color:#6b7280{% endif %}">
                #{{ b.id }}
              </a>
              {% endfor %}
            </div>
            {% else %}
            <span style="color:#9ca3af;font-size:.8rem">No orders</span>
            {% endif %}
          </td>
          <td>
            <div class="action-btns">
              <a href="/admin/customers/{{ c.id }}/edit" class="btn btn-edit">Edit</a>
              <form method="POST" action="/admin/customers/{{ c.id }}/delete" style="display:inline">
                <button class="btn btn-danger" onclick="return confirm('Remove {{ c.full_name }}?')">Remove</button>
              </form>
            </div>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    {% else %}
    <div class="empty-state">No customers yet. Click "+ Add Customer" to get started.</div>
    {% endif %}
  </div>
</div>
<script>
function toggleAdd(){
  var p=document.getElementById('addPanel');
  p.classList.toggle('open');
  if(p.classList.contains('open')) p.querySelector('input').focus();
}
function filterCustomers(){
  var q=document.getElementById('searchBox').value.toLowerCase();
  document.querySelectorAll('.cust-row').forEach(function(row){
    var name=row.querySelector('.cust-name').textContent.toLowerCase();
    var email=row.querySelector('.cust-email').textContent.toLowerCase();
    row.style.display=(name.includes(q)||email.includes(q))?'':'none';
  });
}
</script>
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""

ADMIN_CUSTOMER_IMPORT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Import Customers — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#111827;min-height:100vh}
    .topbar{background:white;border-bottom:1px solid #e5e7eb;padding:.9rem 1.25rem;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:50;flex-wrap:wrap;gap:.4rem}
    .topbar-brand{font-size:1rem;font-weight:700}
    .topbar-nav{display:flex;gap:.25rem;align-items:center;overflow-x:auto;-webkit-overflow-scrolling:touch;flex-wrap:nowrap;max-width:100%}
    .nav-link{color:#6b7280;font-size:.85rem;font-weight:500;text-decoration:none;padding:.38rem .65rem;border-radius:6px;white-space:nowrap}
    .nav-link:hover{background:#f3f4f6;color:#111827}
    .nav-link.active{background:#eff6ff;color:#2563eb;font-weight:600}
    .logout-btn{background:white;border:1px solid #d1d5db;color:#6b7280;padding:.38rem .65rem;border-radius:6px;font-size:.82rem;font-weight:500;text-decoration:none;white-space:nowrap}
    .main{max-width:750px;margin:0 auto;padding:1.75rem}
    .breadcrumb{font-size:.82rem;color:#9ca3af;margin-bottom:1rem}
    .breadcrumb a{color:#2563eb;text-decoration:none}
    .page-title{font-size:1.3rem;font-weight:700;margin-bottom:.35rem}
    .page-sub{font-size:.88rem;color:#6b7280;margin-bottom:1.5rem}
    .flash{padding:.75rem 1rem;border-radius:8px;margin-bottom:1.25rem;font-size:.9rem;font-weight:500}
    .flash-ok{background:#dcfce7;color:#166534;border:1px solid #bbf7d0}
    .flash-err{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}
    .card{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:1.5rem;margin-bottom:1.25rem}
    .card h3{font-size:.9rem;font-weight:700;color:#374151;text-transform:uppercase;letter-spacing:.4px;margin-bottom:1rem;padding-bottom:.5rem;border-bottom:1px solid #f3f4f6}
    .step{display:flex;gap:1rem;margin-bottom:1rem;align-items:flex-start}
    .step-num{width:28px;height:28px;border-radius:50%;background:#2563eb;color:white;font-size:.82rem;font-weight:700;display:flex;align-items:center;justify-content:center;flex-shrink:0;margin-top:.1rem}
    .step-body{flex:1}
    .step-title{font-weight:600;font-size:.9rem;color:#111827;margin-bottom:.25rem}
    .step-desc{font-size:.84rem;color:#6b7280;line-height:1.5}
    code{background:#f3f4f6;padding:.15rem .4rem;border-radius:4px;font-size:.82rem;font-family:monospace}
    .template-box{background:#f8fafc;border:1px solid #e5e7eb;border-radius:7px;padding:.75rem 1rem;font-family:monospace;font-size:.78rem;color:#374151;overflow-x:auto;white-space:nowrap;margin:.5rem 0}
    .upload-area{border:2px dashed #d1d5db;border-radius:10px;padding:2rem;text-align:center;cursor:pointer;transition:all .15s;background:#fafafa}
    .upload-area:hover,.upload-area.drag{border-color:#2563eb;background:#eff6ff}
    .upload-icon{font-size:2rem;margin-bottom:.5rem}
    .upload-label{font-size:.9rem;font-weight:600;color:#374151;margin-bottom:.25rem}
    .upload-sub{font-size:.8rem;color:#9ca3af}
    input[type=file]{display:none}
    .btn{display:inline-block;padding:.5rem 1.1rem;border-radius:7px;font-size:.86rem;font-weight:600;cursor:pointer;border:none;text-decoration:none;transition:all .12s}
    .btn-primary{background:#2563eb;color:white}
    .btn-primary:hover{background:#1d4ed8}
    .btn-outline{background:white;color:#374151;border:1px solid #d1d5db}
    .btn-outline:hover{background:#f3f4f6}
    .btn-sm{padding:.3rem .7rem;font-size:.78rem}
    .actions{display:flex;gap:.75rem;margin-top:1.25rem;align-items:center}
    table{width:100%;border-collapse:collapse;font-size:.83rem;margin-top:.75rem}
    th{padding:.5rem .75rem;text-align:left;font-size:.72rem;font-weight:600;color:#9ca3af;text-transform:uppercase;border-bottom:1px solid #e5e7eb;background:#f9fafb}
    td{padding:.5rem .75rem;border-bottom:1px solid #f3f4f6;color:#374151}
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link active">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>Import Customers</h1>
  <a href="/admin/customers" class="pg-back">← Customers</a>
</div>
<div class="main">
  <div class="breadcrumb"><a href="/admin/customers">Customers</a> › Import</div>
  <div class="page-title">Import Customers from CSV</div>
  <div class="page-sub">Upload a spreadsheet of customers exported from Booqable or filled in manually.</div>

  {% if flash_ok %}<div class="flash flash-ok">✓ {{ flash_ok }}</div>{% endif %}
  {% if flash_err %}<div class="flash flash-err">⚠ {{ flash_err }}</div>{% endif %}
  {% if results %}
  <div class="flash flash-ok">
    ✓ Import complete — {{ results.added }} added, {{ results.updated }} updated, {{ results.skipped }} skipped.
  </div>
  {% endif %}

  <div class="card">
    <h3>How to Import</h3>
    <div class="step">
      <div class="step-num">1</div>
      <div class="step-body">
        <div class="step-title">Download the template</div>
        <div class="step-desc">
          Click below to download a CSV template with the correct column headers.
          <br><br>
          <a href="/admin/customers/import/template" class="btn btn-outline btn-sm">⬇ Download Template</a>
        </div>
      </div>
    </div>
    <div class="step">
      <div class="step-num">2</div>
      <div class="step-body">
        <div class="step-title">Fill it in</div>
        <div class="step-desc">
          Open the template in Google Sheets or Excel. Paste your Booqable customer names, emails, and phone numbers into the matching columns. The required columns are:
          <div class="template-box">full_name, email, phone, company_name, street, city, state, zip, notes</div>
          Only <code>full_name</code> is required. Everything else is optional.
        </div>
      </div>
    </div>
    <div class="step">
      <div class="step-num">3</div>
      <div class="step-body">
        <div class="step-title">Save as CSV and upload</div>
        <div class="step-desc">In Google Sheets: File → Download → CSV. In Excel: Save As → CSV. Then upload below.</div>
      </div>
    </div>
  </div>

  <div class="card">
    <h3>Upload CSV File</h3>
    <form method="POST" action="/admin/customers/import" enctype="multipart/form-data">
      <div class="upload-area" id="dropZone" onclick="document.getElementById('csvFile').click()"
           ondragover="event.preventDefault();this.classList.add('drag')"
           ondragleave="this.classList.remove('drag')"
           ondrop="event.preventDefault();this.classList.remove('drag');handleDrop(event)">
        <div class="upload-icon">📂</div>
        <div class="upload-label" id="uploadLabel">Click to choose a CSV file</div>
        <div class="upload-sub">or drag and drop here</div>
      </div>
      <input type="file" id="csvFile" name="csvfile" accept=".csv,text/csv" onchange="showFileName(this)">
      <div class="actions">
        <a href="/admin/customers" class="btn btn-outline">Cancel</a>
        <button type="submit" class="btn btn-primary">Import Customers</button>
      </div>
    </form>
  </div>

  {% if preview %}
  <div class="card">
    <h3>Preview (first 5 rows)</h3>
    <table>
      <thead><tr><th>Name</th><th>Email</th><th>Phone</th><th>City</th></tr></thead>
      <tbody>
        {% for r in preview %}
        <tr><td>{{ r.full_name }}</td><td>{{ r.email or '—' }}</td><td>{{ r.phone or '—' }}</td><td>{{ r.city or '—' }}</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}
</div>
<script>
function showFileName(input){
  if(input.files&&input.files[0]){
    document.getElementById('uploadLabel').textContent='Selected: '+input.files[0].name;
  }
}
function handleDrop(e){
  const file=e.dataTransfer.files[0];
  if(file){
    const dt=new DataTransfer();dt.items.add(file);
    const inp=document.getElementById('csvFile');
    inp.files=dt.files;
    showFileName(inp);
  }
}
</script>
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""

ADMIN_CUSTOMER_EDIT_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <link rel="manifest" href="/admin-manifest.json">
  <meta name="theme-color" content="#2563eb">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="default">
  <meta name="apple-mobile-web-app-title" content="Rent a Party">
  <link rel="apple-touch-icon" href="/icon-192.png">
  <script>if("serviceWorker"in navigator)navigator.serviceWorker.register("/sw.js");</script>
  <title>Edit Customer — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#111827;min-height:100vh}
    .topbar{background:white;border-bottom:1px solid #e5e7eb;padding:.9rem 1.25rem;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:50;flex-wrap:wrap;gap:.4rem}
    .topbar-brand{font-size:1rem;font-weight:700;color:#111827}
    .topbar-nav{display:flex;gap:.25rem;align-items:center;overflow-x:auto;-webkit-overflow-scrolling:touch;flex-wrap:nowrap;max-width:100%}
    .nav-link{color:#6b7280;font-size:.85rem;font-weight:500;text-decoration:none;padding:.38rem .65rem;border-radius:6px;white-space:nowrap}
    .nav-link:hover{background:#f3f4f6;color:#111827}
    .nav-link.active{background:#eff6ff;color:#2563eb;font-weight:600}
    .logout-btn{background:white;border:1px solid #d1d5db;color:#6b7280;padding:.38rem .65rem;border-radius:6px;font-size:.82rem;font-weight:500;text-decoration:none;white-space:nowrap}
    .main{max-width:700px;margin:0 auto;padding:1.75rem}
    .breadcrumb{font-size:.82rem;color:#9ca3af;margin-bottom:1rem}
    .breadcrumb a{color:#2563eb;text-decoration:none}
    .page-title{font-size:1.3rem;font-weight:700;color:#111827;margin-bottom:1.5rem}
    .card{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:1.5rem;margin-bottom:1.25rem}
    .card h3{font-size:.88rem;font-weight:700;color:#374151;text-transform:uppercase;letter-spacing:.4px;margin-bottom:1rem;padding-bottom:.5rem;border-bottom:1px solid #f3f4f6}
    .form-grid{display:grid;grid-template-columns:1fr 1fr;gap:.75rem}
    .form-group{display:flex;flex-direction:column;gap:.3rem}
    .form-group label{font-size:.75rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.4px}
    .form-group input,.form-group textarea{padding:.45rem .65rem;border:1px solid #d1d5db;border-radius:6px;font-size:.86rem;color:#111827;font-family:inherit}
    .form-group input:focus,.form-group textarea:focus{outline:none;border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,.1)}
    .form-group textarea{resize:vertical;min-height:70px}
    .form-actions{display:flex;gap:.75rem;margin-top:1.25rem}
    .btn{display:inline-block;padding:.5rem 1.1rem;border-radius:7px;font-size:.86rem;font-weight:600;cursor:pointer;border:none;text-decoration:none;transition:all .12s}
    .btn-primary{background:#2563eb;color:white}
    .btn-primary:hover{background:#1d4ed8}
    .btn-outline{background:white;color:#374151;border:1px solid #d1d5db}
    .btn-outline:hover{background:#f3f4f6}
    .btn-danger{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}
    .span2{grid-column:1/-1}
    /* Recent bookings mini-table */
    .bookings-mini{width:100%;border-collapse:collapse;font-size:.84rem}
    .bookings-mini th{padding:.5rem .75rem;text-align:left;color:#9ca3af;font-size:.72rem;font-weight:600;text-transform:uppercase;border-bottom:1px solid #e5e7eb}
    .bookings-mini td{padding:.6rem .75rem;border-bottom:1px solid #f3f4f6;color:#374151}
    .bookings-mini tr:last-child td{border-bottom:none}
    .badge{display:inline-flex;padding:.2rem .55rem;border-radius:20px;font-size:.72rem;font-weight:600}
    .badge-pending{background:#fef9c3;color:#854d0e}
    .badge-agree_to_pay{background:#d1fae5;color:#065f46}
    .badge-accepted{background:#dbeafe;color:#1e40af}
    .badge-confirmed{background:#dbeafe;color:#1e40af}
    .badge-partial{background:#dbeafe;color:#1e40af}
    .badge-denied{background:#fee2e2;color:#991b1b}
    .badge-cancelled{background:#f3f4f6;color:#6b7280}
    .badge-concluded{background:#e5e7eb;color:#374151}
    @media(max-width:600px){.form-grid{grid-template-columns:1fr}.main{padding:1rem}}
  </style>
<style>
/* ── Sidebar (shared) ── */
.sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
.sb-overlay.show{display:block}
.sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
.sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
.sb-brand img{height:1.8rem;width:auto;object-fit:contain}
.sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
.sb-new-btn:hover{background:#15803d}
.sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
.sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
.sb-link:hover{background:#f3f4f6;color:#111827}
.sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
.sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
.sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
.page-content{margin-left:210px;min-height:100vh}
.pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
.pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
.pg-back{font-size:.82rem;color:#6b7280;text-decoration:none;font-weight:500;white-space:nowrap}
.pg-back:hover{color:#111827}
.mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
.mobile-menu-btn:hover{background:#f3f4f6}
@media(max-width:768px){
  .sidebar{transform:translateX(-210px)}
  .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
  .page-content{margin-left:0!important}
  .mobile-menu-btn{display:block}
}
</style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link active">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>{{ c.full_name }}</h1>
  <a href="/admin/customers" class="pg-back">← Customers</a>
</div>
<div class="main">
  <div class="breadcrumb"><a href="/admin/customers">Customers</a> › Edit</div>
  <div class="page-title">{{ c.full_name }}</div>

  <div class="card">
    <h3>Contact Information</h3>
    <form method="POST" action="/admin/customers/{{ c.id }}/save">
      <div class="form-grid">
        <div class="form-group">
          <label>Full Name *</label>
          <input type="text" name="full_name" value="{{ c.full_name or '' }}" required>
        </div>
        <div class="form-group">
          <label>Company Name</label>
          <input type="text" name="company_name" value="{{ c.company_name or '' }}">
        </div>
        <div class="form-group">
          <label>Email</label>
          <input type="email" name="email" value="{{ c.email or '' }}">
        </div>
        <div class="form-group">
          <label>Phone</label>
          <input type="tel" name="phone" value="{{ c.phone or '' }}">
        </div>
        <div class="form-group span2">
          <label>Street Address</label>
          <input type="text" name="street" id="cust_street" value="{{ c.street or '' }}" autocomplete="off" placeholder="123 Main St">
        </div>
        <div class="form-group span2">
          <label>Address Line 2 <span style="color:#9ca3af;font-weight:400;font-size:.8rem">(Apt, Suite, Unit, etc.)</span></label>
          <input type="text" name="street2" value="{{ c.street2 or '' }}" placeholder="Apt 4B">
        </div>
        <div class="form-group">
          <label>City</label>
          <input type="text" name="city" id="cust_city" value="{{ c.city or '' }}">
        </div>
        <div class="form-group">
          <label>State</label>
          <input type="text" name="state" id="cust_state" value="{{ c.state or '' }}" maxlength="2">
        </div>
        <div class="form-group">
          <label>Zip</label>
          <input type="text" name="zip" id="cust_zip" value="{{ c.zip or '' }}" maxlength="10">
        </div>
        <div class="form-group span2">
          <label>Notes</label>
          <textarea name="notes">{{ c.notes or '' }}</textarea>
        </div>
        <div class="form-group span2" style="padding:.75rem 1rem;background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px">
          <label style="display:flex;align-items:center;gap:.6rem;cursor:pointer;font-weight:600;color:#166534;font-size:.95rem;text-transform:none;letter-spacing:0">
            <input type="checkbox" name="tax_exempt" value="1" {% if c.tax_exempt %}checked{% endif %}
                   style="width:18px;height:18px;accent-color:#16a34a;cursor:pointer">
            Tax Exempt
          </label>
          <p style="margin:.3rem 0 0;font-size:.8rem;color:#4b7c5a">
            Check if this customer has a valid CT tax-exempt certificate. Tax (6.35%) will be removed from all their bookings.
          </p>
        </div>
      </div>
      <div class="form-actions">
        <a href="/admin/customers" class="btn btn-outline">Cancel</a>
        <button type="submit" class="btn btn-primary">Save Changes</button>
      </div>
    </form>
    <form method="POST" action="/admin/customers/{{ c.id }}/delete" style="display:inline;margin:0">
      <button class="btn btn-danger" onclick="return confirm('Permanently remove {{ c.full_name }}?')">Remove Customer</button>
    </form>
  </div>

  {% if bookings %}
  <div class="card">
    <h3>Booking History ({{ bookings|length }})</h3>
    <table class="bookings-mini">
      <thead><tr><th>#</th><th>Event Date</th><th>Total</th><th>Status</th></tr></thead>
      <tbody>
        {% for b in bookings %}
        <tr>
          <td><a href="/admin/booking/{{ b.id }}" style="color:#2563eb;font-weight:600">#{{ b.id }}</a></td>
          <td>{{ b.event_start_date.strftime('%m/%d/%Y') if b.event_start_date else '' }}</td>
          <td>${{ "%.2f"|format(b.grand_total or 0) }}</td>
          <td><span class="badge badge-{{ b.status }}">{{ b.status|capitalize }}</span></td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endif %}
</div>
<script>
function initCustStreetAutocomplete() {
  var streetEl = document.getElementById('cust_street');
  if (!streetEl || !window.google) return;
  var ac = new google.maps.places.Autocomplete(streetEl, {
    types: ['address'],
    componentRestrictions: { country: 'us' },
    fields: ['address_components']
  });
  streetEl.addEventListener('keydown', function(e) {
    if (e.key === 'Enter') e.preventDefault();
  });
  ac.addListener('place_changed', function() {
    var place = ac.getPlace();
    if (!place.address_components) return;
    var streetNum = '', route = '', city = '', state = '', zip = '';
    place.address_components.forEach(function(comp) {
      var t = comp.types;
      if (t.includes('street_number'))                    streetNum = comp.long_name;
      else if (t.includes('route'))                       route     = comp.long_name;
      else if (t.includes('locality'))                    city      = comp.long_name;
      else if (t.includes('administrative_area_level_1')) state     = comp.short_name;
      else if (t.includes('postal_code'))                 zip       = comp.long_name;
    });
    streetEl.value = [streetNum, route].filter(Boolean).join(' ');
    document.getElementById('cust_city').value  = city;
    document.getElementById('cust_state').value = state;
    document.getElementById('cust_zip').value   = zip;
  });
}
</script>
{% if google_maps_key %}
<script src="https://maps.googleapis.com/maps/api/js?key={{ google_maps_key }}&libraries=places&callback=initCustStreetAutocomplete" async defer></script>
{% endif %}
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>

<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
<style>
/* ── Mobile horizontal scroll fix ── */
html{overflow-x:auto}
body{overflow-x:auto}
.page-content{overflow-x:auto}
.main{overflow-x:auto}
table{border-collapse:collapse}
.tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;width:100%;display:block}
@media(max-width:768px){
  .card{overflow-x:auto}
  td,th{white-space:nowrap}
}
</style>
<script>
(function(){
  /* Wrap every unwrapped table in a scroll div on mobile */
  function wrapTables(){
    document.querySelectorAll('table').forEach(function(t){
      var p=t.parentElement;
      if(p && p.className && (p.className.indexOf('tbl-wrap')>-1 || p.className.indexOf('table-scroll')>-1)) return;
      var w=document.createElement('div');
      w.className='tbl-wrap';
      p.insertBefore(w,t);
      w.appendChild(t);
    });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',wrapTables);}
  else{wrapTables();}
})();
</script>
</body></html>
"""


# Routes - Customers

@app.route("/admin/customers")
@admin_required
def admin_customers():
    conn = get_db()
    customers = []
    flash_ok = request.args.get("flash_ok", "")
    flash_err = request.args.get("flash_err", "")
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute("SELECT * FROM customers ORDER BY full_name")
            rows = cur.fetchall()
            for r in rows:
                c = dict(r)
                if c.get("email"):
                    cur.execute(
                        "SELECT id, status FROM bookings WHERE LOWER(email)=LOWER(%s) ORDER BY id DESC",
                        (c["email"],)
                    )
                    c["bookings"] = [dict(b) for b in cur.fetchall()]
                else:
                    c["bookings"] = []
                customers.append(c)
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"admin_customers error: {e}")
    return render_template_string(ADMIN_CUSTOMERS_HTML,
        business_name=BUSINESS_NAME,
        customers=customers,
        flash_ok=flash_ok,
        flash_err=flash_err,
    )


@app.route("/admin/customers/add", methods=["POST"])
@admin_required
def add_customer():
    full_name = request.form.get("full_name", "").strip()
    if not full_name:
        return redirect(url_for("admin_customers", flash_err="Name is required"))
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO customers (full_name, company_name, email, phone, street, city, state, zip, notes, tax_exempt) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (
                    full_name,
                    request.form.get("company_name", "").strip(),
                    request.form.get("email", "").strip().lower(),
                    request.form.get("phone", "").strip(),
                    request.form.get("street", "").strip(),
                    request.form.get("city", "").strip(),
                    request.form.get("state", "").strip().upper(),
                    request.form.get("zip", "").strip(),
                    request.form.get("notes", "").strip(),
                    bool(request.form.get("tax_exempt")),
                )
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"add_customer error: {e}")
            return redirect(url_for("admin_customers", flash_err="Error adding customer"))
    return redirect(url_for("admin_customers", flash_ok=f"{full_name} added successfully"))


@app.route("/admin/customers/<int:cid>")
@admin_required
def admin_customer_view(cid):
    return redirect(url_for("admin_customer_edit", cid=cid))


@app.route("/admin/customers/<int:cid>/edit")
@admin_required
def admin_customer_edit(cid):
    conn = get_db()
    c = None
    bookings = []
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute("SELECT * FROM customers WHERE id=%s", (cid,))
            row = cur.fetchone()
            if row:
                c = dict(row)
                if c.get("email"):
                    cur.execute(
                        "SELECT id, status, event_start_date, grand_total FROM bookings WHERE LOWER(email)=LOWER(%s) ORDER BY id DESC",
                        (c["email"],)
                    )
                    bookings = [dict(b) for b in cur.fetchall()]
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"admin_customer_edit error: {e}")
    if not c:
        return "Customer not found", 404
    return render_template_string(ADMIN_CUSTOMER_EDIT_HTML,
        business_name=BUSINESS_NAME,
        c=c,
        bookings=bookings,
        google_maps_key=GOOGLE_MAPS_KEY,
    )


@app.route("/admin/customers/<int:cid>/save", methods=["POST"])
@admin_required
def save_customer(cid):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE customers SET full_name=%s, company_name=%s, email=%s, phone=%s, street=%s, street2=%s, city=%s, state=%s, zip=%s, notes=%s, tax_exempt=%s WHERE id=%s",
                (
                    request.form.get("full_name", "").strip(),
                    request.form.get("company_name", "").strip(),
                    request.form.get("email", "").strip().lower(),
                    request.form.get("phone", "").strip(),
                    request.form.get("street", "").strip(),
                    request.form.get("street2", "").strip(),
                    request.form.get("city", "").strip(),
                    request.form.get("state", "").strip().upper(),
                    request.form.get("zip", "").strip(),
                    request.form.get("notes", "").strip(),
                    bool(request.form.get("tax_exempt")),
                    cid,
                )
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"save_customer error: {e}")
    return redirect(url_for("admin_customers", flash_ok="Customer updated successfully"))


@app.route("/admin/customers/<int:cid>/delete", methods=["POST"])
@admin_required
def delete_customer(cid):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM customers WHERE id=%s", (cid,))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            log.error(f"delete_customer error: {e}")
    return redirect(url_for("admin_customers", flash_ok="Customer removed"))


@app.route("/admin/customers/import", methods=["GET", "POST"])
@admin_required
def import_customers():
    preview = []
    flash_ok = ""
    flash_err = ""
    if request.method == "POST":
        f = request.files.get("csvFile")
        if not f or not f.filename:
            flash_err = "Please select a CSV file"
        else:
            import csv as _csv
            import io as _io
            try:
                content = f.read().decode("utf-8-sig")
                reader = _csv.DictReader(_io.StringIO(content))
                rows = list(reader)
                if request.form.get("preview"):
                    preview = [{"full_name": r.get("full_name",""), "email": r.get("email",""),
                                "phone": r.get("phone",""), "city": r.get("city","")} for r in rows[:10]]
                else:
                    conn = get_db()
                    if conn:
                        cur = conn.cursor()
                        imported = 0
                        skipped  = 0
                        for r in rows:
                            name = (r.get("full_name") or "").strip()
                            if not name:
                                skipped += 1
                                continue
                            email_val = (r.get("email") or "").strip().lower()
                            try:
                                cur.execute(
                                    "INSERT INTO customers "
                                    "(full_name, company_name, email, phone, street, city, state, zip, notes) "
                                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                                    "ON CONFLICT (email) DO NOTHING",
                                    (
                                        name,
                                        (r.get("company_name") or "").strip(),
                                        email_val or None,
                                        (r.get("phone") or "").strip(),
                                        (r.get("street") or "").strip(),
                                        (r.get("city") or "").strip(),
                                        (r.get("state") or "").strip(),
                                        (r.get("zip") or "").strip(),
                                        (r.get("notes") or "").strip(),
                                    ),
                                )
                                imported += 1
                            except Exception:
                                skipped += 1
                        conn.commit()
                        cur.close()
                        conn.close()
                        flash_ok = f"Imported {imported} customers ({skipped} skipped)."
            except Exception as ex:
                flash_err = f"Error reading CSV: {ex}"

    return render_template_string(ADMIN_CUSTOMERS_IMPORT_HTML,
        business_name=BUSINESS_NAME,
        preview=preview,
        flash_ok=flash_ok,
        flash_err=flash_err,
    )


@app.route("/admin/customers/import/template")
@admin_required
def customers_import_template():
    """Download a blank CSV template for customer import."""
    import io as _io, csv as _csv
    output = _io.StringIO()
    writer = _csv.writer(output)
    writer.writerow(["full_name", "company_name", "email", "phone", "street", "city", "state", "zip", "notes"])
    writer.writerow(["Jane Smith", "", "jane@example.com", "860-555-0100", "123 Main St", "Hartford", "CT", "06101", ""])
    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue().encode("utf-8"),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers_template.csv"}
    )



# ══════════════════════════════════════════════════════════════════════════════
#  TAX REPORT
# ══════════════════════════════════════════════════════════════════════════════

ADMIN_TAX_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Tax Report — {{ business_name }}</title>
  <style>
    *{box-sizing:border-box;margin:0;padding:0}
    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f6fa;color:#111827;min-height:100vh}
    /* Sidebar */
    .sb-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:99}
    .sb-overlay.show{display:block}
    .sidebar{width:210px;min-height:100vh;background:#fff;border-right:1px solid #e5e7eb;position:fixed;top:0;left:0;z-index:100;display:flex;flex-direction:column;transition:transform .25s ease}
    .sb-brand{display:flex;align-items:center;gap:.6rem;padding:.9rem 1rem;border-bottom:1px solid #f3f4f6}
    .sb-brand img{height:1.8rem;width:auto;object-fit:contain}
    .sb-brand-name{font-size:.85rem;font-weight:700;color:#111827;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    .sb-new-btn{display:block;margin:.75rem .75rem .25rem;background:#16a34a;color:#fff;text-align:center;padding:.5rem .75rem;border-radius:8px;text-decoration:none;font-size:.82rem;font-weight:600}
    .sb-new-btn:hover{background:#15803d}
    .sb-nav{flex:1;padding:.5rem 0;overflow-y:auto}
    .sb-link{display:flex;align-items:center;gap:.55rem;padding:.55rem 1rem;color:#374151;text-decoration:none;font-size:.85rem;font-weight:500;border-radius:8px;margin:1px .5rem;transition:background .15s,color .15s}
    .sb-link:hover{background:#f3f4f6;color:#111827}
    .sb-link.active{background:#eff6ff;color:#1d4ed8;font-weight:600}
    .sb-bottom{padding:.75rem;border-top:1px solid #f3f4f6}
    .sb-divider{height:1px;background:#f3f4f6;margin:.4rem .75rem}
    .page-content{margin-left:210px;min-height:100vh}
    .pg-hdr{background:#fff;border-bottom:1px solid #e5e7eb;padding:.7rem 1.25rem;display:flex;align-items:center;gap:.75rem;position:sticky;top:0;z-index:50}
    .pg-hdr h1{font-size:1.05rem;font-weight:700;color:#111827;flex:1;margin:0}
    .mobile-menu-btn{display:none;background:none;border:none;font-size:1.35rem;cursor:pointer;color:#374151;padding:.2rem .3rem;line-height:1;border-radius:6px}
    @media(max-width:768px){
      .sidebar{transform:translateX(-210px)}
      .sidebar.open{transform:translateX(0);box-shadow:4px 0 20px rgba(0,0,0,.15)}
      .page-content{margin-left:0!important}
      .mobile-menu-btn{display:block}
    }
    /* Page content */
    .main{max-width:900px;margin:0 auto;padding:1.5rem}
    .flash{padding:.75rem 1rem;border-radius:8px;margin-bottom:1rem;font-size:.9rem;font-weight:500}
    .flash-ok{background:#dcfce7;color:#166534;border:1px solid #bbf7d0}
    .flash-err{background:#fee2e2;color:#991b1b;border:1px solid #fecaca}
    /* Period tabs */
    .period-tabs{display:flex;gap:.4rem;margin-bottom:1.25rem;flex-wrap:wrap}
    .ptab{padding:.4rem .9rem;border-radius:8px;font-size:.83rem;font-weight:600;text-decoration:none;border:1px solid #d1d5db;color:#374151;background:#fff;cursor:pointer}
    .ptab:hover{background:#f3f4f6}
    .ptab.active{background:#166534;color:#fff;border-color:#166534}
    /* Summary cards */
    .summary-row{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1rem;margin-bottom:1.5rem}
    .scard{background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:1.1rem 1.25rem}
    .scard-label{font-size:.75rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.4px;margin-bottom:.35rem}
    .scard-value{font-size:1.6rem;font-weight:700;color:#111827}
    .scard-value.green{color:#16a34a}
    .scard-value.red{color:#dc2626}
    .scard-value.blue{color:#2563eb}
    /* Transfer form */
    .card{background:#fff;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;margin-bottom:1.5rem}
    .card-hdr{padding:.75rem 1.25rem;background:#f9fafb;border-bottom:1px solid #e5e7eb;font-weight:700;font-size:.88rem;color:#374151;display:flex;justify-content:space-between;align-items:center}
    .xfer-form{padding:1rem 1.25rem;display:flex;gap:.75rem;flex-wrap:wrap;align-items:flex-end}
    .fg{display:flex;flex-direction:column;gap:.3rem}
    .fg label{font-size:.75rem;font-weight:600;color:#6b7280;text-transform:uppercase;letter-spacing:.4px}
    .fg input,.fg select,.fg textarea{padding:.42rem .65rem;border:1px solid #d1d5db;border-radius:6px;font-size:.86rem;color:#111827}
    .fg input:focus,.fg select:focus{outline:none;border-color:#16a34a;box-shadow:0 0 0 2px rgba(22,163,74,.1)}
    .btn-green{background:#16a34a;color:#fff;border:none;border-radius:7px;padding:.45rem 1.1rem;font-size:.85rem;font-weight:700;cursor:pointer}
    .btn-green:hover{background:#15803d}
    .btn-danger-sm{background:#fee2e2;color:#991b1b;border:1px solid #fecaca;border-radius:6px;padding:.25rem .6rem;font-size:.75rem;font-weight:600;cursor:pointer}
    .btn-danger-sm:hover{background:#fecaca}
    /* Table */
    table{width:100%;border-collapse:collapse}
    th{padding:.6rem 1rem;text-align:left;font-size:.72rem;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.5px;border-bottom:1px solid #e5e7eb;background:#f9fafb;white-space:nowrap}
    td{padding:.75rem 1rem;border-bottom:1px solid #f3f4f6;font-size:.85rem;vertical-align:middle}
    tr:last-child td{border-bottom:none}
    tbody tr:hover td{background:#fafafa}
    .empty{padding:2.5rem;text-align:center;color:#9ca3af;font-size:.9rem}
    .tbl-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch}
    @media(max-width:600px){.xfer-form{flex-direction:column}.main{padding:1rem}}
  </style>
</head>
<body>
<div class="sb-overlay" id="sb-overlay" onclick="closeSidebar()"></div>
<aside class="sidebar" id="sidebar">
  <div class="sb-brand"><img src="/logo.png" alt=""><span class="sb-brand-name">{{ business_name }}</span></div>
  <a href="/admin/booking/new" class="sb-new-btn">+ New Booking</a>
  <nav class="sb-nav">
    <a href="/admin/dashboard" class="sb-link">🏠 Dashboard</a>
    <div class="sb-divider"></div>
    <a href="/admin/customers" class="sb-link">👥 Clients</a>
    <a href="/admin/inventory" class="sb-link">📦 Inventory</a>
    <a href="/admin/calendar" class="sb-link">📅 Calendar</a>
    <a href="/admin/reports" class="sb-link">📊 Reports</a>
    <a href="/admin/route" class="sb-link">🗺 Route</a>
    <a href="/driver/{{ today }}" class="sb-link">🚚 Driver View</a>
    <a href="/admin/formsite-import" class="sb-link">📥 Import</a>
    <a href="/admin/tax-report" class="sb-link active">💰 Tax Report</a>
  </nav>
  <div class="sb-bottom">
    <a href="/admin/logout" class="sb-link">🚪 Sign Out</a>
  </div>
</aside>
<div class="page-content">
<div class="pg-hdr">
  <button class="mobile-menu-btn" onclick="openSidebar()">&#9776;</button>
  <h1>💰 Tax Report</h1>
</div>
<div class="main">

  {% if flash_ok %}<div class="flash flash-ok">✓ {{ flash_ok }}</div>{% endif %}
  {% if flash_err %}<div class="flash flash-err">⚠ {{ flash_err }}</div>{% endif %}

  <!-- Period selector -->
  <div class="period-tabs">
    <a href="?period=this_month" class="ptab {% if period=='this_month' %}active{% endif %}">This Month</a>
    <a href="?period=last_month" class="ptab {% if period=='last_month' %}active{% endif %}">Last Month</a>
    <a href="?period=this_year"  class="ptab {% if period=='this_year'  %}active{% endif %}">This Year</a>
    <a href="?period=all_time"   class="ptab {% if period=='all_time'   %}active{% endif %}">All Time</a>
  </div>

  <!-- Summary cards -->
  <div class="summary-row">
    <div class="scard">
      <div class="scard-label">Tax Collected</div>
      <div class="scard-value green">${{ '%.2f'|format(tax_collected) }}</div>
      <div style="font-size:.75rem;color:#6b7280;margin-top:.3rem">{{ booking_count }} booking{{ 's' if booking_count != 1 }}</div>
    </div>
    <div class="scard">
      <div class="scard-label">Transferred Out</div>
      <div class="scard-value blue">${{ '%.2f'|format(tax_transferred) }}</div>
      <div style="font-size:.75rem;color:#6b7280;margin-top:.3rem">{{ transfer_count }} transfer{{ 's' if transfer_count != 1 }}</div>
    </div>
    <div class="scard">
      <div class="scard-label">Still Owed</div>
      <div class="scard-value {% if tax_owed > 0 %}red{% else %}green{% endif %}">${{ '%.2f'|format(tax_owed) }}</div>
      <div style="font-size:.75rem;color:#6b7280;margin-top:.3rem">move to tax account</div>
    </div>
  </div>

  <!-- Record a transfer -->
  <div class="card">
    <div class="card-hdr">✅ Record a Tax Transfer</div>
    <form method="POST" action="/admin/tax-report/transfer" class="xfer-form">
      <div class="fg">
        <label>Amount ($)</label>
        <input type="number" name="amount" step="0.01" min="0.01" placeholder="0.00"
               value="{{ '%.2f'|format(tax_owed) if tax_owed > 0 else '' }}"
               style="width:130px" required>
      </div>
      <div class="fg">
        <label>Note (optional)</label>
        <input type="text" name="note" placeholder="e.g. Q2 tax transfer" style="width:220px">
      </div>
      <div class="fg">
        <label>Period Label</label>
        <input type="text" name="period_label" value="{{ period_label }}" style="width:140px">
      </div>
      <button type="submit" class="btn-green">💾 Record Transfer</button>
    </form>
  </div>

  <!-- Transfer history -->
  <div class="card">
    <div class="card-hdr">
      <span>Transfer History</span>
      <span style="font-size:.8rem;color:#6b7280;font-weight:400">all transfers on record</span>
    </div>
    {% if transfers %}
    <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>Date</th><th>Amount</th><th>Period</th><th>Note</th><th></th>
      </tr></thead>
      <tbody>
      {% for t in transfers %}
      <tr>
        <td style="color:#6b7280">{{ t.created_at.strftime('%m/%d/%Y') if t.created_at else '' }}</td>
        <td style="font-weight:700;color:#2563eb">${{ '%.2f'|format(t.amount) }}</td>
        <td>{{ t.period_label or '—' }}</td>
        <td style="color:#6b7280">{{ t.note or '—' }}</td>
        <td>
          <form method="POST" action="/admin/tax-report/transfer/{{ t.id }}/delete" style="display:inline">
            <button type="submit" class="btn-danger-sm" onclick="return confirm('Delete this transfer record?')">✕</button>
          </form>
        </td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
    {% else %}
    <div class="empty">No transfers recorded yet</div>
    {% endif %}
  </div>

  <!-- Bookings with tax -->
  <div class="card">
    <div class="card-hdr">
      <span>Bookings with Tax — {{ period_label }}</span>
      <span style="font-size:.8rem;color:#16a34a;font-weight:700">${{ '%.2f'|format(tax_collected) }} collected</span>
    </div>
    {% if tax_bookings %}
    <div class="tbl-wrap">
    <table>
      <thead><tr>
        <th>#</th><th>Client</th><th>Event Date</th><th>Status</th><th>Grand Total</th><th>Tax (6.35%)</th>
      </tr></thead>
      <tbody>
      {% for b in tax_bookings %}
      <tr>
        <td><a href="/admin/booking/{{ b.id }}" style="color:#2563eb;text-decoration:none;font-weight:600">#{{ b.id }}</a></td>
        <td>{{ b.full_name }}</td>
        <td style="color:#6b7280">{{ b.event_start_date or '—' }}</td>
        <td><span style="font-size:.75rem;font-weight:600;text-transform:capitalize;padding:.15rem .5rem;border-radius:10px;background:#f3f4f6;color:#374151">{{ b.status }}</span></td>
        <td>${{ '%.2f'|format(b.grand_total or 0) }}</td>
        <td style="font-weight:700;color:#16a34a">${{ '%.2f'|format(b.tax_amount or 0) }}</td>
      </tr>
      {% endfor %}
      </tbody>
    </table>
    </div>
    {% else %}
    <div class="empty">No bookings with tax in this period</div>
    {% endif %}
  </div>

</div>
</div>
<script>
function openSidebar(){document.getElementById('sidebar').classList.add('open');document.getElementById('sb-overlay').classList.add('show');}
function closeSidebar(){document.getElementById('sidebar').classList.remove('open');document.getElementById('sb-overlay').classList.remove('show');}
</script>
<button id="back-fab" title="Go back (drag to move)" style="position:fixed;bottom:1.5rem;left:1.5rem;z-index:9999;width:46px;height:46px;border-radius:50%;background:#1e40af;color:white;border:none;cursor:grab;font-size:1.4rem;line-height:1;box-shadow:0 3px 12px rgba(0,0,0,.35);touch-action:none;user-select:none;transition:box-shadow .15s">&#8592;</button>
<script>
(function(){
  var btn = document.getElementById('back-fab');
  if(!btn) return;
  var SK = 'back_fab_pos';
  var dragging = false, didDrag = false;
  var startX, startY, origLeft, origBottom;

  // Restore saved position
  try {
    var saved = JSON.parse(localStorage.getItem(SK));
    if(saved) { btn.style.left = saved.left; btn.style.bottom = saved.bottom; btn.style.top = ''; }
  } catch(e){}

  function savePos() {
    try { localStorage.setItem(SK, JSON.stringify({left: btn.style.left, bottom: btn.style.bottom})); } catch(e){}
  }

  function startDrag(cx, cy) {
    dragging = true; didDrag = false;
    var rect = btn.getBoundingClientRect();
    startX = cx; startY = cy;
    origLeft = rect.left;
    origBottom = window.innerHeight - rect.bottom;
    btn.style.cursor = 'grabbing';
    btn.style.boxShadow = '0 6px 24px rgba(0,0,0,.45)';
    btn.style.transition = 'none';
  }

  function moveDrag(cx, cy) {
    if(!dragging) return;
    var dx = cx - startX, dy = cy - startY;
    if(Math.abs(dx) > 3 || Math.abs(dy) > 3) didDrag = true;
    var newLeft = Math.max(4, Math.min(window.innerWidth - 50, origLeft + dx));
    var newBottom = Math.max(4, Math.min(window.innerHeight - 50, origBottom - dy));
    btn.style.left = newLeft + 'px';
    btn.style.bottom = newBottom + 'px';
    btn.style.top = '';
  }

  function endDrag() {
    if(!dragging) return;
    dragging = false;
    btn.style.cursor = 'grab';
    btn.style.boxShadow = '0 3px 12px rgba(0,0,0,.35)';
    btn.style.transition = 'box-shadow .15s';
    savePos();
  }

  // Mouse
  btn.addEventListener('mousedown', function(e){ e.preventDefault(); startDrag(e.clientX, e.clientY); });
  document.addEventListener('mousemove', function(e){ moveDrag(e.clientX, e.clientY); });
  document.addEventListener('mouseup', function(e){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });

  // Touch
  btn.addEventListener('touchstart', function(e){ e.preventDefault(); startDrag(e.touches[0].clientX, e.touches[0].clientY); }, {passive:false});
  document.addEventListener('touchmove', function(e){ if(dragging){ e.preventDefault(); moveDrag(e.touches[0].clientX, e.touches[0].clientY); } }, {passive:false});
  document.addEventListener('touchend', function(){
    if(!dragging) return;
    var wasDrag = didDrag; endDrag();
    if(!wasDrag) history.back();
  });
})();
</script>
</body></html>
"""


@app.route("/admin/tax-report")
@admin_required
def admin_tax_report():
    from datetime import date as _date
    today = _date.today()
    period = request.args.get("period", "this_month")
    flash_ok  = request.args.get("flash_ok", "")
    flash_err = request.args.get("flash_err", "")

    # Date range for selected period
    if period == "this_month":
        start = today.replace(day=1)
        end   = today
        period_label = today.strftime("%B %Y")
    elif period == "last_month":
        first_this = today.replace(day=1)
        last_last  = first_this - timedelta(days=1)
        start = last_last.replace(day=1)
        end   = last_last
        period_label = last_last.strftime("%B %Y")
    elif period == "this_year":
        start = today.replace(month=1, day=1)
        end   = today
        period_label = str(today.year)
    else:  # all_time
        start = None
        end   = None
        period_label = "All Time"

    conn = get_db()
    tax_bookings   = []
    tax_collected  = 0.0
    transfers      = []
    tax_transferred = 0.0

    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            # Tax from bookings in period
            if start and end:
                cur.execute("""
                    SELECT id, full_name, event_start_date, status, grand_total, tax_amount
                    FROM bookings
                    WHERE status NOT IN ('denied','cancelled')
                      AND (archived IS NULL OR archived = FALSE)
                      AND tax_amount > 0
                      AND created_at::date BETWEEN %s AND %s
                    ORDER BY created_at DESC
                """, (start.isoformat(), end.isoformat()))
            else:
                cur.execute("""
                    SELECT id, full_name, event_start_date, status, grand_total, tax_amount
                    FROM bookings
                    WHERE status NOT IN ('denied','cancelled')
                      AND (archived IS NULL OR archived = FALSE)
                      AND tax_amount > 0
                    ORDER BY created_at DESC
                """)
            tax_bookings = [dict(r) for r in cur.fetchall()]
            tax_collected = sum(float(b.get("tax_amount") or 0) for b in tax_bookings)

            # All transfers (always show full history)
            cur.execute("SELECT * FROM tax_transfers ORDER BY created_at DESC")
            transfers = [dict(r) for r in cur.fetchall()]
            tax_transferred = sum(float(t.get("amount") or 0) for t in transfers)

            cur.close(); conn.close()
        except Exception as e:
            log.error(f"admin_tax_report error: {e}")

    tax_owed = max(0.0, round(tax_collected - tax_transferred, 2))

    return render_template_string(ADMIN_TAX_HTML,
        business_name=BUSINESS_NAME,
        period=period,
        period_label=period_label,
        tax_collected=round(tax_collected, 2),
        tax_transferred=round(tax_transferred, 2),
        tax_owed=tax_owed,
        booking_count=len(tax_bookings),
        transfer_count=len(transfers),
        tax_bookings=tax_bookings,
        transfers=transfers,
        flash_ok=flash_ok,
        flash_err=flash_err,
    )


@app.route("/admin/tax-report/transfer", methods=["POST"])
@admin_required
def admin_tax_transfer():
    amount = request.form.get("amount", "").strip()
    note   = request.form.get("note", "").strip()
    period_label = request.form.get("period_label", "").strip()
    try:
        amt = round(float(amount), 2)
        if amt <= 0:
            raise ValueError("amount must be positive")
    except Exception:
        return redirect(url_for("admin_tax_report", flash_err="Invalid amount"))
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO tax_transfers (amount, note, period_label) VALUES (%s, %s, %s)",
                (amt, note or None, period_label or None)
            )
            conn.commit(); cur.close(); conn.close()
        except Exception as e:
            log.error(f"tax_transfer insert error: {e}")
            return redirect(url_for("admin_tax_report", flash_err="Error saving transfer"))
    return redirect(url_for("admin_tax_report", flash_ok=f"Transfer of ${amt:.2f} recorded"))


@app.route("/admin/tax-report/transfer/<int:transfer_id>/delete", methods=["POST"])
@admin_required
def admin_tax_transfer_delete(transfer_id):
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM tax_transfers WHERE id=%s", (transfer_id,))
            conn.commit(); cur.close(); conn.close()
        except Exception as e:
            log.error(f"tax_transfer delete error: {e}")
    return redirect(url_for("admin_tax_report", flash_ok="Transfer record deleted"))



@app.route("/admin/debug/inv/<int:booking_id>")
@admin_required
def debug_inv_check(booking_id):
    """Debug: show exactly which bookings are counted in the inventory overlap query."""
    products    = get_products()
    prod_totals = {p["id"]: int(p["total"]) for p in products}
    name_to_pid = {p["name"].lower(): p["id"] for p in products}

    conn = get_db()
    if not conn:
        return "<pre>No DB connection</pre>"
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # Fetch this booking
        cur.execute("""SELECT id, full_name, status, items_json,
                              event_start_date, event_end_date, setup_date, delivery_status
                       FROM bookings WHERE id=%s""", (booking_id,))
        brow = cur.fetchone()
        if not brow:
            return f"<pre>Booking #{booking_id} not found</pre>"
        b = _row(brow)

        b_out  = str(b.get("setup_date") or "")[:10]
        b_back = str(b.get("event_end_date") or "")[:10]
        b_items = json.loads(b.get("items_json") or "[]")

        # est. delivery (setup_date) → est. pickup (event_end_date)
        cur.execute("""
            SELECT id, full_name, status, delivery_status,
                   setup_date, event_end_date, items_json
            FROM bookings
            WHERE status = 'accepted' AND payment_status IN ('paid','partial')
              AND (delivery_status IS NULL OR delivery_status != 'picked_up')
              AND id != %s
              AND setup_date     IS NOT NULL
              AND event_end_date IS NOT NULL
              AND setup_date     <= %s
              AND event_end_date >= %s
            ORDER BY setup_date, id
        """, (booking_id, b_back or b_out, b_out or b_back))
        overlapping = [_row(r) for r in cur.fetchall()]
        cur.close(); conn.close()

        # Tally reserved by item
        reserved = {}
        for o in overlapping:
            for it in json.loads(o.get("items_json") or "[]"):
                pid = it.get("id") or name_to_pid.get((it.get("name") or "").lower())
                qty = int(it.get("qty") or 0)
                if pid and qty > 0:
                    reserved[pid] = reserved.get(pid, 0) + qty

        out = ["<pre style='font-family:monospace;padding:1rem;font-size:.83rem;line-height:1.5'>"]
        out.append(f"=== Inventory Debug — Booking #{booking_id} ({b.get('full_name')}) ===\n")
        out.append(f"  Status:          {b.get('status')}\n")
        out.append(f"  Est. Delivery:   {b.get('setup_date')}\n")
        out.append(f"  Est. Pickup:     {b.get('event_end_date')}\n")
        out.append(f"  Checking: confirmed bookings where setup_date <= {b_back or b_out!r} AND event_end_date >= {b_out or b_back!r}\n")
        out.append(f"\n--- This booking needs ---\n")
        for it in b_items:
            out.append(f"  {it.get('name')}: {it.get('qty')}\n")
        out.append(f"\n--- Overlapping bookings ({len(overlapping)}) ---\n")
        for o in overlapping:
            o_start  = str(o.get("setup_date")    or "")[:10]
            o_end    = str(o.get("event_end_date") or "")[:10]
            o_status = o.get("delivery_status") or "not_delivered"
            chairs   = sum(int(it.get("qty",0)) for it in json.loads(o.get("items_json") or "[]")
                           if "chair" in (it.get("name") or "").lower())
            tables   = sum(int(it.get("qty",0)) for it in json.loads(o.get("items_json") or "[]")
                           if "table" in (it.get("name") or "").lower())
            out.append(f"  #{o['id']} {o.get('full_name',''):<25} {o.get('status',''):<12} "
                       f"del:{o_start} pickup:{o_end} [{o_status}]  chairs:{chairs} tables:{tables}\n")
        out.append(f"\n--- Reserved totals ---\n")
        for pid, qty in reserved.items():
            total = prod_totals.get(pid, 0)
            avail = max(0, total - qty)
            out.append(f"  {pid}: {qty} reserved / {total} total → {avail} available\n")

        # Also show bookings on these dates that are NOT being counted
        cur2 = conn.cursor(cursor_factory=psycopg2.extras.DictCursor) if not conn.closed else None
        try:
            conn2 = get_db()
            if conn2:
                cur2 = conn2.cursor(cursor_factory=psycopg2.extras.DictCursor)
                cur2.execute("""
                    SELECT id, full_name, status, payment_status, delivery_status,
                           setup_date, event_end_date, items_json, amount_paid, grand_total
                    FROM bookings
                    WHERE id != %s
                      AND setup_date IS NOT NULL AND event_end_date IS NOT NULL
                      AND setup_date <= %s AND event_end_date >= %s
                      AND NOT (status='accepted' AND payment_status IN ('paid','partial')
                               AND (delivery_status IS NULL OR delivery_status != 'picked_up'))
                      AND status NOT IN ('cancelled','denied','concluded')
                    ORDER BY setup_date, id
                """, (booking_id, b_back or b_out, b_out or b_back))
                excluded = [_row(r) for r in cur2.fetchall()]
                cur2.close(); conn2.close()
                out.append(f"\n--- Bookings on these dates NOT counted in inventory ({len(excluded)}) ---\n")
                out.append(f"  (These are excluded: pending, accepted/waiting, already picked up)\n")
                for o in excluded:
                    o_start  = str(o.get("setup_date")    or "")[:10]
                    o_end    = str(o.get("event_end_date") or "")[:10]
                    pmt      = o.get("payment_status") or "NULL"
                    paid     = o.get("amount_paid") or 0
                    gtotal   = o.get("grand_total") or 0
                    chairs   = sum(int(it.get("qty",0)) for it in json.loads(o.get("items_json") or "[]")
                                   if "chair" in (it.get("name") or "").lower())
                    out.append(f"  #{o['id']} {o.get('full_name',''):<25} status:{o.get('status',''):<10} "
                               f"pmt:{pmt:<8} paid:${paid} total:${gtotal}  "
                               f"del:{o_start} pickup:{o_end}  chairs:{chairs}\n")
        except Exception as exc2:
            out.append(f"  (error fetching excluded: {exc2})\n")

        out.append("</pre>")
        out.append(f'<a href="/admin/booking/{booking_id}" style="margin:1rem;display:inline-block">← Back to Booking #{booking_id}</a>')
        return "".join(out)
    except Exception as e:
        import traceback as _tb
        return f"<pre>Error: {e}\n{_tb.format_exc()}</pre>"


# ══════════════════════════════════════════════════════════════════════════════
#  DATA BACKUP
# ══════════════════════════════════════════════════════════════════════════════

def _build_backup_excel():
    """Return a BytesIO Excel workbook with bookings + customers + inventory."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()

    # ── Bookings sheet ────────────────────────────────────────────────────────
    ws_b = wb.active
    ws_b.title = "Bookings"
    b_cols = ["id","full_name","email","phone","status","payment_status",
              "grand_total","amount_paid","event_start_date","event_end_date",
              "setup_date","setup_time","event_street","event_city","event_state",
              "event_zip","venue_type","delivery_location","delivery_status",
              "items_summary","admin_notes","created_at"]
    hdr_fill = PatternFill("solid", fgColor="1e3a5f")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for ci, col in enumerate(b_cols, 1):
        cell = ws_b.cell(row=1, column=ci, value=col.replace("_"," ").title())
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    conn = get_db()
    if conn:
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("""SELECT """ + ",".join(b_cols) + """ FROM bookings ORDER BY id DESC""")
            for ri, row in enumerate(cur.fetchall(), 2):
                for ci, col in enumerate(b_cols, 1):
                    val = row[col]
                    if hasattr(val, 'isoformat'):
                        val = str(val)
                    ws_b.cell(row=ri, column=ci, value=val)
            cur.close()
        except Exception as e:
            log.error(f"backup bookings: {e}")

    # ── Customers sheet ───────────────────────────────────────────────────────
    ws_c = wb.create_sheet("Customers")
    c_cols = ["id","full_name","email","phone","street","city","state","zip",
              "notes","booking_count","created_at"]
    for ci, col in enumerate(c_cols, 1):
        cell = ws_c.cell(row=1, column=ci, value=col.replace("_"," ").title())
        cell.fill = PatternFill("solid", fgColor="085041")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    if conn:
        try:
            cur2 = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur2.execute("SELECT " + ",".join(c_cols) + " FROM customers ORDER BY id DESC")
            for ri, row in enumerate(cur2.fetchall(), 2):
                for ci, col in enumerate(c_cols, 1):
                    val = row[col]
                    if hasattr(val, 'isoformat'): val = str(val)
                    ws_c.cell(row=ri, column=ci, value=val)
            cur2.close()
        except Exception as e:
            log.error(f"backup customers: {e}")

    # ── Inventory sheet ───────────────────────────────────────────────────────
    ws_i = wb.create_sheet("Inventory")
    i_cols = ["id","name","price","total","sort_order"]
    for ci, col in enumerate(i_cols, 1):
        cell = ws_i.cell(row=1, column=ci, value=col.replace("_"," ").title())
        cell.fill = PatternFill("solid", fgColor="26215C")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
    if conn:
        try:
            cur3 = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur3.execute("SELECT " + ",".join(i_cols) + " FROM inventory ORDER BY sort_order")
            for ri, row in enumerate(cur3.fetchall(), 2):
                for ci, col in enumerate(i_cols, 1):
                    ws_i.cell(row=ri, column=ci, value=row[col])
            cur3.close(); conn.close()
        except Exception as e:
            log.error(f"backup inventory: {e}")

    # Auto-size columns
    for ws in [ws_b, ws_c, ws_i]:
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _send_backup_email(buf, filename):
    """Email the backup Excel file to OWNER_EMAIL."""
    if not all([GMAIL_USER, GMAIL_APP_PASSWORD, OWNER_EMAIL]):
        log.warning("Backup email skipped — Gmail or OWNER_EMAIL not configured")
        return False
    try:
        msg = MIMEMultipart()
        msg["From"]    = f"{BUSINESS_NAME} <{GMAIL_USER}>"
        msg["To"]      = OWNER_EMAIL
        msg["Subject"] = f"📦 {BUSINESS_NAME} — Daily Backup {date.today().strftime('%B %d, %Y')}"
        body = (f"Automated daily backup for {BUSINESS_NAME}.\n\n"
                f"Attached: {filename}\n"
                f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}\n\n"
                "This email is sent automatically every day to protect your data.")
        msg.attach(MIMEText(body, "plain"))
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(buf.read())
        _email_encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            s.send_message(msg)
        log.info(f"Backup email sent to {OWNER_EMAIL}")
        return True
    except Exception as e:
        log.error(f"Backup email error: {e}")
        return False


@app.route("/admin/download-backup")
@admin_required
def admin_download_backup():
    """Download a full Excel backup right now."""
    try:
        buf = _build_backup_excel()
        filename = f"rentaparty_backup_{date.today().isoformat()}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        log.error(f"download_backup error: {e}")
        return f"Backup error: {e}", 500


@app.route("/admin/email-backup", methods=["POST"])
@admin_required
def admin_email_backup():
    """Manually trigger a backup email to OWNER_EMAIL."""
    try:
        filename = f"rentaparty_backup_{date.today().isoformat()}.xlsx"
        buf = _build_backup_excel()
        ok = _send_backup_email(buf, filename)
        if ok:
            return redirect(url_for("admin_dashboard") + "?backup_sent=1")
        else:
            return redirect(url_for("admin_dashboard") + "?backup_err=1")
    except Exception as e:
        log.error(f"email_backup error: {e}")
        return redirect(url_for("admin_dashboard") + "?backup_err=1")


@app.route("/cron/send-reminders")
def cron_send_reminders():
    """
    Called daily by an external cron (e.g. cron-job.org or Render cron).
    Sends three types of automated SMS:
      1. Payment reminder  — delivery in 5 days, still unpaid / not agreed
      2. Delivery confirmation — delivery is tomorrow, accepted
      3. Pickup reminder   — pickup is tomorrow
    Protect with ?secret=CRON_SECRET_KEY
    """
    secret = request.args.get("secret", "")
    if CRON_SECRET_KEY and secret != CRON_SECRET_KEY:
        return "Unauthorized", 401

    today = date.today()
    in_5  = today + timedelta(days=5)
    tmrw  = today + timedelta(days=1)

    sent   = []
    errors = []

    try:
        conn = get_db()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # ── 1. Payment reminders (delivery in 5 days, unpaid) ────────────────
        cur.execute("""
            SELECT id, full_name, phone,
                   COALESCE(delivery_date, setup_date, event_start_date) AS del_date,
                   COALESCE(delivery_time, setup_time) AS del_time
            FROM bookings
            WHERE COALESCE(delivery_date, setup_date, event_start_date) = %s
              AND status = 'accepted'
              AND (payment_status IS NULL OR payment_status NOT IN ('paid'))
              AND (archived IS NULL OR archived = FALSE)
              AND (reminder_sent_at IS NULL
                   OR reminder_sent_at::date < %s - INTERVAL '3 days')
        """, (in_5.isoformat(), today.isoformat()))
        for b in cur.fetchall():
            msg = (
                f"Hi {b['full_name'].split()[0]}! "
                f"This is {BUSINESS_NAME}. Your rental delivery is scheduled for "
                f"{b['del_date'].strftime('%A, %B %-d')}. "
                f"We haven't received your deposit yet — please pay to confirm your booking. "
                f"Questions? Call {BUSINESS_PHONE or 'us'}."
            )
            ok = send_sms(b["phone"], msg)
            if ok:
                cur.execute("UPDATE bookings SET reminder_sent_at = NOW() WHERE id = %s", (b["id"],))
                sent.append(f"payment-reminder → #{b['id']} {b['full_name']}")
            else:
                errors.append(f"payment-reminder FAIL #{b['id']}")

        # ── 2. Delivery confirmation (delivery tomorrow) ─────────────────────
        cur.execute("""
            SELECT id, full_name, phone,
                   COALESCE(delivery_date, setup_date, event_start_date) AS del_date,
                   COALESCE(delivery_time, setup_time) AS del_time,
                   event_start_date
            FROM bookings
            WHERE COALESCE(delivery_date, setup_date, event_start_date) = %s
              AND status = 'accepted'
              AND (archived IS NULL OR archived = FALSE)
              AND (confirmation_sent_at IS NULL
                   OR confirmation_sent_at::date < %s)
        """, (tmrw.isoformat(), today.isoformat()))
        for b in cur.fetchall():
            time_str = ""
            if b["del_time"]:
                try:
                    t = datetime.strptime(b["del_time"], "%H:%M")
                    time_str = f" around {t.strftime('%-I:%M %p')}"
                except Exception:
                    pass
            msg = (
                f"Hi {b['full_name'].split()[0]}! "
                f"Your {BUSINESS_NAME} delivery is tomorrow"
                f"{time_str}. "
                f"Please make sure someone is available at the setup location. "
                f"Questions? Call {BUSINESS_PHONE or 'us'}. See you tomorrow!"
            )
            ok = send_sms(b["phone"], msg)
            if ok:
                cur.execute("UPDATE bookings SET confirmation_sent_at = NOW() WHERE id = %s", (b["id"],))
                sent.append(f"delivery-confirm → #{b['id']} {b['full_name']}")
            else:
                errors.append(f"delivery-confirm FAIL #{b['id']}")

        # ── 3. Pickup reminder (pickup tomorrow) ─────────────────────────────
        cur.execute("""
            SELECT id, full_name, phone, pickup_date, pickup_time
            FROM bookings
            WHERE pickup_date = %s
              AND status = 'accepted'
              AND (archived IS NULL OR archived = FALSE)
              AND (pickup_reminder_sent_at IS NULL
                   OR pickup_reminder_sent_at::date < %s)
        """, (tmrw.isoformat(), today.isoformat()))
        for b in cur.fetchall():
            time_str = ""
            if b["pickup_time"]:
                try:
                    t = datetime.strptime(b["pickup_time"], "%H:%M")
                    time_str = f" around {t.strftime('%-I:%M %p')}"
                except Exception:
                    pass
            msg = (
                f"Hi {b['full_name'].split()[0]}! "
                f"A reminder that {BUSINESS_NAME} will be picking up your rental equipment tomorrow"
                f"{time_str}. "
                f"Please have everything accessible. Thank you for renting with us!"
            )
            ok = send_sms(b["phone"], msg)
            if ok:
                cur.execute("UPDATE bookings SET pickup_reminder_sent_at = NOW() WHERE id = %s", (b["id"],))
                sent.append(f"pickup-reminder → #{b['id']} {b['full_name']}")
            else:
                errors.append(f"pickup-reminder FAIL #{b['id']}")

        conn.commit()
        cur.close()
    except Exception as e:
        log.error(f"cron_send_reminders error: {e}")
        return jsonify({"error": str(e)}), 500

    return jsonify({
        "date": today.isoformat(),
        "sent": sent,
        "errors": errors,
        "total_sent": len(sent),
    })


@app.route("/cron/daily-backup")
def cron_daily_backup():
    """Called by external cron (cron-job.org) every day. Emails backup to owner."""
    secret = request.args.get("secret", "")
    if CRON_SECRET and secret != CRON_SECRET:
        return "Unauthorized", 401
    try:
        filename = f"rentaparty_backup_{date.today().isoformat()}.xlsx"
        buf = _build_backup_excel()
        ok = _send_backup_email(buf, filename)
        return jsonify({"ok": ok, "date": date.today().isoformat()})
    except Exception as e:
        log.error(f"cron_daily_backup error: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
